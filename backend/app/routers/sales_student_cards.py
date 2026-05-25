from io import BytesIO
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi import File, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app import auth
from app.database import get_db
from app.models import Abonement, DiscountType, Student, StudentCard, StudentStatus, User, UserRole
from app.routers.action_log import log_action
from app.schemas.abonements import AbonementResponse
from app.schemas.sales import (
    AnketaConvertRequest,
    AnketaConvertResponse,
    OpenParentCabinetResponse,
    StudentCardCreate,
    StudentCardImportResponse,
    StudentCardResponse,
    StudentCardUpdate,
)
from app.services.parent_invite import create_parent_with_invite
from app.services.person_sync import sync_student_card_person
from app.services.student_card_conversion import (
    StudentCardConvertConflict,
    convert_student_card_to_student as student_card_convert,
)
from app.student_display import get_students_display_names
from app.utils.phone import normalize_phone

router = APIRouter()


def _require_sales_admin_owner(user: User) -> None:
    auth.ensure_permission(user, "sales.access")


def _student_card_response(card: StudentCard, user: User, db: Session) -> StudentCardResponse:
    parent_cabinet_open = False
    if getattr(card, "student_id", None):
        student = db.query(Student).filter(Student.id == card.student_id).first()
        if student and student.parent_id:
            parent_cabinet_open = True
    data = {
        "id": card.id,
        "student_id": getattr(card, "student_id", None),
        "student_full_name": card.student_full_name,
        "parent_cabinet_open": parent_cabinet_open,
        "birth_date": card.birth_date,
        "student_phone": card.student_phone,
        "telegram": card.telegram,
        "gender": card.gender,
        "on_grant": card.on_grant,
        "format_type": card.format_type,
        "city": card.city,
        "school": card.school,
        "grade": card.grade,
        "parent_full_name": card.parent_full_name,
        "parent_phone": card.parent_phone,
        "parent_phone_2": card.parent_phone_2,
        "parent_telegram": getattr(card, "parent_telegram", None),
        "parent_email": getattr(card, "parent_email", None),
        "student_email": getattr(card, "student_email", None),
        "preferred_messenger": getattr(card, "preferred_messenger", None),
        "comment": getattr(card, "comment", None),
        "source": getattr(card, "source", None),
        "payment_link": getattr(card, "payment_link", None),
        "learning_period_start": getattr(card, "learning_period_start", None),
        "next_payment_date": getattr(card, "next_payment_date", None),
        "archived": card.archived,
        "anketa_status": getattr(card, "anketa_status", "converted"),
        "primary_for_bank_payments": getattr(card, "primary_for_bank_payments", False),
        "created_at": card.created_at,
        "updated_at": card.updated_at,
    }
    effective_role = auth.resolve_effective_role(user)
    if effective_role == UserRole.OWNER:
        data["abonement_id"] = card.abonement_id
        data["discount_type"] = card.discount_type
        data["discount_value"] = card.discount_value
        data["abonement"] = AbonementResponse.model_validate(card.abonement) if card.abonement else None
    else:
        data["abonement_id"] = None
        data["discount_type"] = DiscountType.NONE
        data["discount_value"] = 0.0
        data["abonement"] = None
    return StudentCardResponse(**data)


@router.get("/student-cards/{card_id}", response_model=StudentCardResponse)
async def get_student_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")
    return _student_card_response(card, current_user, db)


@router.get("/student-cards", response_model=List[StudentCardResponse])
async def list_student_cards(
    archived: Optional[bool] = None,
    anketa_status: Optional[List[str]] = None,
    student_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    query = db.query(StudentCard)
    if archived is not None:
        query = query.filter(StudentCard.archived == archived)
    if anketa_status:
        query = query.filter(StudentCard.anketa_status.in_(anketa_status))
    if student_id is not None:
        query = query.filter(StudentCard.student_id == student_id)
    items = query.order_by(StudentCard.created_at.desc()).all()
    return [_student_card_response(item, current_user, db) for item in items]


@router.post("/student-cards", response_model=StudentCardResponse, status_code=status.HTTP_201_CREATED)
async def create_student_card(
    payload: StudentCardCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    name = (payload.student_full_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="ФИО ученика обязательно")
    data = payload.model_dump()
    if data.get("student_id") is None and not data.get("anketa_status"):
        data["anketa_status"] = "draft"
    data["phone_normalized"] = normalize_phone(data.get("parent_phone") or data.get("student_phone") or "") or None
    if auth.resolve_effective_role(current_user) != UserRole.OWNER:
        data["abonement_id"] = None
        data["discount_type"] = DiscountType.NONE
        data["discount_value"] = 0.0
    if auth.resolve_effective_role(current_user) not in (UserRole.OWNER, UserRole.ADMIN):
        data["payment_link"] = None
    if data.get("abonement_id"):
        abonement = db.query(Abonement).filter(Abonement.id == data["abonement_id"]).first()
        if not abonement:
            raise HTTPException(status_code=404, detail="Абонемент не найден")
    if data.get("student_id") is not None:
        student = db.query(Student).filter(Student.id == data["student_id"]).first()
        if not student:
            raise HTTPException(status_code=404, detail="Ученик не найден")
    card = StudentCard(**data)
    db.add(card)
    db.flush()
    sync_student_card_person(db, card)
    db.commit()
    db.refresh(card)
    return _student_card_response(card, current_user, db)


@router.get("/student-cards/import-template")
async def download_student_cards_import_template(
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Карточки учеников"
    headers = [
        "ФИО ученика",
        "Дата рождения",
        "Телефон ученика",
        "Телеграм ученика",
        "Пол",
        "На гранте",
        "Формат",
        "Город",
        "Образовательное учреждение",
        "Класс",
        "Email ученика",
        "ФИО родителя",
        "Телефон родителя",
        "Второй телефон родителя",
        "Телеграм родителя",
        "Email родителя",
        "Удобный мессенджер",
        "Комментарий",
        "Откуда пришел",
    ]
    worksheet.append(headers)
    worksheet.append(
        [
            "Иванов Петр Сергеевич",
            "2015-03-15",
            "+7 999 111-22-33",
            "@petr_ivanov",
            "м",
            "нет",
            "группа",
            "Москва",
            "Школа №12",
            "3",
            "petr@example.com",
            "Иванова Анна Петровна",
            "+7 999 111-22-34",
            "+7 900 111-22-44",
            "@anna_ivanova",
            "anna@example.com",
            "Telegram",
            "Записан на пробное занятие",
            "рекомендация",
        ]
    )
    worksheet.freeze_panes = "A2"
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="student_cards_import_template.xlsx"'},
    )


@router.post("/student-cards/import-xlsx", response_model=StudentCardImportResponse)
async def import_student_cards_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Поддерживается только формат .xlsx")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")
    workbook = load_workbook(filename=BytesIO(data), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return StudentCardImportResponse(created=0, skipped=0, errors=["Пустой лист"])
    headers = [str(header).strip().lower() if header is not None else "" for header in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    def cell_value(row, variants: List[str]) -> Optional[str]:
        for key in variants:
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                continue
            raw = row[idx]
            if raw is None:
                continue
            text = str(raw).strip()
            if text:
                return text
        return None

    def parse_date_value(raw_value):
        if raw_value is None:
            return None
        if hasattr(raw_value, "year") and hasattr(raw_value, "month") and hasattr(raw_value, "day"):
            try:
                return raw_value.date() if hasattr(raw_value, "date") else raw_value
            except Exception:
                return raw_value
        text = str(raw_value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                from datetime import datetime

                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            from datetime import datetime

            return datetime.fromisoformat(text[:10]).date()
        except (ValueError, TypeError):
            return None

    created = 0
    skipped = 0
    errors: List[str] = []
    for row in rows[1:]:
        row = list(row) if row else []
        student_full_name = cell_value(row, ["фио ученика", "ученик", "student_full_name"])
        if not student_full_name:
            skipped += 1
            continue
        birth_date_raw = cell_value(row, ["дата рождения", "birth_date"])
        birth_date = parse_date_value(birth_date_raw) if birth_date_raw else None
        student_phone = cell_value(row, ["телефон ученика", "student_phone"])
        telegram = cell_value(row, ["телеграм ученика", "telegram"])
        gender_raw = cell_value(row, ["пол", "gender"])
        gender = (
            gender_raw.lower()
            if gender_raw and gender_raw.lower() in ("м", "ж", "m", "f", "male", "female")
            else (gender_raw or None)
        )
        on_grant_raw = cell_value(row, ["на гранте", "on_grant"])
        on_grant = str(on_grant_raw).strip().lower() in ("да", "yes", "1", "true", "+")
        format_raw = cell_value(row, ["формат", "format_type"])
        format_type = None
        if format_raw:
            lowered = format_raw.lower()
            if "групп" in lowered or lowered == "group":
                format_type = "group"
            elif "индивид" in lowered or lowered == "individual":
                format_type = "individual"
            else:
                format_type = format_raw
        city = cell_value(row, ["город", "city"])
        school = cell_value(row, ["образовательное учреждение", "школа", "school"])
        grade = cell_value(row, ["класс", "grade"])
        student_email = cell_value(row, ["email ученика", "student_email"])
        parent_full_name = cell_value(row, ["фио родителя", "родитель", "parent_full_name"])
        parent_phone = cell_value(row, ["телефон родителя", "parent_phone"])
        parent_phone_2 = cell_value(row, ["второй телефон родителя", "parent_phone_2"])
        parent_telegram = cell_value(row, ["телеграм родителя", "parent_telegram"])
        parent_email = cell_value(row, ["email родителя", "parent_email"])
        preferred_raw = cell_value(row, ["удобный мессенджер", "preferred_messenger"])
        preferred_messenger = None
        if preferred_raw:
            lowered = preferred_raw.lower()
            if "max" in lowered or lowered == "max":
                preferred_messenger = "max"
            elif "telegram" in lowered or "телеграм" in lowered or lowered == "tg":
                preferred_messenger = "telegram"
            elif "sms" in lowered:
                preferred_messenger = "sms"
            else:
                preferred_messenger = preferred_raw
        comment = cell_value(row, ["комментарий", "comment"])
        source = cell_value(row, ["откуда пришел", "источник", "source"])
        card = StudentCard(
            student_full_name=student_full_name,
            birth_date=birth_date,
            student_phone=student_phone or None,
            phone_normalized=normalize_phone(parent_phone or student_phone or "") or None,
            telegram=telegram or None,
            gender=gender,
            on_grant=on_grant,
            format_type=format_type,
            city=city or None,
            school=school or None,
            grade=grade or None,
            parent_full_name=parent_full_name or None,
            parent_phone=parent_phone or None,
            parent_phone_2=parent_phone_2 or None,
            parent_telegram=parent_telegram or None,
            parent_email=parent_email or None,
            student_email=student_email or None,
            preferred_messenger=preferred_messenger,
            comment=comment or None,
            source=source or None,
            discount_type=DiscountType.NONE,
            discount_value=0.0,
            archived=False,
        )
        db.add(card)
        db.flush()
        sync_student_card_person(db, card)
        created += 1
    db.commit()
    log_action(db, current_user.id, "import", "student_card", None, {"created": created, "skipped": skipped})
    return StudentCardImportResponse(created=created, skipped=skipped, errors=errors)


@router.put("/student-cards/{card_id}", response_model=StudentCardResponse)
async def update_student_card(
    card_id: int,
    payload: StudentCardUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")
    data = payload.model_dump(exclude_unset=True)
    if auth.resolve_effective_role(current_user) != UserRole.OWNER:
        data.pop("abonement_id", None)
        data.pop("discount_type", None)
        data.pop("discount_value", None)
    if auth.resolve_effective_role(current_user) not in (UserRole.OWNER, UserRole.ADMIN):
        data.pop("payment_link", None)
    if data.get("abonement_id"):
        abonement = db.query(Abonement).filter(Abonement.id == data["abonement_id"]).first()
        if not abonement:
            raise HTTPException(status_code=404, detail="Абонемент не найден")
    if "student_id" in data and data["student_id"] is not None:
        student = db.query(Student).filter(Student.id == data["student_id"]).first()
        if not student:
            raise HTTPException(status_code=404, detail="Ученик не найден")
    for key, value in data.items():
        setattr(card, key, value)
    sync_student_card_person(db, card)
    db.commit()
    db.refresh(card)
    return _student_card_response(card, current_user, db)


@router.post("/student-cards/{card_id}/convert", response_model=AnketaConvertResponse)
async def convert_anketa_to_student(
    card_id: int,
    payload: Optional[AnketaConvertRequest] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    body = payload or AnketaConvertRequest()
    try:
        result = student_card_convert(
            db,
            card_id,
            use_existing_parent_id=body.use_existing_parent_id,
            use_existing_student_id=body.use_existing_student_id,
        )
    except StudentCardConvertConflict as exc:
        code = exc.detail.get("code", "existing_parent")
        raise HTTPException(status_code=409, detail=exc.detail, headers={"X-Conflict-Code": code})
    except ValueError as exc:
        message = str(exc)
        if "не найден" in message.lower():
            raise HTTPException(status_code=404, detail=message)
        raise HTTPException(status_code=400, detail=message)
    return AnketaConvertResponse(card=_student_card_response(result.card, current_user, db), student_id=result.student_id)


@router.post("/student-cards/{card_id}/archive")
async def archive_student_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")
    card.archived = True
    db.commit()
    return {"archived": True}


@router.post("/student-cards/{card_id}/unarchive")
async def unarchive_student_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")
    card.archived = False
    db.commit()
    return {"archived": False}


@router.post("/student-cards/{card_id}/open-parent-cabinet", response_model=OpenParentCabinetResponse)
async def open_parent_cabinet_from_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Карточка не найдена")
    parent_email = (getattr(card, "parent_email", None) or "").strip().lower()
    if not parent_email:
        raise HTTPException(status_code=400, detail="Укажите email родителя в карточке, чтобы открыть кабинет")
    parent_full_name = (getattr(card, "parent_full_name", None) or "").strip() or "Родитель"

    if not getattr(card, "student_id", None):
        student = Student(full_name=(card.student_full_name or "").strip() or "Ученик", status=StudentStatus.ACTIVE)
        db.add(student)
        db.flush()
        card.student_id = student.id
        db.add(card)
        db.flush()
    else:
        student = db.query(Student).filter(Student.id == card.student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Ученик по карточке не найден")

    if student.parent_id:
        db.commit()
        return OpenParentCabinetResponse(already_open=True, student_id=student.id, parent_id=student.parent_id)

    parent_user = db.query(User).filter(User.email == parent_email, User.role == UserRole.PARENT).first()
    if parent_user:
        student.parent_id = parent_user.id
        db.add(student)
        db.commit()
        return OpenParentCabinetResponse(already_open=False, student_id=student.id, parent_id=parent_user.id)

    try:
        parent_user, invite_link = create_parent_with_invite(db, parent_email, parent_full_name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    student.parent_id = parent_user.id
    db.add(student)
    db.commit()
    log_action(
        db,
        current_user.id,
        "open_parent_cabinet",
        "student_card",
        card_id,
        {"student_id": student.id, "parent_id": parent_user.id},
    )
    return OpenParentCabinetResponse(
        already_open=False,
        student_id=student.id,
        parent_id=parent_user.id,
        invite_link=invite_link,
    )


@router.get("/students-for-cards", response_model=List[Dict])
async def list_students_for_cards(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    students = db.query(Student).filter(Student.status == StudentStatus.ACTIVE).order_by(Student.full_name).all()
    if not students:
        return []
    display_names = get_students_display_names(db, [student.id for student in students])
    return [{"id": student.id, "full_name": display_names.get(student.id, student.full_name)} for student in students]
