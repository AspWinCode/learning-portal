import csv
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import Text, cast, or_
from sqlalchemy.orm import Session, joinedload

from app import auth
from app.database import get_db
from app.models import (
    AccountTemplate,
    Event,
    EventRegistration,
    EventRegistrationStatus,
    Lead,
    LeadInfoTemplate,
    LeadSource,
    LeadStatus,
    LeadStatusOption,
    LeadTask,
    LeadTaskStatus,
    LeadTaskStatusOption as LeadTaskStatusOptionModel,
    LeadTaskTemplate,
    B2BSchool,
    OwnerWorkspaceContact,
    SalesCity,
    SalesClass,
    SalesSchool,
    SalesSchoolContact,
    User,
    UserRole,
)
from app.routers.action_log import log_action
from app.schemas.sales import (
    AccountTemplateCreate,
    AccountTemplateResponse,
    FollowUpItemResponse,
    LeadInfoTemplateCreate,
    LeadInfoTemplateResponse,
    LeadInfoTemplateUpdate,
    LeadPushStatsResponse,
    LeadSourceCreate,
    LeadSourceResponse,
    LeadSourceUpdate,
    LeadStatusOptionCreate,
    LeadStatusOptionResponse,
    LeadStatusOptionUpdate,
    LeadTaskStatusOptionCreate,
    LeadTaskStatusOptionResponse,
    LeadTaskStatusOptionUpdate,
    LeadTaskTemplateCreate,
    LeadTaskTemplateResponse,
    LeadTaskTemplateUpdate,
    SalesCityCreate,
    SalesCityResponse,
    SalesCityUpdate,
    SalesClassCreate,
    SalesClassResponse,
    SalesClassUpdate,
    SalesDashboardResponse,
    SalesQueueRegistrationItem,
    SalesQueueTaskItem,
    SalesSchoolConversionItem,
    SalesSchoolContactCreate,
    SalesSchoolContactResponse,
    SalesSchoolContactUpdate,
    SalesSchoolCreate,
    SalesSchoolResponse,
    SalesSchoolUpdate,
)
from app.utils.datetime import utcnow

router = APIRouter()


def _filter_query_by_role(query, user: User):
    if auth.resolve_effective_role(user) in (UserRole.ADMIN, UserRole.OWNER, UserRole.SALES):
        return query
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")


def _normalize_source_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    return " ".join(name.strip().split())


def _clean_optional_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def _sync_sales_school_to_b2b(db: Session, school: SalesSchool) -> B2BSchool:
    b2b = db.query(B2BSchool).filter(cast(B2BSchool.name, Text).ilike(school.name)).first()
    if not b2b:
        b2b = B2BSchool(name=school.name, pipeline_stage="new")
        db.add(b2b)
    b2b.name = school.name
    b2b.city = getattr(school, "city", None)
    b2b.director = getattr(school, "director", None)
    b2b.email = getattr(school, "email", None)
    b2b.website = getattr(school, "website", None)
    b2b.address = getattr(school, "address", None)
    b2b.phone_school = getattr(school, "phone", None)
    return b2b


def _school_payload_from_row(row: Dict[str, object]) -> Dict[str, Optional[str]]:
    normalized = {str(k).strip().lower(): v for k, v in row.items()}

    def pick(*keys: str) -> Optional[str]:
        for key in keys:
            if key in normalized and normalized[key] is not None:
                text = str(normalized[key]).strip()
                if text:
                    return text
        return None

    return {
        "name": pick("название", "школа", "school", "name"),
        "city": pick("город", "city"),
        "district": pick("район", "district"),
        "director": pick("директор/ио директора", "директор/и.о. директора", "директор", "ио директора", "и.о. директора", "director"),
        "email": pick("почта", "email", "e-mail"),
        "address": pick("адрес", "address"),
        "phone": pick("телефон", "phone"),
        "website": pick("сайт", "website", "site"),
    }


def _parse_school_import(data: bytes, filename: str) -> List[Dict[str, Optional[str]]]:
    lowered = filename.lower().strip()
    if lowered.endswith(".xlsx"):
        wb = load_workbook(filename=BytesIO(data), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # Detect CSV-saved-as-XLSX: all data in column A with delimiter inside cells
        non_none = [v for v in rows[0] if v is not None]
        if len(non_none) == 1:
            first_cell = str(non_none[0])
            tab_c, semi_c, comma_c = first_cell.count("\t"), first_cell.count(";"), first_cell.count(",")
            max_c = max(tab_c, semi_c, comma_c)
            if max_c > 0:
                delim = "\t" if tab_c == max_c else (";" if semi_c == max_c else ",")
                all_text = "\n".join(str(r[0] or "") for r in rows if r and r[0] is not None)
                reader = csv.DictReader(StringIO(all_text), delimiter=delim)
                parsed = [_school_payload_from_row(row) for row in reader]
                if parsed:
                    parsed[0]["_raw_headers"] = first_cell[:200]
                    parsed[0]["_delim"] = repr(delim)
                    parsed[0]["_path"] = f"xlsx-1col-t{tab_c}s{semi_c}c{comma_c}"
                return parsed
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result: List[Dict[str, Optional[str]]] = []
        for values in rows[1:]:
            raw = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            result.append(_school_payload_from_row(raw))
        if result:
            result[0]["_raw_headers"] = str(headers[:8])
            result[0]["_delim"] = "xlsx-normal"
            result[0]["_path"] = f"xlsx-{len(headers)}cols"
        return result

    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("cp1251")
    sample = text[:2048]
    tab_c = sample.count("\t")
    semi_c = sample.count(";")
    comma_c = sample.count(",")
    max_c = max(tab_c, semi_c, comma_c)
    delimiter = "\t" if tab_c == max_c else (";" if semi_c == max_c else ",")
    # Strip outer quotes from each line (Excel wraps every row in one pair of outer quotes
    # when saving tab-delimited text, so csv.DictReader would treat each row as one field)
    lines = text.splitlines(True)
    fixed_lines = []
    for line in lines:
        stripped = line.rstrip("\r\n")
        ending = line[len(stripped):]
        if stripped.startswith('"') and stripped.endswith('"'):
            # Only strip if the content itself doesn't look like a normal quoted CSV field
            # (i.e., the inner content contains the delimiter)
            inner = stripped[1:-1]
            if delimiter in inner:
                stripped = inner
        fixed_lines.append(stripped + ending)
    text = "".join(fixed_lines)
    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    return [_school_payload_from_row(row) for row in reader]


@router.get("/dashboard", response_model=SalesDashboardResponse)
async def get_sales_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    now = utcnow()
    start_month = datetime(now.year, now.month, 1)
    end_month = datetime(now.year + (1 if now.month == 12 else 0), 1 if now.month == 12 else now.month + 1, 1)
    start_today = datetime(now.year, now.month, now.day)
    end_today = start_today + timedelta(days=1)
    next_two_hours = now + timedelta(hours=2)
    next_day = now + timedelta(hours=24)
    week_ago = now - timedelta(days=7)

    leads_q = _filter_query_by_role(db.query(Lead), current_user)
    tasks_q = db.query(LeadTask).join(Lead, Lead.id == LeadTask.lead_id)
    regs_q = db.query(EventRegistration).join(Lead, Lead.id == EventRegistration.lead_id)

    active_lead_statuses = [LeadStatus.NEW, LeadStatus.CONTACTED, LeadStatus.DEMO, LeadStatus.INVOICE_SENT]
    connected_statuses = [LeadStatus.CONTACTED, LeadStatus.DEMO, LeadStatus.INVOICE_SENT, LeadStatus.WON]

    kpi_new_leads = leads_q.filter(Lead.created_at >= start_today, Lead.created_at < end_today).count()
    leads_week_q = leads_q.filter(Lead.created_at >= week_ago)
    leads_week_total = leads_week_q.count()
    leads_week_connected = leads_week_q.filter(Lead.status.in_(connected_statuses)).count()
    kpi_dozvon_percent = round((leads_week_connected / leads_week_total) * 100, 1) if leads_week_total else 0.0
    kpi_info_sent = leads_q.filter(Lead.status.in_([LeadStatus.INVOICE_SENT, LeadStatus.WON])).count()

    kpi_need_push_urgent = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at >= now,
        Lead.next_contact_at < next_two_hours,
    ).count()
    kpi_need_push_today = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at >= start_today,
        Lead.next_contact_at < end_today,
    ).count()
    kpi_need_push_overdue = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at < now,
    ).count()

    kpi_registered_event = regs_q.filter(EventRegistration.status == EventRegistrationStatus.REGISTERED).count()
    kpi_came_count = regs_q.filter(cast(EventRegistration.note, Text).ilike("%[came]%")).count()
    kpi_no_show_count = regs_q.filter(
        or_(
            cast(EventRegistration.note, Text).ilike("%[no-show]%"),
            cast(EventRegistration.note, Text).ilike("%no-show%"),
        )
    ).count()

    overdue_items = (
        tasks_q.filter(LeadTask.status == LeadTaskStatus.OPEN, LeadTask.due_at.isnot(None), LeadTask.due_at < now)
        .order_by(LeadTask.due_at.asc())
        .limit(30)
        .all()
    )
    call_today_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at.isnot(None),
            LeadTask.due_at >= start_today,
            LeadTask.due_at < end_today,
        )
        .order_by(LeadTask.due_at.asc())
        .limit(30)
        .all()
    )
    messenger_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            or_(
                cast(LeadTask.channel, Text).ilike("%messenger%"),
                cast(LeadTask.channel, Text).ilike("%telegram%"),
                cast(LeadTask.note, Text).ilike("%ответ%"),
            ),
        )
        .order_by(LeadTask.created_at.desc())
        .limit(30)
        .all()
    )
    confirm_items = (
        regs_q.join(Event, Event.id == EventRegistration.event_id)
        .filter(
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
            Event.starts_at >= now,
            Event.starts_at < next_day,
        )
        .order_by(Event.starts_at.asc())
        .limit(30)
        .all()
    )

    month_outreach_leads = leads_q.filter(Lead.outreach_at.isnot(None), Lead.outreach_at >= start_month, Lead.outreach_at < end_month).all()
    schools_seen = set()
    classes_seen = set()
    outreach_minutes_month = 0
    school_stats: Dict[str, Dict[str, object]] = {}
    for lead in month_outreach_leads:
        school = (lead.school_name or "").strip()
        class_name = (lead.school_class or "").strip()
        if school:
            schools_seen.add(school)
        if school and class_name:
            classes_seen.add(f"{school}::{class_name}")
        if lead.outreach_minutes and lead.outreach_minutes > 0:
            outreach_minutes_month += lead.outreach_minutes
        if not school:
            continue
        if school not in school_stats:
            school_stats[school] = {"leads": 0, "won": 0, "classes": set(), "minutes": 0}
        school_stats[school]["leads"] = int(school_stats[school]["leads"]) + 1
        if lead.status == LeadStatus.WON:
            school_stats[school]["won"] = int(school_stats[school]["won"]) + 1
        if class_name:
            school_stats[school]["classes"].add(class_name)
        if lead.outreach_minutes and lead.outreach_minutes > 0:
            school_stats[school]["minutes"] = int(school_stats[school]["minutes"]) + lead.outreach_minutes

    top_schools_conversion_month = sorted(
        [
            SalesSchoolConversionItem(
                school_name=school,
                leads_count=int(stats["leads"]),
                won_count=int(stats["won"]),
                conversion_percent=round((int(stats["won"]) / int(stats["leads"])) * 100) if int(stats["leads"]) else 0,
                classes_count=len(stats["classes"]),
                outreach_minutes_total=int(stats["minutes"]),
            )
            for school, stats in school_stats.items()
        ],
        key=lambda item: (item.conversion_percent, item.leads_count),
        reverse=True,
    )[:3]

    def map_task(item: LeadTask) -> SalesQueueTaskItem:
        lead = item.lead
        return SalesQueueTaskItem(task_id=item.id, lead_id=lead.id, lead_name=lead.contact_name, lead_phone=lead.phone, due_at=item.due_at, note=item.note)

    def map_reg(item: EventRegistration) -> SalesQueueRegistrationItem:
        event = db.query(Event).filter(Event.id == item.event_id).first()
        lead = db.query(Lead).filter(Lead.id == item.lead_id).first()
        return SalesQueueRegistrationItem(
            registration_id=item.id,
            event_id=item.event_id,
            event_title=event.title if event else f"Event #{item.event_id}",
            lead_id=item.lead_id,
            lead_name=lead.contact_name if lead else f"Lead #{item.lead_id}",
            starts_at=event.starts_at if event else now,
            note=item.note,
        )

    return SalesDashboardResponse(
        kpi_new_leads=kpi_new_leads,
        kpi_dozvon_percent=kpi_dozvon_percent,
        kpi_info_sent=kpi_info_sent,
        kpi_need_push_urgent=kpi_need_push_urgent,
        kpi_need_push_today=kpi_need_push_today,
        kpi_need_push_overdue=kpi_need_push_overdue,
        kpi_registered_event=kpi_registered_event,
        kpi_came_count=kpi_came_count,
        kpi_no_show_count=kpi_no_show_count,
        overdue_followups=[map_task(item) for item in overdue_items],
        call_today=[map_task(item) for item in call_today_items],
        messenger_replies=[map_task(item) for item in messenger_items],
        confirm_participation=[map_reg(item) for item in confirm_items],
        outreach_schools_month=len(schools_seen),
        outreach_classes_month=len(classes_seen),
        outreach_minutes_month=outreach_minutes_month,
        top_schools_conversion_month=top_schools_conversion_month,
    )


@router.get("/follow-ups", response_model=List[FollowUpItemResponse])
async def list_follow_ups(
    period: str = Query(default="today"),
    source: Optional[str] = None,
    event_id: Optional[int] = None,
    reason: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    now = utcnow()
    start_today = datetime(now.year, now.month, now.day)
    end_today = start_today + timedelta(days=1)
    start_tomorrow = end_today
    end_tomorrow = start_tomorrow + timedelta(days=1)
    end_week = end_today + timedelta(days=7)

    query = db.query(LeadTask, Lead).join(Lead, Lead.id == LeadTask.lead_id).filter(LeadTask.status == LeadTaskStatus.OPEN)
    if period == "overdue":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at < now)
    elif period == "tomorrow":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_tomorrow, LeadTask.due_at < end_tomorrow)
    elif period == "week":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_today, LeadTask.due_at < end_week)
    else:
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_today, LeadTask.due_at < end_today)
    if source:
        query = query.filter(Lead.source.isnot(None), cast(Lead.source, Text).ilike(f"%{source.strip()}%"))
    if reason:
        query = query.filter(LeadTask.note.isnot(None), cast(LeadTask.note, Text).ilike(f"%{reason.strip()}%"))
    if event_id:
        query = query.join(EventRegistration, EventRegistration.lead_id == Lead.id).filter(
            EventRegistration.event_id == event_id,
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
        )
    rows = query.order_by(LeadTask.due_at.asc()).limit(200).all()
    return [
        FollowUpItemResponse(
            task_id=task.id,
            lead_id=lead.id,
            lead_name=lead.contact_name,
            lead_phone=lead.phone,
            lead_source=lead.source,
            due_at=task.due_at,
            status=task.status,
            channel=task.channel,
            note=task.note,
        )
        for task, lead in rows
    ]


@router.get("/leads/push-stats", response_model=List[LeadPushStatsResponse])
async def get_leads_push_stats(
    lead_ids: List[int] = Query(default=[]),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    lead_query = _filter_query_by_role(db.query(Lead), current_user)
    if lead_ids:
        lead_query = lead_query.filter(Lead.id.in_(lead_ids))
    leads = lead_query.all()
    if not leads:
        return []
    allowed_ids = [lead.id for lead in leads]
    tasks = db.query(LeadTask).options(joinedload(LeadTask.template)).filter(LeadTask.lead_id.in_(allowed_ids)).all()
    by_lead: Dict[int, Dict[str, int]] = {lead_id: {"total": 0, "done": 0} for lead_id in allowed_ids}
    for task in tasks:
        template_name = (task.template.name if task.template else "").lower()
        note = (task.note or "").lower()
        is_push = "дожим" in template_name or "дожим" in note or "push" in note
        if not is_push:
            continue
        by_lead[task.lead_id]["total"] += 1
        if task.status == LeadTaskStatus.DONE:
            by_lead[task.lead_id]["done"] += 1

    return [
        LeadPushStatsResponse(
            lead_id=lead_id,
            total_steps=by_lead[lead_id]["total"],
            done_steps=by_lead[lead_id]["done"],
            progress_percent=round((by_lead[lead_id]["done"] / by_lead[lead_id]["total"]) * 100) if by_lead[lead_id]["total"] else 0,
        )
        for lead_id in allowed_ids
    ]


@router.get("/lead-sources", response_model=List[LeadSourceResponse])
async def list_lead_sources(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(LeadSource).order_by(LeadSource.name.asc())
    if active_only:
        query = query.filter(LeadSource.is_active.is_(True))
    return query.all()


@router.post("/lead-sources", response_model=LeadSourceResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_source(
    payload: LeadSourceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Source name is required")
    exists = db.query(LeadSource).filter(cast(LeadSource.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Source already exists")
    source = LeadSource(name=name, is_active=True)
    db.add(source)
    db.commit()
    db.refresh(source)
    log_action(db, current_user.id, "create", "lead_source", source.id, {"name": name})
    return source


@router.put("/lead-sources/{source_id}", response_model=LeadSourceResponse)
async def update_lead_source(
    source_id: int,
    payload: LeadSourceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    source = db.query(LeadSource).filter(LeadSource.id == source_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Source name is required")
        source.name = name
    if "is_active" in data:
        source.is_active = data["is_active"]
    db.commit()
    db.refresh(source)
    log_action(db, current_user.id, "update", "lead_source", source.id, data)
    return source


@router.get("/lead-task-templates", response_model=List[LeadTaskTemplateResponse])
async def list_lead_task_templates(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(LeadTaskTemplate).order_by(LeadTaskTemplate.name.asc())
    if active_only:
        query = query.filter(LeadTaskTemplate.is_active.is_(True))
    return query.all()


@router.post("/lead-task-templates", response_model=LeadTaskTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task_template(
    payload: LeadTaskTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Task template name is required")
    exists = db.query(LeadTaskTemplate).filter(cast(LeadTaskTemplate.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Task template already exists")
    item = LeadTaskTemplate(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_task_template", item.id, {"name": name})
    return item


@router.put("/lead-task-templates/{template_id}", response_model=LeadTaskTemplateResponse)
async def update_lead_task_template(
    template_id: int,
    payload: LeadTaskTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadTaskTemplate).filter(LeadTaskTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Task template not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Task template name is required")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_task_template", item.id, data)
    return item


@router.get("/lead-task-statuses", response_model=List[LeadTaskStatusOptionResponse])
async def list_lead_task_statuses(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(LeadTaskStatusOptionModel).order_by(LeadTaskStatusOptionModel.id.asc())
    if active_only:
        query = query.filter(LeadTaskStatusOptionModel.is_active.is_(True))
    return query.all()


@router.post("/lead-task-statuses", response_model=LeadTaskStatusOptionResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task_status(
    payload: LeadTaskStatusOptionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Task status name is required")
    exists = db.query(LeadTaskStatusOptionModel).filter(cast(LeadTaskStatusOptionModel.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Task status already exists")
    item = LeadTaskStatusOptionModel(name=name, is_active=True, is_closed=payload.is_closed)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_task_status", item.id, {"name": name, "is_closed": payload.is_closed})
    return item


@router.put("/lead-task-statuses/{status_id}", response_model=LeadTaskStatusOptionResponse)
async def update_lead_task_status(
    status_id: int,
    payload: LeadTaskStatusOptionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadTaskStatusOptionModel).filter(LeadTaskStatusOptionModel.id == status_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Task status not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Task status name is required")
        item.name = name
    if "is_closed" in data:
        item.is_closed = data["is_closed"]
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_task_status", item.id, data)
    return item


@router.get("/lead-info-templates", response_model=List[LeadInfoTemplateResponse])
async def list_lead_info_templates(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(LeadInfoTemplate).order_by(LeadInfoTemplate.name.asc())
    if active_only:
        query = query.filter(LeadInfoTemplate.is_active.is_(True))
    return query.all()


@router.post("/lead-info-templates", response_model=LeadInfoTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_info_template(
    payload: LeadInfoTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Template name is required")
    if not (payload.body or "").strip():
        raise HTTPException(status_code=400, detail="Template body is required")
    exists = db.query(LeadInfoTemplate).filter(cast(LeadInfoTemplate.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Template already exists")
    item = LeadInfoTemplate(name=name, body=payload.body.strip(), is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_info_template", item.id, {"name": name})
    return item


@router.put("/lead-info-templates/{template_id}", response_model=LeadInfoTemplateResponse)
async def update_lead_info_template(
    template_id: int,
    payload: LeadInfoTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadInfoTemplate).filter(LeadInfoTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Template not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Template name is required")
        item.name = name
    if "body" in data:
        body = (data["body"] or "").strip()
        if not body:
            raise HTTPException(status_code=400, detail="Template body is required")
        item.body = body
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_info_template", item.id, data)
    return item


@router.get("/cities", response_model=List[SalesCityResponse])
async def list_sales_cities(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(SalesCity).order_by(SalesCity.name.asc())
    if active_only:
        query = query.filter(SalesCity.is_active.is_(True))
    return query.all()


@router.post("/cities", response_model=SalesCityResponse, status_code=status.HTTP_201_CREATED)
async def create_sales_city(
    payload: SalesCityCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="City name is required")
    exists = db.query(SalesCity).filter(cast(SalesCity.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="City already exists")
    item = SalesCity(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "sales_city", item.id, {"name": name})
    return item


@router.put("/cities/{city_id}", response_model=SalesCityResponse)
async def update_sales_city(
    city_id: int,
    payload: SalesCityUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesCity).filter(SalesCity.id == city_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="City not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="City name is required")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "sales_city", item.id, data)
    return item


@router.get("/schools/names", response_model=List[str])
async def list_sales_school_names(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(SalesSchool.name).order_by(SalesSchool.name.asc())
    if active_only:
        query = query.filter(SalesSchool.is_active.is_(True))
    return [row[0] for row in query.all()]


@router.get("/schools", response_model=List[SalesSchoolResponse])
async def list_sales_schools(
    active_only: bool = True,
    search: Optional[str] = Query(default=None),
    limit: int = Query(default=100, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(SalesSchool).options(joinedload(SalesSchool.contacts)).order_by(SalesSchool.name.asc())
    if active_only:
        query = query.filter(SalesSchool.is_active.is_(True))
    if search and search.strip():
        term = f"%{search.strip()}%"
        query = query.filter(
            cast(SalesSchool.name, Text).ilike(term)
            | cast(SalesSchool.city, Text).ilike(term)
            | cast(SalesSchool.director, Text).ilike(term)
        )
    return query.offset(offset).limit(limit).all()


@router.get("/schools/export")
async def export_sales_schools(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Название", "Город", "Район", "Директор/ИО директора", "Почта", "Адрес", "Телефон", "Сайт", "Активна"])
    for school in db.query(SalesSchool).order_by(SalesSchool.name.asc()).all():
        writer.writerow([
            school.name,
            getattr(school, "city", None) or "",
            getattr(school, "district", None) or "",
            getattr(school, "director", None) or "",
            getattr(school, "email", None) or "",
            getattr(school, "address", None) or "",
            getattr(school, "phone", None) or "",
            getattr(school, "website", None) or "",
            "1" if school.is_active else "0",
        ])
    content = "\ufeff" + output.getvalue()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="schools.csv"'},
    )


@router.post("/schools/import")
async def import_sales_schools(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    filename = file.filename or ""
    lowered = filename.lower().strip()
    if not (lowered.endswith(".csv") or lowered.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Поддерживаются только CSV и XLSX")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")

    rows = _parse_school_import(data, filename)
    created = 0
    updated = 0
    skipped = 0
    errors: List[str] = []
    for index, row in enumerate(rows, start=2):
        name = _normalize_source_name(row.get("name"))
        if not name:
            skipped += 1
            continue
        email = _clean_optional_text(row.get("email"))
        if email and ("@" not in email or "." not in email.split("@")[-1]):
            errors.append(f"Строка {index}: некорректная почта")
            skipped += 1
            continue
        school = db.query(SalesSchool).filter(cast(SalesSchool.name, Text).ilike(name)).first()
        if not school:
            school = SalesSchool(name=name, is_active=True)
            db.add(school)
            db.flush()
            created += 1
        else:
            updated += 1
        school.city = _clean_optional_text(row.get("city"))
        school.district = _clean_optional_text(row.get("district"))
        school.director = _clean_optional_text(row.get("director"))
        school.email = email
        school.address = _clean_optional_text(row.get("address"))
        school.phone = _clean_optional_text(row.get("phone"))
        school.website = _clean_optional_text(row.get("website"))
        _sync_sales_school_to_b2b(db, school)

    db.commit()
    log_action(db, current_user.id, "import", "sales_school", None, {"created": created, "updated": updated, "skipped": skipped})
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@router.post("/schools", response_model=SalesSchoolResponse, status_code=status.HTTP_201_CREATED)
async def create_sales_school(
    payload: SalesSchoolCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="School name is required")
    exists = db.query(SalesSchool).filter(cast(SalesSchool.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="School already exists")
    item = SalesSchool(
        name=name,
        city=_clean_optional_text(payload.city),
        district=_clean_optional_text(payload.district),
        director=_clean_optional_text(payload.director),
        email=str(payload.email) if payload.email else None,
        address=_clean_optional_text(payload.address),
        phone=_clean_optional_text(payload.phone),
        website=_clean_optional_text(payload.website),
        is_active=True,
    )
    db.add(item)
    db.flush()
    _sync_sales_school_to_b2b(db, item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "sales_school", item.id, {"name": name})
    return item


@router.put("/schools/{school_id}", response_model=SalesSchoolResponse)
async def update_sales_school(
    school_id: int,
    payload: SalesSchoolUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesSchool).filter(SalesSchool.id == school_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="School not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="School name is required")
        item.name = name
    for field in ("city", "district", "director", "address", "phone", "website"):
        if field in data:
            setattr(item, field, _clean_optional_text(data[field]))
    if "email" in data:
        item.email = str(data["email"]) if data["email"] else None
    if "is_active" in data:
        item.is_active = data["is_active"]
    _sync_sales_school_to_b2b(db, item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "sales_school", item.id, data)
    return item


@router.delete("/schools/{school_id}", status_code=204)
async def delete_sales_school(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesSchool).filter(SalesSchool.id == school_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="School not found")
    name = item.name
    db.delete(item)
    db.commit()
    log_action(db, current_user.id, "delete", "sales_school", school_id, {"name": name})


# ── School contacts ────────────────────────────────────────────────────────────

def _build_ow_comment(phone_extra: Optional[str], note: Optional[str]) -> Optional[str]:
    parts = []
    if phone_extra:
        parts.append(f"Доп. номер: {phone_extra}")
    if note:
        parts.append(note)
    return "\n".join(parts) if parts else None


def _sync_contact_to_ow(db: Session, contact: SalesSchoolContact, school_name: str) -> None:
    """Create or update the mirrored OwnerWorkspaceContact."""
    comment = _build_ow_comment(contact.phone_extra, contact.note)
    if contact.owner_workspace_contact_id:
        ow = db.query(OwnerWorkspaceContact).filter(
            OwnerWorkspaceContact.id == contact.owner_workspace_contact_id
        ).first()
        if ow:
            ow.full_name = contact.full_name
            ow.position = contact.position
            ow.phone = contact.phone
            ow.email = contact.email
            ow.company = school_name
            ow.comment = comment
            return
    # No linked OW contact — create one
    ow = OwnerWorkspaceContact(
        type="individual",
        full_name=contact.full_name,
        position=contact.position,
        phone=contact.phone,
        email=contact.email,
        company=school_name,
        comment=comment,
        source="sales_school",
    )
    db.add(ow)
    db.flush()
    contact.owner_workspace_contact_id = ow.id


@router.get("/schools/{school_id}/contacts", response_model=List[SalesSchoolContactResponse])
async def list_sales_school_contacts(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    school = db.query(SalesSchool).filter(SalesSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    return school.contacts


@router.post("/schools/{school_id}/contacts", response_model=SalesSchoolContactResponse, status_code=201)
async def create_sales_school_contact(
    school_id: int,
    payload: SalesSchoolContactCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    school = db.query(SalesSchool).filter(SalesSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    contact = SalesSchoolContact(school_id=school_id, **payload.dict())
    db.add(contact)
    db.flush()
    _sync_contact_to_ow(db, contact, school.name)
    db.commit()
    db.refresh(contact)
    return contact


@router.put("/schools/{school_id}/contacts/{contact_id}", response_model=SalesSchoolContactResponse)
async def update_sales_school_contact(
    school_id: int,
    contact_id: int,
    payload: SalesSchoolContactUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    contact = db.query(SalesSchoolContact).filter(
        SalesSchoolContact.id == contact_id, SalesSchoolContact.school_id == school_id
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    school = db.query(SalesSchool).filter(SalesSchool.id == school_id).first()
    for field, value in payload.dict(exclude_unset=True).items():
        setattr(contact, field, value)
    _sync_contact_to_ow(db, contact, school.name if school else "")
    db.commit()
    db.refresh(contact)
    return contact


@router.delete("/schools/{school_id}/contacts/{contact_id}", status_code=204)
async def delete_sales_school_contact(
    school_id: int,
    contact_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    contact = db.query(SalesSchoolContact).filter(
        SalesSchoolContact.id == contact_id, SalesSchoolContact.school_id == school_id
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    ow_id = contact.owner_workspace_contact_id
    db.delete(contact)
    db.flush()
    if ow_id:
        ow = db.query(OwnerWorkspaceContact).filter(OwnerWorkspaceContact.id == ow_id).first()
        if ow and ow.source == "sales_school":
            db.delete(ow)
    db.commit()


@router.get("/classes", response_model=List[SalesClassResponse])
async def list_sales_classes(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(SalesClass).order_by(SalesClass.name.asc())
    if active_only:
        query = query.filter(SalesClass.is_active.is_(True))
    return query.all()


@router.post("/classes", response_model=SalesClassResponse, status_code=status.HTTP_201_CREATED)
async def create_sales_class(
    payload: SalesClassCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Название класса обязательно")
    exists = db.query(SalesClass).filter(cast(SalesClass.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Такой класс уже есть")
    item = SalesClass(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "sales_class", item.id, {"name": name})
    return item


@router.put("/classes/{class_id}", response_model=SalesClassResponse)
async def update_sales_class(
    class_id: int,
    payload: SalesClassUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesClass).filter(SalesClass.id == class_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Класс не найден")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = (data["name"] or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Название класса обязательно")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "sales_class", item.id, data)
    return item


@router.get("/account-templates", response_model=List[AccountTemplateResponse])
async def list_account_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    return db.query(AccountTemplate).order_by(AccountTemplate.id.asc()).all()


@router.post("/account-templates", response_model=AccountTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_account_template(
    payload: AccountTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Название счёта обязательно")
    if payload.format not in ("group", "individual"):
        raise HTTPException(status_code=400, detail="Формат должен быть group или individual")
    item = AccountTemplate(name=name, format=payload.format)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "account_template", item.id, {"name": name, "format": payload.format})
    return item


@router.delete("/account-templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_account_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    item = db.query(AccountTemplate).filter(AccountTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Шаблон счёта не найден")
    db.delete(item)
    db.commit()
    log_action(db, current_user.id, "delete", "account_template", template_id, {})
    return None


@router.get("/lead-statuses", response_model=List[LeadStatusOptionResponse])
async def list_lead_statuses(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    query = db.query(LeadStatusOption).order_by(LeadStatusOption.id.asc())
    if active_only:
        query = query.filter(LeadStatusOption.is_active.is_(True))
    return query.all()


@router.post("/lead-statuses", response_model=LeadStatusOptionResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_status(
    payload: LeadStatusOptionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Status name is required")
    exists = db.query(LeadStatusOption).filter(cast(LeadStatusOption.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Lead status already exists")
    base = getattr(payload.base_status, "value", payload.base_status) if hasattr(payload.base_status, "value") else payload.base_status
    item = LeadStatusOption(name=name, base_status=base, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_status_option", item.id, {"name": name, "base_status": base})
    return item


@router.put("/lead-statuses/{status_id}", response_model=LeadStatusOptionResponse)
async def update_lead_status(
    status_id: int,
    payload: LeadStatusOptionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadStatusOption).filter(LeadStatusOption.id == status_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Lead status not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Status name is required")
        item.name = name
    if "base_status" in data:
        base = data["base_status"]
        item.base_status = getattr(base, "value", base) if base is not None else item.base_status
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_status_option", item.id, data)
    return item
