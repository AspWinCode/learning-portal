import hashlib
import os
import re
from datetime import date, datetime, timedelta, time as dt_time
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import cast, Text, or_, update as sa_update
from openpyxl import Workbook, load_workbook

from app import auth
from app.database import get_db
from app.student_display import get_student_display_name, get_students_display_names
from app.services.parent_invite import create_parent_with_invite
from app.services.finance_ledger import ensure_finance_transaction_for_bank_transaction
from app.services.student_card_conversion import (
    convert_student_card_to_student as student_card_convert,
    StudentCardConvertConflict,
)
from app.services.absence_makeup import assign_makeup_for_absence as absence_makeup_assign
from app.services.manual_lesson import create_manual_lesson as manual_lesson_create
from app.services.bank_operation import apply_bank_operation_to_student as bank_operation_apply
from app.services.ai_insights import build_lead_ai_insight
from app.services.payment_status import get_payment_status_list as payment_status_list_svc, get_payment_status_summary as payment_status_summary_svc
from app.services.lead_post_visit import update_lead_post_visit_stage as lead_post_visit_update_stage
from app.services.student_activity import log_student_activity
from app.services.person_sync import sync_lead_person, sync_student_card_person
from app.utils.datetime import utcnow
from app.models import (
    Lead,
    LeadStatus,
    StudentStatus,
    LeadTask,
    LeadTaskStatus,
    LeadSource,
    LeadTaskTemplate,
    LeadTaskStatusOption as LeadTaskStatusOptionModel,
    LeadInfoTemplate,
    LeadStatusOption,
    LeadCommunication,
    Invoice,
    InvoiceStatus,
    Event,
    EventStatus,
    EventRegistration,
    EventRegistrationStatus,
    Abonement,
    User,
    UserRole,
    AccountTemplate,
    SalesCity,
    SalesSchool,
    SalesClass,
    SalesInstruction,
    SalesInstructionImage,
    StudentCard,
    DiscountType,
    AbsenceFollowUp,
    Student,
    Group,
    GroupSchedule,
    LessonAttendance,
    StudentAccount,
    StudentAccountTransaction,
    StudentAccountTransactionKind,
    TochkaAppliedPayment,
    BankTransaction,
    BankTransactionStatus,
    PhonePaymentBinding,
    Program,
    GroupProgram,
    StudentProgram,
    ProgramMakeupCompatibility,
    StudentFreeze,
    CustomLesson,
    CustomLessonStudent,
    CustomLessonType,
    Task,
    TaskStatus,
    LeadActivity,
)
from app.utils.phone import normalize_phone, validate_phone_for_lead
from app.schemas.abonements import AbonementResponse
from app.schemas.sales import (
    LeadCreate,
    LeadUpdate,
    LeadResponse,
    LeadTaskCreate,
    LeadTaskResponse,
    LeadTaskUpdate,
    InvoiceCreate,
    InvoiceResponse,
    EventCreate,
    EventUpdate,
    EventResponse,
    EventRegistrationCreate,
    EventRegistrationResponse,
    LeadSourceCreate,
    LeadSourceResponse,
    LeadSourceUpdate,
    LeadTaskTemplateCreate,
    LeadTaskTemplateResponse,
    LeadTaskTemplateUpdate,
    LeadTaskStatusOptionCreate,
    LeadTaskStatusOptionResponse,
    LeadTaskStatusOptionUpdate,
    LeadImportResponse,
    LeadInfoTemplateCreate,
    LeadInfoTemplateResponse,
    LeadInfoTemplateUpdate,
    LeadStatusOptionCreate,
    LeadStatusOptionResponse,
    LeadStatusOptionUpdate,
    LeadSendInfoRequest,
    SalesCityCreate,
    SalesCityResponse,
    SalesCityUpdate,
    SalesSchoolCreate,
    SalesSchoolResponse,
    SalesSchoolUpdate,
    SalesClassCreate,
    SalesClassResponse,
    SalesClassUpdate,
    AccountTemplateCreate,
    AccountTemplateResponse,
    LeadQuickCommunicationCreate,
    LeadContactResultRequest,
    LeadCommunicationResponse,
    SalesDashboardResponse,
    SalesSchoolConversionItem,
    SalesQueueTaskItem,
    SalesQueueRegistrationItem,
    FollowUpItemResponse,
    LeadPushStatsResponse,
    SalesInstructionCreate,
    SalesInstructionUpdate,
    SalesInstructionResponse,
    StudentCardCreate,
    StudentCardUpdate,
    StudentCardResponse,
    StudentCardImportResponse,
    AbsenceFollowUpResponse,
    AbsenceFollowUpStageUpdate,
    OpenParentCabinetResponse,
    AnketaConvertRequest,
    AnketaConvertResponse,
    LeadConvertToStudentResponse,
    AbsenceMakeupAssign,
    MakeupSuggestionItem,
    PublicMakeupSlotsResponse,
    PublicMakeupSelectionRequest,
    ProgramMakeupCompatibilityResponse,
    ProgramMakeupCompatibilityCreate,
    PaymentStatusItem,
    PaymentStatusSummary,
    StudentFreezeCreate,
    StudentFreezeResponse,
    CloseByFactPreview,
    CloseByFactConfirm,
    LeadPostVisitStageUpdate,
    CustomLessonCreate,
    CustomLessonUpdate,
    CustomLessonResponse,
    SpecialistQuestionnaireRequest,
    SpecialistQuestionnaireResponse,
    TildaLeadRequest,
    TildaLeadResponse,
    LeadActivityCreate,
    LeadActivityResponse,
    LeadNextAction,
    LeadSidebarSummary,
)
from app.schemas.finance import (
    BankPaymentImportResponse,
    BankTransactionApplyRequest,
    BankTransactionExpenseCategoryUpdate,
    BankTransactionResponse,
    PhonePaymentBindingCreate,
    TochkaImportRequest,
)
from app.schemas.groups import LessonCallResultUpdate
from app.routers.action_log import log_action
from app.services.lead_conversion import convert_lead_to_student as lead_conversion_convert
from app.services.makeup_selection import (
    close_send_link_tasks_for_absence,
    create_sales_confirmation_task,
    list_makeup_suggestions_for_absence,
    resolve_absence_by_token,
)
from app.dependencies import require_sales_admin_owner
from app.routers import sales_admin, sales_bank, sales_bank_import, sales_events, sales_invoices, sales_lead_views, sales_lead_workflow, sales_leads, sales_makeups, sales_operations, sales_student_cards, sales_support, sales_tax

router = APIRouter()
router.include_router(sales_admin.router)
router.include_router(sales_bank.router)
router.include_router(sales_bank_import.router)
router.include_router(sales_events.router)
router.include_router(sales_invoices.router)
router.include_router(sales_lead_views.router)
router.include_router(sales_lead_workflow.router)
router.include_router(sales_leads.router)
router.include_router(sales_makeups.router)
router.include_router(sales_operations.router)
router.include_router(sales_student_cards.router)
router.include_router(sales_support.router)
router.include_router(sales_tax.router)


def _lead_eager_options():
    return (
        joinedload(Lead.owner),
        joinedload(Lead.abonement),
        joinedload(Lead.status_option),
        joinedload(Lead.source_ref),
        selectinload(Lead.tasks).joinedload(LeadTask.owner),
        selectinload(Lead.tasks).joinedload(LeadTask.template),
        selectinload(Lead.communications).joinedload(LeadCommunication.sender),
        selectinload(Lead.communications).joinedload(LeadCommunication.template),
        selectinload(Lead.invoices),
    )

# Compatibility layer /api/sales (РўР— СЌС‚Р°Рї 4): CRM + Operations + Finance.
# Р‘РёР·РЅРµСЃ-Р»РѕРіРёРєР° РІ app.services.*; Р·РґРµСЃСЊ вЂ” РїСЂРѕРІРµСЂРєР° РїСЂР°РІ, РІС‹Р·РѕРІ СЃРµСЂРІРёСЃРѕРІ, РјР°РїРїРёРЅРі РѕС‚РІРµС‚РѕРІ.


def _fix_mojibake(s: Optional[str]) -> Optional[str]:
    """
    Р’РѕСЃСЃС‚Р°РЅР°РІР»РёРІР°РµС‚ СЃС‚СЂРѕРєСѓ РёР· Р±РёС‚РѕР№ РєРѕРґРёСЂРѕРІРєРё (UTF-8, РїСЂРѕС‡РёС‚Р°РЅРЅС‹Р№ РєР°Рє Latin-1/CP1252).
    Р”Р»СЏ СѓР¶Рµ РЅРѕСЂРјР°Р»СЊРЅС‹С… СЃС‚СЂРѕРє РѕСЃС‚Р°С‘С‚СЃСЏ Р±РµР· РёР·РјРµРЅРµРЅРёР№.
    """
    if not s or not isinstance(s, str):
        return s
    for enc in ("latin1", "cp1252"):
        try:
            fixed = s.encode(enc).decode("utf-8")
            if fixed != s:
                return fixed
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
    return s


def _fix_lead_strings(lead: Lead) -> Lead:
    """
    РџРѕС‡РёРЅРёС‚СЊ СЃР°РјС‹Рµ Р·Р°РјРµС‚РЅС‹Рµ С‚РµРєСЃС‚РѕРІС‹Рµ РїРѕР»СЏ Р»РёРґР°, РµСЃР»Рё РѕРЅРё Р±С‹Р»Рё СЃРѕС…СЂР°РЅРµРЅС‹ РєСЂР°РєРѕР·СЏР±СЂР°РјРё.
    Р’РѕР·РІСЂР°С‰Р°РµС‚ С‚РѕС‚ Р¶Рµ ORMвЂ‘РѕР±СЉРµРєС‚ (РјСѓС‚Р°Р±РµР»СЊРЅРѕ), С‡С‚РѕР±С‹ Pydantic СѓРІРёРґРµР» СѓР¶Рµ РёСЃРїСЂР°РІР»РµРЅРЅС‹Рµ Р·РЅР°С‡РµРЅРёСЏ.
    """
    for field in [
        "contact_name",
        "phone",
        "parent_full_name",
        "child_full_name",
        "parent_phone",
        "child_phone",
        "email",
        "city",
        "school_name",
        "school_class",
        "source",
        "referral_name",
        "comment",
        "pause_reason",
        "lost_reason",
    ]:
        value = getattr(lead, field, None)
        fixed = _fix_mojibake(value)
        if fixed is not None and fixed != value:
            setattr(lead, field, fixed)
    return lead


def _serialize_time_for_api(t: Optional[dt_time]) -> Optional[str]:
    return t.strftime("%H:%M") if t else None


def _lesson_task_status(lesson_start: datetime, lesson_end: datetime, now: datetime, call_window_min: int = 25) -> str:
    """РЎС‚Р°С‚СѓСЃ РєР°СЂС‚РѕС‡РєРё СѓСЂРѕРєР°: waiting | soon | in_progress | call_round | completed."""
    if now < lesson_start - timedelta(minutes=15):
        return "waiting"
    if now < lesson_start:
        return "soon"
    if now < lesson_start + timedelta(minutes=10):
        return "in_progress"
    call_end = min(lesson_start + timedelta(minutes=call_window_min), lesson_end)
    if now < call_end:
        return "call_round"
    return "completed"


def _lesson_tasks_for_date(
    db: Session,
    target_date: date,
    now: Optional[datetime] = None,
) -> List[dict]:
    """РЎРїРёСЃРѕРє СѓСЂРѕРєРѕРІ РЅР° РѕРґРЅСѓ РґР°С‚Сѓ (РґР»СЏ СЃРµРіРѕРґРЅСЏ/Р·Р°РІС‚СЂР°/РЅРµРґРµР»Рё)."""
    if now is None:
        now = datetime.now()
    weekday = target_date.weekday()
    schedules = (
        db.query(GroupSchedule)
        .join(Group, Group.id == GroupSchedule.group_id)
        .filter(GroupSchedule.day_of_week == weekday, Group.status == "active")
        .order_by(GroupSchedule.start_time)
        .all()
    )
    out: List[dict] = []
    seen_keys = set()
    for sched in schedules:
        group = db.query(Group).options(
            joinedload(Group.trainer),
            joinedload(Group.group_students),
        ).filter(Group.id == sched.group_id).first()
        if not group:
            continue
        key = (group.id, sched.start_time, sched.end_time)
        if key in seen_keys:
            # Р—Р°С‰РёС‚Р° РѕС‚ СЃР»СѓС‡Р°Р№РЅС‹С… РґСѓР±Р»РµР№ СЂР°СЃРїРёСЃР°РЅРёСЏ: РѕРґРёРЅ СЃР»РѕС‚ РІ РґРµРЅСЊ РїРѕРєР°Р·С‹РІР°РµРј РѕРґРёРЅ СЂР°Р·
            continue
        seen_keys.add(key)
        trainer = group.trainer
        lesson_start = datetime.combine(target_date, sched.start_time)
        lesson_end = datetime.combine(target_date, sched.end_time)
        status = _lesson_task_status(lesson_start, lesson_end, now) if target_date == date.today() else "waiting"

        student_ids = [gs.student_id for gs in group.group_students]
        attendance_rows = {}
        if student_ids:
            atts = (
                db.query(LessonAttendance)
                .filter(
                    LessonAttendance.group_id == group.id,
                    LessonAttendance.lesson_date == target_date,
                    LessonAttendance.student_id.in_(student_ids),
                )
                .all()
            )
            for a in atts:
                attendance_rows[a.student_id] = a

        cards = {}
        if student_ids:
            card_list = (
                db.query(StudentCard)
                .filter(
                    StudentCard.student_id.in_(student_ids),
                    StudentCard.archived.is_(False),
                )
                .all()
            )
            for c in card_list:
                if c.student_id:
                    cards[c.student_id] = c

        students_out = []
        students = db.query(Student).filter(Student.id.in_(student_ids)).all() if student_ids else []
        display_names = get_students_display_names(db, student_ids)
        for st in students:
            card = cards.get(st.id)
            att_row = attendance_rows.get(st.id)
            attended = att_row.attended if att_row else None
            late = getattr(att_row, "late", False) if att_row else False
            call_result = getattr(att_row, "call_result", None) if att_row else None
            students_out.append({
                "student_id": st.id,
                "full_name": display_names.get(st.id, st.full_name or "вЂ”"),
                "attended": attended,
                "late": late,
                "call_result": call_result,
                "parent_full_name": (card.parent_full_name if card else None) or None,
                "parent_phone": (card.parent_phone if card else None) or None,
                "parent_phone_2": (card.parent_phone_2 if card else None) or None,
                "parent_telegram": (card.parent_telegram if card else None) or None,
            })

        call_contacted_count = sum(1 for u in students_out if u.get("call_result"))
        out.append({
            "group_id": group.id,
            "group_name": group.name,
            "direction": group.direction,
            "schedule_id": sched.id,
            "lesson_date": target_date.isoformat(),
            "start_time": sched.start_time.strftime("%H:%M") if hasattr(sched.start_time, "strftime") else str(sched.start_time),
            "end_time": sched.end_time.strftime("%H:%M") if hasattr(sched.end_time, "strftime") else str(sched.end_time),
            "status": status,
            "trainer_id": trainer.id if trainer else None,
            "trainer_name": trainer.full_name if trainer else "вЂ”",
            "students": students_out,
            "total": len(students_out),
            "present_count": sum(1 for u in students_out if u.get("attended") is True),
            "absent_count": sum(1 for u in students_out if u.get("attended") is False),
            "unknown_count": sum(1 for u in students_out if u.get("attended") is None),
            "call_contacted_count": call_contacted_count,
        })
    return out


@router.get("/_legacy-disabled/lesson-tasks/today")
async def list_lesson_tasks_today(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    """РЈСЂРѕРєРё РЅР° СЃРµРіРѕРґРЅСЏ РґР»СЏ СЂР°Р·РґРµР»Р° В«РџРѕР·РІР°С‚СЊ РґРµС‚РµР№ РЅР° Р·Р°РЅСЏС‚РёРµВ»."""
    today = date.today()
    out = _lesson_tasks_for_date(db, today)
    # Р”Р»СЏ РІРєР»Р°РґРєРё В«РЎРµРіРѕРґРЅСЏВ» РјРµРЅРµРґР¶РµСЂСѓ РїРѕРєР°Р·С‹РІР°РµРј С‚РѕР»СЊРєРѕ Р°РєС‚СѓР°Р»СЊРЅС‹Рµ СѓСЂРѕРєРё:
    # РѕР¶РёРґР°СЋС‚СЃСЏ / СЃРєРѕСЂРѕ / РёРґСѓС‚ / РёРґС‘С‚ РґРѕР·РІРѕРЅ. Р—Р°РІРµСЂС€С‘РЅРЅС‹Рµ СѓР±РёСЂР°РµРј.
    out = [item for item in out if item.get("status") != "completed"]
    return {"items": out}


@router.get("/_legacy-disabled/lesson-tasks/tomorrow")
async def list_lesson_tasks_tomorrow(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    """РЈСЂРѕРєРё РЅР° Р·Р°РІС‚СЂР°."""
    tomorrow = date.today() + timedelta(days=1)
    out = _lesson_tasks_for_date(db, tomorrow)
    return {"items": out}


@router.get("/_legacy-disabled/lesson-tasks/week")
async def list_lesson_tasks_week(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    """РЈСЂРѕРєРё РЅР° РЅРµРґРµР»СЋ (СЃРµРіРѕРґРЅСЏ + 6 РґРЅРµР№)."""
    today = date.today()
    out = []
    for day_offset in range(7):
        target = today + timedelta(days=day_offset)
        out.extend(_lesson_tasks_for_date(db, target))
    return {"items": out}


VALID_CALL_RESULTS = {"contacted", "no_answer", "cancelled", "technical", "messenger"}


@router.post("/_legacy-disabled/lesson-tasks/call-result")
async def set_lesson_call_result(
    payload: LessonCallResultUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    """РЈСЃС‚Р°РЅРѕРІРёС‚СЊ СЂРµР·СѓР»СЊС‚Р°С‚ РґРѕР·РІРѕРЅР° РїРѕ СѓС‡РµРЅРёРєСѓ (РјРµРЅРµРґР¶РµСЂ): contacted | no_answer | cancelled | technical | messenger."""
    if payload.call_result not in VALID_CALL_RESULTS:
        raise HTTPException(status_code=400, detail=f"call_result must be one of: {sorted(VALID_CALL_RESULTS)}")
    lesson_date = payload.lesson_date if isinstance(payload.lesson_date, date) else date.fromisoformat(str(payload.lesson_date))
    att = (
        db.query(LessonAttendance)
        .filter(
            LessonAttendance.group_id == payload.group_id,
            LessonAttendance.lesson_date == lesson_date,
            LessonAttendance.student_id == payload.student_id,
        )
        .first()
    )
    if not att:
        att = LessonAttendance(
            group_id=payload.group_id,
            lesson_date=lesson_date,
            student_id=payload.student_id,
            attended=False,
        )
        db.add(att)
        db.commit()
        db.refresh(att)
    att.call_result = payload.call_result
    att.call_result_at = utcnow()
    db.commit()
    return {"ok": True}


@router.get("/_legacy-disabled/sales-instructions", response_model=List[SalesInstructionResponse])
async def list_sales_instructions(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    auth.ensure_permission(current_user, "sales.access")
    items = db.query(SalesInstruction).order_by(SalesInstruction.created_at.asc()).all()
    return items


@router.post("/_legacy-disabled/sales-instructions", response_model=SalesInstructionResponse, status_code=status.HTTP_201_CREATED)
async def create_sales_instruction(
    payload: SalesInstructionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    title = (payload.title or "").strip()
    body = (payload.body or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title is required")
    if not body:
        raise HTTPException(status_code=400, detail="Body is required")
    item = SalesInstruction(
        title=title,
        body=body,
        created_by_id=current_user.id,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "sales_instruction", item.id, {"title": title})
    return item


@router.put("/_legacy-disabled/sales-instructions/{instruction_id}", response_model=SalesInstructionResponse)
async def update_sales_instruction(
    instruction_id: int,
    payload: SalesInstructionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesInstruction).filter(SalesInstruction.id == instruction_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Instruction not found")
    data = payload.dict(exclude_unset=True)
    if "title" in data:
        title = (data["title"] or "").strip()
        if not title:
            raise HTTPException(status_code=400, detail="Title is required")
        item.title = title
    if "body" in data:
        body = (data["body"] or "").strip()
        if not body:
            raise HTTPException(status_code=400, detail="Body is required")
        item.body = body
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "sales_instruction", item.id, data)
    return item


@router.delete("/_legacy-disabled/sales-instructions/{instruction_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_sales_instruction(
    instruction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesInstruction).filter(SalesInstruction.id == instruction_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Instruction not found")
    db.delete(item)
    db.commit()
    log_action(db, current_user.id, "delete", "sales_instruction", instruction_id, {})
    return None


@router.post("/_legacy-disabled/instruction-images")
async def upload_sales_instruction_image(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    content_type = file.content_type or "application/octet-stream"
    if not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="РњРѕР¶РЅРѕ Р·Р°РіСЂСѓР¶Р°С‚СЊ С‚РѕР»СЊРєРѕ РёР·РѕР±СЂР°Р¶РµРЅРёСЏ")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    if len(data) > 400 * 1024:
        raise HTTPException(status_code=400, detail="РљР°СЂС‚РёРЅРєР° СЃР»РёС€РєРѕРј Р±РѕР»СЊС€Р°СЏ (Р»РёРјРёС‚ ~400KB)")
    img = SalesInstructionImage(data=data, content_type=content_type)
    db.add(img)
    db.commit()
    db.refresh(img)
    url = f"/api/sales/instruction-images/{img.id}"
    return {"id": img.id, "url": url}


@router.get("/_legacy-disabled/instruction-images/{image_id}")
async def get_sales_instruction_image(
    image_id: int,
    db: Session = Depends(get_db),
):
    img = db.query(SalesInstructionImage).filter(SalesInstructionImage.id == image_id).first()
    if not img:
        raise HTTPException(status_code=404, detail="Image not found")
    return StreamingResponse(BytesIO(img.data), media_type=img.content_type or "application/octet-stream")


def _student_card_response(card: StudentCard, user: User, db: Session) -> StudentCardResponse:
    """РЎРѕР±РёСЂР°РµС‚ РѕС‚РІРµС‚ РїРѕ РєР°СЂС‚РѕС‡РєРµ; РїРѕР»СЏ Р°Р±РѕРЅРµРјРµРЅС‚Р°/СЃРєРёРґРєРё С‚РѕР»СЊРєРѕ РґР»СЏ owner."""
    parent_cabinet_open = False
    if getattr(card, "student_id", None):
        st = db.query(Student).filter(Student.id == card.student_id).first()
        if st and st.parent_id:
            parent_cabinet_open = True
    data = {
        "id": card.id,
        "student_id": getattr(card, "student_id", None),
        "student_full_name": card.student_full_name,
        "parent_cabinet_open": parent_cabinet_open,
        "birth_date": card.birth_date,
        "student_phone": card.student_phone,
        "telegram": card.telegram,
        "gender": card.gender,
        "on_grant": card.on_grant,
        "format_type": card.format_type,
        "city": card.city,
        "school": card.school,
        "grade": card.grade,
        "parent_full_name": card.parent_full_name,
        "parent_phone": card.parent_phone,
        "parent_phone_2": card.parent_phone_2,
        "parent_telegram": getattr(card, "parent_telegram", None),
        "parent_email": getattr(card, "parent_email", None),
        "student_email": getattr(card, "student_email", None),
        "preferred_messenger": getattr(card, "preferred_messenger", None),
        "comment": getattr(card, "comment", None),
        "source": getattr(card, "source", None),
        "payment_link": getattr(card, "payment_link", None),
        "learning_period_start": getattr(card, "learning_period_start", None),
        "next_payment_date": getattr(card, "next_payment_date", None),
        "archived": card.archived,
        "anketa_status": getattr(card, "anketa_status", "converted"),
        "primary_for_bank_payments": getattr(card, "primary_for_bank_payments", False),
        "created_at": card.created_at,
        "updated_at": card.updated_at,
    }
    effective_role = auth.resolve_effective_role(user)
    if effective_role == UserRole.OWNER:
        data["abonement_id"] = card.abonement_id
        data["discount_type"] = card.discount_type
        data["discount_value"] = card.discount_value
        data["abonement"] = AbonementResponse.model_validate(card.abonement) if card.abonement else None
    else:
        data["abonement_id"] = None
        data["discount_type"] = DiscountType.NONE
        data["discount_value"] = 0.0
        data["abonement"] = None
    return StudentCardResponse(**data)


def _require_sales_admin_owner(user: User) -> None:
    auth.ensure_permission(user, "sales.access")


def _normalize_name(s: str) -> str:
    """РќРѕСЂРјР°Р»РёР·Р°С†РёСЏ Р¤РРћ РґР»СЏ СЃСЂР°РІРЅРµРЅРёСЏ: РЅРёР¶РЅРёР№ СЂРµРіРёСЃС‚СЂ, РѕРґРЅР° РїСЂРѕР±РµР»СЊРЅР°СЏ СЃС‚СЂРѕРєР°."""
    if not s or not isinstance(s, str):
        return ""
    return " ".join((s or "").lower().strip().split())


def _payer_matches_parent(payer_name: str, parent_full_name: Optional[str]) -> bool:
    """
    РЎРѕРІРїР°РґР°РµС‚ Р»Рё РїР»Р°С‚РµР»СЊС‰РёРє СЃ Р¤РРћ СЂРѕРґРёС‚РµР»СЏ.
    Р’ РўРѕС‡РєР° Р‘Р°РЅРє С‡Р°СЃС‚Рѕ РїРёС€СѓС‚ В«РРјСЏ РћС‚С‡РµСЃС‚РІРѕ Р¤.В» (РЅР°РїСЂРёРјРµСЂ В«РќР°С‚Р°Р»СЊСЏ Р“РµРѕСЂРіРёРµРІРЅР° Рњ.В»),
    РІ РєР°СЂС‚РѕС‡РєРµ вЂ” В«Р¤Р°РјРёР»РёСЏ РРјСЏ РћС‚С‡РµСЃС‚РІРѕВ» (РЅР°РїСЂРёРјРµСЂ В«РњРµРґРІРµРґРµРІР° РќР°С‚Р°Р»СЊСЏ Р“РµРѕСЂРіРёРµРІРЅР°В»).
    РРЅРёС†РёР°Р» (РѕРґРЅР° Р±СѓРєРІР° РёР»Рё В«РҐ.В») СЃРѕРїРѕСЃС‚Р°РІР»СЏРµС‚СЃСЏ СЃ РїРµСЂРІРѕР№ Р±СѓРєРІРѕР№ Р»СЋР±РѕРіРѕ СЃР»РѕРІР° Сѓ СЂРѕРґРёС‚РµР»СЏ (РѕР±С‹С‡РЅРѕ С„Р°РјРёР»РёРё).
    """
    if not parent_full_name or not payer_name:
        return False
    p = _normalize_name(payer_name)
    parent = _normalize_name(parent_full_name)
    if not p or not parent:
        return False
    if p == parent:
        return True
    if p in parent or parent in p:
        return True
    p_words = p.split()
    parent_words = parent.split()
    parent_word_set = set(parent_words)
    # РЎР»РѕРІР°-РёРЅРёС†РёР°Р»С‹: РѕРґРЅР° Р±СѓРєРІР° РёР»Рё Р±СѓРєРІР° СЃ С‚РѕС‡РєРѕР№ (В«РјВ», В«Рј.В»)
    initials = []
    full_words = []
    for w in p_words:
        w_clean = w.rstrip(".")
        if len(w_clean) == 1 and w_clean.isalpha():
            initials.append(w_clean)
        else:
            full_words.append(w)
    # Р’СЃРµ РЅРµ-РёРЅРёС†РёР°Р»С‹ РёР· РїР»Р°С‚РµР»СЊС‰РёРєР° РґРѕР»Р¶РЅС‹ Р±С‹С‚СЊ РІ Р¤РРћ СЂРѕРґРёС‚РµР»СЏ
    if not all(w in parent_word_set for w in full_words):
        return False
    # РљР°Р¶РґС‹Р№ РёРЅРёС†РёР°Р» РґРѕР»Р¶РµРЅ СЃРѕРІРїР°РґР°С‚СЊ СЃ РїРµСЂРІРѕР№ Р±СѓРєРІРѕР№ РєР°РєРѕРіРѕ-С‚Рѕ СЃР»РѕРІР° Сѓ СЂРѕРґРёС‚РµР»СЏ (С„Р°РјРёР»РёСЏ В«РњРµРґРІРµРґРµРІР°В» в†’ В«РјВ»)
    for letter in initials:
        if not any(pw.startswith(letter) for pw in parent_words):
            return False
    return True



def _resolve_student_for_bank_payment(
    db: Session,
    student_ids: List[int],
    account_id: str,
    tx_date: str,
    amount: float,
    payer_name: str,
) -> Optional[int]:
    """
    РР· РЅРµСЃРєРѕР»СЊРєРёС… СѓС‡РµРЅРёРєРѕРІ (РѕРґРёРЅ СЂРѕРґРёС‚РµР»СЊ) РІС‹Р±РёСЂР°РµРј РѕРґРЅРѕРіРѕ: primary_for_bank_payments, Р·Р°С‚РµРј Р°РєС‚РёРІРЅС‹Р№ Р°Р±РѕРЅРµРјРµРЅС‚, Р·Р°С‚РµРј РѕС‚СЂРёС†Р°С‚РµР»СЊРЅС‹Р№ Р±Р°Р»Р°РЅСЃ.
    """
    if not student_ids:
        return None
    if len(student_ids) == 1:
        return student_ids[0]
    cards = (
        db.query(StudentCard)
        .filter(
            StudentCard.student_id.in_(student_ids),
            StudentCard.archived.is_(False),
        )
        .options(joinedload(StudentCard.student).joinedload(Student.accounts))
        .all()
    )
    by_primary = [c for c in cards if getattr(c, "primary_for_bank_payments", False)]
    if len(by_primary) == 1:
        return by_primary[0].student_id
    by_abonement = [c for c in cards if c.student_id and c.abonement_id]
    if len(by_abonement) == 1:
        return by_abonement[0].student_id
    students_with_negative = []
    for sid in student_ids:
        acc = db.query(StudentAccount).filter(StudentAccount.student_id == sid).first()
        if acc and acc.balance is not None and acc.balance < 0:
            students_with_negative.append(sid)
    if len(students_with_negative) == 1:
        return students_with_negative[0]
    return None


def do_tochka_import_and_apply(
    db: Session,
    account_id: str,
    date_from: date,
    date_to: date,
    actor_user_id: Optional[int] = None,
) -> BankPaymentImportResponse:
    """
    Р—Р°РіСЂСѓР¶Р°РµС‚ РІС‹РїРёСЃРєСѓ РўРѕС‡РєР° Р‘Р°РЅРє Р·Р° РїРµСЂРёРѕРґ. Р”РµРґСѓРїР»РёРєР°С†РёСЏ РїРѕ operation_id (bank_transactions).
    РњР°С‚С‡РёРЅРі РїРѕ РЅРѕСЂРјР°Р»РёР·РѕРІР°РЅРЅРѕРјСѓ С‚РµР»РµС„РѕРЅСѓ РїР»Р°С‚РµР»СЊС‰РёРєР°: РїСЂРёРІСЏР·РєРё (phone_payment_bindings) РёР»Рё СЂРѕРґРёС‚РµР»СЊ РІ РєР°СЂС‚РѕС‡РєРµ.
    РџСЂРё РѕРґРЅРѕРј СЂРѕРґРёС‚РµР»Рµ Рё РЅРµСЃРєРѕР»СЊРєРёС… СѓС‡РµРЅРёРєР°С…: primary_for_bank_payments в†’ Р°РєС‚РёРІРЅС‹Р№ Р°Р±РѕРЅРµРјРµРЅС‚ в†’ РѕС‚СЂРёС†Р°С‚РµР»СЊРЅС‹Р№ Р±Р°Р»Р°РЅСЃ в†’ РёРЅР°С‡Рµ ambiguous.
    """
    from app.services.tochka_client import (
        fetch_statement_ready,
        extract_incoming_transactions,
    )

    statement = fetch_statement_ready(account_id, date_from, date_to)
    transactions = extract_incoming_transactions(statement)
    cards = (
        db.query(StudentCard)
        .filter(StudentCard.archived.is_(False), StudentCard.student_id.isnot(None))
        .options(joinedload(StudentCard.student).joinedload(Student.parent))
        .all()
    )
    bindings = {b.payer_phone_normalized: b.parent_id for b in db.query(PhonePaymentBinding).all()}

    applied: List[dict] = []
    no_match: List[dict] = []
    ambiguous: List[dict] = []

    for idx, tx in enumerate(transactions):
        payer_name = (tx.get("payer_name") or "").strip()
        amount = tx.get("amount") or 0
        tx_date = tx.get("date") or ""
        payer_phone = normalize_phone(tx.get("payer_phone_raw") or "")

        operation_id = (tx.get("operation_id") or "").strip()
        if not operation_id:
            operation_id = hashlib.sha256(
                f"{account_id}|{tx_date}|{amount}|{payer_name}|{payer_phone}|{idx}".encode()
            ).hexdigest()

        existing_bt = db.query(BankTransaction).filter(BankTransaction.operation_id == operation_id).first()
        if existing_bt:
            bt = existing_bt
            ensure_finance_transaction_for_bank_transaction(db, bt, bank_source="tochka")
            if bt.status == BankTransactionStatus.APPLIED.value:
                continue
        else:
            bt = BankTransaction(
                operation_id=operation_id,
                tochka_account_id=account_id,
                amount=amount,
                payer_phone=payer_phone or None,
                payer_name=payer_name[:512] if payer_name else None,
                payment_date=tx_date,
                status=BankTransactionStatus.NEW.value,
            )
            db.add(bt)
            db.flush()
            ensure_finance_transaction_for_bank_transaction(db, bt, bank_source="tochka")

        student_ids: List[int] = []
        if payer_phone:
            parent_id = bindings.get(payer_phone)
            if parent_id is not None:
                student_ids = [
                    row[0]
                    for row in db.query(Student.id).filter(Student.parent_id == parent_id).all()
                ]
            if not student_ids:
                for c in cards:
                    if normalize_phone(c.parent_phone or "") == payer_phone or normalize_phone(
                        getattr(c, "parent_phone_2", None) or ""
                    ) == payer_phone:
                        if c.student_id:
                            student_ids.append(c.student_id)
                student_ids = list(dict.fromkeys(student_ids))

        # Р•СЃР»Рё С‚РµР»РµС„РѕРЅ РІ РІС‹РїРёСЃРєРµ РЅРµ РїСЂРёС€С‘Р» РёР»Рё РїРѕ С‚РµР»РµС„РѕРЅСѓ РЅРµ РЅР°С€Р»Рё вЂ” РїСЂРѕР±СѓРµРј РјР°С‚С‡ РїРѕ Р¤РРћ (fallback)
        if not student_ids and payer_name:
            for c in cards:
                if _payer_matches_parent(payer_name, c.parent_full_name):
                    if c.student_id:
                        student_ids.append(c.student_id)
                elif c.student and c.student.parent and c.student.parent.full_name:
                    if _payer_matches_parent(payer_name, c.student.parent.full_name):
                        if c.student_id:
                            student_ids.append(c.student_id)
            student_ids = list(dict.fromkeys(student_ids))

        if not student_ids:
            bt.status = BankTransactionStatus.NO_MATCH.value
            no_match.append({
                "payer_name": payer_name,
                "amount": amount,
                "date": tx_date,
                "payer_phone": payer_phone or None,
            })
            continue

        chosen_student_id = _resolve_student_for_bank_payment(
            db, student_ids, account_id, tx_date, amount, payer_name
        )
        if chosen_student_id is None:
            bt.status = BankTransactionStatus.AMBIGUOUS.value
            ambiguous.append({
                "payer_name": payer_name,
                "amount": amount,
                "date": tx_date,
                "payer_phone": payer_phone or None,
                "candidates": [
                    {
                        "student_id": sid,
                        "student_name": get_student_display_name(
                            db, db.query(Student).filter(Student.id == sid).first()
                        ),
                        "parent_full_name": next(
                            (c.parent_full_name or "" for c in cards if c.student_id == sid),
                            "",
                        ),
                    }
                    for sid in student_ids
                ],
            })
            continue

        student = db.query(Student).filter(Student.id == chosen_student_id).first()
        if not student:
            bt.status = BankTransactionStatus.NO_MATCH.value
            no_match.append({"payer_name": payer_name, "amount": amount, "date": tx_date, "payer_phone": payer_phone or None})
            continue

        account = (
            db.query(StudentAccount)
            .filter(StudentAccount.student_id == chosen_student_id)
            .order_by(StudentAccount.id)
            .first()
        )
        if not account:
            account = StudentAccount(student_id=chosen_student_id, name="РћСЃРЅРѕРІРЅРѕР№", balance=0.0)
            db.add(account)
            db.flush()

        note = f"РўРѕС‡РєР° Р‘Р°РЅРє, РїР»Р°С‚РµР»СЊС‰РёРє: {payer_name}, РґР°С‚Р°: {tx_date}"
        db.add(
            StudentAccountTransaction(
                account_id=account.id,
                amount=amount,
                kind=StudentAccountTransactionKind.PAYMENT,
                note=note,
            )
        )
        account.balance += amount
        payer_trunc = (payer_name or "")[:512]
        db.add(
            TochkaAppliedPayment(
                tochka_account_id=account_id,
                payment_date=tx_date,
                amount=amount,
                payer_name=payer_trunc,
                student_id=chosen_student_id,
                student_account_id=account.id,
            )
        )
        try:
            pay_date = date.fromisoformat(tx_date[:10]) if tx_date else date.today()
        except (ValueError, TypeError):
            pay_date = date.today()
        from app.services.student_card_period import update_card_payment_dates
        update_card_payment_dates(db, chosen_student_id, pay_date)

        bt.status = BankTransactionStatus.APPLIED.value
        bt.student_id = chosen_student_id
        bt.student_account_id = account.id

        student_name = get_student_display_name(db, student)
        applied.append({
            "payer_name": payer_name,
            "amount": amount,
            "date": tx_date,
            "student_id": chosen_student_id,
            "account_id": account.id,
            "student_name": student_name,
        })

    db.commit()
    if actor_user_id is not None:
        log_action(
            db, actor_user_id, "tochka_import_apply", "sales", None,
            {"applied": len(applied), "no_match": len(no_match), "ambiguous": len(ambiguous)},
        )
    return BankPaymentImportResponse(applied=applied, no_match=no_match, ambiguous=ambiguous)


@router.get("/_legacy-disabled/tochka/status")
async def tochka_bank_status(current_user: User = Depends(require_sales_admin_owner)):
    """РџСЂРѕРІРµСЂРєР°: Р·Р°РґР°РЅС‹ Р»Рё СѓС‡С‘С‚РЅС‹Рµ РґР°РЅРЅС‹Рµ РўРѕС‡РєР° Р‘Р°РЅРє Рё РІРєР»СЋС‡РµРЅРѕ Р»Рё Р°РІС‚РѕР·Р°С‡РёСЃР»РµРЅРёРµ."""
    from app.services.tochka_client import is_configured, is_auto_import_configured
    return {
        "configured": is_configured(),
        "auto_import_configured": is_auto_import_configured(),
    }


@router.get("/_legacy-disabled/tochka/status/public")
async def tochka_bank_status_public():
    """РџСЂРѕРІРµСЂРєР° Р±РµР· Р°РІС‚РѕСЂРёР·Р°С†РёРё (С‚РѕР»СЊРєРѕ configured/auto_import_configured). Р”Р»СЏ РјРѕРЅРёС‚РѕСЂРёРЅРіР° Рё curl СЃ СЃРµСЂРІРµСЂР°."""
    from app.services.tochka_client import is_configured, is_auto_import_configured
    return {
        "configured": is_configured(),
        "auto_import_configured": is_auto_import_configured(),
    }


@router.post("/_legacy-disabled/tochka/import-and-apply", response_model=BankPaymentImportResponse)
async def tochka_import_and_apply(
    payload: TochkaImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """
    Р—Р°РіСЂСѓР·РёС‚СЊ РІС‹РїРёСЃРєСѓ РёР· РўРѕС‡РєР° Р‘Р°РЅРє Р·Р° РїРµСЂРёРѕРґ. РњР°С‚С‡РёРЅРі РїРѕ С‚РµР»РµС„РѕРЅСѓ РїР»Р°С‚РµР»СЊС‰РёРєР° (РїСЂРёРІСЏР·РєРё РёР»Рё РєР°СЂС‚РѕС‡РєР°).
    Р”РµРґСѓРїР»РёРєР°С†РёСЏ РїРѕ operation_id. РЈР¶Рµ РѕР±СЂР°Р±РѕС‚Р°РЅРЅС‹Рµ РѕРїРµСЂР°С†РёРё РїСЂРѕРїСѓСЃРєР°СЋС‚СЃСЏ.
    """
    from app.services.tochka_client import is_configured

    if not is_configured():
        raise HTTPException(status_code=400, detail="РўРѕС‡РєР° Р‘Р°РЅРє РЅРµ РЅР°СЃС‚СЂРѕРµРЅ: Р·Р°РґР°Р№С‚Рµ TOCHKA_CLIENT_ID Рё TOCHKA_CLIENT_SECRET РІ .env")

    account_id = (payload.account_id or "").strip() or (os.getenv("TOCHKA_ACCOUNT_ID") or "").strip()
    if not account_id:
        raise HTTPException(
            status_code=400,
            detail="РЈРєР°Р¶РёС‚Рµ account_id (ID СЃС‡С‘С‚Р° РІ РўРѕС‡РєР° Р‘Р°РЅРє) РІ С‚РµР»Рµ Р·Р°РїСЂРѕСЃР° РёР»Рё Р·Р°РґР°Р№С‚Рµ TOCHKA_ACCOUNT_ID РІ .env",
        )

    try:
        date_from = date.fromisoformat(payload.date_from)
        date_to = date.fromisoformat(payload.date_to)
    except ValueError:
        raise HTTPException(status_code=400, detail="date_from Рё date_to РґРѕР»Р¶РЅС‹ Р±С‹С‚СЊ РІ С„РѕСЂРјР°С‚Рµ YYYY-MM-DD")

    if date_from > date_to:
        raise HTTPException(status_code=400, detail="date_from РЅРµ РјРѕР¶РµС‚ Р±С‹С‚СЊ Р±РѕР»СЊС€Рµ date_to")

    try:
        return do_tochka_import_and_apply(db, account_id, date_from, date_to, actor_user_id=current_user.id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"РћС€РёР±РєР° РІС‹РїРёСЃРєРё РўРѕС‡РєР° Р‘Р°РЅРє: {e!s}")


@router.get("/_legacy-disabled/bank-transactions", response_model=List[BankTransactionResponse])
async def list_bank_transactions(
    status: Optional[List[str]] = Query(None, description="Р¤РёР»СЊС‚СЂ: new, no_match, ambiguous, applied, expense; Р±РµР· РїР°СЂР°РјРµС‚СЂР° вЂ” РІСЃРµ РѕРїРµСЂР°С†РёРё"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """РћС‡РµСЂРµРґСЊ РѕРїРµСЂР°С†РёР№ РёР· Р±Р°РЅРєР° РґР»СЏ СЂСѓС‡РЅРѕРіРѕ СЂР°Р·Р±РѕСЂР° (no_match, ambiguous)."""
    q = db.query(BankTransaction).order_by(BankTransaction.created_at.desc())
    if status:
        q = q.filter(BankTransaction.status.in_(status))
    items = q.limit(500).all()
    return [BankTransactionResponse.model_validate(b) for b in items]


@router.post("/_legacy-disabled/phone-payment-bindings")
async def create_phone_payment_binding(
    payload: PhonePaymentBindingCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """РџСЂРёРІСЏР·Р°С‚СЊ С‚РµР»РµС„РѕРЅ РїР»Р°С‚РµР»СЊС‰РёРєР° Рє СЂРѕРґРёС‚РµР»СЋ: СЃР»РµРґСѓСЋС‰РёРµ РїР»Р°С‚РµР¶Рё СЃ СЌС‚РѕРіРѕ РЅРѕРјРµСЂР° Р·Р°С‡РёСЃР»СЏС‚СЃСЏ Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРё."""
    normalized = normalize_phone(payload.payer_phone)
    if not normalized:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ РЅРѕРјРµСЂ С‚РµР»РµС„РѕРЅР°")
    parent = db.query(User).filter(User.id == payload.parent_id, User.role == UserRole.PARENT).first()
    if not parent:
        raise HTTPException(status_code=404, detail="Р РѕРґРёС‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    existing = db.query(PhonePaymentBinding).filter(PhonePaymentBinding.payer_phone_normalized == normalized).first()
    if existing:
        existing.parent_id = payload.parent_id
        db.commit()
        return {"ok": True, "updated": True}
    db.add(PhonePaymentBinding(payer_phone_normalized=normalized, parent_id=payload.parent_id))
    db.commit()
    return {"ok": True}


@router.post("/_legacy-disabled/bank-transactions/{transaction_id}/apply", response_model=BankTransactionResponse)
async def apply_bank_transaction(
    transaction_id: int,
    payload: BankTransactionApplyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Р—Р°С‡РёСЃР»РёС‚СЊ СЃРїРѕСЂРЅСѓСЋ РѕРїРµСЂР°С†РёСЋ (no_match / ambiguous) РЅР° РІС‹Р±СЂР°РЅРЅРѕРіРѕ СѓС‡РµРЅРёРєР°. РџСЂРё no_match СЃРѕР·РґР°С‘С‚СЃСЏ РїСЂРёРІСЏР·РєР° С‚РµР»РµС„РѕРЅР° Рє СЂРѕРґРёС‚РµР»СЋ."""
    try:
        result = bank_operation_apply(db, transaction_id, payload.student_id)
    except ValueError as e:
        msg = str(e)
        if "РЅРµ РЅР°Р№РґРµРЅ" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)
    return BankTransactionResponse.model_validate(result.transaction)


@router.patch("/_legacy-disabled/bank-transactions/{transaction_id}/expense-category", response_model=BankTransactionResponse)
async def update_bank_transaction_expense_category(
    transaction_id: int,
    payload: BankTransactionExpenseCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """РЈСЃС‚Р°РЅРѕРІРёС‚СЊ РєР°С‚РµРіРѕСЂРёСЋ СЂР°СЃС…РѕРґР° (РєРѕРјРёСЃСЃРёСЏ, С‚РёРїРѕРіСЂР°С„РёСЏ, Р°СЂРµРЅРґР° Рё С‚.Рґ.). РўРѕР»СЊРєРѕ РґР»СЏ РѕРїРµСЂР°С†РёР№ СЃРѕ СЃС‚Р°С‚СѓСЃРѕРј expense."""
    bt = db.query(BankTransaction).filter(BankTransaction.id == transaction_id).first()
    if not bt:
        raise HTTPException(status_code=404, detail="РћРїРµСЂР°С†РёСЏ РЅРµ РЅР°Р№РґРµРЅР°")
    if bt.status != BankTransactionStatus.EXPENSE.value:
        raise HTTPException(status_code=400, detail="РљР°С‚РµРіРѕСЂРёСЋ РјРѕР¶РЅРѕ Р·Р°РґР°С‚СЊ С‚РѕР»СЊРєРѕ РґР»СЏ СЂР°СЃС…РѕРґР° (СЃРїРёСЃР°РЅРёРµ)")
    if payload.expense_category is not None:
        bt.expense_category = (payload.expense_category or "").strip() or None
    db.commit()
    db.refresh(bt)
    return BankTransactionResponse.model_validate(bt)


@router.delete("/_legacy-disabled/bank-transactions/{transaction_id}")
async def delete_bank_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
) -> Dict[str, bool]:
    """РЈРґР°Р»РёС‚СЊ РѕРїРµСЂР°С†РёСЋ Р±Р°РЅРєР° (BankTransaction).

    РСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РґР»СЏ СЂСѓС‡РЅРѕР№ РѕС‡РёСЃС‚РєРё РѕС‡РµСЂРµРґРё РѕРїРµСЂР°С†РёР№ РІ РёРЅС‚РµСЂС„РµР№СЃРµ В«Р”РѕР»РіРё Рё РѕРїР»Р°С‚С‹В» в†’ В«РћРїРµСЂР°С†РёРё Р±Р°РЅРєР°В».
    РџСЂРµРґРїРѕР»Р°РіР°РµС‚СЃСЏ, С‡С‚Рѕ РїРµСЂРµРґ СѓРґР°Р»РµРЅРёРµРј Р°РґРјРёРЅРёСЃС‚СЂР°С†РёСЏ РїСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё РѕС‚РєР°С‚РёР»Р° СЃРІСЏР·Р°РЅРЅС‹Рµ РґРµР№СЃС‚РІРёСЏ РїРѕ СЃС‡РµС‚Р°Рј СѓС‡РµРЅРёРєРѕРІ.
    """
    bt = db.query(BankTransaction).filter(BankTransaction.id == transaction_id).first()
    if not bt:
        raise HTTPException(status_code=404, detail="РћРїРµСЂР°С†РёСЏ РЅРµ РЅР°Р№РґРµРЅР°")
    db.delete(bt)
    db.commit()
    return {"ok": True}


@router.get("/_legacy-disabled/student-cards", response_model=List[StudentCardResponse])
async def list_student_cards(
    archived: Optional[bool] = Query(None, description="Р¤РёР»СЊС‚СЂ РїРѕ Р°СЂС…РёРІСѓ: true/false РёР»Рё РЅРµ РїРµСЂРµРґР°РІР°С‚СЊ вЂ” РІСЃРµ"),
    anketa_status: Optional[List[str]] = Query(None, description="Р¤РёР»СЊС‚СЂ РїРѕ СЃС‚Р°С‚СѓСЃСѓ Р°РЅРєРµС‚С‹: draft, filled, converted, cancelled"),
    student_id: Optional[int] = Query(None, description="Р¤РёР»СЊС‚СЂ РїРѕ РїСЂРёРІСЏР·Р°РЅРЅРѕРјСѓ СѓС‡РµРЅРёРєСѓ"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    query = db.query(StudentCard)
    if archived is not None:
        query = query.filter(StudentCard.archived == archived)
    if anketa_status:
        query = query.filter(StudentCard.anketa_status.in_(anketa_status))
    if student_id is not None:
        query = query.filter(StudentCard.student_id == student_id)
    items = query.order_by(StudentCard.created_at.desc()).all()
    return [_student_card_response(c, current_user, db) for c in items]


@router.post("/_legacy-disabled/student-cards", response_model=StudentCardResponse, status_code=status.HTTP_201_CREATED)
async def create_student_card(
    payload: StudentCardCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    name = (payload.student_full_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Р¤РРћ СѓС‡РµРЅРёРєР° РѕР±СЏР·Р°С‚РµР»СЊРЅРѕ")
    data = payload.model_dump()
    if data.get("student_id") is None and not data.get("anketa_status"):
        data["anketa_status"] = "draft"
    data["phone_normalized"] = normalize_phone(data.get("parent_phone") or data.get("student_phone") or "") or None
    # РђР±РѕРЅРµРјРµРЅС‚ Рё СЃРєРёРґРєР° РјРѕР¶РµС‚ РјРµРЅСЏС‚СЊ С‚РѕР»СЊРєРѕ owner
    if auth.resolve_effective_role(current_user) != UserRole.OWNER:
        data["abonement_id"] = None
        data["discount_type"] = DiscountType.NONE
        data["discount_value"] = 0.0
    # РЎСЃС‹Р»РєСѓ РѕРїР»Р°С‚С‹ РјРѕРіСѓС‚ Р·Р°РґР°РІР°С‚СЊ owner Рё admin; РґР»СЏ РѕСЃС‚Р°Р»СЊРЅС‹С… (sales) РѕС‡РёС‰Р°РµРј
    if auth.resolve_effective_role(current_user) not in (UserRole.OWNER, UserRole.ADMIN):
        data["payment_link"] = None
    if data.get("abonement_id"):
        ab = db.query(Abonement).filter(Abonement.id == data["abonement_id"]).first()
        if not ab:
            raise HTTPException(status_code=404, detail="РђР±РѕРЅРµРјРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    if data.get("student_id") is not None:
        st = db.query(Student).filter(Student.id == data["student_id"]).first()
        if not st:
            raise HTTPException(status_code=404, detail="РЈС‡РµРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    card = StudentCard(**data)
    db.add(card)
    db.flush()
    sync_student_card_person(db, card)
    db.commit()
    db.refresh(card)
    return _student_card_response(card, current_user, db)


@router.get("/_legacy-disabled/student-cards/import-template")
async def download_student_cards_import_template(
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    wb = Workbook()
    ws = wb.active
    ws.title = "РљР°СЂС‚РѕС‡РєРё СѓС‡РµРЅРёРєРѕРІ"
    headers = [
        "Р¤РРћ СѓС‡РµРЅРёРєР°",
        "Р”Р°С‚Р° СЂРѕР¶РґРµРЅРёСЏ",
        "РўРµР»РµС„РѕРЅ СѓС‡РµРЅРёРєР°",
        "РўРµР»РµРіСЂР°Рј СѓС‡РµРЅРёРєР°",
        "РџРѕР»",
        "РќР° РіСЂР°РЅС‚Рµ",
        "Р¤РѕСЂРјР°С‚",
        "Р“РѕСЂРѕРґ",
        "РћР±СЂР°Р·РѕРІР°С‚РµР»СЊРЅРѕРµ СѓС‡СЂРµР¶РґРµРЅРёРµ",
        "РљР»Р°СЃСЃ",
        "Email СѓС‡РµРЅРёРєР°",
        "Р¤РРћ СЂРѕРґРёС‚РµР»СЏ",
        "РўРµР»РµС„РѕРЅ СЂРѕРґРёС‚РµР»СЏ",
        "Р’С‚РѕСЂРѕР№ С‚РµР»РµС„РѕРЅ СЂРѕРґРёС‚РµР»СЏ",
        "РўРµР»РµРіСЂР°Рј СЂРѕРґРёС‚РµР»СЏ",
        "Email СЂРѕРґРёС‚РµР»СЏ",
        "РЈРґРѕР±РЅС‹Р№ РјРµСЃСЃРµРЅРґР¶РµСЂ",
        "РљРѕРјРјРµРЅС‚Р°СЂРёР№",
        "РћС‚РєСѓРґР° РїСЂРёС€РµР»",
    ]
    ws.append(headers)
    ws.append(
        [
            "РРІР°РЅРѕРІ РџРµС‚СЂ РЎРµСЂРіРµРµРІРёС‡",
            "2015-03-15",
            "+7 999 111-22-33",
            "@petr_ivanov",
            "Рј",
            "РЅРµС‚",
            "РіСЂСѓРїРїР°",
            "РњРѕСЃРєРІР°",
            "РЁРєРѕР»Р° в„–12",
            "3",
            "petr@example.com",
            "РРІР°РЅРѕРІР° РђРЅРЅР° РџРµС‚СЂРѕРІРЅР°",
            "+7 999 111-22-34",
            "+7 900 111-22-44",
            "@anna_ivanova",
            "anna@example.com",
            "Telegram",
            "Р—Р°РїРёСЃР°РЅ РЅР° РїСЂРѕР±РЅРѕРµ Р·Р°РЅСЏС‚РёРµ",
            "СЂРµРєРѕРјРµРЅРґР°С†РёСЏ",
        ]
    )
    ws.freeze_panes = "A2"
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "student_cards_import_template.xlsx"
    headers_resp = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


@router.post("/_legacy-disabled/student-cards/import-xlsx", response_model=StudentCardImportResponse)
async def import_student_cards_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="РџРѕРґРґРµСЂР¶РёРІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ С„РѕСЂРјР°С‚ .xlsx")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    wb = load_workbook(filename=BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return StudentCardImportResponse(created=0, skipped=0, errors=["РџСѓСЃС‚РѕР№ Р»РёСЃС‚"])
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    def val(row, key_variants: List[str]) -> Optional[str]:
        for key in key_variants:
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                continue
            raw = row[idx]
            if raw is None:
                continue
            txt = str(raw).strip()
            if txt:
                return txt
        return None

    def parse_date(raw_value) -> Optional[datetime]:
        if raw_value is None:
            return None
        if isinstance(raw_value, datetime):
            return raw_value.date() if hasattr(raw_value, "date") else raw_value
        text = str(raw_value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text[:10]).date()
        except (ValueError, TypeError):
            return None

    created = 0
    skipped = 0
    errors: List[str] = []
    for i, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        student_full_name = val(row, ["С„РёРѕ СѓС‡РµРЅРёРєР°", "СѓС‡РµРЅРёРє", "student_full_name"])
        if not student_full_name:
            skipped += 1
            continue
        birth_date_raw = val(row, ["РґР°С‚Р° СЂРѕР¶РґРµРЅРёСЏ", "birth_date"])
        birth_date = parse_date(birth_date_raw) if birth_date_raw else None
        student_phone = val(row, ["С‚РµР»РµС„РѕРЅ СѓС‡РµРЅРёРєР°", "student_phone"])
        telegram = val(row, ["С‚РµР»РµРіСЂР°Рј СѓС‡РµРЅРёРєР°", "telegram"])
        gender_raw = val(row, ["РїРѕР»", "gender"])
        gender = gender_raw.lower() if gender_raw and gender_raw.lower() in ("Рј", "Р¶", "m", "f", "male", "female") else (gender_raw or None)
        on_grant_raw = val(row, ["РЅР° РіСЂР°РЅС‚Рµ", "on_grant"])
        on_grant = str(on_grant_raw).strip().lower() in ("РґР°", "yes", "1", "true", "+")
        format_raw = val(row, ["С„РѕСЂРјР°С‚", "format_type"])
        format_type = None
        if format_raw:
            f = format_raw.lower()
            if "РіСЂСѓРїРї" in f or f == "group":
                format_type = "group"
            elif "РёРЅРґРёРІРёРґ" in f or f == "individual":
                format_type = "individual"
            else:
                format_type = format_raw
        city = val(row, ["РіРѕСЂРѕРґ", "city"])
        school = val(row, ["РѕР±СЂР°Р·РѕРІР°С‚РµР»СЊРЅРѕРµ СѓС‡СЂРµР¶РґРµРЅРёРµ", "С€РєРѕР»Р°", "school"])
        grade = val(row, ["РєР»Р°СЃСЃ", "grade"])
        student_email = val(row, ["email СѓС‡РµРЅРёРєР°", "student_email"])
        parent_full_name = val(row, ["С„РёРѕ СЂРѕРґРёС‚РµР»СЏ", "СЂРѕРґРёС‚РµР»СЊ", "parent_full_name"])
        parent_phone = val(row, ["С‚РµР»РµС„РѕРЅ СЂРѕРґРёС‚РµР»СЏ", "parent_phone"])
        parent_phone_2 = val(row, ["РІС‚РѕСЂРѕР№ С‚РµР»РµС„РѕРЅ СЂРѕРґРёС‚РµР»СЏ", "parent_phone_2"])
        parent_telegram = val(row, ["С‚РµР»РµРіСЂР°Рј СЂРѕРґРёС‚РµР»СЏ", "parent_telegram"])
        parent_email = val(row, ["email СЂРѕРґРёС‚РµР»СЏ", "parent_email"])
        preferred_raw = val(row, ["СѓРґРѕР±РЅС‹Р№ РјРµСЃСЃРµРЅРґР¶РµСЂ", "preferred_messenger"])
        preferred_messenger = None
        if preferred_raw:
            p = preferred_raw.lower()
            if "max" in p or p == "max":
                preferred_messenger = "max"
            elif "telegram" in p or "С‚РµР»РµРіСЂР°Рј" in p or p == "tg":
                preferred_messenger = "telegram"
            elif "sms" in p:
                preferred_messenger = "sms"
            else:
                preferred_messenger = preferred_raw
        comment = val(row, ["РєРѕРјРјРµРЅС‚Р°СЂРёР№", "comment"])
        source = val(row, ["РѕС‚РєСѓРґР° РїСЂРёС€РµР»", "РёСЃС‚РѕС‡РЅРёРє", "source"])
        abonement_id = None
        discount_type = DiscountType.NONE
        discount_value = 0.0
        if auth.resolve_effective_role(current_user) == UserRole.OWNER:
            ab_id_raw = val(row, ["Р°Р±РѕРЅРµРјРµРЅС‚", "abonement_id"])
            if ab_id_raw:
                try:
                    abonement_id = int(float(ab_id_raw))
                except (ValueError, TypeError):
                    pass
        card = StudentCard(
            student_full_name=student_full_name,
            birth_date=birth_date,
            student_phone=student_phone or None,
            phone_normalized=normalize_phone(parent_phone or student_phone or "") or None,
            telegram=telegram or None,
            gender=gender,
            on_grant=on_grant,
            format_type=format_type,
            city=city or None,
            school=school or None,
            grade=grade or None,
            parent_full_name=parent_full_name or None,
            parent_phone=parent_phone or None,
            parent_phone_2=parent_phone_2 or None,
            parent_telegram=parent_telegram or None,
            parent_email=parent_email or None,
            student_email=student_email or None,
            preferred_messenger=preferred_messenger,
            comment=comment or None,
            source=source or None,
            abonement_id=abonement_id,
            discount_type=discount_type,
            discount_value=discount_value,
            archived=False,
        )
        db.add(card)
        db.flush()
        sync_student_card_person(db, card)
        created += 1
    db.commit()
    log_action(db, current_user.id, "import", "student_card", None, {"created": created, "skipped": skipped})
    return StudentCardImportResponse(created=created, skipped=skipped, errors=errors)


def _parse_vertical_date(text: str, today: date) -> Optional[str]:
    """РџР°СЂСЃРёС‚ РґР°С‚Сѓ РёР· СЃС‚СЂРѕРєРё РІРёРґР° 'РЎРµРіРѕРґРЅСЏ, 26 С„РµРІСЂР°Р»СЏ', 'Р’С‡РµСЂР°, 25 С„РµРІСЂР°Р»СЏ', '25 С„РµРІСЂР°Р»СЏ, 19:51', '24 С„РµРІСЂР°Р»СЏ'."""
    if not text or not str(text).strip():
        return None
    s = str(text).strip()
    year = today.year
    months_ru = {
        "СЏРЅРІР°СЂСЏ": 1, "С„РµРІСЂР°Р»СЏ": 2, "РјР°СЂС‚Р°": 3, "Р°РїСЂРµР»СЏ": 4, "РјР°СЏ": 5, "РёСЋРЅСЏ": 6,
        "РёСЋР»СЏ": 7, "Р°РІРіСѓСЃС‚Р°": 8, "СЃРµРЅС‚СЏР±СЂСЏ": 9, "РѕРєС‚СЏР±СЂСЏ": 10, "РЅРѕСЏР±СЂСЏ": 11, "РґРµРєР°Р±СЂСЏ": 12,
    }
    if "СЃРµРіРѕРґРЅСЏ" in s.lower():
        return today.isoformat()
    if "РІС‡РµСЂР°" in s.lower():
        from datetime import timedelta
        return (today - timedelta(days=1)).isoformat()
    for month_name, month_num in months_ru.items():
        if month_name in s.lower():
            parts = re.findall(r"\d+", s)
            if parts:
                day = int(parts[0])
                try:
                    return date(year, month_num, day).isoformat()
                except ValueError:
                    return None
            return None
    return None


def _import_bank_transactions_vertical(rows: list, db: Session) -> dict:
    """РРјРїРѕСЂС‚ РІС‹РїРёСЃРєРё РІ РІРµСЂС‚РёРєР°Р»СЊРЅРѕРј С„РѕСЂРјР°С‚Рµ (РѕРґРЅР° РєРѕР»РѕРЅРєР°, Р±Р»РѕРєРё: РґР°С‚Р°, СЃСѓРјРјР°, СЃС‚Р°С‚СѓСЃ, РєРѕРЅС‚СЂР°РіРµРЅС‚, РѕРїРёСЃР°РЅРёРµ)."""
    lines = []
    for row in rows:
        val = row[0] if row and len(row) > 0 else None
        lines.append(str(val).strip() if val is not None else "")
    today = date.today()
    amount_re = re.compile(r"^[+\-вЂ“]\s*([\d\s,]+)\s*[в‚ЅСЂ]", re.IGNORECASE)
    imported = 0
    skipped = 0
    last_date_str = None
    for i, line in enumerate(lines):
        if not line:
            continue
        line_norm = line.replace("\xa0", " ")
        if "в‚Ѕ" not in line_norm and " СЂ" not in line_norm.lower():
            maybe_date = _parse_vertical_date(line, today)
            if maybe_date:
                last_date_str = maybe_date
            continue
        m = amount_re.match(line_norm)
        if not m:
            continue
        amount_str = m.group(1).replace(" ", "").replace(",", ".")
        try:
            amount_val = float(amount_str)
        except ValueError:
            skipped += 1
            continue
        if amount_val <= 0:
            amount_val = -abs(amount_val)
        else:
            amount_val = abs(amount_val)
        if line.strip().startswith(("-", "вЂ“")):
            amount_val = -abs(amount_val)
        date_str = last_date_str
        counterparty = ""
        if i + 1 < len(lines):
            counterparty = (lines[i + 2] or "").strip() if i + 2 < len(lines) else ""
            (lines[i + 3] or "").strip() if i + 3 < len(lines) else ""
            time_line = (lines[i + 4] or "").strip() if i + 4 < len(lines) else ""
            if time_line:
                parsed = _parse_vertical_date(time_line, today)
                if parsed:
                    date_str = parsed
        if not date_str:
            date_str = today.isoformat()
        payer_name = counterparty or "РЎРїРёСЃР°РЅРёРµ" if amount_val < 0 else "РР· РІС‹РїРёСЃРєРё (Р±РµР· Р¤РРћ)"
        payer_phone = None
        if amount_val > 0 and counterparty:
            phone_m = re.search(r"\+7\s*\(?\d{3}\)?\s*\d{3}[- ]?\d{2}[- ]?\d{2}", counterparty)
            if phone_m:
                payer_phone = normalize_phone(re.sub(r"\D", "", phone_m.group(0)))
            if "," in counterparty:
                rest = counterparty.split(",", 1)[1].strip()
                if rest and len(rest) >= 2:
                    payer_name = rest[:512]
        if amount_val > 0 and not payer_name:
            payer_name = counterparty or "РР· РІС‹РїРёСЃРєРё (Р±РµР· Р¤РРћ)"
        op_id_source = f"vertical_xlsx|{date_str}|{amount_val}|{payer_name}|{payer_phone or ''}|{i}"
        operation_id = hashlib.sha256(op_id_source.encode("utf-8")).hexdigest()
        if db.query(BankTransaction.id).filter(BankTransaction.operation_id == operation_id).first():
            skipped += 1
            continue
        status = BankTransactionStatus.EXPENSE.value if amount_val < 0 else BankTransactionStatus.NEW.value
        bt = BankTransaction(
            operation_id=operation_id,
            tochka_account_id=None,
            amount=amount_val,
            payer_phone=payer_phone,
            payer_name=(payer_name or "")[:512] or None,
            payment_date=date_str,
            status=status,
            expense_category=None,
        )
        db.add(bt)
        ensure_finance_transaction_for_bank_transaction(db, bt, bank_source="import_xlsx")
        imported += 1
    db.commit()
    errors = []
    if imported == 0 and skipped == 0 and len(lines) > 1:
        errors.append("Р’ С„Р°Р№Р»Рµ РЅРµ РЅР°Р№РґРµРЅРѕ СЃС‚СЂРѕРє СЃ СЃСѓРјРјРѕР№ РІРёРґР° В«+ 3 200 в‚ЅВ» РёР»Рё В«вЂ“ 102 в‚ЅВ».")
    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.post("/_legacy-disabled/bank-transactions/import-xlsx")
async def import_bank_transactions_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """
    РРјРїРѕСЂС‚ Р±Р°РЅРєРѕРІСЃРєРѕР№ РІС‹РїРёСЃРєРё РёР· .xlsx, РєРѕРіРґР° С„Р°Р№Р» СЃРєР°С‡Р°РЅ РІСЂСѓС‡РЅСѓСЋ (РўРѕС‡РєР° РёР»Рё РґСЂСѓРіРѕР№ Р±Р°РЅРє).
    РЎРѕР·РґР°С‘С‚ Р·Р°РїРёСЃРё РІ bank_transactions СЃРѕ СЃС‚Р°С‚СѓСЃРѕРј new; РґР°Р»СЊС€Рµ РјРµРЅРµРґР¶РµСЂ СЂР°СЃРїСЂРµРґРµР»СЏРµС‚ РёС… РїРѕ СѓС‡РµРЅРёРєР°Рј РІРѕ РІРєР»Р°РґРєРµ В«РћРїРµСЂР°С†РёРё Р±Р°РЅРєР°В».
    РџРѕРґРґРµСЂР¶РёРІР°СЋС‚СЃСЏ РґРІР° С„РѕСЂРјР°С‚Р°:
    1) РЈРЅРёРІРµСЂСЃР°Р»СЊРЅС‹Р№: РєРѕР»РѕРЅРєРё Р”Р°С‚Р°, РЎСѓРјРјР°, Р¤РРћ РїР»Р°С‚РµР»СЊС‰РёРєР°, РўРµР»РµС„РѕРЅ (СЂРµРіРёСЃС‚СЂ РЅРµ РІР°Р¶РµРЅ).
    2) Р’С‹РїРёСЃРєР° РўРѕС‡РєР° Р‘Р°РЅРє: Р”Р°С‚Р° РѕРїРµСЂР°С†РёРё/Р”Р°С‚Р° РґРѕРєСѓРјРµРЅС‚Р°, Р—Р°С‡РёСЃР»РµРЅРёРµ (РїСЂРёС…РѕРґ) РёР»Рё РЎСѓРјРјР°, РќР°Р·РЅР°С‡РµРЅРёРµ РїР»Р°С‚РµР¶Р° (РёР· РЅРµРіРѕ РёР·РІР»РµРєР°СЋС‚СЃСЏ С‚РµР»РµС„РѕРЅ Рё В«РџРѕР»СѓС‡Р°С‚РµР»СЊ вЂ¦В»/В«РџР»Р°С‚РµР»СЊС‰РёРє вЂ¦В»); РїСЂРё РѕС‚СЃСѓС‚СЃС‚РІРёРё Р¤РРћ/С‚РµР»РµС„РѕРЅР° РІ РєРѕР»РѕРЅРєР°С… РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ СЂР°Р·Р±РѕСЂ РЅР°Р·РЅР°С‡РµРЅРёСЏ.
    """
    _require_sales_admin_owner(current_user)
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="РџРѕРґРґРµСЂР¶РёРІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ С„РѕСЂРјР°С‚ .xlsx")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    wb = load_workbook(filename=BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": ["РџСѓСЃС‚РѕР№ Р»РёСЃС‚"]}
    first_row = rows[0]
    is_vertical = (
        len(first_row) == 1
        and first_row[0] is not None
        and not any(
            k in str(first_row[0]).strip().lower()
            for k in ("РґР°С‚Р°", "СЃСѓРјРјР°", "Р·Р°С‡РёСЃР»РµРЅРёРµ", "СЃРїРёСЃР°РЅРёРµ", "РєСЂРµРґРёС‚", "РґРµР±РµС‚", "РЅР°Р·РЅР°С‡РµРЅРёРµ")
        )
    )
    if is_vertical:
        return _import_bank_transactions_vertical(rows, db)
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    def col(row, keys: List[str], allow_number: bool = True):
        for key in keys:
            for name, idx in header_map.items():
                if key in name and idx < len(row):
                    raw = row[idx] if idx < len(row) else None
                    if raw is None:
                        continue
                    if allow_number and isinstance(raw, (int, float)):
                        return str(raw).strip() if raw != 0 else None
                    txt = str(raw).strip()
                    if txt:
                        return txt
        return None

    def parse_date_any(raw_value) -> Optional[str]:
        if raw_value is None:
            return None
        if isinstance(raw_value, datetime):
            try:
                d = raw_value.date() if hasattr(raw_value, "date") else raw_value
                return d.isoformat()
            except Exception:
                pass
        if isinstance(raw_value, (int, float)):
            try:
                from datetime import timedelta
                base = date(1899, 12, 30)
                d = base + timedelta(days=int(float(raw_value)))
                return d.isoformat()
            except (ValueError, TypeError, OverflowError):
                pass
        text = str(raw_value).strip()
        try:
            serial = float(text)
            if 1000 < serial < 100000:
                from datetime import timedelta
                base = date(1899, 12, 30)
                d = base + timedelta(days=int(serial))
                return d.isoformat()
        except (ValueError, TypeError):
            pass
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text[:10]).date().isoformat()
        except (ValueError, TypeError):
            return text  # СЃРѕС…СЂР°РЅСЏРµРј РєР°Рє РµСЃС‚СЊ

    def parse_amount(raw_value) -> Optional[float]:
        if raw_value is None:
            return None
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        text = str(raw_value).strip().replace(" ", "").replace("\u00a0", "")
        if not text:
            return None
        text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    def parse_payment_purpose(purpose: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
        """РР· В«РќР°Р·РЅР°С‡РµРЅРёРµ РїР»Р°С‚РµР¶Р°В» РёР·РІР»РµС‡СЊ С‚РµР»РµС„РѕРЅ Рё Р¤РРћ. РџРѕРґРґРµСЂР¶РёРІР°РµС‚:
        - В«РџРѕР»СѓС‡Р°С‚РµР»СЊ XВ» / В«РџР»Р°С‚РµР»СЊС‰РёРє XВ»;
        - С„РѕСЂРјР°С‚ РўРѕС‡РєР°/РЎР‘Рџ: В«+7 (950) 112-78-38 Р”РјРёС‚СЂРёР№ РђРЅРґСЂРµРµРІРёС‡ Рџ. Р—Р°РєР°Р· 767: вЂ¦В» (Р¤РРћ СЃСЂР°Р·Сѓ РїРѕСЃР»Рµ С‚РµР»РµС„РѕРЅР°).
        Р’РѕР·РІСЂР°С‰Р°РµС‚ (payer_name, payer_phone_raw)."""
        if not purpose or not str(purpose).strip():
            return None, None
        text = str(purpose).strip()
        phone_raw = None
        phone_end = 0
        for pattern in (
            r"\+7\s*\(?\d{3}\)?\s*\d{3}[- ]?\d{2}[- ]?\d{2}",
            r"8\s*\(?\d{3}\)?\s*\d{3}[- ]?\d{2}[- ]?\d{2}",
            r"\+7\d{10}",
            r"8\d{10}",
        ):
            m = re.search(pattern, text)
            if m:
                phone_raw = re.sub(r"\D", "", m.group(0))
                if phone_raw.startswith("8") and len(phone_raw) == 11:
                    phone_raw = "7" + phone_raw[1:]
                elif phone_raw.startswith("7") and len(phone_raw) == 11:
                    pass
                elif len(phone_raw) == 10:
                    phone_raw = "7" + phone_raw
                else:
                    phone_raw = None
                if phone_raw:
                    phone_end = m.end()
                    break
        name = None
        for prefix in ("РџРѕР»СѓС‡Р°С‚РµР»СЊ ", "РџР»Р°С‚РµР»СЊС‰РёРє "):
            if prefix in text:
                start = text.find(prefix) + len(prefix)
                end = text.find(" С‡РµСЂРµР·", start)
                if end == -1:
                    end = text.find(".", start)
                if end == -1:
                    end = len(text)
                name = text[start:end].strip()
                if name:
                    break
        if not name and phone_end > 0:
            rest = text[phone_end:].strip()
            for stop in ("Р—Р°РєР°Р· ", "Order ", " Р·Р°РєР°Р· ", " в„–", " N "):
                if stop in rest:
                    idx = rest.find(stop)
                    candidate = rest[:idx].strip()
                    if candidate and len(candidate) >= 3 and re.search(r"[\u0400-\u04FF]", candidate):
                        name = candidate
                        break
            if not name and rest and re.search(r"[\u0400-\u04FF]", rest):
                name = rest[:80].strip()
        return name or None, phone_raw

    imported = 0
    skipped = 0
    for idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        date_str = parse_date_any(col(row, ["РґР°С‚Р°", "date"]))
        if not date_str:
            for name, idx in header_map.items():
                if "РґР°С‚Р°" in name and idx < len(row) and row[idx] is not None:
                    date_str = parse_date_any(row[idx])
                    if date_str:
                        break
        amount_raw = col(row, ["СЃСѓРјРјР°", "amount", "Р·Р°С‡РёСЃР»РµРЅРёРµ", "РєСЂРµРґРёС‚"])
        amount = parse_amount(amount_raw)
        if amount is None or amount <= 0:
            amount = parse_amount(col(row, ["СЃРїРёСЃР°РЅРёРµ", "РґРµР±РµС‚"]))
            if amount is not None and amount > 0:
                amount = -amount
        payer_name = col(row, ["С„РёРѕ", "РїР»Р°С‚РµР»СЊС‰РёРє", "payer"])
        payer_phone_raw = col(row, ["С‚РµР»РµС„РѕРЅ", "phone"])
        purpose = col(row, ["РЅР°Р·РЅР°С‡РµРЅРёРµ", "payment purpose", "РЅР°Р·РЅР°С‡РµРЅРёРµ РїР»Р°С‚РµР¶Р°"])
        if (not payer_name or not payer_phone_raw) and purpose:
            name_from_purpose, phone_from_purpose = parse_payment_purpose(purpose)
            if not payer_name and name_from_purpose:
                payer_name = name_from_purpose
            if not payer_phone_raw and phone_from_purpose:
                payer_phone_raw = phone_from_purpose
        if not payer_name and amount and amount > 0 and date_str:
            payer_name = "РР· РІС‹РїРёСЃРєРё (Р±РµР· Р¤РРћ)"
        if amount is None or not date_str:
            skipped += 1
            continue
        is_expense = amount < 0
        if is_expense:
            payer_name = payer_name or col(row, ["РєРѕРЅС‚СЂР°РіРµРЅС‚", "counterparty"]) or (purpose[:512] if purpose else "РЎРїРёСЃР°РЅРёРµ")
        else:
            if not payer_name:
                skipped += 1
                continue

        payer_phone = (normalize_phone(payer_phone_raw or "") or None) if not is_expense else None
        op_id_source = f"manual_xlsx|{date_str}|{amount}|{payer_name}|{payer_phone}|{idx}"
        operation_id = hashlib.sha256(op_id_source.encode("utf-8")).hexdigest()

        exists = db.query(BankTransaction.id).filter(BankTransaction.operation_id == operation_id).first()
        if exists:
            skipped += 1
            continue

        bt = BankTransaction(
            operation_id=operation_id,
            tochka_account_id=None,
            amount=amount,
            payer_phone=payer_phone or None if not is_expense else None,
            payer_name=(payer_name or "")[:512] or None,
            payment_date=date_str,
            status=BankTransactionStatus.EXPENSE.value if is_expense else BankTransactionStatus.NEW.value,
            expense_category=None,
        )
        db.add(bt)
        imported += 1

    db.commit()
    errors: List[str] = []
    if imported == 0 and skipped > 0:
        errors.append(
            "РќРё РѕРґРЅР° СЃС‚СЂРѕРєР° РЅРµ РїРѕРґРѕС€Р»Р°. РџСЂРѕРІРµСЂСЊС‚Рµ: РІ РїРµСЂРІРѕР№ СЃС‚СЂРѕРєРµ вЂ” Р·Р°РіРѕР»РѕРІРєРё (Р”Р°С‚Р°, Р—Р°С‡РёСЃР»РµРЅРёРµ/РЎРїРёСЃР°РЅРёРµ РёР»Рё РљСЂРµРґРёС‚/Р”РµР±РµС‚, РќР°Р·РЅР°С‡РµРЅРёРµ/РљРѕРЅС‚СЂР°РіРµРЅС‚); РґР°С‚С‹ РІ С„РѕСЂРјР°С‚Рµ Р”Р”.РњРњ.Р“Р“Р“Р“ РёР»Рё С‡РёСЃР»Рѕ Excel; СЃСѓРјРјС‹ вЂ” С‡РёСЃР»Р° СЃ Р·Р°РїСЏС‚РѕР№ РёР»Рё С‚РѕС‡РєРѕР№."
        )
    return {"imported": imported, "skipped": skipped, "errors": errors}

@router.get("/_legacy-disabled/student-cards/{card_id}", response_model=StudentCardResponse)
async def get_student_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="РљР°СЂС‚РѕС‡РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    return _student_card_response(card, current_user, db)


@router.put("/_legacy-disabled/student-cards/{card_id}", response_model=StudentCardResponse)
async def update_student_card(
    card_id: int,
    payload: StudentCardUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="РљР°СЂС‚РѕС‡РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    data = payload.model_dump(exclude_unset=True)
    # РђР±РѕРЅРµРјРµРЅС‚ Рё СЃРєРёРґРєСѓ РјРѕР¶РµС‚ РјРµРЅСЏС‚СЊ С‚РѕР»СЊРєРѕ owner
    if auth.resolve_effective_role(current_user) != UserRole.OWNER:
        data.pop("abonement_id", None)
        data.pop("discount_type", None)
        data.pop("discount_value", None)
    # payment_link РјРѕР¶РµС‚ РјРµРЅСЏС‚СЊ С‚РѕР»СЊРєРѕ owner Рё admin (sales РЅРµ РјРѕР¶РµС‚)
    if auth.resolve_effective_role(current_user) not in (UserRole.OWNER, UserRole.ADMIN):
        data.pop("payment_link", None)
    if data.get("abonement_id"):
        ab = db.query(Abonement).filter(Abonement.id == data["abonement_id"]).first()
        if not ab:
            raise HTTPException(status_code=404, detail="РђР±РѕРЅРµРјРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    if "student_id" in data and data["student_id"] is not None:
        st = db.query(Student).filter(Student.id == data["student_id"]).first()
        if not st:
            raise HTTPException(status_code=404, detail="РЈС‡РµРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    if any(field in data for field in ("parent_phone", "student_phone")):
        data["phone_normalized"] = normalize_phone(
            data.get("parent_phone", card.parent_phone) or data.get("student_phone", card.student_phone) or ""
        ) or None
    for k, v in data.items():
        setattr(card, k, v)
    sync_student_card_person(db, card)
    db.commit()
    db.refresh(card)
    return _student_card_response(card, current_user, db)


@router.post("/_legacy-disabled/student-cards/{card_id}/convert", response_model=AnketaConvertResponse)
async def convert_anketa_to_student(
    card_id: int,
    payload: Optional[AnketaConvertRequest] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """
    РљРѕРЅРІРµСЂСЃРёСЏ Р°РЅРєРµС‚С‹ РІ СѓС‡РµРЅРёРєР°: СЃРѕР·РґР°С‘С‚/РїСЂРёРІСЏР·С‹РІР°РµС‚ Student, РЅРµ СЃРѕР·РґР°С‘С‚ РєР°Р±РёРЅРµС‚ СЂРѕРґРёС‚РµР»СЏ.
    РџСЂРё РґСѓР±Р»СЏС… РїРѕ email СЂРѕРґРёС‚РµР»СЏ РёР»Рё Р¤РРћ СѓС‡РµРЅРёРєР° РІРѕР·РІСЂР°С‰Р°РµС‚ 409 СЃ РІС‹Р±РѕСЂРѕРј РїСЂРёРІСЏР·РєРё Рє СЃСѓС‰РµСЃС‚РІСѓСЋС‰РµРјСѓ.
    """
    _require_sales_admin_owner(current_user)
    body = payload or AnketaConvertRequest()
    try:
        result = student_card_convert(
            db,
            card_id,
            use_existing_parent_id=body.use_existing_parent_id,
            use_existing_student_id=body.use_existing_student_id,
        )
    except StudentCardConvertConflict as e:
        code = e.detail.get("code", "existing_parent")
        raise HTTPException(
            status_code=409,
            detail=e.detail,
            headers={"X-Conflict-Code": code},
        )
    except ValueError as e:
        msg = str(e)
        if "РЅРµ РЅР°Р№РґРµРЅ" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)
    return AnketaConvertResponse(
        card=_student_card_response(result.card, current_user, db),
        student_id=result.student_id,
    )


@router.post("/_legacy-disabled/student-cards/{card_id}/archive")
async def archive_student_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="РљР°СЂС‚РѕС‡РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    card.archived = True
    db.commit()
    return {"archived": True}


@router.post("/_legacy-disabled/student-cards/{card_id}/unarchive")
async def unarchive_student_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="РљР°СЂС‚РѕС‡РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    card.archived = False
    db.commit()
    return {"archived": False}


@router.post("/_legacy-disabled/student-cards/{card_id}/open-parent-cabinet", response_model=OpenParentCabinetResponse)
async def open_parent_cabinet_from_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """
    РљР°СЂС‚РѕС‡РєР° РІРѕ РіР»Р°РІРµ: РїРѕ РєР°СЂС‚РѕС‡РєРµ СЃРѕР·РґР°С‘Рј/РїСЂРёРІСЏР·С‹РІР°РµРј СѓС‡РµРЅРёРєР° Рё СЂРѕРґРёС‚РµР»СЏ, РѕС‚РєСЂС‹РІР°РµРј РєР°Р±РёРЅРµС‚.
    Р•СЃР»Рё Сѓ РєР°СЂС‚РѕС‡РєРё РЅРµС‚ student_id вЂ” СЃРѕР·РґР°С‘С‚СЃСЏ Student РёР· Р¤РРћ РєР°СЂС‚РѕС‡РєРё.
    Р•СЃР»Рё Сѓ СѓС‡РµРЅРёРєР° РЅРµС‚ parent_id вЂ” СЃРѕР·РґР°С‘С‚СЃСЏ РёР»Рё РЅР°С…РѕРґРёС‚СЃСЏ СЂРѕРґРёС‚РµР»СЊ РїРѕ email РєР°СЂС‚РѕС‡РєРё, РїСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё РѕС‚РїСЂР°РІР»СЏРµС‚СЃСЏ РїСЂРёРіР»Р°С€РµРЅРёРµ.
    Р‘Р°Р·Р° (РѕС†РµРЅРєРё, С…Р°СЂР°РєС‚РµСЂРёСЃС‚РёРєРё) РЅРµ РјРµРЅСЏРµС‚СЃСЏ.
    """
    _require_sales_admin_owner(current_user)
    card = db.query(StudentCard).filter(StudentCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="РљР°СЂС‚РѕС‡РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    parent_email = (getattr(card, "parent_email", None) or "").strip().lower()
    if not parent_email:
        raise HTTPException(
            status_code=400,
            detail="РЈРєР°Р¶РёС‚Рµ email СЂРѕРґРёС‚РµР»СЏ РІ РєР°СЂС‚РѕС‡РєРµ (РїРѕР»Рµ В«Email СЂРѕРґРёС‚РµР»СЏВ»), С‡С‚РѕР±С‹ РѕС‚РєСЂС‹С‚СЊ РєР°Р±РёРЅРµС‚.",
        )
    parent_full_name = (getattr(card, "parent_full_name", None) or "").strip() or "Р РѕРґРёС‚РµР»СЊ"

    if not getattr(card, "student_id", None):
        student = Student(
            full_name=(card.student_full_name or "").strip() or "РЈС‡РµРЅРёРє",
            status=StudentStatus.ACTIVE,
        )
        db.add(student)
        db.flush()
        card.student_id = student.id
        db.add(card)
        db.flush()
    else:
        student = db.query(Student).filter(Student.id == card.student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="РЈС‡РµРЅРёРє РїРѕ РєР°СЂС‚РѕС‡РєРµ РЅРµ РЅР°Р№РґРµРЅ")

    if student.parent_id:
        db.commit()
        return OpenParentCabinetResponse(
            already_open=True,
            student_id=student.id,
            parent_id=student.parent_id,
        )

    parent_user = db.query(User).filter(
        User.email == parent_email,
        User.role == UserRole.PARENT,
    ).first()
    if parent_user:
        student.parent_id = parent_user.id
        db.add(student)
        db.commit()
        return OpenParentCabinetResponse(
            already_open=False,
            student_id=student.id,
            parent_id=parent_user.id,
        )

    try:
        parent_user, invite_link = create_parent_with_invite(db, parent_email, parent_full_name)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    student.parent_id = parent_user.id
    db.add(student)
    db.commit()
    log_action(db, current_user.id, "open_parent_cabinet", "student_card", card_id, {"student_id": student.id, "parent_id": parent_user.id})
    return OpenParentCabinetResponse(
        already_open=False,
        student_id=student.id,
        parent_id=parent_user.id,
        invite_link=invite_link,
    )


@router.get("/_legacy-disabled/students-for-cards", response_model=List[Dict])
async def list_students_for_cards(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РЎРїРёСЃРѕРє СѓС‡РµРЅРёРєРѕРІ (id, full_name) РґР»СЏ РїСЂРёРІСЏР·РєРё Рє РєР°СЂС‚РѕС‡РєРµ. Р¤РРћ РёР· РєР°СЂС‚РѕС‡РєРё, РµСЃР»Рё РїСЂРёРІСЏР·Р°РЅР°."""
    _require_sales_admin_owner(current_user)
    students = db.query(Student).filter(Student.status == StudentStatus.ACTIVE).order_by(Student.full_name).all()
    if not students:
        return []
    display_names = get_students_display_names(db, [s.id for s in students])
    return [{"id": s.id, "full_name": display_names.get(s.id, s.full_name)} for s in students]


def _get_student_program_name(db: Session, student_id: int, fallback_group_id: Optional[int] = None) -> Optional[str]:
    """РџСЂРѕРіСЂР°РјРјР° СѓС‡РµРЅРёРєР° (РёСЃС‚РѕС‡РЅРёРє РїСЂР°РІРґС‹ РґР»СЏ РѕС‚СЂР°Р±РѕС‚РѕРє). РЎРЅР°С‡Р°Р»Р° student_programs, РёРЅР°С‡Рµ РїСЂРѕРіСЂР°РјРјР° РіСЂСѓРїРїС‹."""
    sp = db.query(StudentProgram).filter(
        StudentProgram.student_id == student_id,
        StudentProgram.status == "active",
    ).first()
    if sp and sp.program:
        return sp.program.name
    if fallback_group_id:
        gp = db.query(GroupProgram).filter(GroupProgram.group_id == fallback_group_id).first()
        if gp and gp.program:
            return gp.program.name
    return None


def _absence_to_response(db: Session, a: AbsenceFollowUp) -> AbsenceFollowUpResponse:
    student = db.query(Student).filter(Student.id == a.student_id).first()
    group = db.query(Group).filter(Group.id == a.group_id).first()
    makeup_group = db.query(Group).filter(Group.id == a.makeup_group_id).first() if a.makeup_group_id else None
    makeup_custom_lesson = None
    if getattr(a, "makeup_custom_lesson_id", None):
        makeup_custom_lesson = db.query(CustomLesson).filter(CustomLesson.id == a.makeup_custom_lesson_id).first()
    return AbsenceFollowUpResponse(
        id=a.id,
        lesson_attendance_id=a.lesson_attendance_id,
        student_id=a.student_id,
        group_id=a.group_id,
        lesson_date=a.lesson_date,
        stage=a.stage,
        absence_reason=getattr(a, "absence_reason", None),
        absence_comment=getattr(a, "absence_comment", None),
        makeup_group_id=getattr(a, "makeup_group_id", None),
        makeup_lesson_date=getattr(a, "makeup_lesson_date", None),
        makeup_custom_lesson_id=getattr(a, "makeup_custom_lesson_id", None),
        makeup_custom_lesson_title=makeup_custom_lesson.title if makeup_custom_lesson else None,
        created_at=a.created_at,
        updated_at=a.updated_at,
        student_name=get_student_display_name(db, student) if student else None,
        group_name=group.name if group else None,
        program_name=_get_student_program_name(db, a.student_id, a.group_id),
        makeup_group_name=makeup_group.name if makeup_group else None,
    )


@router.get("/_legacy-disabled/absences", response_model=List[AbsenceFollowUpResponse])
async def list_absences(
    stage: Optional[str] = Query(None, description="Р¤РёР»СЊС‚СЂ РїРѕ СЌС‚Р°РїСѓ: missed, assigned, made_up, missed_makeup"),
    student_id: Optional[int] = Query(None, description="Р¤РёР»СЊС‚СЂ РїРѕ СѓС‡РµРЅРёРєСѓ"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    query = db.query(AbsenceFollowUp).order_by(AbsenceFollowUp.lesson_date.desc(), AbsenceFollowUp.id.desc())
    if stage:
        query = query.filter(AbsenceFollowUp.stage == stage)
    if student_id is not None:
        query = query.filter(AbsenceFollowUp.student_id == student_id)
    items = query.all()
    return [_absence_to_response(db, a) for a in items]


@router.patch("/_legacy-disabled/absences/{absence_id}", response_model=AbsenceFollowUpResponse)
async def update_absence_stage(
    absence_id: int,
    payload: AbsenceFollowUpStageUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    valid_stages = ("missed", "assigned", "link_sent", "made_up", "missed_makeup")
    if payload.stage not in valid_stages:
        raise HTTPException(status_code=400, detail=f"stage РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РѕРґРёРЅ РёР·: {valid_stages}")
    absence = db.query(AbsenceFollowUp).filter(AbsenceFollowUp.id == absence_id).first()
    if not absence:
        raise HTTPException(status_code=404, detail="РџСЂРѕРїСѓСЃРє РЅРµ РЅР°Р№РґРµРЅ")
    absence.stage = payload.stage
    if payload.stage == "missed_makeup":
        absence.makeup_group_id = None
        absence.makeup_lesson_date = None
    db.commit()
    db.refresh(absence)
    return _absence_to_response(db, absence)


@router.post("/_legacy-disabled/absences/{absence_id}/assign-makeup", response_model=AbsenceFollowUpResponse)
async def assign_makeup(
    absence_id: int,
    payload: AbsenceMakeupAssign,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РќР°Р·РЅР°С‡РёС‚СЊ РѕС‚СЂР°Р±РѕС‚РєСѓ РЅР° РіСЂСѓРїРїСѓ Рё РґР°С‚Сѓ Р·Р°РЅСЏС‚РёСЏ (РўР— Рї.5.5)."""
    _require_sales_admin_owner(current_user)
    try:
        result = absence_makeup_assign(
            db,
            absence_id,
            makeup_group_id=payload.makeup_group_id,
            makeup_lesson_date=payload.makeup_lesson_date,
        )
    except ValueError as e:
        msg = str(e)
        if "РЅРµ РЅР°Р№РґРµРЅ" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)
    absence = result.absence
    log_student_activity(
        db,
        student_id=absence.student_id,
        activity_type="makeup_scheduled",
        title="РќР°Р·РЅР°С‡РµРЅР° РѕС‚СЂР°Р±РѕС‚РєР°",
        description=f"Р”Р°С‚Р°: {absence.makeup_lesson_date}" if getattr(absence, "makeup_lesson_date", None) else "РќР°Р·РЅР°С‡РµРЅР° РѕС‚СЂР°Р±РѕС‚РєР°",
        created_by=current_user.id,
        payload_json={"absence_id": absence.id, "makeup_group_id": absence.makeup_group_id, "makeup_lesson_date": str(getattr(absence, "makeup_lesson_date", None) or "")},
    )
    db.commit()
    db.refresh(absence)
    return _absence_to_response(db, absence)


@router.get("/_legacy-disabled/absences/{absence_id}/suggest-makeups", response_model=List[MakeupSuggestionItem])
async def suggest_makeups(
    absence_id: int,
    days_ahead: int = Query(30, ge=7, le=60),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РџРѕРґР±РѕСЂ РІР°СЂРёР°РЅС‚РѕРІ РѕС‚СЂР°Р±РѕС‚РєРё РїРѕ СЃРѕРІРјРµСЃС‚РёРјРѕСЃС‚Рё РїСЂРѕРіСЂР°РјРј (РўР— Рї.5). РћРєРЅРѕ 14вЂ“30 РґРЅРµР№ РїРѕ СѓРјРѕР»С‡Р°РЅРёСЋ 30."""
    _require_sales_admin_owner(current_user)
    absence = db.query(AbsenceFollowUp).filter(AbsenceFollowUp.id == absence_id).first()
    if not absence:
        raise HTTPException(status_code=404, detail="РџСЂРѕРїСѓСЃРє РЅРµ РЅР°Р№РґРµРЅ")
    student_id = absence.student_id
    today = date.today()
    end_date = today + timedelta(days=days_ahead)
    student_programs = db.query(StudentProgram).filter(
        StudentProgram.student_id == student_id,
        StudentProgram.status == "active",
    ).all()
    source_program_ids = [sp.program_id for sp in student_programs if sp.program_id]
    if not source_program_ids:
        group_prog = db.query(GroupProgram).filter(GroupProgram.group_id == absence.group_id).first()
        if group_prog:
            source_program_ids = [group_prog.program_id]
    allowed_target_ids = set()
    for pid in source_program_ids:
        compats = db.query(ProgramMakeupCompatibility).filter(
            ProgramMakeupCompatibility.source_program_id == pid,
        ).all()
        for c in compats:
            allowed_target_ids.add(c.target_program_id)
    if not allowed_target_ids and source_program_ids:
        allowed_target_ids = set(source_program_ids)
    group_ids = list({
        row[0] for row in db.query(GroupProgram.group_id).filter(
            GroupProgram.program_id.in_(allowed_target_ids),
        ).distinct().all()
    }) if allowed_target_ids else []
    groups_active = db.query(Group).filter(Group.id.in_(group_ids), Group.status == "active").all() if group_ids else []
    # РСЃРєР»СЋС‡Р°РµРј РёРЅРґРёРІРёРґСѓР°Р»СЊРЅС‹Рµ РіСЂСѓРїРїС‹ РёР· РІР°СЂРёР°РЅС‚РѕРІ РѕС‚СЂР°Р±РѕС‚РєРё (РўР—)
    groups_active = [g for g in groups_active if "РёРЅРґРёРІРёРґ" not in (g.name or "").lower()]
    group_ids = [g.id for g in groups_active]
    slots = []
    for sched in db.query(GroupSchedule).filter(GroupSchedule.group_id.in_(group_ids)).all():
        for d in range((end_date - today).days + 1):
            lesson_date = today + timedelta(days=d)
            if lesson_date.weekday() == sched.day_of_week:
                slots.append((sched.group_id, lesson_date, sched.start_time))
    result = []
    seen = set()
    for group_id, lesson_date, start_time in sorted(slots, key=lambda x: (x[1], x[2] or dt_time(0, 0))):
        if (group_id, lesson_date) in seen:
            continue
        seen.add((group_id, lesson_date))
        group = next((g for g in groups_active if g.id == group_id), None) or db.query(Group).filter(Group.id == group_id).first()
        if not group:
            continue
        gp = db.query(GroupProgram).filter(GroupProgram.group_id == group_id).first()
        program_name = gp.program.name if gp and gp.program else None
        result.append(MakeupSuggestionItem(
            group_id=group_id,
            group_name=group.name,
            program_name=program_name,
            lesson_date=lesson_date,
            day_of_week=lesson_date.weekday(),
            start_time=start_time.strftime("%H:%M") if start_time else None,
        ))
    return result[:50]


@router.get("/_legacy-disabled/public/makeup-selection", response_model=PublicMakeupSlotsResponse)
async def get_public_makeup_selection(
    token: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
):
    try:
        absence = resolve_absence_by_token(db, token)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    student = db.query(Student).filter(Student.id == absence.student_id).first()
    group = db.query(Group).filter(Group.id == absence.group_id).first()
    return PublicMakeupSlotsResponse(
        absence_id=absence.id,
        student_id=absence.student_id,
        student_name=get_student_display_name(db, student) if student else None,
        original_group_name=group.name if group else None,
        missed_lesson_date=absence.lesson_date,
        available_slots=list_makeup_suggestions_for_absence(db, absence),
    )


@router.post("/_legacy-disabled/public/makeup-selection/confirm", response_model=AbsenceFollowUpResponse)
async def confirm_public_makeup_selection(
    payload: PublicMakeupSelectionRequest,
    db: Session = Depends(get_db),
):
    try:
        absence = resolve_absence_by_token(db, payload.token)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    try:
        result = absence_makeup_assign(
            db,
            absence.id,
            makeup_group_id=payload.makeup_group_id,
            makeup_lesson_date=payload.makeup_lesson_date,
        )
    except ValueError as exc:
        message = str(exc)
        if "РЅРµ РЅР°Р№РґРµРЅ" in message.lower():
            raise HTTPException(status_code=404, detail=message)
        raise HTTPException(status_code=400, detail=message)

    absence = result.absence
    log_student_activity(
        db,
        student_id=absence.student_id,
        activity_type="makeup_scheduled",
        title="РќР°Р·РЅР°С‡РµРЅР° РѕС‚СЂР°Р±РѕС‚РєР°",
        description=f"Р”Р°С‚Р°: {absence.makeup_lesson_date}" if getattr(absence, "makeup_lesson_date", None) else "РќР°Р·РЅР°С‡РµРЅР° РѕС‚СЂР°Р±РѕС‚РєР°",
        created_by=None,
        payload_json={"absence_id": absence.id, "makeup_group_id": absence.makeup_group_id, "makeup_lesson_date": str(getattr(absence, "makeup_lesson_date", None) or "")},
    )
    selected_group = db.query(Group).filter(Group.id == absence.makeup_group_id).first()
    create_sales_confirmation_task(
        db,
        absence=absence,
        selected_group_name=selected_group.name if selected_group else f"#{absence.makeup_group_id}",
        created_by_id=None,
    )
    close_send_link_tasks_for_absence(db, absence_id=absence.id)
    db.commit()
    db.refresh(absence)
    return _absence_to_response(db, absence)


def _require_owner(user: User) -> None:
    """РўРѕР»СЊРєРѕ owner (РўР—: Р·Р°РјРѕСЂРѕР·РєР°, Р·Р°РєСЂС‹С‚РёРµ РїРѕ С„Р°РєС‚Сѓ)."""
    if auth.resolve_effective_role(user) != UserRole.OWNER:
        raise HTTPException(status_code=403, detail="РўРѕР»СЊРєРѕ owner")


def _require_owner_or_admin_settings(user: User) -> None:
    """Owner РёР»Рё admin (РЅР°СЃС‚СЂРѕР№РєРё РѕС‚СЂР°Р±РѕС‚РѕРє, Р±РµР· lead)."""
    if auth.resolve_effective_role(user) not in (UserRole.OWNER, UserRole.ADMIN):
        raise HTTPException(status_code=403, detail="РўРѕР»СЊРєРѕ owner РёР»Рё admin")


@router.get("/_legacy-disabled/payment-status", response_model=List[PaymentStatusItem])
async def list_payment_status(
    status_filter: Optional[str] = Query(None, description="overdue | due_soon | ok | РІСЃРµ"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Р Р°Р·РґРµР» В«Р”РѕР»РіРёВ» РґР»СЏ РјРµРЅРµРґР¶РµСЂР° (Рї.8.2, 12.2): СѓС‡РµРЅРёРєРё СЃ РґР°С‚РѕР№ СЃР»РµРґСѓСЋС‰РµР№ РѕРїР»Р°С‚С‹ Рё СЃС‚Р°С‚СѓСЃРѕРј."""
    items = payment_status_list_svc(db, status_filter=status_filter)
    return [PaymentStatusItem(**item) for item in items]


@router.get("/_legacy-disabled/payment-status-summary", response_model=PaymentStatusSummary)
async def get_payment_status_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """РЎРІРѕРґРєР° РїРѕ РїСЂРѕСЃСЂРѕС‡РєР°Рј: С‡РёСЃР»Рѕ СѓС‡РµРЅРёРєРѕРІ СЃ РґРѕР»РіРѕРј 3+ Рё 10+ РґРЅРµР№ (РґР»СЏ KPI РЅР° СЃС‚СЂР°РЅРёС†Рµ В«Р”РѕР»РіРёВ»)."""
    data = payment_status_summary_svc(db)
    return PaymentStatusSummary(**data)


# --- Custom (manual) lessons for sales/admin/owner ---


def _custom_lesson_to_response(db: Session, lesson: CustomLesson) -> CustomLessonResponse:
    trainer = db.query(User).filter(User.id == lesson.trainer_id).first()
    students_rows = (
        db.query(CustomLessonStudent, Student)
        .join(Student, Student.id == CustomLessonStudent.student_id)
        .filter(CustomLessonStudent.lesson_id == lesson.id)
        .all()
    )
    students = []
    for cls, s in students_rows:
        students.append(
            {
                "id": cls.id,
                "student_id": cls.student_id,
                "student_name": get_student_display_name(db, s) if s else None,
                "planned_absence_id": cls.planned_absence_id,
                "attended": bool(getattr(cls, "attended", False)),
                "absence_reason": getattr(cls, "absence_reason", None),
                "absence_comment": getattr(cls, "absence_comment", None),
            }
        )
    return CustomLessonResponse(
        id=lesson.id,
        title=lesson.title,
        lesson_date=lesson.lesson_date,
        start_time=_serialize_time_for_api(lesson.start_time) or "",
        end_time=_serialize_time_for_api(lesson.end_time),
        trainer_id=lesson.trainer_id,
        trainer_name=trainer.full_name if trainer else None,
        lesson_type=lesson.lesson_type.value if isinstance(lesson.lesson_type, CustomLessonType) else str(lesson.lesson_type),
        comment=lesson.comment,
        students=students,
    )


@router.post("/_legacy-disabled/custom-lessons", response_model=CustomLessonResponse, status_code=status.HTTP_201_CREATED)
async def create_custom_lesson(
    payload: CustomLessonCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("lessons.manage")),
):
    """РЎРѕР·РґР°С‚СЊ СЂСѓС‡РЅРѕР№ СѓСЂРѕРє Р±РµР· РіСЂСѓРїРїС‹ (РѕС‚СЂР°Р±РѕС‚РєР° / РґРѕРї.СѓСЂРѕРє / РїСЂРѕР±РЅРѕРµ). РўРѕР»СЊРєРѕ admin/owner/sales."""
    try:
        start_t = datetime.strptime(payload.start_time.strip(), "%H:%M").time()
        end_t = None
        if payload.end_time:
            end_t = datetime.strptime(payload.end_time.strip(), "%H:%M").time()
    except ValueError:
        raise HTTPException(status_code=400, detail="Р’СЂРµРјСЏ СѓРєР°Р¶РёС‚Рµ РІ С„РѕСЂРјР°С‚Рµ HH:MM")

    lesson_type_value = payload.lesson_type or "makeup"
    if lesson_type_value not in {t.value for t in CustomLessonType}:
        raise HTTPException(status_code=400, detail=f"lesson_type РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РѕРґРЅРёРј РёР·: {[t.value for t in CustomLessonType]}")

    students_tuples = [
        (item.student_id, item.planned_absence_id)
        for item in (payload.students or [])
    ]
    try:
        result = manual_lesson_create(
            db,
            title=payload.title,
            lesson_date=payload.lesson_date,
            start_time=start_t,
            end_time=end_t,
            trainer_id=payload.trainer_id,
            lesson_type=lesson_type_value,
            comment=payload.comment,
            students=students_tuples,
            created_by_id=current_user.id,
        )
    except ValueError as e:
        msg = str(e)
        if "РЅРµ РЅР°Р№РґРµРЅ" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)
    return _custom_lesson_to_response(db, result.lesson)


@router.get("/_legacy-disabled/custom-lessons", response_model=List[CustomLessonResponse])
async def list_custom_lessons(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    trainer_id: Optional[int] = Query(None),
    student_id: Optional[int] = Query(None),
    lesson_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("lessons.access")),
):
    """РЎРїРёСЃРѕРє СЂСѓС‡РЅС‹С… СѓСЂРѕРєРѕРІ РґР»СЏ РјРµРЅРµРґР¶РµСЂР° (admin/owner/sales)."""
    query = db.query(CustomLesson)
    if date_from:
        query = query.filter(CustomLesson.lesson_date >= date_from)
    if date_to:
        query = query.filter(CustomLesson.lesson_date <= date_to)
    if trainer_id:
        query = query.filter(CustomLesson.trainer_id == trainer_id)
    if lesson_type:
        query = query.filter(CustomLesson.lesson_type == lesson_type)
    lessons = query.order_by(CustomLesson.lesson_date.desc(), CustomLesson.start_time.asc()).all()

    # Р¤РёР»СЊС‚СЂ РїРѕ СѓС‡РµРЅРёРєСѓ С‡РµСЂРµР· СЃРІСЏР·РєСѓ
    if student_id is not None:
        lesson_ids = {
            ls.lesson_id
            for ls in db.query(CustomLessonStudent).filter(CustomLessonStudent.student_id == student_id).all()
        }
        lessons = [lesson for lesson in lessons if lesson.id in lesson_ids]

    return [_custom_lesson_to_response(db, lesson) for lesson in lessons]


@router.get("/_legacy-disabled/custom-lessons/{lesson_id}", response_model=CustomLessonResponse)
async def get_custom_lesson(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("lessons.access")),
):
    lesson = db.query(CustomLesson).filter(CustomLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Р СѓС‡РЅРѕР№ СѓСЂРѕРє РЅРµ РЅР°Р№РґРµРЅ")
    return _custom_lesson_to_response(db, lesson)


@router.put("/_legacy-disabled/custom-lessons/{lesson_id}", response_model=CustomLessonResponse)
async def update_custom_lesson(
    lesson_id: int,
    payload: CustomLessonUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("lessons.manage")),
):
    """Р РµРґР°РєС‚РёСЂРѕРІР°РЅРёРµ СЂСѓС‡РЅРѕРіРѕ СѓСЂРѕРєР°. РўРѕР»СЊРєРѕ admin/owner/sales."""
    lesson = db.query(CustomLesson).filter(CustomLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Р СѓС‡РЅРѕР№ СѓСЂРѕРє РЅРµ РЅР°Р№РґРµРЅ")

    if payload.title is not None:
        lesson.title = payload.title.strip()
    if payload.lesson_date is not None:
        lesson.lesson_date = payload.lesson_date
    if payload.start_time is not None:
        try:
            lesson.start_time = datetime.strptime(payload.start_time.strip(), "%H:%M").time()
        except ValueError:
            raise HTTPException(status_code=400, detail="start_time: С„РѕСЂРјР°С‚ HH:MM")
    if payload.end_time is not None:
        if payload.end_time == "":
            lesson.end_time = None
        else:
            try:
                lesson.end_time = datetime.strptime(payload.end_time.strip(), "%H:%M").time()
            except ValueError:
                raise HTTPException(status_code=400, detail="end_time: С„РѕСЂРјР°С‚ HH:MM")
    if payload.trainer_id is not None:
        trainer = db.query(User).filter(User.id == payload.trainer_id).first()
        if not trainer:
            raise HTTPException(status_code=404, detail="РўСЂРµРЅРµСЂ РЅРµ РЅР°Р№РґРµРЅ")
        lesson.trainer_id = payload.trainer_id
    if payload.lesson_type is not None:
        if payload.lesson_type not in {t.value for t in CustomLessonType}:
            raise HTTPException(status_code=400, detail=f"lesson_type РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РѕРґРЅРёРј РёР·: {[t.value for t in CustomLessonType]}")
        lesson.lesson_type = CustomLessonType(payload.lesson_type)
    if payload.comment is not None:
        lesson.comment = payload.comment.strip() or None

    # РћР±РЅРѕРІР»РµРЅРёРµ СЃРїРёСЃРєР° СѓС‡РµРЅРёРєРѕРІ: РµСЃР»Рё РїРµСЂРµРґР°РЅ payload.students вЂ” РїРµСЂРµСЃРѕР±РёСЂР°РµРј СЃРїРёСЃРѕРє
    if payload.students is not None:
        existing = db.query(CustomLessonStudent).filter(CustomLessonStudent.lesson_id == lesson.id).all()
        {(e.student_id, e.planned_absence_id): e for e in existing}
        db.query(CustomLessonStudent).filter(CustomLessonStudent.lesson_id == lesson.id).delete(synchronize_session=False)

        for item in payload.students:
            student = db.query(Student).filter(Student.id == item.student_id).first()
            if not student:
                raise HTTPException(status_code=404, detail=f"РЈС‡РµРЅРёРє {item.student_id} РЅРµ РЅР°Р№РґРµРЅ")
            planned_absence_id = item.planned_absence_id
            if planned_absence_id is not None:
                absence = db.query(AbsenceFollowUp).filter(AbsenceFollowUp.id == planned_absence_id).first()
                if not absence or absence.student_id != item.student_id:
                    raise HTTPException(status_code=400, detail=f"РџСЂРѕРїСѓСЃРє {planned_absence_id} РЅРµ РЅР°Р№РґРµРЅ РґР»СЏ СЌС‚РѕРіРѕ СѓС‡РµРЅРёРєР°")
            # РЎР±СЂР°СЃС‹РІР°РµРј РїРѕСЃРµС‰Р°РµРјРѕСЃС‚СЊ РїСЂРё РїРѕР»РЅРѕРј РїРµСЂРµСЃР±РѕСЂРµ СЃРїРёСЃРєР°
            cls = CustomLessonStudent(
                lesson_id=lesson.id,
                student_id=item.student_id,
                planned_absence_id=planned_absence_id,
                attended=False,
            )
            db.add(cls)

    db.commit()
    db.refresh(lesson)
    return _custom_lesson_to_response(db, lesson)


@router.delete("/_legacy-disabled/custom-lessons/{lesson_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_custom_lesson(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("lessons.manage")),
):
    """РЈРґР°Р»РёС‚СЊ СЂСѓС‡РЅРѕР№ СѓСЂРѕРє. РўРѕР»СЊРєРѕ admin/owner/sales."""
    lesson = db.query(CustomLesson).filter(CustomLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Р СѓС‡РЅРѕР№ СѓСЂРѕРє РЅРµ РЅР°Р№РґРµРЅ")
    db.query(CustomLessonStudent).filter(CustomLessonStudent.lesson_id == lesson.id).delete(synchronize_session=False)
    db.delete(lesson)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/_legacy-disabled/program-makeup-compatibility", response_model=List[ProgramMakeupCompatibilityResponse])
async def list_program_makeup_compatibility(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РЎРїРёСЃРѕРє РїСЂР°РІРёР» СЃРѕРІРјРµСЃС‚РёРјРѕСЃС‚Рё РїСЂРѕРіСЂР°РјРј РґР»СЏ РѕС‚СЂР°Р±РѕС‚РѕРє (РўР— Рї.5.3). Owner/admin."""
    _require_owner_or_admin_settings(current_user)
    items = db.query(ProgramMakeupCompatibility).order_by(
        ProgramMakeupCompatibility.source_program_id,
        ProgramMakeupCompatibility.target_program_id,
    ).all()
    result = []
    for c in items:
        src = db.query(Program).filter(Program.id == c.source_program_id).first()
        tgt = db.query(Program).filter(Program.id == c.target_program_id).first()
        result.append(ProgramMakeupCompatibilityResponse(
            id=c.id,
            source_program_id=c.source_program_id,
            target_program_id=c.target_program_id,
            source_program_name=src.name if src else None,
            target_program_name=tgt.name if tgt else None,
        ))
    return result


@router.post("/_legacy-disabled/program-makeup-compatibility", response_model=ProgramMakeupCompatibilityResponse, status_code=status.HTTP_201_CREATED)
async def create_program_makeup_compatibility(
    payload: ProgramMakeupCompatibilityCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """Р”РѕР±Р°РІРёС‚СЊ РїСЂР°РІРёР»Рѕ: РїСЂРѕРіСЂР°РјРјР° source РјРѕР¶РµС‚ РѕС‚СЂР°Р±Р°С‚С‹РІР°С‚СЊ РІ РїСЂРѕРіСЂР°РјРјРµ target. Owner/admin."""
    _require_owner_or_admin_settings(current_user)
    for pid in (payload.source_program_id, payload.target_program_id):
        if not db.query(Program).filter(Program.id == pid).first():
            raise HTTPException(status_code=404, detail=f"РџСЂРѕРіСЂР°РјРјР° {pid} РЅРµ РЅР°Р№РґРµРЅР°")
    compat = ProgramMakeupCompatibility(
        source_program_id=payload.source_program_id,
        target_program_id=payload.target_program_id,
    )
    db.add(compat)
    db.commit()
    db.refresh(compat)
    src = db.query(Program).filter(Program.id == compat.source_program_id).first()
    tgt = db.query(Program).filter(Program.id == compat.target_program_id).first()
    return ProgramMakeupCompatibilityResponse(
        id=compat.id,
        source_program_id=compat.source_program_id,
        target_program_id=compat.target_program_id,
        source_program_name=src.name if src else None,
        target_program_name=tgt.name if tgt else None,
    )


@router.delete("/_legacy-disabled/program-makeup-compatibility/{compat_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_program_makeup_compatibility(
    compat_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РЈРґР°Р»РёС‚СЊ РїСЂР°РІРёР»Рѕ СЃРѕРІРјРµСЃС‚РёРјРѕСЃС‚Рё. Owner/admin."""
    _require_owner_or_admin_settings(current_user)
    compat = db.query(ProgramMakeupCompatibility).filter(ProgramMakeupCompatibility.id == compat_id).first()
    if not compat:
        raise HTTPException(status_code=404, detail="РџСЂР°РІРёР»Рѕ РЅРµ РЅР°Р№РґРµРЅРѕ")
    db.delete(compat)
    db.commit()


@router.get("/_legacy-disabled/students/{student_id}/freezes", response_model=List[StudentFreezeResponse])
async def list_student_freezes(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РЎРїРёСЃРѕРє Р·Р°РјРѕСЂРѕР·РѕРє СѓС‡РµРЅРёРєР°. Sales/admin/owner."""
    _require_sales_admin_owner(current_user)
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="РЈС‡РµРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    freezes = db.query(StudentFreeze).filter(StudentFreeze.student_id == student_id).order_by(StudentFreeze.freeze_start.desc()).all()
    return [StudentFreezeResponse(id=f.id, student_id=f.student_id, freeze_start=f.freeze_start, freeze_end=f.freeze_end, created_at=f.created_at) for f in freezes]


@router.post("/_legacy-disabled/students/{student_id}/freezes", response_model=StudentFreezeResponse, status_code=status.HTTP_201_CREATED)
async def create_student_freeze(
    student_id: int,
    payload: StudentFreezeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РџРѕСЃС‚Р°РІРёС‚СЊ Р·Р°РјРѕСЂРѕР·РєСѓ (РўР— Рї.7). РўРѕР»СЊРєРѕ owner. РЎРґРІРёРі РїРµСЂРёРѕРґР° вЂ” РїСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё РѕР±РЅРѕРІРёС‚СЊ next_payment_date РІСЂСѓС‡РЅСѓСЋ РёР»Рё РѕС‚РґРµР»СЊРЅС‹Рј РїСЂР°РІРёР»РѕРј."""
    _require_owner(current_user)
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="РЈС‡РµРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    if payload.freeze_end <= payload.freeze_start:
        raise HTTPException(status_code=400, detail="freeze_end РґРѕР»Р¶РЅР° Р±С‹С‚СЊ Р±РѕР»СЊС€Рµ freeze_start")
    freeze = StudentFreeze(student_id=student_id, freeze_start=payload.freeze_start, freeze_end=payload.freeze_end)
    db.add(freeze)
    card = db.query(StudentCard).filter(StudentCard.student_id == student_id).first()
    if card and getattr(card, "next_payment_date", None):
        delta = (payload.freeze_end - payload.freeze_start).days
        card.next_payment_date = card.next_payment_date + timedelta(days=delta)
    log_student_activity(
        db,
        student_id=student_id,
        activity_type="freeze_set",
        title="РџРѕСЃС‚Р°РІР»РµРЅР° Р·Р°РјРѕСЂРѕР·РєР°",
        description=f"{payload.freeze_start} вЂ” {payload.freeze_end}",
        created_by=current_user.id,
        payload_json={"freeze_start": payload.freeze_start.isoformat(), "freeze_end": payload.freeze_end.isoformat()},
    )
    db.commit()
    db.refresh(freeze)
    return StudentFreezeResponse(id=freeze.id, student_id=freeze.student_id, freeze_start=freeze.freeze_start, freeze_end=freeze.freeze_end, created_at=freeze.created_at)


@router.delete("/_legacy-disabled/students/{student_id}/freezes/{freeze_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_student_freeze(
    student_id: int,
    freeze_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РЎРЅСЏС‚СЊ Р·Р°РјРѕСЂРѕР·РєСѓ. РўРѕР»СЊРєРѕ owner."""
    _require_owner(current_user)
    freeze = db.query(StudentFreeze).filter(StudentFreeze.id == freeze_id, StudentFreeze.student_id == student_id).first()
    if not freeze:
        raise HTTPException(status_code=404, detail="Р—Р°РјРѕСЂРѕР·РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    db.delete(freeze)
    db.commit()


@router.get("/_legacy-disabled/students/{student_id}/close-by-fact-preview", response_model=CloseByFactPreview)
async def close_by_fact_preview(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """РџСЂРµРґРїСЂРѕСЃРјРѕС‚СЂ: СЃРєРѕР»СЊРєРѕ Р·Р°РЅСЏС‚РёР№ РїРѕСЃРµС‰РµРЅРѕ РІ С‚РµРєСѓС‰РµРј РїРµСЂРёРѕРґРµ Рё СЃСѓРјРјР° Рє РѕРїР»Р°С‚Рµ (РўР— Рї.9). РўРѕР»СЊРєРѕ owner."""
    _require_owner(current_user)
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="РЈС‡РµРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    card = db.query(StudentCard).filter(StudentCard.student_id == student_id).first()
    period_start = getattr(card, "learning_period_start", None) if card else None
    if not period_start:
        period_start = date.today() - timedelta(days=30)
    period_end = date.today()
    attended = db.query(LessonAttendance).filter(
        LessonAttendance.student_id == student_id,
        LessonAttendance.attended.is_(True),
        LessonAttendance.lesson_date >= period_start,
        LessonAttendance.lesson_date <= period_end,
    ).count()
    abonement = student.abonement or (db.query(Abonement).filter(Abonement.id == student.abonement_id).first() if student.abonement_id else None)
    if not card and student.abonement_id:
        abonement = db.query(Abonement).filter(Abonement.id == student.abonement_id).first()
    if card and getattr(card, "abonement_id", None):
        abonement = db.query(Abonement).filter(Abonement.id == card.abonement_id).first() or abonement
    price_per_lesson = 0.0
    if abonement and (abonement.lessons_count or 8) > 0:
        price_per_lesson = float(abonement.price or 0) / (abonement.lessons_count or 8)
    amount = round(price_per_lesson * attended, 2)
    return CloseByFactPreview(
        lessons_attended_in_period=attended,
        amount=amount,
        period_start=period_start,
        period_end=period_end,
    )


@router.post("/_legacy-disabled/students/{student_id}/close-by-fact")
async def close_by_fact_confirm(
    student_id: int,
    payload: CloseByFactConfirm,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """Р—Р°РєСЂС‹С‚СЊ СѓС‡РµРЅРёРєР° СЃ РѕРїР»Р°С‚РѕР№ РїРѕ С„Р°РєС‚Сѓ (РўР— Рї.9). РўРѕР»СЊРєРѕ owner. Р¤РёРєСЃРёСЂСѓРµС‚ РѕРїР»Р°С‚Сѓ, Р°СЂС…РёРІРёСЂСѓРµС‚, СЃРЅРёРјР°РµС‚ РїСЂРѕРїСѓСЃРєРё СЃ РѕС‡РµСЂРµРґРё."""
    _require_owner(current_user)
    if not payload.confirm:
        raise HTTPException(status_code=400, detail="РџРѕРґС‚РІРµСЂРґРёС‚Рµ Р·Р°РєСЂС‹С‚РёРµ")
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="РЈС‡РµРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    card = db.query(StudentCard).filter(StudentCard.student_id == student_id).first()
    period_start = getattr(card, "learning_period_start", None) if card else None
    if not period_start:
        period_start = date.today() - timedelta(days=30)
    period_end = date.today()
    attended = db.query(LessonAttendance).filter(
        LessonAttendance.student_id == student_id,
        LessonAttendance.attended.is_(True),
        LessonAttendance.lesson_date >= period_start,
        LessonAttendance.lesson_date <= period_end,
    ).count()
    abonement = student.abonement or (db.query(Abonement).filter(Abonement.id == student.abonement_id).first() if student.abonement_id else None)
    if card and getattr(card, "abonement_id", None):
        abonement = db.query(Abonement).filter(Abonement.id == card.abonement_id).first() or abonement
    price_per_lesson = float(abonement.price or 0) / (abonement.lessons_count or 8) if abonement and (abonement.lessons_count or 8) > 0 else 0.0
    amount = round(price_per_lesson * attended, 2)
    account = db.query(StudentAccount).filter(StudentAccount.student_id == student_id).order_by(StudentAccount.id).first()
    if account and amount > 0:
        db.add(StudentAccountTransaction(
            account_id=account.id,
            amount=amount,
            kind=StudentAccountTransactionKind.PAYMENT,
            note=f"Р—Р°РєСЂС‹С‚РёРµ РїРѕ С„Р°РєС‚Сѓ: {attended} Р·Р°РЅСЏС‚РёР№ Р·Р° РїРµСЂРёРѕРґ {period_start}вЂ“{period_end}",
        ))
        account.balance += amount
    for a in db.query(AbsenceFollowUp).filter(AbsenceFollowUp.student_id == student_id).all():
        a.stage = "made_up"
    if card:
        card.archived = True
    student.status = StudentStatus.ARCHIVED
    db.commit()
    return {"ok": True, "student_id": student_id, "amount": amount, "lessons_attended": attended}


def _require_owner_or_admin(lead: Lead, user: User) -> None:
    if auth.resolve_effective_role(user) in (UserRole.ADMIN, UserRole.OWNER, UserRole.SALES):
        return
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")


def _add_activity(
    db: Session,
    lead_id: int,
    actor_id: int,
    type: str,
    title: str,
    description: Optional[str] = None,
    channel: Optional[str] = None,
    status_effect_from: Optional[str] = None,
    status_effect_to: Optional[str] = None,
    related_task_id: Optional[int] = None,
    related_invoice_id: Optional[int] = None,
    payload_json: Optional[dict] = None,
) -> LeadActivity:
    """Create a LeadActivity record without committing."""
    activity = LeadActivity(
        lead_id=lead_id,
        type=type,
        title=title,
        description=description,
        channel=channel,
        created_by=actor_id,
        status_effect_from=status_effect_from,
        status_effect_to=status_effect_to,
        related_task_id=related_task_id,
        related_invoice_id=related_invoice_id,
        payload_json=payload_json,
    )
    db.add(activity)
    return activity


def _filter_query_by_role(query, user: User):
    if auth.resolve_effective_role(user) in (UserRole.ADMIN, UserRole.OWNER, UserRole.SALES):
        return query
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")


def _compute_price(abonement: Abonement) -> float:
    price = abonement.price or 0.0
    if abonement.discount_type is None:
        return price
    if hasattr(abonement, "discount_type"):
        dt = getattr(abonement.discount_type, "value", abonement.discount_type)
        if dt == "amount":
            price = max(price - (abonement.discount_value or 0.0), 0.0)
        elif dt == "percent":
            price = price * (1 - (abonement.discount_value or 0.0) / 100)
    return round(price, 2)


def _normalize_source_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    return " ".join(name.strip().split())


def _is_referral_source(name: Optional[str]) -> bool:
    return (name or "").strip().lower() == "СЂРµРєРѕРјРµРЅРґР°С†РёСЏ"


def _resolve_source(
    db: Session,
    source_id: Optional[int] = None,
    source_name: Optional[str] = None,
) -> Tuple[Optional[int], Optional[str]]:
    normalized = _normalize_source_name(source_name)
    if source_id:
        source_obj = db.query(LeadSource).filter(LeadSource.id == source_id).first()
        if not source_obj:
            raise HTTPException(status_code=404, detail="Lead source not found")
        return source_obj.id, source_obj.name
    if normalized:
        source_obj = db.query(LeadSource).filter(cast(LeadSource.name, Text).ilike(normalized)).first()
        if source_obj:
            return source_obj.id, source_obj.name
        return None, normalized
    return None, None


def _append_note_tag(note: Optional[str], tag: str) -> str:
    source = (note or "").strip()
    if tag in source:
        return source
    return f"{tag} {source}".strip()


def _remove_note_tag(note: Optional[str], tag: str) -> str:
    """РЈРґР°Р»РёС‚СЊ С‚РµРі РёР· Р·Р°РјРµС‚РєРё (Р±РµР· СѓС‡С‘С‚Р° СЂРµРіРёСЃС‚СЂР°), СЃРѕС…СЂР°РЅРёРІ РѕСЃС‚Р°Р»СЊРЅРѕР№ С‚РµРєСЃС‚."""
    if not note:
        return ""
    lower_tag = tag.lower()
    parts = [p for p in note.split() if p.lower() != lower_tag]
    return " ".join(parts).strip()


def _get_default_open_status_option_id(db: Session) -> Optional[int]:
    default_open = (
        db.query(LeadTaskStatusOptionModel)
        .filter(
            LeadTaskStatusOptionModel.is_active.is_(True),
            LeadTaskStatusOptionModel.is_closed.is_(False),
        )
        .order_by(LeadTaskStatusOptionModel.id.asc())
        .first()
    )
    return default_open.id if default_open else None


def _get_default_lead_status_option_id(db: Session, base_status: LeadStatus) -> Optional[int]:
    """Р’РѕР·РІСЂР°С‰Р°РµС‚ id РїРµСЂРІРѕР№ Р°РєС‚РёРІРЅРѕР№ РѕРїС†РёРё СЃС‚Р°С‚СѓСЃР° Р»РёРґР° РґР»СЏ РґР°РЅРЅРѕРіРѕ base_status."""
    status_str = base_status.value if hasattr(base_status, "value") else str(base_status)
    opt = (
        db.query(LeadStatusOption)
        .filter(
            LeadStatusOption.base_status == status_str,
            LeadStatusOption.is_active.is_(True),
        )
        .order_by(LeadStatusOption.id.asc())
        .first()
    )
    return opt.id if opt else None


def _create_auto_event_task(
    db: Session,
    *,
    lead: Lead,
    owner_id: int,
    note: str,
    due_at: datetime,
    preferred_template_keywords: List[str],
) -> LeadTask:
    template = None
    templates = (
        db.query(LeadTaskTemplate)
        .filter(LeadTaskTemplate.is_active.is_(True))
        .order_by(LeadTaskTemplate.id.asc())
        .all()
    )
    for tpl in templates:
        name = (tpl.name or "").lower()
        if any(keyword in name for keyword in preferred_template_keywords):
            template = tpl
            break
    task = LeadTask(
        lead_id=lead.id,
        owner_id=owner_id,
        template_id=template.id if template else None,
        status_option_id=_get_default_open_status_option_id(db),
        note=note,
        channel="call",
        due_at=due_at,
        status=LeadTaskStatus.OPEN,
    )
    db.add(task)
    return task


def _has_open_task_like(db: Session, lead_id: int, marker: str) -> bool:
    return (
        db.query(LeadTask)
        .filter(
            LeadTask.lead_id == lead_id,
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.note.isnot(None),
            cast(LeadTask.note, Text).ilike(f"%{marker}%"),
        )
        .first()
        is not None
    )


ALLOWED_PAUSE_REASONS = {"Р¶РґС‘Рј РѕС‚РІРµС‚", "РїРѕРґСѓРјР°С‚СЊ", "РЅРµС‚ РІСЂРµРјРµРЅРё"}


@router.get("/_legacy-disabled/dashboard", response_model=SalesDashboardResponse)
async def get_sales_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    now = utcnow()
    start_month = datetime(now.year, now.month, 1)
    end_month = datetime(now.year + (1 if now.month == 12 else 0), 1 if now.month == 12 else now.month + 1, 1)
    start_today = datetime(now.year, now.month, now.day)
    end_today = start_today + timedelta(days=1)
    next_two_hours = now + timedelta(hours=2)
    next_day = now + timedelta(hours=24)
    week_ago = now - timedelta(days=7)

    leads_q = _filter_query_by_role(db.query(Lead), current_user)
    tasks_q = db.query(LeadTask).join(Lead, Lead.id == LeadTask.lead_id)
    regs_q = db.query(EventRegistration).join(Lead, Lead.id == EventRegistration.lead_id)

    active_lead_statuses = [LeadStatus.NEW, LeadStatus.CONTACTED, LeadStatus.DEMO, LeadStatus.INVOICE_SENT]
    connected_statuses = [LeadStatus.CONTACTED, LeadStatus.DEMO, LeadStatus.INVOICE_SENT, LeadStatus.WON]

    kpi_new_leads = leads_q.filter(Lead.created_at >= start_today, Lead.created_at < end_today).count()
    leads_week_q = leads_q.filter(Lead.created_at >= week_ago)
    leads_week_total = leads_week_q.count()
    leads_week_connected = leads_week_q.filter(Lead.status.in_(connected_statuses)).count()
    kpi_dozvon_percent = round((leads_week_connected / leads_week_total) * 100, 1) if leads_week_total else 0.0
    kpi_info_sent = leads_q.filter(Lead.status.in_([LeadStatus.INVOICE_SENT, LeadStatus.WON])).count()

    kpi_need_push_urgent = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at >= now,
        Lead.next_contact_at < next_two_hours,
    ).count()
    kpi_need_push_today = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at >= start_today,
        Lead.next_contact_at < end_today,
    ).count()
    kpi_need_push_overdue = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at < now,
    ).count()

    kpi_registered_event = regs_q.filter(EventRegistration.status == EventRegistrationStatus.REGISTERED).count()
    # Until dedicated attendance statuses are introduced, derive rough metrics from note tags.
    kpi_came_count = regs_q.filter(cast(EventRegistration.note, Text).ilike("%[came]%")).count()
    kpi_no_show_count = regs_q.filter(
        or_(
            cast(EventRegistration.note, Text).ilike("%[no-show]%"),
            cast(EventRegistration.note, Text).ilike("%no-show%"),
        )
    ).count()

    overdue_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at.isnot(None),
            LeadTask.due_at < now,
        )
        .order_by(LeadTask.due_at.asc())
        .limit(30)
        .all()
    )
    call_today_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at.isnot(None),
            LeadTask.due_at >= start_today,
            LeadTask.due_at < end_today,
        )
        .order_by(LeadTask.due_at.asc())
        .limit(30)
        .all()
    )
    messenger_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            or_(
                cast(LeadTask.channel, Text).ilike("%messenger%"),
                cast(LeadTask.channel, Text).ilike("%telegram%"),
                cast(LeadTask.note, Text).ilike("%РѕС‚РІРµС‚%"),
            ),
        )
        .order_by(LeadTask.created_at.desc())
        .limit(30)
        .all()
    )
    confirm_items = (
        regs_q.join(Event, Event.id == EventRegistration.event_id)
        .filter(
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
            Event.starts_at >= now,
            Event.starts_at < next_day,
        )
        .order_by(Event.starts_at.asc())
        .limit(30)
        .all()
    )

    month_outreach_leads = (
        leads_q.filter(
            Lead.outreach_at.isnot(None),
            Lead.outreach_at >= start_month,
            Lead.outreach_at < end_month,
        )
        .all()
    )
    schools_seen = set()
    classes_seen = set()
    outreach_minutes_month = 0
    school_stats: Dict[str, Dict[str, object]] = {}
    for lead in month_outreach_leads:
        school = (lead.school_name or "").strip()
        class_name = (lead.school_class or "").strip()
        if school:
            schools_seen.add(school)
        if school and class_name:
            classes_seen.add(f"{school}::{class_name}")
        if lead.outreach_minutes and lead.outreach_minutes > 0:
            outreach_minutes_month += lead.outreach_minutes
        if not school:
            continue
        if school not in school_stats:
            school_stats[school] = {"leads": 0, "won": 0, "classes": set(), "minutes": 0}
        school_stats[school]["leads"] = int(school_stats[school]["leads"]) + 1
        if lead.status == LeadStatus.WON:
            school_stats[school]["won"] = int(school_stats[school]["won"]) + 1
        if class_name:
            school_stats[school]["classes"].add(class_name)
        if lead.outreach_minutes and lead.outreach_minutes > 0:
            school_stats[school]["minutes"] = int(school_stats[school]["minutes"]) + lead.outreach_minutes

    top_schools_conversion_month = sorted(
        [
            SalesSchoolConversionItem(
                school_name=school,
                leads_count=int(stats["leads"]),
                won_count=int(stats["won"]),
                conversion_percent=round((int(stats["won"]) / int(stats["leads"])) * 100) if int(stats["leads"]) else 0,
                classes_count=len(stats["classes"]),
                outreach_minutes_total=int(stats["minutes"]),
            )
            for school, stats in school_stats.items()
        ],
        key=lambda item: (item.conversion_percent, item.leads_count),
        reverse=True,
    )[:3]

    def map_task(item: LeadTask) -> SalesQueueTaskItem:
        lead = item.lead
        return SalesQueueTaskItem(
            task_id=item.id,
            lead_id=lead.id,
            lead_name=lead.contact_name,
            lead_phone=lead.phone,
            due_at=item.due_at,
            note=item.note,
        )

    def map_reg(item: EventRegistration) -> SalesQueueRegistrationItem:
        event = db.query(Event).filter(Event.id == item.event_id).first()
        lead = db.query(Lead).filter(Lead.id == item.lead_id).first()
        return SalesQueueRegistrationItem(
            registration_id=item.id,
            event_id=item.event_id,
            event_title=event.title if event else f"Event #{item.event_id}",
            lead_id=item.lead_id,
            lead_name=lead.contact_name if lead else f"Lead #{item.lead_id}",
            starts_at=event.starts_at if event else now,
            note=item.note,
        )

    return SalesDashboardResponse(
        kpi_new_leads=kpi_new_leads,
        kpi_dozvon_percent=kpi_dozvon_percent,
        kpi_info_sent=kpi_info_sent,
        kpi_need_push_urgent=kpi_need_push_urgent,
        kpi_need_push_today=kpi_need_push_today,
        kpi_need_push_overdue=kpi_need_push_overdue,
        kpi_registered_event=kpi_registered_event,
        kpi_came_count=kpi_came_count,
        kpi_no_show_count=kpi_no_show_count,
        overdue_followups=[map_task(i) for i in overdue_items],
        call_today=[map_task(i) for i in call_today_items],
        messenger_replies=[map_task(i) for i in messenger_items],
        confirm_participation=[map_reg(i) for i in confirm_items],
        outreach_schools_month=len(schools_seen),
        outreach_classes_month=len(classes_seen),
        outreach_minutes_month=outreach_minutes_month,
        top_schools_conversion_month=top_schools_conversion_month,
    )


@router.get("/_legacy-disabled/follow-ups", response_model=List[FollowUpItemResponse])
async def list_follow_ups(
    period: str = Query(default="today"),
    source: Optional[str] = None,
    event_id: Optional[int] = None,
    reason: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    now = utcnow()
    start_today = datetime(now.year, now.month, now.day)
    end_today = start_today + timedelta(days=1)
    start_tomorrow = end_today
    end_tomorrow = start_tomorrow + timedelta(days=1)
    end_week = end_today + timedelta(days=7)

    query = db.query(LeadTask, Lead).join(Lead, Lead.id == LeadTask.lead_id)
    query = query.filter(LeadTask.status == LeadTaskStatus.OPEN)

    if period == "overdue":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at < now)
    elif period == "tomorrow":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_tomorrow, LeadTask.due_at < end_tomorrow)
    elif period == "week":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_today, LeadTask.due_at < end_week)
    else:
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_today, LeadTask.due_at < end_today)

    if source:
        query = query.filter(Lead.source.isnot(None), cast(Lead.source, Text).ilike(f"%{source.strip()}%"))
    if reason:
        query = query.filter(LeadTask.note.isnot(None), cast(LeadTask.note, Text).ilike(f"%{reason.strip()}%"))
    if event_id:
        query = query.join(EventRegistration, EventRegistration.lead_id == Lead.id).filter(
            EventRegistration.event_id == event_id,
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
        )

    rows = query.order_by(LeadTask.due_at.asc()).limit(200).all()
    return [
        FollowUpItemResponse(
            task_id=task.id,
            lead_id=lead.id,
            lead_name=lead.contact_name,
            lead_phone=lead.phone,
            lead_source=lead.source,
            due_at=task.due_at,
            status=task.status,
            channel=task.channel,
            note=task.note,
        )
        for task, lead in rows
    ]


@router.get("/_legacy-disabled/leads/push-stats", response_model=List[LeadPushStatsResponse])
async def get_leads_push_stats(
    lead_ids: List[int] = Query(default=[]),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead_query = _filter_query_by_role(db.query(Lead), current_user)
    if lead_ids:
        lead_query = lead_query.filter(Lead.id.in_(lead_ids))
    leads = lead_query.all()
    if not leads:
        return []

    allowed_ids = [lead.id for lead in leads]
    tasks = (
        db.query(LeadTask)
        .options(joinedload(LeadTask.template))
        .filter(LeadTask.lead_id.in_(allowed_ids))
        .all()
    )

    by_lead: Dict[int, Dict[str, int]] = {lead_id: {"total": 0, "done": 0} for lead_id in allowed_ids}
    for task in tasks:
        template_name = (task.template.name if task.template else "").lower()
        note = (task.note or "").lower()
        is_push = "РґРѕР¶РёРј" in template_name or "РґРѕР¶РёРј" in note or "push" in note
        if not is_push:
            continue
        by_lead[task.lead_id]["total"] += 1
        if task.status == LeadTaskStatus.DONE:
            by_lead[task.lead_id]["done"] += 1

    response: List[LeadPushStatsResponse] = []
    for lead_id in allowed_ids:
        total = by_lead[lead_id]["total"]
        done = by_lead[lead_id]["done"]
        pct = round((done / total) * 100) if total else 0
        response.append(
            LeadPushStatsResponse(
                lead_id=lead_id,
                total_steps=total,
                done_steps=done,
                progress_percent=pct,
            )
        )
    return response


@router.get("/_legacy-disabled/lead-sources", response_model=List[LeadSourceResponse])
async def list_lead_sources(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(LeadSource).order_by(LeadSource.name.asc())
    if active_only:
        query = query.filter(LeadSource.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/lead-sources", response_model=LeadSourceResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_source(
    payload: LeadSourceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Source name is required")
    exists = db.query(LeadSource).filter(cast(LeadSource.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Source already exists")
    source = LeadSource(name=name, is_active=True)
    db.add(source)
    db.commit()
    db.refresh(source)
    log_action(db, current_user.id, "create", "lead_source", source.id, {"name": name})
    return source


@router.put("/_legacy-disabled/lead-sources/{source_id}", response_model=LeadSourceResponse)
async def update_lead_source(
    source_id: int,
    payload: LeadSourceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    source = db.query(LeadSource).filter(LeadSource.id == source_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Source name is required")
        source.name = name
    if "is_active" in data:
        source.is_active = data["is_active"]
    db.commit()
    db.refresh(source)
    log_action(db, current_user.id, "update", "lead_source", source.id, data)
    return source


@router.get("/_legacy-disabled/lead-task-templates", response_model=List[LeadTaskTemplateResponse])
async def list_lead_task_templates(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(LeadTaskTemplate).order_by(LeadTaskTemplate.name.asc())
    if active_only:
        query = query.filter(LeadTaskTemplate.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/lead-task-templates", response_model=LeadTaskTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task_template(
    payload: LeadTaskTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Task template name is required")
    exists = db.query(LeadTaskTemplate).filter(cast(LeadTaskTemplate.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Task template already exists")
    item = LeadTaskTemplate(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_task_template", item.id, {"name": name})
    return item


@router.put("/_legacy-disabled/lead-task-templates/{template_id}", response_model=LeadTaskTemplateResponse)
async def update_lead_task_template(
    template_id: int,
    payload: LeadTaskTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadTaskTemplate).filter(LeadTaskTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Task template not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Task template name is required")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_task_template", item.id, data)
    return item


@router.get("/_legacy-disabled/lead-task-statuses", response_model=List[LeadTaskStatusOptionResponse])
async def list_lead_task_statuses(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(LeadTaskStatusOptionModel).order_by(LeadTaskStatusOptionModel.id.asc())
    if active_only:
        query = query.filter(LeadTaskStatusOptionModel.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/lead-task-statuses", response_model=LeadTaskStatusOptionResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task_status(
    payload: LeadTaskStatusOptionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Task status name is required")
    exists = db.query(LeadTaskStatusOptionModel).filter(cast(LeadTaskStatusOptionModel.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Task status already exists")
    item = LeadTaskStatusOptionModel(name=name, is_active=True, is_closed=payload.is_closed)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_task_status", item.id, {"name": name, "is_closed": payload.is_closed})
    return item


@router.put("/_legacy-disabled/lead-task-statuses/{status_id}", response_model=LeadTaskStatusOptionResponse)
async def update_lead_task_status(
    status_id: int,
    payload: LeadTaskStatusOptionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadTaskStatusOptionModel).filter(LeadTaskStatusOptionModel.id == status_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Task status not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Task status name is required")
        item.name = name
    if "is_closed" in data:
        item.is_closed = data["is_closed"]
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_task_status", item.id, data)
    return item


@router.get("/_legacy-disabled/lead-info-templates", response_model=List[LeadInfoTemplateResponse])
async def list_lead_info_templates(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(LeadInfoTemplate).order_by(LeadInfoTemplate.name.asc())
    if active_only:
        query = query.filter(LeadInfoTemplate.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/lead-info-templates", response_model=LeadInfoTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_info_template(
    payload: LeadInfoTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Template name is required")
    if not (payload.body or "").strip():
        raise HTTPException(status_code=400, detail="Template body is required")
    exists = db.query(LeadInfoTemplate).filter(cast(LeadInfoTemplate.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Template already exists")
    item = LeadInfoTemplate(name=name, body=payload.body.strip(), is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_info_template", item.id, {"name": name})
    return item


@router.put("/_legacy-disabled/lead-info-templates/{template_id}", response_model=LeadInfoTemplateResponse)
async def update_lead_info_template(
    template_id: int,
    payload: LeadInfoTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadInfoTemplate).filter(LeadInfoTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Template not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Template name is required")
        item.name = name
    if "body" in data:
        body = (data["body"] or "").strip()
        if not body:
            raise HTTPException(status_code=400, detail="Template body is required")
        item.body = body
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_info_template", item.id, data)
    return item


# --- Sales cities (СЃРїСЂР°РІРѕС‡РЅРёРє РіРѕСЂРѕРґРѕРІ РґР»СЏ Sales) ---
@router.get("/_legacy-disabled/cities", response_model=List[SalesCityResponse])
async def list_sales_cities(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(SalesCity).order_by(SalesCity.name.asc())
    if active_only:
        query = query.filter(SalesCity.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/cities", response_model=SalesCityResponse, status_code=status.HTTP_201_CREATED)
async def create_sales_city(
    payload: SalesCityCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="City name is required")
    exists = db.query(SalesCity).filter(cast(SalesCity.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="City already exists")
    item = SalesCity(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "sales_city", item.id, {"name": name})
    return item


@router.put("/_legacy-disabled/cities/{city_id}", response_model=SalesCityResponse)
async def update_sales_city(
    city_id: int,
    payload: SalesCityUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesCity).filter(SalesCity.id == city_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="City not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="City name is required")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "sales_city", item.id, data)
    return item


# --- Sales schools (СЃРїСЂР°РІРѕС‡РЅРёРє С€РєРѕР» РґР»СЏ Sales) ---
@router.get("/_legacy-disabled/schools", response_model=List[SalesSchoolResponse])
async def list_sales_schools(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(SalesSchool).order_by(SalesSchool.name.asc())
    if active_only:
        query = query.filter(SalesSchool.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/schools", response_model=SalesSchoolResponse, status_code=status.HTTP_201_CREATED)
async def create_sales_school(
    payload: SalesSchoolCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="School name is required")
    exists = db.query(SalesSchool).filter(cast(SalesSchool.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="School already exists")
    item = SalesSchool(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "sales_school", item.id, {"name": name})
    return item


@router.put("/_legacy-disabled/schools/{school_id}", response_model=SalesSchoolResponse)
async def update_sales_school(
    school_id: int,
    payload: SalesSchoolUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesSchool).filter(SalesSchool.id == school_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="School not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="School name is required")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "sales_school", item.id, data)
    return item


# --- Sales classes (СЃРїСЂР°РІРѕС‡РЅРёРє РєР»Р°СЃСЃРѕРІ РґР»СЏ Р»РёРґРѕРІ) ---
@router.get("/_legacy-disabled/classes", response_model=List[SalesClassResponse])
async def list_sales_classes(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(SalesClass).order_by(SalesClass.name.asc())
    if active_only:
        query = query.filter(SalesClass.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/classes", response_model=SalesClassResponse, status_code=status.HTTP_201_CREATED)
async def create_sales_class(
    payload: SalesClassCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="РќР°Р·РІР°РЅРёРµ РєР»Р°СЃСЃР° РѕР±СЏР·Р°С‚РµР»СЊРЅРѕ")
    exists = db.query(SalesClass).filter(cast(SalesClass.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="РўР°РєРѕР№ РєР»Р°СЃСЃ СѓР¶Рµ РµСЃС‚СЊ")
    item = SalesClass(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "sales_class", item.id, {"name": name})
    return item


@router.put("/_legacy-disabled/classes/{class_id}", response_model=SalesClassResponse)
async def update_sales_class(
    class_id: int,
    payload: SalesClassUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(SalesClass).filter(SalesClass.id == class_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="РљР»Р°СЃСЃ РЅРµ РЅР°Р№РґРµРЅ")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = (data["name"] or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="РќР°Р·РІР°РЅРёРµ РєР»Р°СЃСЃР° РѕР±СЏР·Р°С‚РµР»СЊРЅРѕ")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "sales_class", item.id, data)
    return item


# --- РЁР°Р±Р»РѕРЅС‹ СЃС‡РµС‚РѕРІ (РЅР°Р·РІР°РЅРёРµ + С„РѕСЂРјР°С‚: РіСЂСѓРїРїРѕРІРѕР№/РёРЅРґРёРІРёРґСѓР°Р»СЊРЅС‹Р№) ---
@router.get("/_legacy-disabled/account-templates", response_model=List[AccountTemplateResponse])
async def list_account_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    return db.query(AccountTemplate).order_by(AccountTemplate.id.asc()).all()


@router.post("/_legacy-disabled/account-templates", response_model=AccountTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_account_template(
    payload: AccountTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="РќР°Р·РІР°РЅРёРµ СЃС‡С‘С‚Р° РѕР±СЏР·Р°С‚РµР»СЊРЅРѕ")
    if payload.format not in ("group", "individual"):
        raise HTTPException(status_code=400, detail="Р¤РѕСЂРјР°С‚ РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ group РёР»Рё individual")
    item = AccountTemplate(name=name, format=payload.format)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "account_template", item.id, {"name": name, "format": payload.format})
    return item


@router.delete("/_legacy-disabled/account-templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_account_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    item = db.query(AccountTemplate).filter(AccountTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="РЁР°Р±Р»РѕРЅ СЃС‡С‘С‚Р° РЅРµ РЅР°Р№РґРµРЅ")
    db.delete(item)
    db.commit()
    log_action(db, current_user.id, "delete", "account_template", template_id, {})
    return None


# --- Lead status options (РєР°СЃС‚РѕРјРЅС‹Рµ СЃС‚Р°С‚СѓСЃС‹ Р»РёРґР°) ---
@router.get("/_legacy-disabled/lead-statuses", response_model=List[LeadStatusOptionResponse])
async def list_lead_statuses(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(LeadStatusOption).order_by(LeadStatusOption.id.asc())
    if active_only:
        query = query.filter(LeadStatusOption.is_active.is_(True))
    return query.all()


@router.post("/_legacy-disabled/lead-statuses", response_model=LeadStatusOptionResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_status(
    payload: LeadStatusOptionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Status name is required")
    exists = db.query(LeadStatusOption).filter(cast(LeadStatusOption.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Lead status already exists")
    base = getattr(payload.base_status, "value", payload.base_status) if hasattr(payload.base_status, "value") else payload.base_status
    item = LeadStatusOption(name=name, base_status=base, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_status_option", item.id, {"name": name, "base_status": base})
    return item


@router.put("/_legacy-disabled/lead-statuses/{status_id}", response_model=LeadStatusOptionResponse)
async def update_lead_status(
    status_id: int,
    payload: LeadStatusOptionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("settings.manage")),
):
    item = db.query(LeadStatusOption).filter(LeadStatusOption.id == status_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Lead status not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Status name is required")
        item.name = name
    if "base_status" in data:
        base = data["base_status"]
        item.base_status = getattr(base, "value", base) if base is not None else item.base_status
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_status_option", item.id, data)
    return item


# --- List leads (must be before /leads/{lead_id} to avoid 404 "Not Found") ---
@router.get("/_legacy-disabled/leads", response_model=List[LeadResponse])
async def list_leads(
    status_filter: Optional[LeadStatus] = None,
    questionnaire_filled: Optional[bool] = None,
    q: Optional[str] = None,
    source: Optional[str] = None,
    tag: Optional[str] = None,
    overdue_only: bool = False,
    created_from: Optional[datetime] = Query(default=None),
    created_to: Optional[datetime] = Query(default=None),
    next_contact_from: Optional[datetime] = Query(default=None),
    next_contact_to: Optional[datetime] = Query(default=None),
    limit: int = Query(default=500, ge=1, le=2000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    auth.ensure_permission(current_user, "sales.access")

    query = _filter_query_by_role(
        db.query(Lead).options(*_lead_eager_options()).order_by(Lead.created_at.desc()),
        current_user,
    )
    if status_filter:
        query = query.filter(Lead.status == status_filter)
    if questionnaire_filled is not None:
        query = query.filter(Lead.questionnaire_filled == questionnaire_filled)
    if q:
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                cast(Lead.contact_name, Text).ilike(search),
                cast(Lead.phone, Text).ilike(search),
                cast(Lead.parent_full_name, Text).ilike(search),
                cast(Lead.child_full_name, Text).ilike(search),
                cast(Lead.parent_phone, Text).ilike(search),
                cast(Lead.child_phone, Text).ilike(search),
                cast(Lead.school_name, Text).ilike(search),
                cast(Lead.school_class, Text).ilike(search),
            )
        )
    if source:
        query = query.filter(Lead.source.ilike(f"%{source.strip()}%"))
    if tag:
        # tags are stored as JSON array; text-search keeps compatibility across DB backends.
        query = query.filter(Lead.tags.isnot(None), cast(Lead.tags, Text).ilike(f"%{tag.strip()}%"))
    if overdue_only:
        now = utcnow()
        query = query.filter(Lead.next_contact_at.isnot(None), Lead.next_contact_at < now)
    if created_from:
        query = query.filter(Lead.created_at >= created_from)
    if created_to:
        query = query.filter(Lead.created_at <= created_to)
    if next_contact_from:
        query = query.filter(Lead.next_contact_at >= next_contact_from)
    if next_contact_to:
        query = query.filter(Lead.next_contact_at <= next_contact_to)
    leads = query.offset(offset).limit(limit).all()

    # One-time/gradual Р±СЌРєР°РїРѕРІРєР° СЃС‚Р°СЂС‹С… Р»РёРґРѕРІ РІ РІРѕСЂРѕРЅРєСѓ В«Р”РѕР¶Р°С‚СЊ РЅР° РѕР±СѓС‡РµРЅРёРµВ»:
    # РµСЃР»Рё СЃС‚Р°С‚СѓСЃ demo, РїРѕ Р»РёРґСѓ СѓР¶Рµ РµСЃС‚СЊ [came] РІ СЂРµРіРёСЃС‚СЂР°С†РёСЏС… СЃРѕР±С‹С‚РёСЏ, РЅРѕ post_visit_stage РµС‰С‘ РЅРµ Р·Р°РґР°РЅР°,
    # РїСЂРѕСЃС‚Р°РІР»СЏРµРј stage='new', С‡С‚РѕР±С‹ С‚Р°РєРёРµ Р»РёРґС‹ РїРѕСЏРІРёР»РёСЃСЊ РЅР° СЃС‚СЂР°РЅРёС†Рµ РґРѕР¶РёРјР°.
    if status_filter == LeadStatus.DEMO:
        candidate_ids = [lead.id for lead in leads if not getattr(lead, "post_visit_stage", None)]
        if candidate_ids:
            # РћРґРёРЅ Р·Р°РїСЂРѕСЃ РІРјРµСЃС‚Рѕ N: РёС‰РµРј РІСЃРµ lead_id СЃ [came] СЃСЂРµРґРё РєР°РЅРґРёРґР°С‚РѕРІ
            came_lead_ids = {
                row[0]
                for row in db.query(EventRegistration.lead_id)
                .filter(
                    EventRegistration.lead_id.in_(candidate_ids),
                    cast(EventRegistration.note, Text).ilike("%[came]%"),
                )
                .distinct()
                .all()
            }
            if came_lead_ids:
                db.execute(
                    sa_update(Lead)
                    .where(Lead.id.in_(came_lead_ids))
                    .values(post_visit_stage="new")
                    .execution_options(synchronize_session=False)
                )
                db.commit()
                # РћРґРЅРёРј Р·Р°РїСЂРѕСЃРѕРј РїРµСЂРµР·Р°РіСЂСѓР¶Р°РµРј С‚РѕР»СЊРєРѕ РёР·РјРµРЅС‘РЅРЅС‹Рµ Р»РёРґС‹
                refreshed = {
                    r.id: r
                    for r in db.query(Lead).options(*_lead_eager_options()).filter(Lead.id.in_(came_lead_ids)).all()
                }
                leads = [refreshed.get(lead.id, lead) for lead in leads]

    result: List[Lead] = []
    for lead in leads:
        fixed = _fix_lead_strings(lead)
        fixed.ai_insight = build_lead_ai_insight(fixed)
        result.append(fixed)
    return result


@router.get("/_legacy-disabled/leads/{lead_id}/communications", response_model=List[LeadCommunicationResponse])
async def list_lead_communications(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).options(*_lead_eager_options()).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    return (
        db.query(LeadCommunication)
        .filter(LeadCommunication.lead_id == lead_id)
        .order_by(LeadCommunication.created_at.desc())
        .all()
    )


@router.post("/_legacy-disabled/leads/{lead_id}/communications", response_model=LeadCommunicationResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_communication(
    lead_id: int,
    payload: LeadQuickCommunicationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    channel = (payload.channel or "messenger").strip().lower()
    if channel not in {"messenger", "call", "email"}:
        raise HTTPException(status_code=400, detail="Unsupported channel")
    message = (payload.message or "").strip() or f"[quick-{channel}]"
    follow_up_at = payload.follow_up_at or utcnow()

    comm = LeadCommunication(
        lead_id=lead.id,
        sent_by=current_user.id,
        template_id=None,
        channel=channel,
        message=message,
        pause_reason=None,
        follow_up_at=follow_up_at,
    )
    db.add(comm)
    if payload.follow_up_at:
        lead.next_contact_at = payload.follow_up_at
    if lead.status == LeadStatus.NEW:
        lead.status = LeadStatus.CONTACTED
    db.commit()
    db.refresh(comm)
    log_action(
        db,
        current_user.id,
        "log_communication",
        "lead",
        lead.id,
        {
            "channel": channel,
            "follow_up_at": follow_up_at.isoformat(),
        },
    )
    return comm


@router.post("/_legacy-disabled/leads/{lead_id}/send-info", response_model=LeadCommunicationResponse, status_code=status.HTTP_201_CREATED)
async def send_info_for_lead(
    lead_id: int,
    payload: LeadSendInfoRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    if payload.follow_up_at <= utcnow():
        raise HTTPException(status_code=400, detail="follow_up_at must be in the future")
    message = (payload.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="Message is required")
    if payload.pause_reason and payload.pause_reason not in ALLOWED_PAUSE_REASONS:
        raise HTTPException(status_code=400, detail="Unsupported pause reason")

    template_id = payload.template_id
    if template_id:
        template = db.query(LeadInfoTemplate).filter(LeadInfoTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")

    comm = LeadCommunication(
        lead_id=lead.id,
        sent_by=current_user.id,
        template_id=template_id,
        channel=(payload.channel or "messenger").strip(),
        message=message,
        pause_reason=payload.pause_reason,
        follow_up_at=payload.follow_up_at,
    )
    db.add(comm)

    auto_task = LeadTask(
        lead_id=lead.id,
        owner_id=current_user.id,
        note=f"[auto-follow-up] {payload.pause_reason or 'Р±РµР· РїСЂРёС‡РёРЅС‹'}",
        channel=(payload.channel or "messenger").strip(),
        due_at=payload.follow_up_at,
        status=LeadTaskStatus.OPEN,
    )
    db.add(auto_task)
    lead.next_contact_at = payload.follow_up_at
    lead.pause_reason = payload.pause_reason
    if lead.status == LeadStatus.NEW:
        lead.status = LeadStatus.CONTACTED
    db.commit()
    db.refresh(comm)
    log_action(
        db,
        current_user.id,
        "send_info",
        "lead",
        lead.id,
        {
            "template_id": template_id,
            "channel": payload.channel,
            "follow_up_at": payload.follow_up_at.isoformat(),
            "pause_reason": payload.pause_reason,
        },
    )
    return comm


@router.post("/_legacy-disabled/leads/{lead_id}/contact-result", response_model=LeadCommunicationResponse, status_code=status.HTTP_201_CREATED)
async def save_lead_contact_result(
    lead_id: int,
    payload: LeadContactResultRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    outcome = (payload.outcome or "").strip().lower()
    if outcome not in {"connected", "no_answer", "callback"}:
        raise HTTPException(status_code=400, detail="Unsupported contact outcome")

    if outcome in {"no_answer", "callback"} and payload.follow_up_at is None:
        raise HTTPException(status_code=400, detail="follow_up_at is required for this outcome")
    if payload.follow_up_at and payload.follow_up_at <= utcnow():
        raise HTTPException(status_code=400, detail="follow_up_at must be in the future")

    label_map = {
        "connected": "РґРѕР·РІРѕРЅ",
        "no_answer": "РЅРµ РґРѕР·РІРѕРЅРёР»РёСЃСЊ",
        "callback": "РїРµСЂРµР·РІРѕРЅРёС‚СЊ",
    }
    message = f"[contact-result] {label_map[outcome]}"
    if payload.note and payload.note.strip():
        message = f"{message}: {payload.note.strip()}"

    comm = LeadCommunication(
        lead_id=lead.id,
        sent_by=current_user.id,
        template_id=None,
        channel="call",
        message=message,
        pause_reason=None,
        follow_up_at=payload.follow_up_at or utcnow(),
    )
    db.add(comm)

    if outcome in {"no_answer", "callback"} and payload.follow_up_at:
        auto_task = LeadTask(
            lead_id=lead.id,
            owner_id=current_user.id,
            note=f"[auto-follow-up] {label_map[outcome]}",
            channel="call",
            due_at=payload.follow_up_at,
            status=LeadTaskStatus.OPEN,
        )
        db.add(auto_task)
        lead.next_contact_at = payload.follow_up_at

    if outcome in {"connected", "callback"} and lead.status == LeadStatus.NEW:
        lead.status = LeadStatus.CONTACTED

    db.commit()
    db.refresh(comm)
    log_action(
        db,
        current_user.id,
        "contact_result",
        "lead",
        lead.id,
        {
            "outcome": outcome,
            "follow_up_at": payload.follow_up_at.isoformat() if payload.follow_up_at else None,
        },
    )
    return comm


@router.post("/_legacy-disabled/leads/import-xlsx", response_model=LeadImportResponse)
async def import_leads_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="РџРѕРґРґРµСЂР¶РёРІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ С„РѕСЂРјР°С‚ .xlsx")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")

    wb = load_workbook(filename=BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return LeadImportResponse(created=0, skipped=0, errors=["РџСѓСЃС‚РѕР№ Р»РёСЃС‚"])

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    def val(row, key_variants: List[str]) -> Optional[str]:
        for key in key_variants:
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                continue
            raw = row[idx]
            if raw is None:
                continue
            txt = str(raw).strip()
            if txt:
                return txt
        return None

    def parse_row_datetime(raw_value) -> Optional[datetime]:
        if raw_value is None:
            return None
        if isinstance(raw_value, datetime):
            return raw_value
        text = str(raw_value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None

    def parse_row_minutes(raw_value) -> Optional[int]:
        if raw_value is None:
            return None
        text = str(raw_value).strip()
        if not text:
            return None
        try:
            minutes = int(float(text.replace(",", ".")))
            return minutes if minutes >= 0 else None
        except ValueError:
            return None

    created = 0
    skipped = 0
    errors: List[str] = []
    for i, row in enumerate(rows[1:], start=2):
        parent_name = val(row, ["С„РёРѕ СЂРѕРґРёС‚РµР»СЏ", "СЂРѕРґРёС‚РµР»СЊ", "parent_full_name"])
        child_name = val(row, ["С„РёРѕ СЂРµР±РµРЅРєР°", "С„РёРѕ СЂРµР±С‘РЅРєР°", "СЂРµР±РµРЅРѕРє", "СЂРµР±С‘РЅРѕРє", "child_full_name"])
        parent_phone = val(row, ["С‚РµР»РµС„РѕРЅ СЂРѕРґРёС‚РµР»СЏ", "parent_phone"])
        child_phone = val(row, ["С‚РµР»РµС„РѕРЅ С€РєРѕР»СЊРЅРёРєР°", "С‚РµР»РµС„РѕРЅ СЂРµР±РµРЅРєР°", "С‚РµР»РµС„РѕРЅ СЂРµР±С‘РЅРєР°", "child_phone"])
        source_name_raw = val(row, ["РёСЃС‚РѕС‡РЅРёРє", "source"])
        referral_name = val(row, ["РєС‚Рѕ РїСЂРёРіР»Р°СЃРёР»", "СЂРµРєРѕРјРµРЅРґРѕРІР°Р»", "referral_name"])
        comment = val(row, ["РєРѕРјРјРµРЅС‚Р°СЂРёР№", "comment"])
        school_name = val(row, ["С€РєРѕР»Р°", "school", "school_name"])
        school_class = val(row, ["РєР»Р°СЃСЃ", "class", "school_class"])
        outreach_date_raw = val(row, ["РґР°С‚Р° РѕР±С…РѕРґР°", "outreach_date", "outreach_at"])
        outreach_minutes_raw = val(row, ["РІСЂРµРјСЏ РѕР±С…РѕРґР° (РјРёРЅ)", "РІСЂРµРјСЏ РѕР±С…РѕРґР°", "outreach_minutes"])
        outreach_at = parse_row_datetime(outreach_date_raw)
        outreach_minutes = parse_row_minutes(outreach_minutes_raw)

        if not any([parent_name, child_name, parent_phone, child_phone, source_name_raw, comment]):
            skipped += 1
            continue

        source_id, source_name = _resolve_source(db, None, source_name_raw)
        if _is_referral_source(source_name) and not (referral_name or "").strip():
            errors.append("РЎС‚СЂРѕРєР° {0}: РґР»СЏ РёСЃС‚РѕС‡РЅРёРєР° 'СЂРµРєРѕРјРµРЅРґР°С†РёСЏ' РЅРµ СѓРєР°Р·Р°РЅ РїСЂРёРіР»Р°СЃРёРІС€РёР№".format(i))
            skipped += 1
            continue

        contact_name = parent_name or child_name or "Р‘РµР· РёРјРµРЅРё"
        phone = parent_phone or child_phone or "РЅРµ СѓРєР°Р·Р°РЅ"
        lead = Lead(
            owner_id=current_user.id,
            contact_name=contact_name,
            phone=phone,
            phone_normalized=normalize_phone(parent_phone or child_phone or phone) or None,
            parent_full_name=parent_name,
            child_full_name=child_name,
            parent_phone=parent_phone,
            child_phone=child_phone,
            source=source_name,
            source_id=source_id,
            referral_name=referral_name,
            comment=comment,
            school_name=school_name,
            school_class=school_class,
            outreach_at=outreach_at,
            outreach_minutes=outreach_minutes,
            status=LeadStatus.NEW,
        )
        db.add(lead)
        db.flush()
        sync_lead_person(db, lead)
        created += 1

    db.commit()
    log_action(db, current_user.id, "import", "lead", None, {"created": created, "skipped": skipped})
    return LeadImportResponse(created=created, skipped=skipped, errors=errors)


@router.get("/_legacy-disabled/leads/import-template")
async def download_leads_import_template(
    current_user: User = Depends(require_sales_admin_owner),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "LeadsImport"
    headers = [
        "Р¤РРћ СЂРѕРґРёС‚РµР»СЏ",
        "Р¤РРћ СЂРµР±РµРЅРєР°",
        "РўРµР»РµС„РѕРЅ СЂРѕРґРёС‚РµР»СЏ",
        "РўРµР»РµС„РѕРЅ С€РєРѕР»СЊРЅРёРєР°",
        "РЁРєРѕР»Р°",
        "РљР»Р°СЃСЃ",
        "Р”Р°С‚Р° РѕР±С…РѕРґР°",
        "Р’СЂРµРјСЏ РѕР±С…РѕРґР° (РјРёРЅ)",
        "РСЃС‚РѕС‡РЅРёРє",
        "РљС‚Рѕ РїСЂРёРіР»Р°СЃРёР»",
        "РљРѕРјРјРµРЅС‚Р°СЂРёР№",
    ]
    ws.append(headers)
    ws.append(
        [
            "РРІР°РЅРѕРІР° РђРЅРЅР° РџРµС‚СЂРѕРІРЅР°",
            "РРІР°РЅРѕРІ РџРµС‚СЂ",
            "+7 999 111-22-33",
            "+7 900 111-22-44",
            "РЁРєРѕР»Р° в„–12",
            "7Рђ",
            "2026-02-01",
            "35",
            "СЂРµРєРѕРјРµРЅРґР°С†РёСЏ",
            "РњР°СЂРёСЏ РЎРёРґРѕСЂРѕРІР°",
            "РРЅС‚РµСЂРµСЃ Рє Р·Р°РЅСЏС‚РёСЏРј РїРѕСЃР»Рµ РїСЂРѕР±РЅРѕРіРѕ СѓСЂРѕРєР°",
        ]
    )
    ws.freeze_panes = "A2"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "leads_import_template.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post(
    "/_legacy-disabled/public/leads/specialist-questionnaire",
    response_model=SpecialistQuestionnaireResponse,
    status_code=status.HTTP_201_CREATED,
)
async def submit_specialist_questionnaire(
    payload: SpecialistQuestionnaireRequest,
    db: Session = Depends(get_db),
):
    """
    РџСѓР±Р»РёС‡РЅР°СЏ Р°РЅРєРµС‚Р° РґР»СЏ РЅР°РїСЂР°РІР»РµРЅРёСЏ В«РЎРїРµС†РёР°Р»РёСЃС‚В».
    Р”РѕСЃС‚СѓРїРЅР° Р±РµР· Р°РІС‚РѕСЂРёР·Р°С†РёРё, СЃРѕР·РґР°С‘С‚ Р»РёРґР° СЃ questionnaire_filled=True.
    """
    owner = (
        db.query(User)
        .filter(User.role.in_([UserRole.SALES, UserRole.OWNER, UserRole.ADMIN]))
        .order_by(User.id)
        .first()
    )
    if not owner:
        raise HTTPException(status_code=500, detail="No sales/owner/admin user configured")

    # РЎРѕР·РґР°С‘Рј Р»РёС‡РЅСѓСЋ Р°РЅРєРµС‚Сѓ СѓС‡РµРЅРёРєР° (StudentCard) СЃРѕ РІСЃРµРјРё РїРѕР»СЏРјРё
    card = StudentCard(
        student_full_name=payload.child_full_name,
        birth_date=payload.birth_date,
        student_phone=payload.child_phone,
        phone_normalized=normalize_phone(payload.parent_phone or payload.child_phone or "") or None,
        telegram=payload.child_telegram,
        gender=payload.gender,
        on_grant=False,
        format_type=None,
        city=payload.city,
        school=payload.school_name,
        grade=payload.school_class,
        parent_full_name=payload.parent_full_name,
        parent_phone=payload.parent_phone,
        parent_phone_2=payload.parent_phone_2,
        parent_telegram=payload.parent_telegram,
        parent_email=payload.parent_email,
        student_email=payload.student_email,
        preferred_messenger=payload.preferred_messenger,
        comment=payload.comment,
        source=payload.source or "РђРЅРєРµС‚Р° РЎРїРµС†РёР°Р»РёСЃС‚",
        discount_type=DiscountType.NONE,
        discount_value=0.0,
        anketa_status="filled",
    )

    extra_parts: List[str] = []
    if payload.birth_date:
        extra_parts.append(f"Р”Р°С‚Р° СЂРѕР¶РґРµРЅРёСЏ: {payload.birth_date.isoformat()}")
    if payload.child_phone:
        extra_parts.append(f"РўРµР»РµС„РѕРЅ СѓС‡РµРЅРёРєР°: {payload.child_phone}")
    if payload.child_telegram:
        extra_parts.append(f"РўРµР»РµРіСЂР°Рј СѓС‡РµРЅРёРєР°: {payload.child_telegram}")
    if payload.gender:
        extra_parts.append(f"РџРѕР»: {payload.gender}")
    if payload.parent_phone_2:
        extra_parts.append(f"Р’С‚РѕСЂРѕР№ С‚РµР»РµС„РѕРЅ СЂРѕРґРёС‚РµР»СЏ: {payload.parent_phone_2}")
    if payload.parent_telegram:
        extra_parts.append(f"РўРµР»РµРіСЂР°Рј СЂРѕРґРёС‚РµР»СЏ: {payload.parent_telegram}")
    if payload.student_email:
        extra_parts.append(f"Email СѓС‡РµРЅРёРєР°: {payload.student_email}")
    if payload.preferred_messenger:
        extra_parts.append(f"РњРµСЃСЃРµРЅРґР¶РµСЂ: {payload.preferred_messenger}")

    base_comment = payload.comment or ""
    extras_str = "\n".join(extra_parts) if extra_parts else ""
    full_comment = base_comment
    if extras_str:
        full_comment = (base_comment + "\n\n" if base_comment else "") + extras_str

    questionnaire_data = payload.model_dump(mode="json")  # РІСЃРµ РїРѕР»СЏ С„РѕСЂРјС‹ РґР»СЏ РѕС‚РѕР±СЂР°Р¶РµРЅРёСЏ РІ РєР°СЂС‚РѕС‡РєРµ Р»РёРґР°
    lead = Lead(
        owner_id=owner.id,
        contact_name=payload.parent_full_name,
        phone=payload.parent_phone,
        phone_normalized=normalize_phone(payload.parent_phone or payload.child_phone or "") or None,
        parent_full_name=payload.parent_full_name,
        child_full_name=payload.child_full_name,
        parent_phone=payload.parent_phone,
        child_phone=payload.child_phone,
        email=payload.parent_email or payload.student_email,
        city=payload.city,
        school_name=payload.school_name,
        school_class=payload.school_class,
        comment=full_comment or None,
        source=payload.source or "РђРЅРєРµС‚Р° РЎРїРµС†РёР°Р»РёСЃС‚",
        tags=["direction:specialist"],
        status=LeadStatus.NEW,
        questionnaire_filled=True,
        questionnaire_data=questionnaire_data,
    )
    db.add(card)
    db.add(lead)
    db.flush()
    sync_student_card_person(db, card)
    lead.student_card_id = card.id
    sync_lead_person(db, lead)
    db.commit()
    db.refresh(lead)
    return SpecialistQuestionnaireResponse(lead_id=lead.id)


TILDA_SOURCE_START = "РўРёР»СЊРґР°_РџРµСЂРІС‹Р№ РЁР°Рі"
TILDA_SOURCE_BASE = "РўРёР»СЊРґР°_РЎРїРµС†РёР°Р»РёСЃС‚"
TILDA_SOURCE_PRO = "РўРёР»СЊРґР°_Р­РєСЃРїРµСЂС‚"


@router.post(
    "/_legacy-disabled/public/leads/tilda-lead",
    response_model=TildaLeadResponse,
    status_code=status.HTTP_201_CREATED,
)
async def submit_tilda_lead(
    payload: TildaLeadRequest,
    db: Session = Depends(get_db),
):
    """
    РџСѓР±Р»РёС‡РЅР°СЏ Р°РЅРєРµС‚Р° Р»РёРґР° СЃ СЃР°Р№С‚Р° Tilda.
    РЎРѕР·РґР°С‘С‚ Р»РёРґР° РІ СЃС‚Р°С‚СѓСЃРµ В«РЅРѕРІС‹Р№В» СЃ РїРѕРјРµС‚РєРѕР№ РёСЃС‚РѕС‡РЅРёРєР° РІ Р·Р°РІРёСЃРёРјРѕСЃС‚Рё РѕС‚ РЅР°РїСЂР°РІР»РµРЅРёСЏ:
    - РўРёР»СЊРґР°_РџРµСЂРІС‹Р№ РЁР°Рі (kind=start)
    - РўРёР»СЊРґР°_РЎРїРµС†РёР°Р»РёСЃС‚ (kind=base)
    - РўРёР»СЊРґР°_Р­РєСЃРїРµСЂС‚ (kind=pro)
    """
    parent_name = (payload.parent_full_name or "").strip()
    child_name = (payload.child_full_name or "").strip()
    if not parent_name:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ Р¤РРћ СЂРѕРґРёС‚РµР»СЏ")
    if not child_name:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ Р¤РРћ СѓС‡РµРЅРёРєР°")

    normalized_phone, phone_error = validate_phone_for_lead(payload.parent_phone)
    if phone_error:
        raise HTTPException(status_code=400, detail=phone_error)

    owner = (
        db.query(User)
        .filter(User.role.in_([UserRole.SALES, UserRole.OWNER, UserRole.ADMIN]))
        .order_by(User.id)
        .first()
    )
    if not owner:
        raise HTTPException(status_code=500, detail="Р’ СЃРёСЃС‚РµРјРµ РЅРµ РЅР°СЃС‚СЂРѕРµРЅ РїРѕР»СЊР·РѕРІР°С‚РµР»СЊ РґР»СЏ РїСЂРёС‘РјР° Р·Р°СЏРІРѕРє")

    kind = (payload.kind or "start").strip()
    if kind == "base":
        src_label = TILDA_SOURCE_BASE
        tag = "tilda_base_lead"
    elif kind == "pro":
        src_label = TILDA_SOURCE_PRO
        tag = "tilda_pro_lead"
    else:
        src_label = TILDA_SOURCE_START
        tag = "tilda_start_lead"

    source_id, source_name = _resolve_source(db, None, src_label)
    status_option_id = _get_default_lead_status_option_id(db, LeadStatus.NEW)

    lead = Lead(
        owner_id=owner.id,
        contact_name=parent_name,
        phone=normalized_phone,
        phone_normalized=normalized_phone,
        parent_full_name=parent_name,
        parent_phone=normalized_phone,
        child_full_name=child_name,
        source=source_name or src_label,
        source_id=source_id,
        status=LeadStatus.NEW,
        status_option_id=status_option_id,
        tags=[tag],
    )
    db.add(lead)
    db.flush()
    sync_lead_person(db, lead)
    db.commit()
    db.refresh(lead)
    return TildaLeadResponse(lead_id=lead.id)


@router.post("/_legacy-disabled/leads", response_model=LeadResponse, status_code=status.HTTP_201_CREATED)
async def create_lead(
    payload: LeadCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    effective_role = auth.resolve_effective_role(current_user)
    owner_id = payload.owner_id if (effective_role in (UserRole.ADMIN, UserRole.OWNER) and payload.owner_id) else current_user.id
    source_id, source_name = _resolve_source(db, payload.source_id, payload.source)
    if _is_referral_source(source_name) and not (payload.referral_name or "").strip():
        raise HTTPException(status_code=400, detail="Р”Р»СЏ РёСЃС‚РѕС‡РЅРёРєР° 'СЂРµРєРѕРјРµРЅРґР°С†РёСЏ' СѓРєР°Р¶РёС‚Рµ, РєС‚Рѕ РїСЂРёРіР»Р°СЃРёР»")

    abonement = None
    if payload.abonement_id:
        abonement = db.query(Abonement).filter(Abonement.id == payload.abonement_id).first()
        if not abonement:
            raise HTTPException(status_code=404, detail="Abonement not found")

    lead = Lead(
        owner_id=owner_id,
        contact_name=payload.contact_name,
        phone=payload.phone,
        phone_normalized=normalize_phone(payload.parent_phone or payload.phone or payload.child_phone or "") or None,
        parent_full_name=payload.parent_full_name,
        child_full_name=payload.child_full_name,
        parent_phone=payload.parent_phone,
        child_phone=payload.child_phone,
        email=payload.email,
        city=payload.city,
        school_name=payload.school_name,
        school_class=payload.school_class,
        outreach_at=payload.outreach_at,
        outreach_minutes=payload.outreach_minutes,
        source=source_name,
        source_id=source_id,
        referral_name=payload.referral_name,
        tags=payload.tags,
        abonement_id=payload.abonement_id,
        desired_slot=payload.desired_slot,
        comment=payload.comment,
        next_contact_at=payload.next_contact_at,
        status=LeadStatus.NEW,
    )
    db.add(lead)
    db.flush()
    sync_lead_person(db, lead)
    _add_activity(
        db, lead.id, current_user.id,
        type="lead_created",
        title="Р›РёРґ СЃРѕР·РґР°РЅ",
        description=f"РСЃС‚РѕС‡РЅРёРє: {source_name or 'вЂ”'}",
    )
    db.commit()
    db.refresh(lead)

    log_action(db, current_user.id, "create", "lead", lead.id, {"owner_id": owner_id})
    return lead


_SEND_INFO_TASK_MARKER = "РѕС‚РїСЂР°РІРёС‚СЊ РёРЅС„РѕСЂРјР°С†РёСЋ"


@router.get("/_legacy-disabled/leads/send-info-status")
async def get_leads_send_info_status(
    lead_ids: str = Query(..., description="Comma-separated lead IDs"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Р’РѕР·РІСЂР°С‰Р°РµС‚ РґР»СЏ РєР°Р¶РґРѕРіРѕ lead_id СЃС‚Р°С‚СѓСЃ Р·Р°РґР°С‡Рё В«РћС‚РїСЂР°РІРёС‚СЊ РёРЅС„РѕСЂРјР°С†РёСЋВ»: open, done, none."""
    if not lead_ids.strip():
        return {}
    ids = [int(x.strip()) for x in lead_ids.split(",") if x.strip()]
    if not ids:
        return {}
    base = _filter_query_by_role(db.query(Lead), current_user)
    allowed_ids = {lead.id for lead in base.filter(Lead.id.in_(ids)).all()}
    tasks = (
        db.query(LeadTask)
        .options(joinedload(LeadTask.template))
        .outerjoin(LeadTaskTemplate, LeadTask.template_id == LeadTaskTemplate.id)
        .filter(
            LeadTask.lead_id.in_(allowed_ids),
            or_(
                cast(LeadTask.note, Text).ilike(f"%{_SEND_INFO_TASK_MARKER}%"),
                cast(LeadTaskTemplate.name, Text).ilike(f"%{_SEND_INFO_TASK_MARKER}%"),
            ),
        )
        .all()
    )
    result: Dict[str, str] = {str(i): "none" for i in ids if i in allowed_ids}

    # 1) РЎС‚Р°С‚СѓСЃС‹ РїРѕ LeadTask (СЃС‚Р°СЂС‹Р№ РјРµС…Р°РЅРёР·Рј)
    for t in tasks:
        if t.lead_id not in allowed_ids:
            continue
        key = str(t.lead_id)
        if t.status == LeadTaskStatus.OPEN:
            result[key] = "open"
        elif result.get(key) != "open":
            result[key] = "done"

    # 2) Р”РѕРїРѕР»РЅРёС‚РµР»СЊРЅРѕ СѓС‡РёС‚С‹РІР°РµРј РѕР±С‰РёРµ Р·Р°РґР°С‡Рё Task СЃ С‚РµРіР°РјРё ["send_info", f"lead:{id}"].
    #    РћРЅРё Р”РћР›Р–РќР« РїРµСЂРµРѕРїСЂРµРґРµР»СЏС‚СЊ СЃС‚Р°С‚СѓСЃ, РґР°Р¶Рµ РµСЃР»Рё LeadTask РµС‰С‘ РѕС‚РєСЂС‹С‚:
    #    - РµСЃС‚СЊ С…РѕС‚СЏ Р±С‹ РѕРґРЅР° Р°РєС‚РёРІРЅР°СЏ Task в†’ open
    #    - Р°РєС‚РёРІРЅС‹С… РЅРµС‚, РЅРѕ РµСЃС‚СЊ Р°СЂС…РёРІРЅС‹Рµ в†’ done
    common_tasks = (
        db.query(Task)
        .filter(Task.category == "leads", Task.tags.isnot(None))
        .all()
    )
    per_lead_flags: Dict[int, Dict[str, bool]] = {}
    for ct in common_tasks:
        tags = ct.tags or []
        lead_tag = next((t for t in tags if isinstance(t, str) and t.startswith("lead:")), None)
        if not lead_tag:
            continue
        try:
            lead_id = int(lead_tag.split(":", 1)[1])
        except (ValueError, IndexError):
            continue
        if lead_id not in allowed_ids:
            continue
        flags = per_lead_flags.setdefault(lead_id, {"has_active": False, "has_any": False})
        flags["has_any"] = True
        if ct.status == TaskStatus.ACTIVE.value:
            flags["has_active"] = True

    for lead_id, flags in per_lead_flags.items():
        key = str(lead_id)
        if flags["has_active"]:
            result[key] = "open"
        else:
            result[key] = "done"

    return result


@router.get("/_legacy-disabled/leads/badges")
async def get_leads_badges(
    lead_ids: str = Query(..., description="Comma-separated lead IDs"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Returns badge flags (has_invoice, has_task_today, is_overdue) for kanban cards."""
    if not lead_ids.strip():
        return {}
    ids = [int(x.strip()) for x in lead_ids.split(",") if x.strip()]
    if not ids:
        return {}
    base = _filter_query_by_role(db.query(Lead), current_user)
    allowed_ids = {lead.id for lead in base.filter(Lead.id.in_(ids)).all()}

    today_start = datetime.combine(date.today(), dt_time.min)
    today_end = datetime.combine(date.today(), dt_time.max)

    # Leads with open invoices
    invoice_lead_ids = set(
        r[0] for r in db.query(Invoice.lead_id).filter(
            Invoice.lead_id.in_(allowed_ids),
            Invoice.status != InvoiceStatus.PAID,
        ).distinct().all()
    )

    # Leads with tasks due today
    task_today_lead_ids = set(
        r[0] for r in db.query(LeadTask.lead_id).filter(
            LeadTask.lead_id.in_(allowed_ids),
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at >= today_start,
            LeadTask.due_at <= today_end,
        ).distinct().all()
    )

    # Leads with overdue tasks
    overdue_lead_ids = set(
        r[0] for r in db.query(LeadTask.lead_id).filter(
            LeadTask.lead_id.in_(allowed_ids),
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at < today_start,
            LeadTask.due_at.isnot(None),
        ).distinct().all()
    )

    result = {}
    for lid in allowed_ids:
        result[str(lid)] = {
            "has_invoice": lid in invoice_lead_ids,
            "has_task_today": lid in task_today_lead_ids,
            "is_overdue": lid in overdue_lead_ids,
        }
    return result


@router.get("/_legacy-disabled/leads/no-show-ids", response_model=List[int])
async def list_no_show_lead_ids(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Return IDs of leads that have a no-show registration. Single query, replaces N+1 on frontend."""
    rows = (
        db.query(EventRegistration.lead_id)
        .filter(
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
            or_(
                cast(EventRegistration.note, Text).ilike("%[no-show]%"),
                cast(EventRegistration.note, Text).ilike("%no-show%"),
            ),
        )
        .distinct()
        .all()
    )
    return [r[0] for r in rows]


@router.get("/_legacy-disabled/leads/{lead_id}", response_model=LeadResponse)
async def get_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    lead = _fix_lead_strings(lead)
    lead.ai_insight = build_lead_ai_insight(lead)
    return lead


@router.put("/_legacy-disabled/leads/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: int,
    payload: LeadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    update_data = payload.dict(exclude_unset=True)
    old_status = lead.status.value
    # Prevent sales from changing owner/status to restricted values
    if "status" in update_data and update_data["status"] is not None:
        lead.status = update_data["status"]
    if "lost_reason" in update_data:
        lead.lost_reason = update_data["lost_reason"]

    if "source_id" in update_data or "source" in update_data:
        source_id, source_name = _resolve_source(db, update_data.get("source_id"), update_data.get("source"))
        if _is_referral_source(source_name) and not (update_data.get("referral_name") or lead.referral_name or "").strip():
            raise HTTPException(status_code=400, detail="Р”Р»СЏ РёСЃС‚РѕС‡РЅРёРєР° 'СЂРµРєРѕРјРµРЅРґР°С†РёСЏ' СѓРєР°Р¶РёС‚Рµ, РєС‚Рѕ РїСЂРёРіР»Р°СЃРёР»")
        lead.source_id = source_id
        lead.source = source_name

    for field in [
        "contact_name",
        "phone",
        "parent_full_name",
        "child_full_name",
        "parent_phone",
        "child_phone",
        "email",
        "city",
        "school_name",
        "school_class",
        "outreach_at",
        "outreach_minutes",
        "referral_name",
        "tags",
        "abonement_id",
        "desired_slot",
        "comment",
        "next_contact_at",
        "pause_reason",
        "questionnaire_filled",
        "no_answer_attempt",
        "max_user_id",
    ]:
        if field in update_data:
            setattr(lead, field, update_data[field])
    if any(field in update_data for field in ("phone", "parent_phone", "child_phone")):
        lead.phone_normalized = normalize_phone(
            update_data.get("parent_phone", lead.parent_phone)
            or update_data.get("phone", lead.phone)
            or update_data.get("child_phone", lead.child_phone)
            or ""
        ) or None
    sync_lead_person(db, lead)

    new_status = lead.status.value
    if "status" in update_data and old_status != new_status:
        _add_activity(
            db, lead_id, current_user.id,
            type="status_changed",
            title="РЎС‚Р°С‚СѓСЃ РёР·РјРµРЅС‘РЅ",
            status_effect_from=old_status,
            status_effect_to=new_status,
        )

    db.commit()
    db.refresh(lead)
    log_action(db, current_user.id, "update", "lead", lead.id, update_data)
    return lead


@router.delete("/_legacy-disabled/leads/{lead_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """
    РџРѕР»РЅРѕРµ СѓРґР°Р»РµРЅРёРµ Р»РёРґР° Рё СЃРІСЏР·Р°РЅРЅС‹С… СЃСѓС‰РЅРѕСЃС‚РµР№ (Р·Р°РґР°С‡Рё, РєРѕРјРјСѓРЅРёРєР°С†РёРё, СЃС‡РµС‚Р° Рё С‚.Рґ.).
    РСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ, РєРѕРіРґР° Р»РёРґ СЃРѕР·РґР°РЅ РїРѕ РѕС€РёР±РєРµ РёР»Рё СЏРІРЅРѕ РЅРµ РЅСѓР¶РµРЅ РІ СЃРёСЃС‚РµРјРµ.
    """
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    log_action(db, current_user.id, "delete", "lead", lead.id, None)
    db.delete(lead)
    db.commit()
    return None


@router.post("/_legacy-disabled/leads/{lead_id}/convert-to-student", response_model=LeadConvertToStudentResponse)
async def convert_lead_to_student(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    """
    РџРµСЂРµРІРѕРґРёС‚ Р»РёРґР° РІ СѓС‡РµРЅРёРєР°: СЃРѕР·РґР°С‘С‚ СЂРѕРґРёС‚РµР»СЏ (РµСЃР»Рё РЅРµС‚ РїРѕ email) Рё СѓС‡РµРЅРёРєР° СЃ from_lead_id,
    РїСЂРёРІСЏР·С‹РІР°РµС‚ РёР»Рё СЃРѕР·РґР°С‘С‚ Р°РЅРєРµС‚Сѓ (StudentCard) РёР· РґР°РЅРЅС‹С… Р»РёРґР°/РѕРїСЂРѕСЃР°.
    РћР±РЅРѕРІР»СЏРµС‚ Р»РёРґР° (status=WON, converted_to_student_id).
    Р‘РёР·РЅРµСЃ-Р»РѕРіРёРєР° РІС‹РЅРµСЃРµРЅР° РІ app.services.lead_conversion.
    """
    _require_sales_admin_owner(current_user)
    try:
        result = lead_conversion_convert(db, lead_id, actor_user_id=current_user.id)
    except ValueError as e:
        msg = str(e)
        if "РЅРµ РЅР°Р№РґРµРЅ" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)
    return LeadConvertToStudentResponse(
        student_id=result.student_id,
        lead=_fix_lead_strings(result.lead),
    )


@router.post("/_legacy-disabled/leads/{lead_id}/tasks", response_model=LeadTaskResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task(
    lead_id: int,
    payload: LeadTaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    template = None
    if payload.template_id:
        template = (
            db.query(LeadTaskTemplate)
            .filter(LeadTaskTemplate.id == payload.template_id, LeadTaskTemplate.is_active.is_(True))
            .first()
        )
        if not template:
            raise HTTPException(status_code=404, detail="Task template not found")

    status_option_id = payload.status_option_id
    if status_option_id:
        status_option = (
            db.query(LeadTaskStatusOptionModel)
            .filter(LeadTaskStatusOptionModel.id == status_option_id, LeadTaskStatusOptionModel.is_active.is_(True))
            .first()
        )
        if not status_option:
            raise HTTPException(status_code=404, detail="Task status not found")
    else:
        default_open = (
            db.query(LeadTaskStatusOptionModel)
            .filter(
                LeadTaskStatusOptionModel.is_active.is_(True),
                LeadTaskStatusOptionModel.is_closed.is_(False),
            )
            .order_by(LeadTaskStatusOptionModel.id.asc())
            .first()
        )
        status_option_id = default_open.id if default_open else None

    task = LeadTask(
        lead_id=lead_id,
        owner_id=current_user.id,
        template_id=payload.template_id,
        status_option_id=status_option_id,
        note=payload.note,
        channel=payload.channel,
        due_at=payload.due_at,
        status=LeadTaskStatus.OPEN,
    )
    db.add(task)

    # Р•СЃР»Рё СЌС‚Рѕ Р·Р°РґР°С‡Р° В«РћС‚РїСЂР°РІРёС‚СЊ РёРЅС„РѕСЂРјР°С†РёСЋВ» вЂ” СЃРѕР·РґР°С‘Рј С‚Р°РєР¶Рµ РѕР±С‰СѓСЋ Р·Р°РґР°С‡Сѓ РІ /tasks
    # (РєР°С‚РµРіРѕСЂРёСЏ leads), С‡С‚РѕР±С‹ РјРµРЅРµРґР¶РµСЂ СѓРІРёРґРµР» РµС‘ РІ В«РџР»Р°РЅРµ РЅР° СЃРµРіРѕРґРЅСЏВ».
    note_lower = (payload.note or "").strip().lower() if payload.note else ""
    if _SEND_INFO_TASK_MARKER in note_lower:
        from app.models import Task, TaskStatus  # Р»РѕРєР°Р»СЊРЅС‹Р№ РёРјРїРѕСЂС‚, С‡С‚РѕР±С‹ РЅРµ РїР»РѕРґРёС‚СЊ Р·Р°РІРёСЃРёРјРѕСЃС‚Рё РЅР°РІРµСЂС…Сѓ

        title = f"РћС‚РїСЂР°РІРёС‚СЊ РёРЅС„РѕСЂРјР°С†РёСЋ: {lead.parent_full_name or lead.contact_name or lead.phone or f'Р›РёРґ #{lead.id}'}"
        description = f"РћС‚РїСЂР°РІРёС‚СЊ РёРЅС„РѕСЂРјР°С†РёСЋ РїРѕ Р»РёРґСѓ #{lead.id}"
        assignee_id = lead.owner_id or current_user.id
        common_task = Task(
            title=title,
            description=description,
            created_by_id=current_user.id,
            assigned_to_id=assignee_id,
            category="leads",
            status=TaskStatus.ACTIVE.value,
            due_at=utcnow(),
            priority="normal",
            tags=["send_info", f"lead:{lead.id}"],
        )
        db.add(common_task)

    _add_activity(
        db, lead_id, current_user.id,
        type="task_created",
        title=f"РЎРѕР·РґР°РЅР° Р·Р°РґР°С‡Р°: {payload.note or 'Р±РµР· РЅР°Р·РІР°РЅРёСЏ'}",
        description=f"РЎСЂРѕРє: {payload.due_at.strftime('%d.%m.%Y %H:%M') if payload.due_at else 'РЅРµ СѓРєР°Р·Р°РЅ'}",
        related_task_id=task.id,
    )
    db.commit()
    db.refresh(task)
    log_action(db, current_user.id, "create", "lead_task", task.id, {"lead_id": lead_id})
    return task


@router.get("/_legacy-disabled/leads/{lead_id}/tasks", response_model=List[LeadTaskResponse])
async def list_lead_tasks(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    tasks = (
        db.query(LeadTask)
        .filter(LeadTask.lead_id == lead_id)
        .order_by(LeadTask.created_at.desc())
        .all()
    )
    return tasks


@router.post("/_legacy-disabled/leads/{lead_id}/tasks/{task_id}/close", response_model=LeadTaskResponse)
async def close_lead_task(
    lead_id: int,
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    task = (
        db.query(LeadTask)
        .options(joinedload(LeadTask.template))
        .filter(LeadTask.id == task_id, LeadTask.lead_id == lead_id)
        .first()
    )
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    task.status = LeadTaskStatus.DONE
    closed_status = (
        db.query(LeadTaskStatusOptionModel)
        .filter(
            LeadTaskStatusOptionModel.is_active.is_(True),
            LeadTaskStatusOptionModel.is_closed.is_(True),
        )
        .order_by(LeadTaskStatusOptionModel.id.asc())
        .first()
    )
    if closed_status:
        task.status_option_id = closed_status.id
    task.updated_at = utcnow()
    db.commit()
    db.refresh(task)

    # Р•СЃР»Рё Р·Р°РєСЂС‹Р»Рё Р·Р°РґР°С‡Сѓ В«РћС‚РїСЂР°РІРёС‚СЊ РёРЅС„РѕСЂРјР°С†РёСЋВ» вЂ” РїРµСЂРµРІРѕРґРёРј Р»РёРґР° РІ В«РџРѕРґСѓРјР°СЋС‚В» Рё СЃРѕР·РґР°С‘Рј Р·Р°РґР°С‡Сѓ В«РџРѕР·РІРѕРЅРёС‚СЊ Р»РёРґСѓ Рё СѓР·РЅР°С‚СЊ СЂРµС€РµРЅРёРµВ» С‡РµСЂРµР· 2 РґРЅСЏ
    _FOLLOW_UP_NOTE = "РџРѕР·РІРѕРЅРёС‚СЊ Р»РёРґСѓ Рё СѓР·РЅР°С‚СЊ СЂРµС€РµРЅРёРµ"
    template_name = (task.template.name if task.template else "") or ""
    note_lower = (task.note or "").lower()
    is_send_info = _SEND_INFO_TASK_MARKER in template_name.lower() or _SEND_INFO_TASK_MARKER in note_lower
    if is_send_info:
        lead.status = LeadStatus.THINKING
        default_open = (
            db.query(LeadTaskStatusOptionModel)
            .filter(
                LeadTaskStatusOptionModel.is_active.is_(True),
                LeadTaskStatusOptionModel.is_closed.is_(False),
            )
            .order_by(LeadTaskStatusOptionModel.id.asc())
            .first()
        )
        follow_up_due = utcnow() + timedelta(days=2)
        follow_up = LeadTask(
            lead_id=lead_id,
            owner_id=current_user.id,
            template_id=None,
            status_option_id=default_open.id if default_open else None,
            note=_FOLLOW_UP_NOTE,
            channel=task.channel,
            due_at=follow_up_due,
            status=LeadTaskStatus.OPEN,
        )
        db.add(follow_up)
        db.commit()

    _add_activity(
        db, lead_id, current_user.id,
        type="task_done",
        title=f"Р—Р°РґР°С‡Р° РІС‹РїРѕР»РЅРµРЅР°: {task.note or 'Р±РµР· РЅР°Р·РІР°РЅРёСЏ'}",
        related_task_id=task.id,
    )
    if not is_send_info:
        db.commit()

    log_action(db, current_user.id, "close", "lead_task", task.id, {"lead_id": lead_id})
    return task


@router.put("/_legacy-disabled/leads/{lead_id}/tasks/{task_id}", response_model=LeadTaskResponse)
async def update_lead_task(
    lead_id: int,
    task_id: int,
    payload: LeadTaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    task = db.query(LeadTask).filter(LeadTask.id == task_id, LeadTask.lead_id == lead_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    update_data = payload.dict(exclude_unset=True)
    if "status" in update_data and update_data["status"] is not None:
        task.status = update_data["status"]
    if "note" in update_data:
        task.note = update_data["note"]
    if "channel" in update_data:
        task.channel = update_data["channel"]
    if "due_at" in update_data:
        task.due_at = update_data["due_at"]
    task.updated_at = utcnow()
    db.commit()
    db.refresh(task)
    log_action(db, current_user.id, "update", "lead_task", task.id, update_data)
    return task


@router.post("/_legacy-disabled/leads/{lead_id}/invoices", response_model=InvoiceResponse, status_code=status.HTTP_201_CREATED)
async def create_invoice_for_lead(
    lead_id: int,
    payload: InvoiceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    abonement = db.query(Abonement).filter(Abonement.id == payload.abonement_id).first()
    if not abonement:
        raise HTTPException(status_code=404, detail="Abonement not found")

    amount = _compute_price(abonement)
    invoice = Invoice(
        lead_id=lead_id,
        abonement_id=abonement.id,
        amount=amount,
        currency=payload.currency or "RUB",
        status=InvoiceStatus.DRAFT,
        email_to=payload.email_to or lead.email,
        link=None,
    )
    db.add(invoice)
    old_status = lead.status.value
    if lead.status not in (LeadStatus.WON, LeadStatus.LOST):
        lead.status = LeadStatus.INVOICE_SENT
    db.flush()  # get invoice.id before activity
    _add_activity(
        db, lead_id, current_user.id,
        type="invoice_created",
        title=f"Р’С‹СЃС‚Р°РІР»РµРЅ СЃС‡С‘С‚ РЅР° {amount} {payload.currency or 'RUB'}",
        status_effect_from=old_status if old_status != lead.status.value else None,
        status_effect_to=lead.status.value if old_status != lead.status.value else None,
        related_invoice_id=invoice.id,
    )
    db.commit()
    db.refresh(invoice)
    log_action(db, current_user.id, "create", "invoice", invoice.id, {"lead_id": lead_id, "amount": amount})
    return invoice


@router.get("/_legacy-disabled/leads/{lead_id}/invoices", response_model=List[InvoiceResponse])
async def list_invoices_for_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    return (
        db.query(Invoice)
        .filter(Invoice.lead_id == lead_id)
        .order_by(Invoice.created_at.desc())
        .all()
    )


@router.get("/_legacy-disabled/leads/{lead_id}/card")
async def get_lead_card(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Aggregated endpoint for lead card first render."""
    lead = db.query(Lead).options(*_lead_eager_options()).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    # Next action вЂ” closest open task
    next_task = (
        db.query(LeadTask)
        .options(joinedload(LeadTask.owner))
        .filter(LeadTask.lead_id == lead_id, LeadTask.status == LeadTaskStatus.OPEN)
        .order_by(LeadTask.due_at.asc().nullslast())
        .first()
    )
    now = utcnow()
    today_start = datetime.combine(date.today(), dt_time.min)
    today_end = datetime.combine(date.today(), dt_time.max)
    if next_task:
        if next_task.due_at:
            if next_task.due_at < now:
                state = "overdue"
            elif today_start <= next_task.due_at <= today_end:
                state = "today"
            else:
                state = "on_time"
        else:
            state = "on_time"
        next_action = LeadNextAction(
            type="task",
            title=next_task.note or "Р—Р°РґР°С‡Р°",
            due_at=next_task.due_at,
            owner_name=next_task.owner.full_name if next_task.owner else None,
            task_id=next_task.id,
            state=state,
        )
    elif lead.next_contact_at:
        if lead.next_contact_at < now:
            state = "overdue"
        elif today_start <= lead.next_contact_at <= today_end:
            state = "today"
        else:
            state = "on_time"
        next_action = LeadNextAction(
            type="contact",
            title="РЎРІСЏР·Р°С‚СЊСЃСЏ",
            due_at=lead.next_contact_at,
            owner_name=lead.owner.full_name if lead.owner else None,
            state=state,
        )
    else:
        next_action = LeadNextAction(state="none")

    # Pinned comment
    pinned_comment = lead.comment if lead.comment else None

    # Sidebar summary
    upcoming_tasks = (
        db.query(LeadTask)
        .filter(LeadTask.lead_id == lead_id, LeadTask.status == LeadTaskStatus.OPEN)
        .order_by(LeadTask.due_at.asc().nullslast())
        .limit(3)
        .all()
    )
    latest_invoice = (
        db.query(Invoice)
        .filter(Invoice.lead_id == lead_id)
        .order_by(Invoice.created_at.desc())
        .first()
    )
    sidebar = LeadSidebarSummary(
        contacts={
            "phone": lead.parent_phone or lead.phone,
            "child_phone": lead.child_phone,
            "email": lead.email,
            "telegram": (lead.questionnaire_data or {}).get("parent_telegram") if lead.questionnaire_data else None,
            "communication_channel": lead.communication_channel,
        },
        source=lead.source,
        owner_name=lead.owner.full_name if lead.owner else None,
        created_at=lead.created_at,
        updated_at=lead.updated_at,
        upcoming_tasks=[
            {"id": t.id, "note": t.note, "due_at": t.due_at.isoformat() if t.due_at else None, "status": t.status.value}
            for t in upcoming_tasks
        ],
        latest_invoice={
            "id": latest_invoice.id,
            "amount": latest_invoice.amount,
            "currency": latest_invoice.currency,
            "status": latest_invoice.status.value,
        } if latest_invoice else None,
    )

    # Timeline preview (last 10 activities)
    activities = (
        db.query(LeadActivity)
        .options(joinedload(LeadActivity.creator))
        .filter(LeadActivity.lead_id == lead_id)
        .order_by(LeadActivity.created_at.desc())
        .limit(10)
        .all()
    )
    timeline_preview = [
        {
            "id": a.id,
            "type": a.type,
            "title": a.title,
            "description": a.description,
            "channel": a.channel,
            "created_at": a.created_at.isoformat() if a.created_at else None,
            "creator_name": a.creator.full_name if a.creator else None,
            "status_effect_from": a.status_effect_from,
            "status_effect_to": a.status_effect_to,
        }
        for a in activities
    ]

    lead = _fix_lead_strings(lead)
    lead.ai_insight = build_lead_ai_insight(lead)
    return {
        "lead": lead,
        "next_action": next_action,
        "pinned_comment": pinned_comment,
        "sidebar": sidebar,
        "timeline_preview": timeline_preview,
    }


@router.get("/_legacy-disabled/leads/{lead_id}/timeline", response_model=List[LeadActivityResponse])
async def get_lead_timeline(
    lead_id: int,
    offset: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Full timeline with pagination."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    activities = (
        db.query(LeadActivity)
        .options(joinedload(LeadActivity.creator))
        .filter(LeadActivity.lead_id == lead_id)
        .order_by(LeadActivity.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return [
        LeadActivityResponse(
            id=a.id,
            lead_id=a.lead_id,
            type=a.type,
            title=a.title,
            description=a.description,
            channel=a.channel,
            created_at=a.created_at,
            created_by=a.created_by,
            creator_name=a.creator.full_name if a.creator else None,
            payload_json=a.payload_json,
            status_effect_from=a.status_effect_from,
            status_effect_to=a.status_effect_to,
            related_task_id=a.related_task_id,
            related_invoice_id=a.related_invoice_id,
        )
        for a in activities
    ]


@router.post("/_legacy-disabled/leads/{lead_id}/activities", response_model=LeadActivityResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_activity(
    lead_id: int,
    payload: LeadActivityCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Create a new activity entry for a lead."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    activity = LeadActivity(
        lead_id=lead_id,
        type=payload.type,
        title=payload.title,
        description=payload.description,
        channel=payload.channel,
        created_by=current_user.id,
        payload_json=payload.payload_json,
        status_effect_from=payload.status_effect_from,
        status_effect_to=payload.status_effect_to,
        related_task_id=payload.related_task_id,
        related_invoice_id=payload.related_invoice_id,
    )
    db.add(activity)

    # Apply status effect if specified
    if payload.status_effect_to and payload.status_effect_to != lead.status.value:
        lead.status = LeadStatus(payload.status_effect_to)

    # Update last_contact_at for contact-type activities
    if payload.type in ("call", "no_answer", "info_sent"):
        lead.last_contact_at = utcnow()

    db.commit()
    db.refresh(activity)
    creator = db.query(User).filter(User.id == activity.created_by).first()
    return LeadActivityResponse(
        id=activity.id,
        lead_id=activity.lead_id,
        type=activity.type,
        title=activity.title,
        description=activity.description,
        channel=activity.channel,
        created_at=activity.created_at,
        created_by=activity.created_by,
        creator_name=creator.full_name if creator else None,
        payload_json=activity.payload_json,
        status_effect_from=activity.status_effect_from,
        status_effect_to=activity.status_effect_to,
        related_task_id=activity.related_task_id,
        related_invoice_id=activity.related_invoice_id,
    )


@router.post("/_legacy-disabled/leads/{lead_id}/invoices/{invoice_id}/mark-paid", response_model=InvoiceResponse)
async def mark_invoice_paid(
    lead_id: int,
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """Mark invoice as paid and create invoice_paid activity."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    invoice = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.lead_id == lead_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.status == InvoiceStatus.PAID:
        return invoice

    invoice.status = InvoiceStatus.PAID
    _add_activity(
        db, lead_id, current_user.id,
        type="invoice_paid",
        title=f"РЎС‡С‘С‚ РѕРїР»Р°С‡РµРЅ: {invoice.amount} {invoice.currency}",
        related_invoice_id=invoice.id,
    )
    db.commit()
    db.refresh(invoice)
    log_action(db, current_user.id, "mark_paid", "invoice", invoice.id, {"lead_id": lead_id})
    return invoice


@router.post("/_legacy-disabled/invoices/{invoice_id}/send-email", response_model=InvoiceResponse)
async def send_invoice_email(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    lead = db.query(Lead).filter(Lead.id == invoice.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    if not invoice.email_to and not lead.email:
        raise HTTPException(status_code=400, detail="No email to send invoice")

    # Stub: actual email sending integration should be implemented separately.
    invoice.status = InvoiceStatus.SENT
    invoice.sent_at = utcnow()
    if lead.status not in (LeadStatus.WON, LeadStatus.LOST):
        lead.status = LeadStatus.INVOICE_SENT
    db.commit()
    db.refresh(invoice)
    log_action(db, current_user.id, "send_email", "invoice", invoice.id, {"lead_id": lead.id})
    return invoice


@router.get("/_legacy-disabled/invoices", response_model=List[InvoiceResponse])
async def list_invoices(
    status_filter: Optional[InvoiceStatus] = None,
    lead_id: Optional[int] = None,
    created_from: Optional[datetime] = Query(default=None),
    created_to: Optional[datetime] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = (
        db.query(Invoice)
        .join(Lead, Lead.id == Invoice.lead_id)
        .options(joinedload(Invoice.lead))
        .order_by(Invoice.created_at.desc())
    )

    if status_filter:
        query = query.filter(Invoice.status == status_filter)
    if lead_id:
        query = query.filter(Invoice.lead_id == lead_id)
    if created_from:
        query = query.filter(Invoice.created_at >= created_from)
    if created_to:
        query = query.filter(Invoice.created_at <= created_to)

    return query.all()


@router.get("/_legacy-disabled/events", response_model=List[EventResponse])
async def list_events(
    status_filter: Optional[EventStatus] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    query = db.query(Event).order_by(Event.starts_at.asc())
    if status_filter:
        query = query.filter(Event.status == status_filter)
    return query.all()


@router.post("/_legacy-disabled/events", response_model=EventResponse, status_code=status.HTTP_201_CREATED)
async def create_event(
    payload: EventCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    event = Event(
        title=payload.title,
        description=payload.description,
        starts_at=payload.starts_at,
        ends_at=payload.ends_at,
        location=payload.location,
        capacity=payload.capacity,
        status=EventStatus.ACTIVE,
        created_by=current_user.id,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    log_action(db, current_user.id, "create", "event", event.id)
    return event


@router.put("/_legacy-disabled/events/{event_id}", response_model=EventResponse)
async def update_event(
    event_id: int,
    payload: EventUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    update_data = payload.dict(exclude_unset=True)
    for k, v in update_data.items():
        setattr(event, k, v)
    db.commit()
    db.refresh(event)
    log_action(db, current_user.id, "update", "event", event.id, update_data)
    return event


@router.get("/_legacy-disabled/events/{event_id}/registrations", response_model=List[EventRegistrationResponse])
async def list_event_registrations(
    event_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    query = db.query(EventRegistration).filter(EventRegistration.event_id == event_id).order_by(EventRegistration.created_at.desc())
    return query.all()


@router.post("/_legacy-disabled/events/{event_id}/registrations", response_model=EventRegistrationResponse, status_code=status.HTTP_201_CREATED)
async def create_event_registration(
    event_id: int,
    payload: EventRegistrationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    event = db.query(Event).filter(Event.id == event_id, Event.status == EventStatus.ACTIVE).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found or inactive")
    if event.capacity is not None:
        current_registered = (
            db.query(EventRegistration)
            .filter(
                EventRegistration.event_id == event_id,
                EventRegistration.status == EventRegistrationStatus.REGISTERED,
            )
            .count()
        )
        if current_registered >= event.capacity:
            raise HTTPException(status_code=400, detail="РЎРІРѕР±РѕРґРЅС‹С… СЃР»РѕС‚РѕРІ РЅРµС‚")

    lead = db.query(Lead).filter(Lead.id == payload.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    existing = db.query(EventRegistration).filter(
        EventRegistration.event_id == event_id,
        EventRegistration.lead_id == payload.lead_id
    ).first()
    if existing and existing.status == EventRegistrationStatus.REGISTERED:
        raise HTTPException(status_code=400, detail="Lead already registered")

    if existing:
        existing.status = EventRegistrationStatus.REGISTERED
        existing.note = payload.note
        existing.owner_id = current_user.id
        db.commit()
        db.refresh(existing)
        log_action(db, current_user.id, "restore", "event_registration", existing.id, {"event_id": event_id, "lead_id": lead.id})
        return existing

    reg = EventRegistration(
        event_id=event_id,
        lead_id=lead.id,
        owner_id=current_user.id,
        status=EventRegistrationStatus.REGISTERED,
        note=payload.note,
    )
    db.add(reg)
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "create", "event_registration", reg.id, {"event_id": event_id, "lead_id": lead.id})
    return reg


@router.post("/_legacy-disabled/events/{event_id}/registrations/{registration_id}/cancel", response_model=EventRegistrationResponse)
async def cancel_event_registration(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")

    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    reg.status = EventRegistrationStatus.CANCELLED
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "cancel", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    return reg


@router.post("/_legacy-disabled/events/{event_id}/registrations/{registration_id}/confirm", response_model=EventRegistrationResponse)
async def confirm_event_registration(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id,
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    reg.note = _append_note_tag(reg.note, "[confirmed]")
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "confirm", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    return reg


@router.post("/_legacy-disabled/events/{event_id}/registrations/{registration_id}/mark-came", response_model=EventRegistrationResponse)
async def mark_event_registration_came(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id,
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    # Р•СЃР»Рё СЂР°РЅРµРµ РїРѕРјРµС‚РёР»Рё РєР°Рє no-show, СѓР±РёСЂР°РµРј СЌС‚РѕС‚ С‚РµРі Рё СЃС‚Р°РІРёРј came
    cleaned_note = _remove_note_tag(reg.note, "[no-show]")
    reg.note = _append_note_tag(cleaned_note, "[came]")
    # РћР±РЅРѕРІР»СЏРµРј СЃС‚Р°С‚СѓСЃ Р»РёРґР° Рё СЃС‚Р°СЂС‚РѕРІСѓСЋ СЃС‚Р°РґРёСЋ РІРѕСЂРѕРЅРєРё В«Р”РѕР¶Р°С‚СЊ РЅР° РѕР±СѓС‡РµРЅРёРµВ»
    lead.status = LeadStatus.DEMO
    lead.status_option_id = _get_default_lead_status_option_id(db, LeadStatus.DEMO)
    if not getattr(lead, "post_visit_stage", None):
        lead.post_visit_stage = "new"
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "mark_came", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    # UX automation: after attendance create follow-up "offer course" (do not fail the request if this fails)
    try:
        if not _has_open_task_like(db, lead.id, "[auto_attended_offer]"):
            due_at = utcnow() + timedelta(hours=24)
            auto_task = _create_auto_event_task(
                db,
                lead=lead,
                owner_id=lead.owner_id,
                note="[auto_attended_offer] РџРѕСЃР»Рµ РјРµСЂРѕРїСЂРёСЏС‚РёСЏ: РїСЂРµРґР»РѕР¶РёС‚СЊ РєСѓСЂСЃ",
                due_at=due_at,
                preferred_template_keywords=["РєСѓСЂСЃ", "РїСЂРµРґР»РѕР¶", "РґРѕР¶РёРј", "offer"],
            )
            db.add(auto_task)
            db.flush()
            log_action(
                db,
                current_user.id,
                "create",
                "lead_task",
                auto_task.id,
                {"lead_id": lead.id, "type": "auto_attended_offer"},
            )
            if lead.next_contact_at is None or (auto_task.due_at and lead.next_contact_at > auto_task.due_at):
                lead.next_contact_at = auto_task.due_at
            db.commit()
    except Exception:
        db.rollback()
    return reg


@router.post("/_legacy-disabled/events/{event_id}/registrations/{registration_id}/mark-no-show", response_model=EventRegistrationResponse)
async def mark_event_registration_no_show(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id,
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    # Р•СЃР»Рё СЂР°РЅРµРµ РїРѕРјРµС‚РёР»Рё РєР°Рє РїСЂРёС€С‘Р», СѓР±РёСЂР°РµРј СЌС‚РѕС‚ С‚РµРі Рё СЃС‚Р°РІРёРј no-show
    cleaned_note = _remove_note_tag(reg.note, "[came]")
    reg.note = _append_note_tag(cleaned_note, "[no-show]")
    # UX automation: after no-show create reactivation follow-up.
    if not _has_open_task_like(db, lead.id, "[auto_no_show_reactivate]"):
        due_at = utcnow() + timedelta(hours=24)
        auto_task = _create_auto_event_task(
            db,
            lead=lead,
            owner_id=lead.owner_id,
            note="[auto_no_show_reactivate] No-show: СЂРµР°РєС‚РёРІР°С†РёСЏ Рё РїРµСЂРµРЅРѕСЃ",
            due_at=due_at,
            preferred_template_keywords=["СЂРµР°РєС‚РёРІР°С†", "РїРµСЂРµР·Р°Рї", "no-show", "РґРѕР¶РёРј"],
        )
        db.flush()
        log_action(
            db,
            current_user.id,
            "create",
            "lead_task",
            auto_task.id,
            {"lead_id": lead.id, "type": "auto_no_show_reactivate"},
        )
        if lead.next_contact_at is None or (auto_task.due_at and lead.next_contact_at > auto_task.due_at):
            lead.next_contact_at = auto_task.due_at
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "mark_no_show", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    return reg


@router.get("/_legacy-disabled/leads/{lead_id}/event-registrations", response_model=List[EventRegistrationResponse])
async def list_lead_event_registrations(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    return (
        db.query(EventRegistration)
        .filter(EventRegistration.lead_id == lead_id)
        .order_by(EventRegistration.created_at.desc())
        .all()
    )


@router.post("/_legacy-disabled/leads/{lead_id}/post-visit-stage", response_model=LeadResponse)
async def update_lead_post_visit_stage(
    lead_id: int,
    payload: LeadPostVisitStageUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """РћР±РЅРѕРІР»РµРЅРёРµ СЃС‚Р°РґРёРё РІРѕСЂРѕРЅРєРё В«Р”РѕР¶Р°С‚СЊ РЅР° РѕР±СѓС‡РµРЅРёРµВ» РїРѕСЃР»Рµ РјРµСЂРѕРїСЂРёСЏС‚РёСЏ."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    try:
        result = lead_post_visit_update_stage(
            db,
            lead_id,
            stage=payload.stage,
            review=payload.review,
            project_date=payload.project_date,
            decline_reason=payload.decline_reason,
            get_default_lead_status_option_id=_get_default_lead_status_option_id,
        )
    except ValueError as e:
        msg = str(e)
        if "not found" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)

    lead = result.lead
    if result.need_auto_task and not _has_open_task_like(db, lead.id, "[auto_post_visit_agreed]"):
        due_at = utcnow() + timedelta(hours=48)
        auto_task = _create_auto_event_task(
            db,
            lead=lead,
            owner_id=lead.owner_id,
            note="[auto_post_visit_agreed] РљРѕРЅС‚СЂРѕР»СЊ РѕРїР»Р°С‚С‹",
            due_at=due_at,
            preferred_template_keywords=["РѕРїР»Р°С‚", "РєРѕРЅС‚СЂРѕР»СЊ", "РґРѕР¶РёРј"],
        )
        db.add(auto_task)
        db.flush()
        log_action(db, current_user.id, "create", "lead_task", auto_task.id, {"lead_id": lead.id, "type": "auto_post_visit_agreed"})
        if lead.next_contact_at is None or (auto_task.due_at and lead.next_contact_at > auto_task.due_at):
            lead.next_contact_at = auto_task.due_at

    db.commit()
    db.refresh(lead)
    log_action(
        db,
        current_user.id,
        "post_visit_stage",
        "lead",
        lead.id,
        {
            "stage": payload.stage,
            "review": payload.review,
            "project_date": payload.project_date.isoformat() if payload.project_date else None,
            "decline_reason": payload.decline_reason,
        },
    )
    return _fix_lead_strings(lead)


@router.get("/_legacy-disabled/post-visit/leads", response_model=List[LeadResponse])
async def list_post_visit_leads(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sales_admin_owner),
):
    """
    Р›РёРґС‹ РґР»СЏ СЃС‚СЂР°РЅРёС†С‹ В«Р”РѕР¶Р°С‚СЊ РЅР° РѕР±СѓС‡РµРЅРёРµВ».

    Р›РѕРіРёРєР°:
    - Р±РµСЂС‘Рј Р»РёРґРѕРІ, РїРѕ РєРѕС‚РѕСЂС‹Рј РµСЃС‚СЊ СЂРµРіРёСЃС‚СЂР°С†РёСЏ РЅР° РјРµСЂРѕРїСЂРёСЏС‚РёРµ СЃ С‚РµРіРѕРј [came] (РЅР°Р¶Р°Р»Рё В«РџСЂРёС€РµР»В»);
    - РѕРіСЂР°РЅРёС‡РёРІР°РµРј РїРѕ РїСЂР°РІР°Рј (admin/owner/sales РІРёРґСЏС‚ С‚РѕР»СЊРєРѕ СЃРІРѕРё Р»РёРґС‹);
    - РµСЃР»Рё post_visit_stage РµС‰С‘ РЅРµ Р·Р°РґР°РЅР° вЂ” РїСЂРѕСЃС‚Р°РІР»СЏРµРј 'new' (РѕРґРЅРѕРєСЂР°С‚РЅРѕ).
    """
    came_lead_ids = (
        db.query(EventRegistration.lead_id)
        .filter(cast(EventRegistration.note, Text).ilike("%[came]%"))
        .distinct()
    )
    leads_q = (
        _filter_query_by_role(db.query(Lead).options(*_lead_eager_options()), current_user)
        .filter(Lead.id.in_(came_lead_ids))
        .order_by(Lead.created_at.desc())
    )
    leads = leads_q.all()

    updated = False
    for lead in leads:
        if not getattr(lead, "post_visit_stage", None):
            lead.post_visit_stage = "new"
            updated = True
    if updated:
        db.commit()
        for lead in leads:
            db.refresh(lead)

    return [_fix_lead_strings(lead) for lead in leads]
