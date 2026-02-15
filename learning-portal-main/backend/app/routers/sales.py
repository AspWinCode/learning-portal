from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import cast, Text, or_
from openpyxl import Workbook, load_workbook

from app import auth
from app.database import get_db
from app.models import (
    Lead,
    LeadStatus,
    LeadTask,
    LeadTaskStatus,
    LeadSource,
    LeadTaskTemplate,
    LeadTaskStatusOption as LeadTaskStatusOptionModel,
    LeadInfoTemplate,
    LeadCommunication,
    Invoice,
    InvoiceStatus,
    Event,
    EventStatus,
    EventRegistration,
    EventRegistrationStatus,
    Abonement,
    User,
    UserRole,
)
from app.schemas import (
    LeadCreate,
    LeadUpdate,
    LeadResponse,
    LeadTaskCreate,
    LeadTaskResponse,
    LeadTaskUpdate,
    InvoiceCreate,
    InvoiceResponse,
    EventCreate,
    EventUpdate,
    EventResponse,
    EventRegistrationCreate,
    EventRegistrationResponse,
    LeadSourceCreate,
    LeadSourceResponse,
    LeadSourceUpdate,
    LeadTaskTemplateCreate,
    LeadTaskTemplateResponse,
    LeadTaskTemplateUpdate,
    LeadTaskStatusOptionCreate,
    LeadTaskStatusOptionResponse,
    LeadTaskStatusOptionUpdate,
    LeadImportResponse,
    LeadInfoTemplateCreate,
    LeadInfoTemplateResponse,
    LeadInfoTemplateUpdate,
    LeadSendInfoRequest,
    LeadQuickCommunicationCreate,
    LeadContactResultRequest,
    LeadCommunicationResponse,
    SalesDashboardResponse,
    SalesSchoolConversionItem,
    SalesQueueTaskItem,
    SalesQueueRegistrationItem,
    FollowUpItemResponse,
    LeadPushStatsResponse,
)
from app.routers.action_log import log_action

router = APIRouter()


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _to_utc(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _require_owner_or_admin(lead: Lead, user: User) -> None:
    if user.role in (UserRole.ADMIN, UserRole.OWNER):
        return
    if user.role == UserRole.SALES and lead.owner_id == user.id:
        return
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")


def _filter_query_by_role(query, user: User):
    if user.role in (UserRole.ADMIN, UserRole.OWNER):
        return query
    if user.role == UserRole.SALES:
        return query.filter(Lead.owner_id == user.id)
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")


def _compute_price(abonement: Abonement) -> float:
    price = abonement.price or 0.0
    if abonement.discount_type is None:
        return price
    if hasattr(abonement, "discount_type"):
        dt = getattr(abonement.discount_type, "value", abonement.discount_type)
        if dt == "amount":
            price = max(price - (abonement.discount_value or 0.0), 0.0)
        elif dt == "percent":
            price = price * (1 - (abonement.discount_value or 0.0) / 100)
    return round(price, 2)


def _normalize_source_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    return " ".join(name.strip().split())


def _is_referral_source(name: Optional[str]) -> bool:
    return (name or "").strip().lower() == "рекомендация"


def _resolve_source(
    db: Session,
    source_id: Optional[int] = None,
    source_name: Optional[str] = None,
) -> Tuple[Optional[int], Optional[str]]:
    normalized = _normalize_source_name(source_name)
    if source_id:
        source_obj = db.query(LeadSource).filter(LeadSource.id == source_id).first()
        if not source_obj:
            raise HTTPException(status_code=404, detail="Lead source not found")
        return source_obj.id, source_obj.name
    if normalized:
        source_obj = db.query(LeadSource).filter(cast(LeadSource.name, Text).ilike(normalized)).first()
        if source_obj:
            return source_obj.id, source_obj.name
        return None, normalized
    return None, None


def _append_note_tag(note: Optional[str], tag: str) -> str:
    source = (note or "").strip()
    if tag in source:
        return source
    return f"{tag} {source}".strip()


def _get_default_open_status_option_id(db: Session) -> Optional[int]:
    default_open = (
        db.query(LeadTaskStatusOptionModel)
        .filter(
            LeadTaskStatusOptionModel.is_active.is_(True),
            LeadTaskStatusOptionModel.is_closed.is_(False),
        )
        .order_by(LeadTaskStatusOptionModel.id.asc())
        .first()
    )
    return default_open.id if default_open else None


def _create_auto_event_task(
    db: Session,
    *,
    lead: Lead,
    owner_id: int,
    note: str,
    due_at: datetime,
    preferred_template_keywords: List[str],
) -> LeadTask:
    template = None
    templates = (
        db.query(LeadTaskTemplate)
        .filter(LeadTaskTemplate.is_active.is_(True))
        .order_by(LeadTaskTemplate.id.asc())
        .all()
    )
    for tpl in templates:
        name = (tpl.name or "").lower()
        if any(keyword in name for keyword in preferred_template_keywords):
            template = tpl
            break
    task = LeadTask(
        lead_id=lead.id,
        owner_id=owner_id,
        template_id=template.id if template else None,
        status_option_id=_get_default_open_status_option_id(db),
        note=note,
        channel="call",
        due_at=due_at,
        status=LeadTaskStatus.OPEN,
    )
    db.add(task)
    return task


def _has_open_task_like(db: Session, lead_id: int, marker: str) -> bool:
    return (
        db.query(LeadTask)
        .filter(
            LeadTask.lead_id == lead_id,
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.note.isnot(None),
            cast(LeadTask.note, Text).ilike(f"%{marker}%"),
        )
        .first()
        is not None
    )


ALLOWED_PAUSE_REASONS = {"ждём ответ", "подумать", "нет времени"}


@router.get("/sales/dashboard", response_model=SalesDashboardResponse)
async def get_sales_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    now = datetime.utcnow()
    start_month = datetime(now.year, now.month, 1)
    end_month = datetime(now.year + (1 if now.month == 12 else 0), 1 if now.month == 12 else now.month + 1, 1)
    start_today = datetime(now.year, now.month, now.day)
    end_today = start_today + timedelta(days=1)
    next_two_hours = now + timedelta(hours=2)
    next_day = now + timedelta(hours=24)
    week_ago = now - timedelta(days=7)

    leads_q = _filter_query_by_role(db.query(Lead), current_user)
    tasks_q = db.query(LeadTask).join(Lead, Lead.id == LeadTask.lead_id)
    regs_q = db.query(EventRegistration).join(Lead, Lead.id == EventRegistration.lead_id)
    if current_user.role == UserRole.SALES:
        tasks_q = tasks_q.filter(Lead.owner_id == current_user.id)
        regs_q = regs_q.filter(Lead.owner_id == current_user.id)

    active_lead_statuses = [LeadStatus.NEW, LeadStatus.CONTACTED, LeadStatus.DEMO, LeadStatus.INVOICE_SENT]
    connected_statuses = [LeadStatus.CONTACTED, LeadStatus.DEMO, LeadStatus.INVOICE_SENT, LeadStatus.WON]

    kpi_new_leads = leads_q.filter(Lead.created_at >= start_today, Lead.created_at < end_today).count()
    leads_week_q = leads_q.filter(Lead.created_at >= week_ago)
    leads_week_total = leads_week_q.count()
    leads_week_connected = leads_week_q.filter(Lead.status.in_(connected_statuses)).count()
    kpi_dozvon_percent = round((leads_week_connected / leads_week_total) * 100, 1) if leads_week_total else 0.0
    kpi_info_sent = leads_q.filter(Lead.status.in_([LeadStatus.INVOICE_SENT, LeadStatus.WON])).count()

    kpi_need_push_urgent = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at >= now,
        Lead.next_contact_at < next_two_hours,
    ).count()
    kpi_need_push_today = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at >= start_today,
        Lead.next_contact_at < end_today,
    ).count()
    kpi_need_push_overdue = leads_q.filter(
        Lead.status.in_(active_lead_statuses),
        Lead.next_contact_at.isnot(None),
        Lead.next_contact_at < now,
    ).count()

    kpi_registered_event = regs_q.filter(EventRegistration.status == EventRegistrationStatus.REGISTERED).count()
    # Until dedicated attendance statuses are introduced, derive rough metrics from note tags.
    kpi_came_count = regs_q.filter(cast(EventRegistration.note, Text).ilike("%[came]%")).count()
    kpi_no_show_count = regs_q.filter(
        or_(
            cast(EventRegistration.note, Text).ilike("%[no-show]%"),
            cast(EventRegistration.note, Text).ilike("%no-show%"),
        )
    ).count()

    overdue_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at.isnot(None),
            LeadTask.due_at < now,
        )
        .order_by(LeadTask.due_at.asc())
        .limit(30)
        .all()
    )
    call_today_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at.isnot(None),
            LeadTask.due_at >= start_today,
            LeadTask.due_at < end_today,
        )
        .order_by(LeadTask.due_at.asc())
        .limit(30)
        .all()
    )
    messenger_items = (
        tasks_q.filter(
            LeadTask.status == LeadTaskStatus.OPEN,
            or_(
                cast(LeadTask.channel, Text).ilike("%messenger%"),
                cast(LeadTask.channel, Text).ilike("%telegram%"),
                cast(LeadTask.note, Text).ilike("%ответ%"),
            ),
        )
        .order_by(LeadTask.created_at.desc())
        .limit(30)
        .all()
    )
    confirm_items = (
        regs_q.join(Event, Event.id == EventRegistration.event_id)
        .filter(
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
            Event.starts_at >= now,
            Event.starts_at < next_day,
        )
        .order_by(Event.starts_at.asc())
        .limit(30)
        .all()
    )

    month_outreach_leads = (
        leads_q.filter(
            Lead.outreach_at.isnot(None),
            Lead.outreach_at >= start_month,
            Lead.outreach_at < end_month,
        )
        .all()
    )
    schools_seen = set()
    classes_seen = set()
    outreach_minutes_month = 0
    school_stats: Dict[str, Dict[str, object]] = {}
    for lead in month_outreach_leads:
        school = (lead.school_name or "").strip()
        class_name = (lead.school_class or "").strip()
        if school:
            schools_seen.add(school)
        if school and class_name:
            classes_seen.add(f"{school}::{class_name}")
        if lead.outreach_minutes and lead.outreach_minutes > 0:
            outreach_minutes_month += lead.outreach_minutes
        if not school:
            continue
        if school not in school_stats:
            school_stats[school] = {"leads": 0, "won": 0, "classes": set(), "minutes": 0}
        school_stats[school]["leads"] = int(school_stats[school]["leads"]) + 1
        if lead.status == LeadStatus.WON:
            school_stats[school]["won"] = int(school_stats[school]["won"]) + 1
        if class_name:
            school_stats[school]["classes"].add(class_name)
        if lead.outreach_minutes and lead.outreach_minutes > 0:
            school_stats[school]["minutes"] = int(school_stats[school]["minutes"]) + lead.outreach_minutes

    top_schools_conversion_month = sorted(
        [
            SalesSchoolConversionItem(
                school_name=school,
                leads_count=int(stats["leads"]),
                won_count=int(stats["won"]),
                conversion_percent=round((int(stats["won"]) / int(stats["leads"])) * 100) if int(stats["leads"]) else 0,
                classes_count=len(stats["classes"]),
                outreach_minutes_total=int(stats["minutes"]),
            )
            for school, stats in school_stats.items()
        ],
        key=lambda item: (item.conversion_percent, item.leads_count),
        reverse=True,
    )[:3]

    def map_task(item: LeadTask) -> SalesQueueTaskItem:
        lead = item.lead
        return SalesQueueTaskItem(
            task_id=item.id,
            lead_id=lead.id,
            lead_name=lead.contact_name,
            lead_phone=lead.phone,
            due_at=item.due_at,
            note=item.note,
        )

    def map_reg(item: EventRegistration) -> SalesQueueRegistrationItem:
        event = db.query(Event).filter(Event.id == item.event_id).first()
        lead = db.query(Lead).filter(Lead.id == item.lead_id).first()
        return SalesQueueRegistrationItem(
            registration_id=item.id,
            event_id=item.event_id,
            event_title=event.title if event else f"Event #{item.event_id}",
            lead_id=item.lead_id,
            lead_name=lead.contact_name if lead else f"Lead #{item.lead_id}",
            starts_at=event.starts_at if event else now,
            note=item.note,
        )

    return SalesDashboardResponse(
        kpi_new_leads=kpi_new_leads,
        kpi_dozvon_percent=kpi_dozvon_percent,
        kpi_info_sent=kpi_info_sent,
        kpi_need_push_urgent=kpi_need_push_urgent,
        kpi_need_push_today=kpi_need_push_today,
        kpi_need_push_overdue=kpi_need_push_overdue,
        kpi_registered_event=kpi_registered_event,
        kpi_came_count=kpi_came_count,
        kpi_no_show_count=kpi_no_show_count,
        overdue_followups=[map_task(i) for i in overdue_items],
        call_today=[map_task(i) for i in call_today_items],
        messenger_replies=[map_task(i) for i in messenger_items],
        confirm_participation=[map_reg(i) for i in confirm_items],
        outreach_schools_month=len(schools_seen),
        outreach_classes_month=len(classes_seen),
        outreach_minutes_month=outreach_minutes_month,
        top_schools_conversion_month=top_schools_conversion_month,
    )


@router.get("/follow-ups", response_model=List[FollowUpItemResponse])
async def list_follow_ups(
    period: str = Query(default="today"),
    source: Optional[str] = None,
    event_id: Optional[int] = None,
    reason: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    now = datetime.utcnow()
    start_today = datetime(now.year, now.month, now.day)
    end_today = start_today + timedelta(days=1)
    start_tomorrow = end_today
    end_tomorrow = start_tomorrow + timedelta(days=1)
    end_week = end_today + timedelta(days=7)

    query = db.query(LeadTask, Lead).join(Lead, Lead.id == LeadTask.lead_id)
    if current_user.role == UserRole.SALES:
        query = query.filter(Lead.owner_id == current_user.id)
    query = query.filter(LeadTask.status == LeadTaskStatus.OPEN)

    if period == "overdue":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at < now)
    elif period == "tomorrow":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_tomorrow, LeadTask.due_at < end_tomorrow)
    elif period == "week":
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_today, LeadTask.due_at < end_week)
    else:
        query = query.filter(LeadTask.due_at.isnot(None), LeadTask.due_at >= start_today, LeadTask.due_at < end_today)

    if source:
        query = query.filter(Lead.source.isnot(None), cast(Lead.source, Text).ilike(f"%{source.strip()}%"))
    if reason:
        query = query.filter(LeadTask.note.isnot(None), cast(LeadTask.note, Text).ilike(f"%{reason.strip()}%"))
    if event_id:
        query = query.join(EventRegistration, EventRegistration.lead_id == Lead.id).filter(
            EventRegistration.event_id == event_id,
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
        )

    rows = query.order_by(LeadTask.due_at.asc()).limit(200).all()
    return [
        FollowUpItemResponse(
            task_id=task.id,
            lead_id=lead.id,
            lead_name=lead.contact_name,
            lead_phone=lead.phone,
            lead_source=lead.source,
            due_at=task.due_at,
            status=task.status,
            channel=task.channel,
            note=task.note,
        )
        for task, lead in rows
    ]


@router.get("/leads/push-stats", response_model=List[LeadPushStatsResponse])
async def get_leads_push_stats(
    lead_ids: List[int] = Query(default=[]),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    lead_query = _filter_query_by_role(db.query(Lead), current_user)
    if lead_ids:
        lead_query = lead_query.filter(Lead.id.in_(lead_ids))
    leads = lead_query.all()
    if not leads:
        return []

    allowed_ids = [lead.id for lead in leads]
    tasks = (
        db.query(LeadTask)
        .options(joinedload(LeadTask.template))
        .filter(LeadTask.lead_id.in_(allowed_ids))
        .all()
    )

    by_lead: Dict[int, Dict[str, int]] = {lead_id: {"total": 0, "done": 0} for lead_id in allowed_ids}
    for task in tasks:
        template_name = (task.template.name if task.template else "").lower()
        note = (task.note or "").lower()
        is_push = "дожим" in template_name or "дожим" in note or "push" in note
        if not is_push:
            continue
        by_lead[task.lead_id]["total"] += 1
        if task.status == LeadTaskStatus.DONE:
            by_lead[task.lead_id]["done"] += 1

    response: List[LeadPushStatsResponse] = []
    for lead_id in allowed_ids:
        total = by_lead[lead_id]["total"]
        done = by_lead[lead_id]["done"]
        pct = round((done / total) * 100) if total else 0
        response.append(
            LeadPushStatsResponse(
                lead_id=lead_id,
                total_steps=total,
                done_steps=done,
                progress_percent=pct,
            )
        )
    return response


@router.get("/lead-sources", response_model=List[LeadSourceResponse])
async def list_lead_sources(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    query = db.query(LeadSource).order_by(LeadSource.name.asc())
    if active_only:
        query = query.filter(LeadSource.is_active.is_(True))
    return query.all()


@router.post("/lead-sources", response_model=LeadSourceResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_source(
    payload: LeadSourceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Source name is required")
    exists = db.query(LeadSource).filter(cast(LeadSource.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Source already exists")
    source = LeadSource(name=name, is_active=True)
    db.add(source)
    db.commit()
    db.refresh(source)
    log_action(db, current_user.id, "create", "lead_source", source.id, {"name": name})
    return source


@router.put("/lead-sources/{source_id}", response_model=LeadSourceResponse)
async def update_lead_source(
    source_id: int,
    payload: LeadSourceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    source = db.query(LeadSource).filter(LeadSource.id == source_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Source name is required")
        source.name = name
    if "is_active" in data:
        source.is_active = data["is_active"]
    db.commit()
    db.refresh(source)
    log_action(db, current_user.id, "update", "lead_source", source.id, data)
    return source


@router.get("/lead-task-templates", response_model=List[LeadTaskTemplateResponse])
async def list_lead_task_templates(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    query = db.query(LeadTaskTemplate).order_by(LeadTaskTemplate.name.asc())
    if active_only:
        query = query.filter(LeadTaskTemplate.is_active.is_(True))
    return query.all()


@router.post("/lead-task-templates", response_model=LeadTaskTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task_template(
    payload: LeadTaskTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Task template name is required")
    exists = db.query(LeadTaskTemplate).filter(cast(LeadTaskTemplate.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Task template already exists")
    item = LeadTaskTemplate(name=name, is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_task_template", item.id, {"name": name})
    return item


@router.put("/lead-task-templates/{template_id}", response_model=LeadTaskTemplateResponse)
async def update_lead_task_template(
    template_id: int,
    payload: LeadTaskTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    item = db.query(LeadTaskTemplate).filter(LeadTaskTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Task template not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Task template name is required")
        item.name = name
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_task_template", item.id, data)
    return item


@router.get("/lead-task-statuses", response_model=List[LeadTaskStatusOptionResponse])
async def list_lead_task_statuses(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    query = db.query(LeadTaskStatusOptionModel).order_by(LeadTaskStatusOptionModel.id.asc())
    if active_only:
        query = query.filter(LeadTaskStatusOptionModel.is_active.is_(True))
    return query.all()


@router.post("/lead-task-statuses", response_model=LeadTaskStatusOptionResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task_status(
    payload: LeadTaskStatusOptionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Task status name is required")
    exists = db.query(LeadTaskStatusOptionModel).filter(cast(LeadTaskStatusOptionModel.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Task status already exists")
    item = LeadTaskStatusOptionModel(name=name, is_active=True, is_closed=payload.is_closed)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_task_status", item.id, {"name": name, "is_closed": payload.is_closed})
    return item


@router.put("/lead-task-statuses/{status_id}", response_model=LeadTaskStatusOptionResponse)
async def update_lead_task_status(
    status_id: int,
    payload: LeadTaskStatusOptionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    item = db.query(LeadTaskStatusOptionModel).filter(LeadTaskStatusOptionModel.id == status_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Task status not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Task status name is required")
        item.name = name
    if "is_closed" in data:
        item.is_closed = data["is_closed"]
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_task_status", item.id, data)
    return item


@router.get("/lead-info-templates", response_model=List[LeadInfoTemplateResponse])
async def list_lead_info_templates(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    query = db.query(LeadInfoTemplate).order_by(LeadInfoTemplate.name.asc())
    if active_only:
        query = query.filter(LeadInfoTemplate.is_active.is_(True))
    return query.all()


@router.post("/lead-info-templates", response_model=LeadInfoTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_info_template(
    payload: LeadInfoTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    name = _normalize_source_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="Template name is required")
    if not (payload.body or "").strip():
        raise HTTPException(status_code=400, detail="Template body is required")
    exists = db.query(LeadInfoTemplate).filter(cast(LeadInfoTemplate.name, Text).ilike(name)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Template already exists")
    item = LeadInfoTemplate(name=name, body=payload.body.strip(), is_active=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "create", "lead_info_template", item.id, {"name": name})
    return item


@router.put("/lead-info-templates/{template_id}", response_model=LeadInfoTemplateResponse)
async def update_lead_info_template(
    template_id: int,
    payload: LeadInfoTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner"])),
):
    item = db.query(LeadInfoTemplate).filter(LeadInfoTemplate.id == template_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Template not found")
    data = payload.dict(exclude_unset=True)
    if "name" in data:
        name = _normalize_source_name(data["name"])
        if not name:
            raise HTTPException(status_code=400, detail="Template name is required")
        item.name = name
    if "body" in data:
        body = (data["body"] or "").strip()
        if not body:
            raise HTTPException(status_code=400, detail="Template body is required")
        item.body = body
    if "is_active" in data:
        item.is_active = data["is_active"]
    db.commit()
    db.refresh(item)
    log_action(db, current_user.id, "update", "lead_info_template", item.id, data)
    return item


@router.get("/leads/{lead_id}/communications", response_model=List[LeadCommunicationResponse])
async def list_lead_communications(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    return (
        db.query(LeadCommunication)
        .filter(LeadCommunication.lead_id == lead_id)
        .order_by(LeadCommunication.created_at.desc())
        .all()
    )


@router.post("/leads/{lead_id}/communications", response_model=LeadCommunicationResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_communication(
    lead_id: int,
    payload: LeadQuickCommunicationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    channel = (payload.channel or "messenger").strip().lower()
    if channel not in {"messenger", "call", "email"}:
        raise HTTPException(status_code=400, detail="Unsupported channel")
    message = (payload.message or "").strip() or f"[quick-{channel}]"
    follow_up_at = _to_utc(payload.follow_up_at) or _utcnow()

    comm = LeadCommunication(
        lead_id=lead.id,
        sent_by=current_user.id,
        template_id=None,
        channel=channel,
        message=message,
        pause_reason=None,
        follow_up_at=follow_up_at,
    )
    db.add(comm)
    if follow_up_at:
        lead.next_contact_at = follow_up_at
    if lead.status == LeadStatus.NEW:
        lead.status = LeadStatus.CONTACTED
    db.commit()
    db.refresh(comm)
    log_action(
        db,
        current_user.id,
        "log_communication",
        "lead",
        lead.id,
        {
            "channel": channel,
            "follow_up_at": follow_up_at.isoformat(),
        },
    )
    return comm


@router.post("/leads/{lead_id}/send-info", response_model=LeadCommunicationResponse, status_code=status.HTTP_201_CREATED)
async def send_info_for_lead(
    lead_id: int,
    payload: LeadSendInfoRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    follow_up_at = _to_utc(payload.follow_up_at)
    if follow_up_at is None:
        raise HTTPException(status_code=400, detail="follow_up_at is required")
    if follow_up_at <= _utcnow():
        raise HTTPException(status_code=400, detail="follow_up_at must be in the future")
    message = (payload.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="Message is required")
    if payload.pause_reason and payload.pause_reason not in ALLOWED_PAUSE_REASONS:
        raise HTTPException(status_code=400, detail="Unsupported pause reason")

    template_id = payload.template_id
    if template_id:
        template = db.query(LeadInfoTemplate).filter(LeadInfoTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")

    comm = LeadCommunication(
        lead_id=lead.id,
        sent_by=current_user.id,
        template_id=template_id,
        channel=(payload.channel or "messenger").strip(),
        message=message,
        pause_reason=payload.pause_reason,
        follow_up_at=follow_up_at,
    )
    db.add(comm)

    auto_task = LeadTask(
        lead_id=lead.id,
        owner_id=current_user.id,
        note=f"[auto-follow-up] {payload.pause_reason or 'без причины'}",
        channel=(payload.channel or "messenger").strip(),
        due_at=follow_up_at,
        status=LeadTaskStatus.OPEN,
    )
    db.add(auto_task)
    lead.next_contact_at = follow_up_at
    lead.pause_reason = payload.pause_reason
    if lead.status == LeadStatus.NEW:
        lead.status = LeadStatus.CONTACTED
    db.commit()
    db.refresh(comm)
    log_action(
        db,
        current_user.id,
        "send_info",
        "lead",
        lead.id,
        {
            "template_id": template_id,
            "channel": payload.channel,
            "follow_up_at": follow_up_at.isoformat(),
            "pause_reason": payload.pause_reason,
        },
    )
    return comm


@router.post("/leads/{lead_id}/contact-result", response_model=LeadCommunicationResponse, status_code=status.HTTP_201_CREATED)
async def save_lead_contact_result(
    lead_id: int,
    payload: LeadContactResultRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    outcome = (payload.outcome or "").strip().lower()
    if outcome not in {"connected", "no_answer", "callback"}:
        raise HTTPException(status_code=400, detail="Unsupported contact outcome")

    follow_up_at = _to_utc(payload.follow_up_at)
    if outcome in {"no_answer", "callback"} and follow_up_at is None:
        raise HTTPException(status_code=400, detail="follow_up_at is required for this outcome")
    if follow_up_at and follow_up_at <= _utcnow():
        raise HTTPException(status_code=400, detail="follow_up_at must be in the future")

    label_map = {
        "connected": "дозвон",
        "no_answer": "не дозвонились",
        "callback": "перезвонить",
    }
    message = f"[contact-result] {label_map[outcome]}"
    if payload.note and payload.note.strip():
        message = f"{message}: {payload.note.strip()}"

    comm = LeadCommunication(
        lead_id=lead.id,
        sent_by=current_user.id,
        template_id=None,
        channel="call",
        message=message,
        pause_reason=None,
        follow_up_at=follow_up_at or _utcnow(),
    )
    db.add(comm)

    if outcome in {"no_answer", "callback"} and follow_up_at:
        auto_task = LeadTask(
            lead_id=lead.id,
            owner_id=current_user.id,
            note=f"[auto-follow-up] {label_map[outcome]}",
            channel="call",
            due_at=follow_up_at,
            status=LeadTaskStatus.OPEN,
        )
        db.add(auto_task)
        lead.next_contact_at = follow_up_at

    if outcome in {"connected", "callback"} and lead.status == LeadStatus.NEW:
        lead.status = LeadStatus.CONTACTED

    db.commit()
    db.refresh(comm)
    log_action(
        db,
        current_user.id,
        "contact_result",
        "lead",
        lead.id,
        {
            "outcome": outcome,
            "follow_up_at": follow_up_at.isoformat() if follow_up_at else None,
        },
    )
    return comm


@router.post("/leads/import-xlsx", response_model=LeadImportResponse)
async def import_leads_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Поддерживается только формат .xlsx")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")

    wb = load_workbook(filename=BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return LeadImportResponse(created=0, skipped=0, errors=["Пустой лист"])

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    def val(row, key_variants: List[str]) -> Optional[str]:
        for key in key_variants:
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                continue
            raw = row[idx]
            if raw is None:
                continue
            txt = str(raw).strip()
            if txt:
                return txt
        return None

    def parse_row_datetime(raw_value) -> Optional[datetime]:
        if raw_value is None:
            return None
        if isinstance(raw_value, datetime):
            return raw_value
        text = str(raw_value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None

    def parse_row_minutes(raw_value) -> Optional[int]:
        if raw_value is None:
            return None
        text = str(raw_value).strip()
        if not text:
            return None
        try:
            minutes = int(float(text.replace(",", ".")))
            return minutes if minutes >= 0 else None
        except ValueError:
            return None

    created = 0
    skipped = 0
    errors: List[str] = []
    for i, row in enumerate(rows[1:], start=2):
        parent_name = val(row, ["фио родителя", "родитель", "parent_full_name"])
        child_name = val(row, ["фио ребенка", "фио ребёнка", "ребенок", "ребёнок", "child_full_name"])
        parent_phone = val(row, ["телефон родителя", "parent_phone"])
        child_phone = val(row, ["телефон школьника", "телефон ребенка", "телефон ребёнка", "child_phone"])
        email = val(row, ["email", "e-mail", "почта"])
        city = val(row, ["город", "city"])
        source_name_raw = val(row, ["источник", "source"])
        referral_name = val(row, ["кто пригласил", "рекомендовал", "referral_name"])
        comment = val(row, ["комментарий", "comment"])
        school_name = val(row, ["школа", "school", "school_name"])
        school_class = val(row, ["класс", "class", "school_class"])
        desired_slot = val(row, ["желаемое время занятий", "желаемое время", "время занятий", "desired_slot"])
        tags_raw = val(row, ["теги", "tags", "теги (через запятую)"])
        outreach_date_raw = val(row, ["дата обхода", "outreach_date", "outreach_at"])
        outreach_minutes_raw = val(row, ["время обхода (мин)", "время обхода", "outreach_minutes"])
        outreach_at = parse_row_datetime(outreach_date_raw)
        outreach_minutes = parse_row_minutes(outreach_minutes_raw)
        
        # Парсим теги из строки
        tags_list = None
        if tags_raw:
            tags_list = [t.strip() for t in tags_raw.split(",") if t.strip()]

        if not any([parent_name, child_name, parent_phone, child_phone, source_name_raw, comment]):
            skipped += 1
            continue

        source_id, source_name = _resolve_source(db, None, source_name_raw)
        if _is_referral_source(source_name) and not (referral_name or "").strip():
            errors.append("Строка {0}: для источника 'рекомендация' не указан пригласивший".format(i))
            skipped += 1
            continue

        contact_name = parent_name or child_name or "Без имени"
        phone = parent_phone or child_phone or "не указан"
        lead = Lead(
            owner_id=current_user.id,
            contact_name=contact_name,
            phone=phone,
            parent_full_name=parent_name,
            child_full_name=child_name,
            parent_phone=parent_phone,
            child_phone=child_phone,
            email=email,
            city=city,
            source=source_name,
            source_id=source_id,
            referral_name=referral_name,
            comment=comment,
            school_name=school_name,
            school_class=school_class,
            desired_slot=desired_slot,
            tags=tags_list,
            outreach_at=outreach_at,
            outreach_minutes=outreach_minutes,
            status=LeadStatus.NEW,
        )
        db.add(lead)
        created += 1

    db.commit()
    log_action(db, current_user.id, "import", "lead", None, {"created": created, "skipped": skipped})
    return LeadImportResponse(created=created, skipped=skipped, errors=errors)


@router.get("/leads/import-template")
async def download_leads_import_template(
    current_user: User = Depends(auth.require_role(["admin", "owner", "sales"])),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "LeadsImport"
    headers = [
        "ФИО родителя",
        "ФИО ребенка",
        "Телефон родителя",
        "Телефон школьника",
        "Email",
        "Город",
        "Школа",
        "Класс",
        "Дата обхода",
        "Время обхода (мин)",
        "Желаемое время занятий",
        "Источник",
        "Кто пригласил",
        "Теги (через запятую)",
        "Комментарий",
    ]
    ws.append(headers)
    ws.append(
        [
            "Иванова Анна Петровна",
            "Иванов Петр",
            "+7 999 111-22-33",
            "+7 900 111-22-44",
            "ivanova@example.com",
            "Москва",
            "Школа №12",
            "7А",
            "2026-02-01",
            "35",
            "Будни 17:00-19:00",
            "рекомендация",
            "Мария Сидорова",
            "горячий, интересуется",
            "Интерес к занятиям после пробного урока",
        ]
    )
    ws.freeze_panes = "A2"
    
    # Добавим ширину колонок для удобства
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['O'].width = 40

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "leads_import_template.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/leads", response_model=List[LeadResponse])
async def list_leads(
    status_filter: Optional[LeadStatus] = None,
    q: Optional[str] = None,
    source: Optional[str] = None,
    tag: Optional[str] = None,
    overdue_only: bool = False,
    created_from: Optional[datetime] = Query(default=None),
    created_to: Optional[datetime] = Query(default=None),
    next_contact_from: Optional[datetime] = Query(default=None),
    next_contact_to: Optional[datetime] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    if current_user.role not in (UserRole.ADMIN, UserRole.OWNER, UserRole.SALES):
        raise HTTPException(status_code=403, detail="Not enough permissions")

    query = _filter_query_by_role(db.query(Lead).order_by(Lead.created_at.desc()), current_user)
    if status_filter:
        query = query.filter(Lead.status == status_filter)
    if q:
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                cast(Lead.contact_name, Text).ilike(search),
                cast(Lead.phone, Text).ilike(search),
                cast(Lead.parent_full_name, Text).ilike(search),
                cast(Lead.child_full_name, Text).ilike(search),
                cast(Lead.parent_phone, Text).ilike(search),
                cast(Lead.child_phone, Text).ilike(search),
                cast(Lead.school_name, Text).ilike(search),
                cast(Lead.school_class, Text).ilike(search),
            )
        )
    if source:
        query = query.filter(Lead.source.ilike(f"%{source.strip()}%"))
    if tag:
        # tags are stored as JSON array; text-search keeps compatibility across DB backends.
        query = query.filter(Lead.tags.isnot(None), cast(Lead.tags, Text).ilike(f"%{tag.strip()}%"))
    if overdue_only:
        now = datetime.utcnow()
        query = query.filter(Lead.next_contact_at.isnot(None), Lead.next_contact_at < now)
    if created_from:
        query = query.filter(Lead.created_at >= created_from)
    if created_to:
        query = query.filter(Lead.created_at <= created_to)
    if next_contact_from:
        query = query.filter(Lead.next_contact_at >= next_contact_from)
    if next_contact_to:
        query = query.filter(Lead.next_contact_at <= next_contact_to)
    return query.all()


@router.post("/leads", response_model=LeadResponse, status_code=status.HTTP_201_CREATED)
async def create_lead(
    payload: LeadCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    owner_id = payload.owner_id if (current_user.role in (UserRole.ADMIN, UserRole.OWNER) and payload.owner_id) else current_user.id
    source_id, source_name = _resolve_source(db, payload.source_id, payload.source)
    if _is_referral_source(source_name) and not (payload.referral_name or "").strip():
        raise HTTPException(status_code=400, detail="Для источника 'рекомендация' укажите, кто пригласил")

    abonement = None
    if payload.abonement_id:
        abonement = db.query(Abonement).filter(Abonement.id == payload.abonement_id).first()
        if not abonement:
            raise HTTPException(status_code=404, detail="Abonement not found")

    lead = Lead(
        owner_id=owner_id,
        contact_name=payload.contact_name,
        phone=payload.phone,
        parent_full_name=payload.parent_full_name,
        child_full_name=payload.child_full_name,
        parent_phone=payload.parent_phone,
        child_phone=payload.child_phone,
        email=payload.email,
        city=payload.city,
        school_name=payload.school_name,
        school_class=payload.school_class,
        outreach_at=payload.outreach_at,
        outreach_minutes=payload.outreach_minutes,
        source=source_name,
        source_id=source_id,
        referral_name=payload.referral_name,
        tags=payload.tags,
        abonement_id=payload.abonement_id,
        desired_slot=payload.desired_slot,
        comment=payload.comment,
        next_contact_at=payload.next_contact_at,
        status=LeadStatus.NEW,
    )
    db.add(lead)
    db.commit()
    db.refresh(lead)

    log_action(db, current_user.id, "create", "lead", lead.id, {"owner_id": owner_id})
    return lead


@router.get("/leads/{lead_id}", response_model=LeadResponse)
async def get_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    return lead


@router.put("/leads/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: int,
    payload: LeadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    update_data = payload.dict(exclude_unset=True)
    # Prevent sales from changing owner/status to restricted values
    if "status" in update_data and update_data["status"] is not None:
        lead.status = update_data["status"]
    if "lost_reason" in update_data:
        lead.lost_reason = update_data["lost_reason"]

    if "source_id" in update_data or "source" in update_data:
        source_id, source_name = _resolve_source(db, update_data.get("source_id"), update_data.get("source"))
        if _is_referral_source(source_name) and not (update_data.get("referral_name") or lead.referral_name or "").strip():
            raise HTTPException(status_code=400, detail="Для источника 'рекомендация' укажите, кто пригласил")
        lead.source_id = source_id
        lead.source = source_name

    for field in [
        "contact_name",
        "phone",
        "parent_full_name",
        "child_full_name",
        "parent_phone",
        "child_phone",
        "email",
        "city",
        "school_name",
        "school_class",
        "outreach_at",
        "outreach_minutes",
        "referral_name",
        "tags",
        "abonement_id",
        "desired_slot",
        "comment",
        "next_contact_at",
        "pause_reason",
    ]:
        if field in update_data:
            setattr(lead, field, update_data[field])

    db.commit()
    db.refresh(lead)
    log_action(db, current_user.id, "update", "lead", lead.id, update_data)
    return lead


@router.post("/leads/{lead_id}/tasks", response_model=LeadTaskResponse, status_code=status.HTTP_201_CREATED)
async def create_lead_task(
    lead_id: int,
    payload: LeadTaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    template = None
    if payload.template_id:
        template = (
            db.query(LeadTaskTemplate)
            .filter(LeadTaskTemplate.id == payload.template_id, LeadTaskTemplate.is_active.is_(True))
            .first()
        )
        if not template:
            raise HTTPException(status_code=404, detail="Task template not found")

    status_option_id = payload.status_option_id
    if status_option_id:
        status_option = (
            db.query(LeadTaskStatusOptionModel)
            .filter(LeadTaskStatusOptionModel.id == status_option_id, LeadTaskStatusOptionModel.is_active.is_(True))
            .first()
        )
        if not status_option:
            raise HTTPException(status_code=404, detail="Task status not found")
    else:
        default_open = (
            db.query(LeadTaskStatusOptionModel)
            .filter(
                LeadTaskStatusOptionModel.is_active.is_(True),
                LeadTaskStatusOptionModel.is_closed.is_(False),
            )
            .order_by(LeadTaskStatusOptionModel.id.asc())
            .first()
        )
        status_option_id = default_open.id if default_open else None

    task = LeadTask(
        lead_id=lead_id,
        owner_id=current_user.id,
        template_id=payload.template_id,
        status_option_id=status_option_id,
        note=payload.note,
        channel=payload.channel,
        due_at=payload.due_at,
        status=LeadTaskStatus.OPEN,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    log_action(db, current_user.id, "create", "lead_task", task.id, {"lead_id": lead_id})
    return task


@router.get("/leads/{lead_id}/tasks", response_model=List[LeadTaskResponse])
async def list_lead_tasks(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    tasks = (
        db.query(LeadTask)
        .filter(LeadTask.lead_id == lead_id)
        .order_by(LeadTask.created_at.desc())
        .all()
    )
    return tasks


@router.post("/leads/{lead_id}/tasks/{task_id}/close", response_model=LeadTaskResponse)
async def close_lead_task(
    lead_id: int,
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    task = (
        db.query(LeadTask)
        .filter(LeadTask.id == task_id, LeadTask.lead_id == lead_id)
        .first()
    )
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    task.status = LeadTaskStatus.DONE
    closed_status = (
        db.query(LeadTaskStatusOptionModel)
        .filter(
            LeadTaskStatusOptionModel.is_active.is_(True),
            LeadTaskStatusOptionModel.is_closed.is_(True),
        )
        .order_by(LeadTaskStatusOptionModel.id.asc())
        .first()
    )
    if closed_status:
        task.status_option_id = closed_status.id
    task.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(task)
    log_action(db, current_user.id, "close", "lead_task", task.id, {"lead_id": lead_id})
    return task


@router.put("/leads/{lead_id}/tasks/{task_id}", response_model=LeadTaskResponse)
async def update_lead_task(
    lead_id: int,
    task_id: int,
    payload: LeadTaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    task = db.query(LeadTask).filter(LeadTask.id == task_id, LeadTask.lead_id == lead_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    update_data = payload.dict(exclude_unset=True)
    if "status" in update_data and update_data["status"] is not None:
        task.status = update_data["status"]
    if "note" in update_data:
        task.note = update_data["note"]
    if "channel" in update_data:
        task.channel = update_data["channel"]
    if "due_at" in update_data:
        task.due_at = update_data["due_at"]
    task.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(task)
    log_action(db, current_user.id, "update", "lead_task", task.id, update_data)
    return task


@router.post("/leads/{lead_id}/invoices", response_model=InvoiceResponse, status_code=status.HTTP_201_CREATED)
async def create_invoice_for_lead(
    lead_id: int,
    payload: InvoiceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    abonement = db.query(Abonement).filter(Abonement.id == payload.abonement_id).first()
    if not abonement:
        raise HTTPException(status_code=404, detail="Abonement not found")

    amount = _compute_price(abonement)
    invoice = Invoice(
        lead_id=lead_id,
        abonement_id=abonement.id,
        amount=amount,
        currency=payload.currency or "RUB",
        status=InvoiceStatus.DRAFT,
        email_to=payload.email_to or lead.email,
        link=None,
    )
    db.add(invoice)
    if lead.status not in (LeadStatus.WON, LeadStatus.LOST):
        lead.status = LeadStatus.INVOICE_SENT
    db.commit()
    db.refresh(invoice)
    log_action(db, current_user.id, "create", "invoice", invoice.id, {"lead_id": lead_id, "amount": amount})
    return invoice


@router.get("/leads/{lead_id}/invoices", response_model=List[InvoiceResponse])
async def list_invoices_for_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    return (
        db.query(Invoice)
        .filter(Invoice.lead_id == lead_id)
        .order_by(Invoice.created_at.desc())
        .all()
    )


@router.post("/invoices/{invoice_id}/send-email", response_model=InvoiceResponse)
async def send_invoice_email(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    lead = db.query(Lead).filter(Lead.id == invoice.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    if not invoice.email_to and not lead.email:
        raise HTTPException(status_code=400, detail="No email to send invoice")

    # Stub: actual email sending integration should be implemented separately.
    invoice.status = InvoiceStatus.SENT
    invoice.sent_at = datetime.utcnow()
    if lead.status not in (LeadStatus.WON, LeadStatus.LOST):
        lead.status = LeadStatus.INVOICE_SENT
    db.commit()
    db.refresh(invoice)
    log_action(db, current_user.id, "send_email", "invoice", invoice.id, {"lead_id": lead.id})
    return invoice


@router.get("/invoices", response_model=List[InvoiceResponse])
async def list_invoices(
    status_filter: Optional[InvoiceStatus] = None,
    lead_id: Optional[int] = None,
    created_from: Optional[datetime] = Query(default=None),
    created_to: Optional[datetime] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    query = (
        db.query(Invoice)
        .join(Lead, Lead.id == Invoice.lead_id)
        .options(joinedload(Invoice.lead))
        .order_by(Invoice.created_at.desc())
    )

    if current_user.role == UserRole.SALES:
        query = query.filter(Lead.owner_id == current_user.id)

    if status_filter:
        query = query.filter(Invoice.status == status_filter)
    if lead_id:
        query = query.filter(Invoice.lead_id == lead_id)
    if created_from:
        query = query.filter(Invoice.created_at >= created_from)
    if created_to:
        query = query.filter(Invoice.created_at <= created_to)

    return query.all()


@router.get("/events", response_model=List[EventResponse])
async def list_events(
    status_filter: Optional[EventStatus] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    query = db.query(Event).order_by(Event.starts_at.asc())
    if status_filter:
        query = query.filter(Event.status == status_filter)
    return query.all()


@router.post("/events", response_model=EventResponse, status_code=status.HTTP_201_CREATED)
async def create_event(
    payload: EventCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    event = Event(
        title=payload.title,
        description=payload.description,
        starts_at=payload.starts_at,
        ends_at=payload.ends_at,
        location=payload.location,
        capacity=payload.capacity,
        status=EventStatus.ACTIVE,
        created_by=current_user.id,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    log_action(db, current_user.id, "create", "event", event.id)
    return event


@router.put("/events/{event_id}", response_model=EventResponse)
async def update_event(
    event_id: int,
    payload: EventUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    update_data = payload.dict(exclude_unset=True)
    for k, v in update_data.items():
        setattr(event, k, v)
    db.commit()
    db.refresh(event)
    log_action(db, current_user.id, "update", "event", event.id, update_data)
    return event


@router.get("/events/{event_id}/registrations", response_model=List[EventRegistrationResponse])
async def list_event_registrations(
    event_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    query = db.query(EventRegistration).filter(EventRegistration.event_id == event_id).order_by(EventRegistration.created_at.desc())
    if current_user.role == UserRole.SALES:
        query = query.filter(EventRegistration.owner_id == current_user.id)
    return query.all()


@router.post("/events/{event_id}/registrations", response_model=EventRegistrationResponse, status_code=status.HTTP_201_CREATED)
async def create_event_registration(
    event_id: int,
    payload: EventRegistrationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    event = db.query(Event).filter(Event.id == event_id, Event.status == EventStatus.ACTIVE).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found or inactive")
    if event.capacity is not None:
        current_registered = (
            db.query(EventRegistration)
            .filter(
                EventRegistration.event_id == event_id,
                EventRegistration.status == EventRegistrationStatus.REGISTERED,
            )
            .count()
        )
        if current_registered >= event.capacity:
            raise HTTPException(status_code=400, detail="Свободных слотов нет")

    lead = db.query(Lead).filter(Lead.id == payload.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    existing = db.query(EventRegistration).filter(
        EventRegistration.event_id == event_id,
        EventRegistration.lead_id == payload.lead_id
    ).first()
    if existing and existing.status == EventRegistrationStatus.REGISTERED:
        raise HTTPException(status_code=400, detail="Lead already registered")

    if existing:
        existing.status = EventRegistrationStatus.REGISTERED
        existing.note = payload.note
        existing.owner_id = current_user.id
        db.commit()
        db.refresh(existing)
        log_action(db, current_user.id, "restore", "event_registration", existing.id, {"event_id": event_id, "lead_id": lead.id})
        return existing

    reg = EventRegistration(
        event_id=event_id,
        lead_id=lead.id,
        owner_id=current_user.id,
        status=EventRegistrationStatus.REGISTERED,
        note=payload.note,
    )
    db.add(reg)
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "create", "event_registration", reg.id, {"event_id": event_id, "lead_id": lead.id})
    return reg


@router.post("/events/{event_id}/registrations/{registration_id}/cancel", response_model=EventRegistrationResponse)
async def cancel_event_registration(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")

    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    reg.status = EventRegistrationStatus.CANCELLED
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "cancel", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    return reg


@router.post("/events/{event_id}/registrations/{registration_id}/confirm", response_model=EventRegistrationResponse)
async def confirm_event_registration(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id,
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    reg.note = _append_note_tag(reg.note, "[confirmed]")
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "confirm", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    return reg


@router.post("/events/{event_id}/registrations/{registration_id}/mark-came", response_model=EventRegistrationResponse)
async def mark_event_registration_came(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id,
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    reg.note = _append_note_tag(reg.note, "[came]")
    # UX automation: after attendance create follow-up "offer course".
    if not _has_open_task_like(db, lead.id, "[auto_attended_offer]"):
        due_at = datetime.utcnow() + timedelta(hours=24)
        auto_task = _create_auto_event_task(
            db,
            lead=lead,
            owner_id=lead.owner_id,
            note="[auto_attended_offer] После мероприятия: предложить курс",
            due_at=due_at,
            preferred_template_keywords=["курс", "предлож", "дожим", "offer"],
        )
        db.flush()
        log_action(
            db,
            current_user.id,
            "create",
            "lead_task",
            auto_task.id,
            {"lead_id": lead.id, "type": "auto_attended_offer"},
        )
        if lead.next_contact_at is None or (auto_task.due_at and lead.next_contact_at > auto_task.due_at):
            lead.next_contact_at = auto_task.due_at
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "mark_came", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    return reg


@router.post("/events/{event_id}/registrations/{registration_id}/mark-no-show", response_model=EventRegistrationResponse)
async def mark_event_registration_no_show(
    event_id: int,
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    reg = db.query(EventRegistration).filter(
        EventRegistration.id == registration_id,
        EventRegistration.event_id == event_id,
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    lead = db.query(Lead).filter(Lead.id == reg.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    reg.note = _append_note_tag(reg.note, "[no-show]")
    # UX automation: after no-show create reactivation follow-up.
    if not _has_open_task_like(db, lead.id, "[auto_no_show_reactivate]"):
        due_at = datetime.utcnow() + timedelta(hours=24)
        auto_task = _create_auto_event_task(
            db,
            lead=lead,
            owner_id=lead.owner_id,
            note="[auto_no_show_reactivate] No-show: реактивация и перенос",
            due_at=due_at,
            preferred_template_keywords=["реактивац", "перезап", "no-show", "дожим"],
        )
        db.flush()
        log_action(
            db,
            current_user.id,
            "create",
            "lead_task",
            auto_task.id,
            {"lead_id": lead.id, "type": "auto_no_show_reactivate"},
        )
        if lead.next_contact_at is None or (auto_task.due_at and lead.next_contact_at > auto_task.due_at):
            lead.next_contact_at = auto_task.due_at
    db.commit()
    db.refresh(reg)
    log_action(db, current_user.id, "mark_no_show", "event_registration", reg.id, {"event_id": event_id, "lead_id": reg.lead_id})
    return reg


@router.get("/leads/{lead_id}/event-registrations", response_model=List[EventRegistrationResponse])
async def list_lead_event_registrations(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["admin", "sales"])),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    return (
        db.query(EventRegistration)
        .filter(EventRegistration.lead_id == lead_id)
        .order_by(EventRegistration.created_at.desc())
        .all()
    )
