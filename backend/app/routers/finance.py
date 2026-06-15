from typing import Dict, List, Optional

import csv
import hashlib
from datetime import date, datetime, timedelta, timezone
from io import StringIO, BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File, Form
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from app import auth
from app.database import db_transaction, get_db
from app.models import (
    User,
    BankTransaction,
    BankTransactionStatus,
    FinanceTransaction,
    FinanceAccount,
    FinanceAccountOwnerScope,
    FinanceTarget,
    FinanceArticle,
    FinanceArticleDirection,
    FinanceArticleCostKind,
    FinanceArticleScope,
    FinanceTransactionDirection,
    FinanceTransactionStatus,
    FinanceModel,
    BudgetEntry,
    MetricDefinition,
    DashboardWidget,
)
from app.schemas.finance import (
    BankTransactionApplyRequest,
    BankTransactionResponse,
    FinanceAccountBalance,
    FinanceAccountCreate,
    FinanceAccountResponse,
    FinanceTargetCreate,
    FinanceAnalyticsExpenseBreakdownRow,
    FinanceAnalyticsKpiBlock,
    FinanceAnalyticsSummaryResponse,
    FinanceAnalyticsTargetBreakdownRow,
    FinanceArticleCreate,
    FinanceArticleResponse,
    FinanceArticleTreeItem,
    FinanceArticleUpdate,
    FinanceModelCreate,
    FinanceModelResponse,
    FinanceModelUpdate,
    BudgetEntryResponse,
    BudgetSaveRequest,
    MetricComputeResponse,
    MetricDefinitionCreate,
    MetricDefinitionResponse,
    MetricDefinitionUpdate,
    DashboardWidgetCreate,
    DashboardWidgetComputedResponse,
    DashboardWidgetResponse,
    DashboardWidgetUpdate,
    FinanceLedgerBankRow,
    FinanceLedgerTransactionRow,
    FinanceManualTransactionCreate,
    FinanceModelTemplateCreate,
    FinanceModelTemplateResponse,
    FinanceModelTemplateUpdate,
    FinancePnlRow,
    FinanceStudentAccountCreate,
    PLReportRow,
    PLReportResponse,
    CommandCenterResponse,
    CommandCenterKpi,
    CommandCenterAccount,
    CommandCenterProjectColumn,
    CommandCenterCashFlowPoint,
    CommandCenterTransaction,
    CommandCenterAlert,
    FinanceTargetResponse,
    FinanceTransactionApplyStudentRequest,
    FinanceTransactionUpdate,
    StudentAccountResponse,
)
from app.services.finance_metric_engine import compute_metric_formula
from app.services.finance_templates import get_template
from app.services.finance_ledger import apply_recognition_rules
from app.services.student_account_finance import create_student_account as finance_create_student_account
from app.services.bank_operation import apply_bank_operation_to_student as bank_operation_apply
from app.services.payment_status import get_payment_status_summary
from app.dependencies import require_finance_access as dep_require_finance_access


router = APIRouter()


def _require_finance_access(user: User) -> None:
    """Права доступа к финансовому журналу: admin / owner / sales."""
    if not auth.has_permission(user, "finance.access"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Недостаточно прав для работы с финансовым журналом")


def _slug(value: str) -> str:
    cleaned = "".join(ch.lower() if ch.isalnum() else "_" for ch in str(value or "").strip())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_")[:96] or "item"


def _article_response(article: FinanceArticle) -> FinanceArticleResponse:
    return FinanceArticleResponse(
        id=article.id,
        code=article.code,
        target_id=article.target_id,
        parent_id=article.parent_id,
        name=article.name,
        direction=str(getattr(article.direction, "value", article.direction)),
        cost_kind=str(getattr(article.cost_kind, "value", article.cost_kind)),
        scope=str(getattr(article.scope, "value", article.scope)),
        color=article.color,
        sort_order=int(article.sort_order or 0),
        is_active=bool(article.is_active),
    )


def _model_response(model: FinanceModel) -> FinanceModelResponse:
    target = getattr(model, "target", None)
    return FinanceModelResponse(
        id=model.id,
        target_id=model.target_id,
        target_code=getattr(target, "code", None),
        target_name=getattr(target, "name", None),
        name=model.name,
        template_key=model.template_key,
        currency=model.currency,
        period_type=model.period_type,
        settings_json=model.settings_json,
        created_at=model.created_at,
        updated_at=model.updated_at,
    )


def _metric_response(metric: MetricDefinition) -> MetricDefinitionResponse:
    target = getattr(metric, "target", None)
    return MetricDefinitionResponse(
        id=metric.id,
        target_id=metric.target_id,
        target_name=getattr(target, "name", None),
        name=metric.name,
        formula=metric.formula,
        unit=metric.unit,
        goal_value=metric.goal_value,
        sort_order=int(metric.sort_order or 0),
        created_at=metric.created_at,
        updated_at=metric.updated_at,
    )


def _widget_response(widget: DashboardWidget) -> DashboardWidgetResponse:
    metric = getattr(widget, "metric", None)
    target = getattr(widget, "target", None)
    return DashboardWidgetResponse(
        id=widget.id,
        owner_id=widget.owner_id,
        metric_id=widget.metric_id,
        metric_name=getattr(metric, "name", None),
        target_id=widget.target_id,
        target_name=getattr(target, "name", None),
        widget_type=widget.widget_type,
        period_type=widget.period_type,
        position_x=int(widget.position_x or 0),
        position_y=int(widget.position_y or 0),
        width=int(widget.width or 1),
        title_override=widget.title_override,
    )


def _create_article_from_template(
    db: Session,
    target_id: int,
    item: Dict[str, object],
    parent_id: Optional[int] = None,
    sort_order: int = 0,
) -> FinanceArticle:
    code = _slug(str(item.get("code") or item.get("name") or "article"))
    existing = (
        db.query(FinanceArticle)
        .filter(FinanceArticle.target_id == target_id, FinanceArticle.code == code)
        .first()
    )
    if existing:
        article = existing
    else:
        direction = str(item.get("direction") or "expense")
        cost_kind = str(item.get("cost_kind") or "none")
        article = FinanceArticle(
            target_id=target_id,
            parent_id=parent_id,
            code=code,
            name=str(item.get("name") or code),
            direction=FinanceArticleDirection(direction),
            cost_kind=FinanceArticleCostKind(cost_kind),
            scope=FinanceArticleScope.ANY,
            sort_order=sort_order,
            is_active=True,
        )
        db.add(article)
        db.flush()
    children = item.get("children") or []
    if isinstance(children, list):
        for index, child in enumerate(children):
            if isinstance(child, dict):
                _create_article_from_template(db, target_id, child, article.id, index)
    return article


@router.get("/accounts", response_model=List[FinanceAccountResponse])
async def list_finance_accounts(
    only_active: bool = Query(True, description="Только активные счета"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceAccountResponse]:
    """Справочник счетов (банк/карта) для единого финансового журнала."""
    _require_finance_access(current_user)
    q = db.query(FinanceAccount)
    if only_active:
        q = q.filter(FinanceAccount.is_active.is_(True))
    accounts = q.order_by(FinanceAccount.code).all()
    return [
        FinanceAccountResponse(
            id=a.id,
            code=a.code,
            name=a.name,
            owner_scope=str(getattr(a.owner_scope, "value", a.owner_scope)),
            is_active=bool(a.is_active),
        )
        for a in accounts
    ]


@router.post("/accounts", response_model=FinanceAccountResponse, status_code=201)
async def create_finance_account(
    payload: FinanceAccountCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceAccountResponse:
    """Создать новый счёт для импорта выписок."""
    _require_finance_access(current_user)
    code = payload.code.strip().lower().replace(" ", "_")
    if db.query(FinanceAccount).filter(FinanceAccount.code == code).first():
        raise HTTPException(status_code=400, detail=f"Счёт с кодом '{code}' уже существует")
    try:
        scope = FinanceAccountOwnerScope(payload.owner_scope)
    except ValueError:
        scope = FinanceAccountOwnerScope.BUSINESS
    account = FinanceAccount(code=code, name=payload.name.strip(), owner_scope=scope, is_active=True)
    db.add(account)
    db.commit()
    db.refresh(account)
    return FinanceAccountResponse(
        id=account.id,
        code=account.code,
        name=account.name,
        owner_scope=str(getattr(account.owner_scope, "value", account.owner_scope)),
        is_active=bool(account.is_active),
    )


@router.delete("/accounts/{account_id}", status_code=204)
async def delete_finance_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> None:
    """Удалить (деактивировать) счёт для импорта."""
    _require_finance_access(current_user)
    account = db.query(FinanceAccount).filter(FinanceAccount.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Счёт не найден")
    account.is_active = False
    db.commit()


@router.get("/targets", response_model=List[FinanceTargetResponse])
async def list_finance_targets(
    only_active: bool = Query(True, description="Только активные цели/проекты"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceTargetResponse]:
    """Справочник targets (academy, personal, leninets, gogol_mogol и т.п.)."""
    _require_finance_access(current_user)
    q = db.query(FinanceTarget)
    if only_active:
        q = q.filter(FinanceTarget.is_active.is_(True))
    targets = q.order_by(FinanceTarget.code).all()
    return [
        FinanceTargetResponse(
            id=t.id,
            code=t.code,
            name=t.name,
            is_active=bool(t.is_active),
        )
        for t in targets
    ]


@router.post("/targets", response_model=FinanceTargetResponse, status_code=201)
async def create_finance_target(
    payload: FinanceTargetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceTargetResponse:
    """Создать новый проект/цель."""
    _require_finance_access(current_user)
    code = payload.code.strip().lower().replace(" ", "_")
    if db.query(FinanceTarget).filter(FinanceTarget.code == code).first():
        raise HTTPException(status_code=400, detail=f"Проект с кодом '{code}' уже существует")
    target = FinanceTarget(code=code, name=payload.name.strip(), is_active=True)
    db.add(target)
    db.commit()
    db.refresh(target)
    return FinanceTargetResponse(id=target.id, code=target.code, name=target.name, is_active=bool(target.is_active))


@router.delete("/targets/{target_id}", status_code=204)
async def delete_finance_target(
    target_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> None:
    """Деактивировать проект/цель."""
    _require_finance_access(current_user)
    target = db.query(FinanceTarget).filter(FinanceTarget.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Проект не найден")
    target.is_active = False
    db.commit()


@router.get("/model-templates", response_model=List[FinanceModelTemplateResponse])
async def list_finance_model_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceModelTemplateResponse]:
    _require_finance_access(current_user)
    from app.models import FinanceModelTemplate
    rows = db.query(FinanceModelTemplate).order_by(FinanceModelTemplate.sort_order, FinanceModelTemplate.id).all()
    return rows


@router.post("/model-templates", response_model=FinanceModelTemplateResponse)
async def create_finance_model_template(
    payload: FinanceModelTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceModelTemplateResponse:
    _require_finance_access(current_user)
    from app.models import FinanceModelTemplate
    key = payload.key.strip().lower().replace(" ", "_")
    if not key:
        raise HTTPException(status_code=400, detail="Ключ шаблона не может быть пустым")
    if db.query(FinanceModelTemplate).filter(FinanceModelTemplate.key == key).first():
        raise HTTPException(status_code=400, detail="Шаблон с таким ключом уже существует")
    row = FinanceModelTemplate(
        key=key,
        name=payload.name.strip(),
        articles_json=[a.model_dump(exclude_none=True) for a in payload.articles],
        metrics_json=[m.model_dump(exclude_none=True) for m in payload.metrics],
        is_system=False,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.patch("/model-templates/{template_id}", response_model=FinanceModelTemplateResponse)
async def update_finance_model_template(
    template_id: int,
    payload: FinanceModelTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceModelTemplateResponse:
    _require_finance_access(current_user)
    from app.models import FinanceModelTemplate
    row = db.query(FinanceModelTemplate).filter(FinanceModelTemplate.id == template_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Шаблон не найден")
    if payload.name is not None:
        row.name = payload.name.strip() or row.name
    if payload.articles is not None:
        row.articles_json = [a.model_dump(exclude_none=True) for a in payload.articles]
    if payload.metrics is not None:
        row.metrics_json = [m.model_dump(exclude_none=True) for m in payload.metrics]
    db.commit()
    db.refresh(row)
    return row


@router.delete("/model-templates/{template_id}", status_code=204)
async def delete_finance_model_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> None:
    _require_finance_access(current_user)
    from app.models import FinanceModelTemplate
    row = db.query(FinanceModelTemplate).filter(FinanceModelTemplate.id == template_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Шаблон не найден")
    if row.is_system:
        raise HTTPException(status_code=400, detail="Системный шаблон нельзя удалить")
    db.delete(row)
    db.commit()


@router.get("/models", response_model=List[FinanceModelResponse])
async def list_finance_models(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceModelResponse]:
    _require_finance_access(current_user)
    rows = (
        db.query(FinanceModel)
        .options(joinedload(FinanceModel.target))
        .order_by(FinanceModel.created_at.desc(), FinanceModel.id.desc())
        .all()
    )
    return [_model_response(row) for row in rows]


@router.post("/models", response_model=FinanceModelResponse, status_code=201)
async def create_finance_model(
    payload: FinanceModelCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceModelResponse:
    _require_finance_access(current_user)
    target: Optional[FinanceTarget] = None
    if payload.target_id is not None:
        target = db.query(FinanceTarget).filter(FinanceTarget.id == payload.target_id).first()
    elif payload.target_code:
        code = _slug(payload.target_code)
        target = db.query(FinanceTarget).filter(FinanceTarget.code == code).first()
        if target is None:
            target = FinanceTarget(code=code, name=(payload.target_name or payload.name).strip(), is_active=True)
            db.add(target)
            db.flush()
    model = FinanceModel(
        target_id=target.id if target else None,
        name=payload.name.strip(),
        template_key=payload.template_key or "blank",
        currency=(payload.currency or "RUB").strip().upper()[:8],
        period_type=payload.period_type or "month",
        settings_json=payload.settings_json,
    )
    template = get_template(model.template_key, db=db)
    try:
        with db_transaction(db):
            db.add(model)
            db.flush()
            if target is not None:
                for index, item in enumerate(template.get("articles") or []):
                    if isinstance(item, dict):
                        _create_article_from_template(db, target.id, item, None, index)
                for index, metric_data in enumerate(template.get("metrics") or []):
                    if isinstance(metric_data, dict):
                        metric = MetricDefinition(
                            target_id=target.id,
                            name=str(metric_data.get("name") or "Metric"),
                            formula=str(metric_data.get("formula") or "0"),
                            unit=metric_data.get("unit") if metric_data.get("unit") is not None else None,
                            sort_order=index,
                        )
                        db.add(metric)
                        db.flush()
                        db.add(
                            DashboardWidget(
                                owner_id=current_user.id,
                                metric_id=metric.id,
                                target_id=target.id,
                                widget_type="number",
                                period_type="current_month",
                                position_x=index % 4,
                                position_y=index // 4,
                                width=1,
                            )
                        )
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Finance model with this name already exists for target")
    db.refresh(model)
    db.refresh(model, ["target"])
    return _model_response(model)


@router.get("/models/{model_id}", response_model=FinanceModelResponse)
async def get_finance_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceModelResponse:
    _require_finance_access(current_user)
    row = db.query(FinanceModel).options(joinedload(FinanceModel.target)).filter(FinanceModel.id == model_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Finance model not found")
    return _model_response(row)


@router.patch("/models/{model_id}", response_model=FinanceModelResponse)
async def update_finance_model(
    model_id: int,
    payload: FinanceModelUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceModelResponse:
    _require_finance_access(current_user)
    row = db.query(FinanceModel).options(joinedload(FinanceModel.target)).filter(FinanceModel.id == model_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Finance model not found")
    if payload.name is not None:
        row.name = payload.name.strip() or row.name
    if payload.template_key is not None:
        row.template_key = payload.template_key or row.template_key
    if payload.currency is not None:
        row.currency = payload.currency.strip().upper()[:8] or row.currency
    if payload.period_type is not None:
        row.period_type = payload.period_type or row.period_type
    if payload.settings_json is not None:
        row.settings_json = payload.settings_json
    db.commit()
    db.refresh(row)
    db.refresh(row, ["target"])
    return _model_response(row)


@router.delete("/models/{model_id}", status_code=204)
async def delete_finance_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> None:
    _require_finance_access(current_user)
    row = db.query(FinanceModel).filter(FinanceModel.id == model_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Finance model not found")
    db.delete(row)
    db.commit()
    return None


@router.get("/articles/tree", response_model=List[FinanceArticleTreeItem])
async def get_finance_articles_tree(
    target_id: int = Query(...),
    only_active: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceArticleTreeItem]:
    _require_finance_access(current_user)
    q = db.query(FinanceArticle).filter(FinanceArticle.target_id == target_id)
    if only_active:
        q = q.filter(FinanceArticle.is_active.is_(True))
    articles = q.order_by(FinanceArticle.sort_order.asc(), FinanceArticle.name.asc()).all()
    by_parent: Dict[Optional[int], List[FinanceArticle]] = {}
    for article in articles:
        by_parent.setdefault(article.parent_id, []).append(article)

    def build(parent_id: Optional[int]) -> List[FinanceArticleTreeItem]:
        nodes: List[FinanceArticleTreeItem] = []
        for article in by_parent.get(parent_id, []):
            data = _article_response(article).model_dump()
            data["children"] = build(article.id)
            nodes.append(FinanceArticleTreeItem(**data))
        return nodes

    return build(None)


@router.get("/articles", response_model=List[FinanceArticleResponse])
async def list_finance_articles(
    only_active: bool = Query(True, description="Только активные статьи"),
    scope: Optional[str] = Query(None, description="Фильтр по scope: academy | personal | any"),
    direction: Optional[str] = Query(None, description="Фильтр по direction: income | expense"),
    target_id: Optional[int] = Query(None, description="Filter by finance_targets.id"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceArticleResponse]:
    """Единый справочник статей доходов/расходов."""
    _require_finance_access(current_user)
    q = db.query(FinanceArticle)
    if only_active:
        q = q.filter(FinanceArticle.is_active.is_(True))
    if scope:
        q = q.filter(FinanceArticle.scope == scope)
    if direction:
        q = q.filter(FinanceArticle.direction == direction)
    if target_id is not None:
        q = q.filter(FinanceArticle.target_id == target_id)
    arts = q.order_by(FinanceArticle.sort_order.asc(), FinanceArticle.name.asc()).all()
    return [_article_response(a) for a in arts]


@router.post("/articles", response_model=FinanceArticleResponse)
async def create_finance_article(
    payload: FinanceArticleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceArticleResponse:
    """Создание статьи дохода/расхода в едином справочнике."""
    _require_finance_access(current_user)

    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    direction_value = payload.direction
    if direction_value not in {d.value for d in FinanceArticleDirection}:  # type: ignore[attr-defined]
        raise HTTPException(status_code=400, detail="Некорректное значение direction")
    direction = FinanceArticleDirection(direction_value)  # type: ignore[call-arg]

    scope_value = payload.scope or "personal"
    if scope_value not in {s.value for s in FinanceArticleScope}:  # type: ignore[attr-defined]
        raise HTTPException(status_code=400, detail="Некорректное значение scope")
    scope = FinanceArticleScope(scope_value)  # type: ignore[call-arg]

    cost_kind_value = payload.cost_kind or "none"
    if cost_kind_value not in {c.value for c in FinanceArticleCostKind}:  # type: ignore[attr-defined]
        raise HTTPException(status_code=400, detail="Некорректное значение cost_kind")
    cost_kind = FinanceArticleCostKind(cost_kind_value)  # type: ignore[call-arg]

    art = FinanceArticle(
        code=_slug(payload.code or name),
        target_id=payload.target_id,
        parent_id=payload.parent_id,
        name=name,
        direction=direction,
        cost_kind=cost_kind,
        scope=scope,
        color=payload.color,
        sort_order=payload.sort_order or 0,
        is_active=True,
    )
    db.add(art)
    db.commit()
    db.refresh(art)

    return _article_response(art)


@router.patch("/articles/{article_id}", response_model=FinanceArticleResponse)
async def update_finance_article(
    article_id: int,
    payload: FinanceArticleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceArticleResponse:
    """Частичное обновление статьи (переименование, смена типа, деактивация)."""
    _require_finance_access(current_user)

    art = db.query(FinanceArticle).filter(FinanceArticle.id == article_id).first()
    if not art:
        raise HTTPException(status_code=404, detail="Статья не найдена")

    if payload.name is not None:
        art.name = payload.name.strip() or art.name
    if payload.code is not None:
        art.code = _slug(payload.code) if payload.code else None
    if payload.target_id is not None:
        if payload.target_id and not db.query(FinanceTarget.id).filter(FinanceTarget.id == payload.target_id).first():
            raise HTTPException(status_code=400, detail="target_id not found")
        art.target_id = payload.target_id
    if payload.parent_id is not None:
        if payload.parent_id == art.id:
            raise HTTPException(status_code=400, detail="Article cannot be its own parent")
        if payload.parent_id and not db.query(FinanceArticle.id).filter(FinanceArticle.id == payload.parent_id).first():
            raise HTTPException(status_code=400, detail="parent_id not found")
        art.parent_id = payload.parent_id

    if payload.direction is not None:
        if payload.direction not in {d.value for d in FinanceArticleDirection}:  # type: ignore[attr-defined]
            raise HTTPException(status_code=400, detail="Некорректное значение direction")
        art.direction = FinanceArticleDirection(payload.direction)  # type: ignore[call-arg]

    if payload.scope is not None:
        if payload.scope not in {s.value for s in FinanceArticleScope}:  # type: ignore[attr-defined]
            raise HTTPException(status_code=400, detail="Некорректное значение scope")
        art.scope = FinanceArticleScope(payload.scope)  # type: ignore[call-arg]

    if payload.cost_kind is not None:
        if payload.cost_kind not in {c.value for c in FinanceArticleCostKind}:  # type: ignore[attr-defined]
            raise HTTPException(status_code=400, detail="Некорректное значение cost_kind")
        art.cost_kind = FinanceArticleCostKind(payload.cost_kind)  # type: ignore[call-arg]
    if payload.color is not None:
        art.color = payload.color[:16] if payload.color else None
    if payload.sort_order is not None:
        art.sort_order = int(payload.sort_order)

    if payload.is_active is not None:
        art.is_active = bool(payload.is_active)

    db.commit()
    db.refresh(art)

    return _article_response(art)


@router.delete("/articles/{article_id}")
async def delete_finance_article(
    article_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> Dict[str, bool]:
    """
    Удаление статьи.
    По умолчанию делаем мягкое удаление (is_active = False),
    чтобы не ломать ссылки из существующих транзакций.
    """
    _require_finance_access(current_user)

    art = db.query(FinanceArticle).filter(FinanceArticle.id == article_id).first()
    if not art:
        raise HTTPException(status_code=404, detail="Статья не найдена")

    art.is_active = False
    db.commit()
    return {"ok": True}


@router.get("/budget", response_model=List[BudgetEntryResponse])
async def list_budget_entries(
    target_id: int = Query(...),
    period: str = Query(..., min_length=7, max_length=7),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[BudgetEntryResponse]:
    _require_finance_access(current_user)
    rows = (
        db.query(BudgetEntry)
        .options(joinedload(BudgetEntry.article))
        .filter(BudgetEntry.target_id == target_id, BudgetEntry.period == period)
        .order_by(BudgetEntry.id.asc())
        .all()
    )

    # Считаем факт из транзакций за выбранный период (YYYY-MM → первый/последний день месяца)
    try:
        year, month = int(period[:4]), int(period[5:7])
        period_start = date(year, month, 1)
        if month == 12:
            period_end = date(year + 1, 1, 1)
        else:
            period_end = date(year, month + 1, 1)
    except (ValueError, IndexError):
        period_start = period_end = None

    fact_by_article: dict = {}
    if period_start and period_end:
        from sqlalchemy import func as sqlfunc
        fact_rows = (
            db.query(
                FinanceTransaction.article_id,
                sqlfunc.sum(FinanceTransaction.amount).label("total"),
            )
            .filter(
                FinanceTransaction.target_id == target_id,
                FinanceTransaction.article_id.isnot(None),
                FinanceTransaction.direction != FinanceTransactionDirection.TRANSFER,
                FinanceTransaction.occurred_at >= datetime.combine(period_start, datetime.min.time(), tzinfo=timezone.utc),
                FinanceTransaction.occurred_at < datetime.combine(period_end, datetime.min.time(), tzinfo=timezone.utc),
            )
            .group_by(FinanceTransaction.article_id)
            .all()
        )
        fact_by_article = {r.article_id: float(r.total or 0) for r in fact_rows}

    return [
        BudgetEntryResponse(
            id=row.id,
            target_id=row.target_id,
            article_id=row.article_id,
            article_name=getattr(row.article, "name", None),
            period=row.period,
            amount_plan=float(row.amount_plan or 0),
            amount_fact=fact_by_article.get(row.article_id, 0.0),
        )
        for row in rows
    ]


@router.put("/budget", response_model=List[BudgetEntryResponse])
async def save_budget_entries(
    payload: BudgetSaveRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[BudgetEntryResponse]:
    _require_finance_access(current_user)
    if len(payload.period) != 7:
        raise HTTPException(status_code=400, detail="period must have YYYY-MM format")
    with db_transaction(db):
        for item in payload.entries:
            article = (
                db.query(FinanceArticle)
                .filter(FinanceArticle.id == item.article_id, FinanceArticle.target_id == payload.target_id)
                .first()
            )
            if not article:
                raise HTTPException(status_code=400, detail=f"article_id {item.article_id} not found for target")
            row = (
                db.query(BudgetEntry)
                .filter(
                    BudgetEntry.target_id == payload.target_id,
                    BudgetEntry.article_id == item.article_id,
                    BudgetEntry.period == payload.period,
                )
                .first()
            )
            if row is None:
                row = BudgetEntry(target_id=payload.target_id, article_id=item.article_id, period=payload.period)
                db.add(row)
            row.amount_plan = item.amount_plan
    return await list_budget_entries(payload.target_id, payload.period, db, current_user)


@router.get("/metrics", response_model=List[MetricDefinitionResponse])
async def list_metrics(
    target_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[MetricDefinitionResponse]:
    _require_finance_access(current_user)
    q = db.query(MetricDefinition).options(joinedload(MetricDefinition.target))
    if target_id is not None:
        q = q.filter(MetricDefinition.target_id == target_id)
    rows = q.order_by(MetricDefinition.sort_order.asc(), MetricDefinition.name.asc()).all()
    return [_metric_response(row) for row in rows]


@router.post("/metrics", response_model=MetricDefinitionResponse, status_code=201)
async def create_metric(
    payload: MetricDefinitionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> MetricDefinitionResponse:
    _require_finance_access(current_user)
    if not db.query(FinanceTarget.id).filter(FinanceTarget.id == payload.target_id).first():
        raise HTTPException(status_code=400, detail="target_id not found")
    row = MetricDefinition(
        target_id=payload.target_id,
        name=payload.name.strip(),
        formula=payload.formula.strip(),
        unit=payload.unit,
        goal_value=payload.goal_value,
        sort_order=payload.sort_order,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    db.refresh(row, ["target"])
    return _metric_response(row)


@router.patch("/metrics/{metric_id}", response_model=MetricDefinitionResponse)
async def update_metric(
    metric_id: int,
    payload: MetricDefinitionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> MetricDefinitionResponse:
    _require_finance_access(current_user)
    row = db.query(MetricDefinition).options(joinedload(MetricDefinition.target)).filter(MetricDefinition.id == metric_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Metric not found")
    if payload.name is not None:
        row.name = payload.name.strip() or row.name
    if payload.formula is not None:
        row.formula = payload.formula.strip() or row.formula
    if payload.unit is not None:
        row.unit = payload.unit or None
    if payload.goal_value is not None:
        row.goal_value = payload.goal_value
    if payload.sort_order is not None:
        row.sort_order = int(payload.sort_order)
    db.commit()
    db.refresh(row)
    db.refresh(row, ["target"])
    return _metric_response(row)


@router.get("/metrics/{metric_id}/compute", response_model=MetricComputeResponse)
async def compute_metric(
    metric_id: int,
    period: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> MetricComputeResponse:
    _require_finance_access(current_user)
    row = db.query(MetricDefinition).filter(MetricDefinition.id == metric_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Metric not found")
    try:
        value = compute_metric_formula(db, row.target_id, row.formula, period)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return MetricComputeResponse(
        metric_id=row.id,
        name=row.name,
        formula=row.formula,
        value=value,
        unit=row.unit,
        goal_value=row.goal_value,
        period=period,
    )


@router.delete("/metrics/{metric_id}", status_code=204)
async def delete_metric(
    metric_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> None:
    _require_finance_access(current_user)
    row = db.query(MetricDefinition).filter(MetricDefinition.id == metric_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Metric not found")
    db.delete(row)
    db.commit()
    return None


@router.get("/dashboard", response_model=List[DashboardWidgetResponse])
async def list_dashboard_widgets(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[DashboardWidgetResponse]:
    _require_finance_access(current_user)
    rows = (
        db.query(DashboardWidget)
        .options(joinedload(DashboardWidget.metric), joinedload(DashboardWidget.target))
        .filter(DashboardWidget.owner_id == current_user.id)
        .order_by(DashboardWidget.position_y.asc(), DashboardWidget.position_x.asc(), DashboardWidget.id.asc())
        .all()
    )
    return [_widget_response(row) for row in rows]


@router.post("/dashboard/widgets", response_model=DashboardWidgetResponse, status_code=201)
async def create_dashboard_widget(
    payload: DashboardWidgetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> DashboardWidgetResponse:
    _require_finance_access(current_user)
    row = DashboardWidget(
        owner_id=current_user.id,
        metric_id=payload.metric_id,
        target_id=payload.target_id,
        widget_type=payload.widget_type,
        period_type=payload.period_type,
        position_x=payload.position_x,
        position_y=payload.position_y,
        width=payload.width,
        title_override=payload.title_override,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    db.refresh(row, ["metric", "target"])
    return _widget_response(row)


@router.patch("/dashboard/widgets/{widget_id}", response_model=DashboardWidgetResponse)
async def update_dashboard_widget(
    widget_id: int,
    payload: DashboardWidgetUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> DashboardWidgetResponse:
    _require_finance_access(current_user)
    row = (
        db.query(DashboardWidget)
        .options(joinedload(DashboardWidget.metric), joinedload(DashboardWidget.target))
        .filter(DashboardWidget.id == widget_id, DashboardWidget.owner_id == current_user.id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Widget not found")
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(row, key, value)
    db.commit()
    db.refresh(row)
    db.refresh(row, ["metric", "target"])
    return _widget_response(row)


@router.delete("/dashboard/widgets/{widget_id}", status_code=204)
async def delete_dashboard_widget(
    widget_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> None:
    _require_finance_access(current_user)
    row = db.query(DashboardWidget).filter(DashboardWidget.id == widget_id, DashboardWidget.owner_id == current_user.id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Widget not found")
    db.delete(row)
    db.commit()
    return None


@router.get("/dashboard/compute", response_model=List[DashboardWidgetComputedResponse])
async def compute_dashboard_widgets(
    period: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[DashboardWidgetComputedResponse]:
    _require_finance_access(current_user)
    rows = (
        db.query(DashboardWidget)
        .options(joinedload(DashboardWidget.metric), joinedload(DashboardWidget.target))
        .filter(DashboardWidget.owner_id == current_user.id)
        .order_by(DashboardWidget.position_y.asc(), DashboardWidget.position_x.asc(), DashboardWidget.id.asc())
        .all()
    )
    result: List[DashboardWidgetComputedResponse] = []
    for row in rows:
        base = _widget_response(row).model_dump()
        metric = getattr(row, "metric", None)
        value = None
        unit = None
        goal_value = None
        if metric is not None:
            try:
                value = compute_metric_formula(db, metric.target_id, metric.formula, period)
            except ValueError:
                value = None
            unit = metric.unit
            goal_value = metric.goal_value
        result.append(DashboardWidgetComputedResponse(**base, value=value, unit=unit, goal_value=goal_value))
    return result


@router.get("/ledger/transactions", response_model=List[FinanceLedgerTransactionRow])
async def list_finance_ledger_transactions(
    target_codes: Optional[List[str]] = Query(
        None,
        description="Фильтр по кодам target (personal, leninets, gogol_mogol, academy). Пусто = все эти цели.",
    ),
    date_from: Optional[date] = Query(None, description="Начало периода (включительно)"),
    date_to: Optional[date] = Query(None, description="Конец периода (включительно)"),
    limit: int = Query(5000, ge=1, le=50000),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceLedgerTransactionRow]:
    """
    Список транзакций единого журнала для дашборда личных финансов.
    По умолчанию возвращаются операции с любым target (personal, leninets, gogol_mogol, academy).
    """
    _require_finance_access(current_user)

    q = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .filter(
            FinanceTransaction.direction.in_(
                [FinanceTransactionDirection.INCOME, FinanceTransactionDirection.EXPENSE]
            )
        )
    )

    if target_codes:
        targets = db.query(FinanceTarget.id).filter(FinanceTarget.code.in_(target_codes)).all()
        target_ids = [t[0] for t in targets]
        if target_ids:
            q = q.filter(FinanceTransaction.target_id.in_(target_ids))
        else:
            return []
    else:
        # по умолчанию — только «личные» цели (для дашборда)
        default_codes = ["personal", "leninets", "gogol_mogol", "academy"]
        targets = db.query(FinanceTarget.id).filter(FinanceTarget.code.in_(default_codes)).all()
        target_ids = [t[0] for t in targets]
        if target_ids:
            q = q.filter(FinanceTransaction.target_id.in_(target_ids))
        else:
            return []

    if date_from is not None:
        q = q.filter(FinanceTransaction.occurred_at >= datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc))
    if date_to is not None:
        q = q.filter(FinanceTransaction.occurred_at <= datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc))

    q = q.order_by(FinanceTransaction.occurred_at.desc())
    items: List[FinanceTransaction] = q.limit(limit).all()

    rows: List[FinanceLedgerTransactionRow] = []
    for tx in items:
        target = getattr(tx, "target", None)
        article = getattr(tx, "article", None)
        rows.append(
            FinanceLedgerTransactionRow(
                id=tx.id,
                occurred_at=tx.occurred_at,
                amount=float(tx.amount or 0.0),
                direction=str(getattr(tx.direction, "value", tx.direction)),
                target_code=getattr(target, "code", None) if target else None,
                target_name=getattr(target, "name", None) if target else None,
                article_id=tx.article_id,
                article_name=getattr(article, "name", None) if article else None,
                counterparty_name=tx.counterparty_name,
                description_raw=tx.description_raw,
            )
        )
    return rows


@router.get("/balances", response_model=List[FinanceAccountBalance])
async def get_finance_account_balances(
    as_of: Optional[date] = Query(None, description="Дата, на которую считать остатки (по occurred_at <= as_of)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceAccountBalance]:
    """
    Остатки по счетам (finance_accounts) на указанную дату с оборотами доходов/расходов.
    Transfer-операции учитываются как перевод между своими счетами и не попадают в обороты.
    """
    _require_finance_access(current_user)

    # Базовый набор счетов
    accounts = db.query(FinanceAccount).filter(FinanceAccount.is_active.is_(True)).all()
    balance_map: Dict[int, Dict[str, float]] = {}
    for acc in accounts:
        balance_map[acc.id] = {
            "income": 0.0,
            "expense": 0.0,
            "balance": 0.0,
        }

    # Все транзакции до указанной даты
    q = db.query(FinanceTransaction)
    if as_of is not None:
        q = q.filter(FinanceTransaction.occurred_at <= as_of)
    txs: List[FinanceTransaction] = q.all()

    for tx in txs:
        # Доход / расход: учитываем только по account_id
        if tx.direction == FinanceTransactionDirection.INCOME:
            if tx.account_id and tx.account_id in balance_map:
                balance_map[tx.account_id]["income"] += float(tx.amount or 0.0)
                balance_map[tx.account_id]["balance"] += float(tx.amount or 0.0)
        elif tx.direction == FinanceTransactionDirection.EXPENSE:
            if tx.account_id and tx.account_id in balance_map:
                amt = float(tx.amount or 0.0)
                # amount для расхода может быть отрицательным; расход считаем по модулю
                balance_map[tx.account_id]["expense"] += abs(amt)
                balance_map[tx.account_id]["balance"] += amt
        elif tx.direction == FinanceTransactionDirection.TRANSFER:
            # Перевод между своими счетами: не считаем в доход/расход, только в балансах
            amt = abs(float(tx.amount or 0.0))
            if tx.account_id and tx.account_id in balance_map:
                balance_map[tx.account_id]["balance"] -= amt
            if tx.to_account_id and tx.to_account_id in balance_map:
                balance_map[tx.to_account_id]["balance"] += amt

    result: List[FinanceAccountBalance] = []
    for acc in accounts:
        data = balance_map.get(acc.id) or {"income": 0.0, "expense": 0.0, "balance": 0.0}
        result.append(
            FinanceAccountBalance(
                account_id=acc.id,
                account_code=acc.code,
                account_name=acc.name,
                income_total=round(data["income"], 2),
                expense_total=round(data["expense"], 2),
                balance=round(data["balance"], 2),
            )
        )
    # Сортируем по коду счета для стабильного вывода
    result.sort(key=lambda x: x.account_code)
    return result


@router.get("/pnl", response_model=List[FinancePnlRow])
async def get_finance_pnl(
    target_code: str = Query("academy", description="Код проекта/кошелька (finance_targets.code)"),
    date_from: Optional[date] = Query(None, description="Начало периода (включительно)"),
    date_to: Optional[date] = Query(None, description="Конец периода (включительно)"),
    group_by: str = Query("month", description="Группировка: 'month' или 'date'"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinancePnlRow]:
    """
    P&L по операциям журнала для выбранного проекта (target).
    Учитываются только direction = income/expense; transfer-операции исключаются.
    """
    _require_finance_access(current_user)

    target = db.query(FinanceTarget).filter(FinanceTarget.code == target_code).first()
    if not target:
        raise HTTPException(status_code=404, detail="Target с таким кодом не найден")

    q = db.query(FinanceTransaction).filter(
        FinanceTransaction.target_id == target.id,
        FinanceTransaction.direction.in_(
            [FinanceTransactionDirection.INCOME, FinanceTransactionDirection.EXPENSE]
        ),
    )
    if date_from is not None:
        q = q.filter(FinanceTransaction.occurred_at >= date_from)
    if date_to is not None:
        q = q.filter(FinanceTransaction.occurred_at <= date_to)

    txs: List[FinanceTransaction] = q.all()

    buckets: Dict[str, Dict[str, float]] = {}
    for tx in txs:
        if not tx.occurred_at:
            continue
        if group_by == "date":
            key = tx.occurred_at.date().isoformat()
        else:
            # month: YYYY-MM
            key = f"{tx.occurred_at.year:04d}-{tx.occurred_at.month:02d}"

        if key not in buckets:
            buckets[key] = {"income": 0.0, "expense": 0.0}

        amt = float(tx.amount or 0.0)
        if tx.direction == FinanceTransactionDirection.INCOME:
            buckets[key]["income"] += amt
        elif tx.direction == FinanceTransactionDirection.EXPENSE:
            buckets[key]["expense"] += abs(amt)

    rows: List[FinancePnlRow] = []
    for key in sorted(buckets.keys()):
        income = buckets[key]["income"]
        expense = buckets[key]["expense"]
        rows.append(
            FinancePnlRow(
                period=key,
                income=round(income, 2),
                expense=round(expense, 2),
                profit=round(income - expense, 2),
            )
        )
    return rows


# ─── Finance Command Center ──────────────────────────────────────────────────

@router.get("/command-center", response_model=CommandCenterResponse)
async def get_command_center(
    period: Optional[str] = Query(None, description="YYYY-MM, default = current month"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> CommandCenterResponse:
    _require_finance_access(current_user)

    from sqlalchemy import func as sqlfunc
    import calendar as _cal

    today = date.today()
    if period:
        y, m = int(period[:4]), int(period[5:7])
    else:
        y, m = today.year, today.month

    period_str = f"{y:04d}-{m:02d}"
    period_start = datetime(y, m, 1, tzinfo=timezone.utc)
    last_day = _cal.monthrange(y, m)[1]
    period_end = datetime(y, m, last_day, 23, 59, 59, tzinfo=timezone.utc)

    # previous month boundaries
    pm = m - 1 if m > 1 else 12
    py = y if m > 1 else y - 1
    pm_last = _cal.monthrange(py, pm)[1]
    prev_start = datetime(py, pm, 1, tzinfo=timezone.utc)
    prev_end = datetime(py, pm, pm_last, 23, 59, 59, tzinfo=timezone.utc)

    # ── helpers ───────────────────────────────────────────────────────────────
    def _txn_agg(start: datetime, end: datetime, target_id: Optional[int] = None):
        q = (
            db.query(FinanceTransaction.direction, sqlfunc.sum(FinanceTransaction.amount))
            .filter(
                FinanceTransaction.occurred_at >= start,
                FinanceTransaction.occurred_at <= end,
                FinanceTransaction.direction.in_([
                    FinanceTransactionDirection.INCOME,
                    FinanceTransactionDirection.EXPENSE,
                ]),
            )
        )
        if target_id is not None:
            q = q.filter(FinanceTransaction.target_id == target_id)
        q = q.group_by(FinanceTransaction.direction).all()
        income = expense = 0.0
        for direction, total in q:
            if direction == FinanceTransactionDirection.INCOME:
                income = float(total or 0)
            else:
                expense = abs(float(total or 0))
        return income, expense

    # ── KPI ───────────────────────────────────────────────────────────────────
    cur_inc, cur_exp = _txn_agg(period_start, period_end)
    prv_inc, prv_exp = _txn_agg(prev_start, prev_end)

    cur_profit = cur_inc - cur_exp
    prv_profit = prv_inc - prv_exp
    cur_margin = round(cur_profit / cur_inc * 100, 1) if cur_inc else 0.0

    def _pct_delta(cur: float, prv: float) -> float:
        return round((cur - prv) / prv * 100, 1) if prv else 0.0

    # ── Account balances ──────────────────────────────────────────────────────
    accounts = db.query(FinanceAccount).filter(FinanceAccount.is_active.is_(True)).order_by(FinanceAccount.code).all()

    def _account_balances_up_to(up_to: datetime) -> Dict[int, float]:
        txns = db.query(
            FinanceTransaction.account_id,
            FinanceTransaction.to_account_id,
            FinanceTransaction.direction,
            FinanceTransaction.amount,
        ).filter(FinanceTransaction.occurred_at <= up_to).all()
        bmap: Dict[int, float] = {a.id: 0.0 for a in accounts}
        for acc_id, to_acc_id, direction, amount in txns:
            amt = float(amount or 0)
            if direction == FinanceTransactionDirection.INCOME:
                if acc_id in bmap:
                    bmap[acc_id] += amt
            elif direction == FinanceTransactionDirection.EXPENSE:
                if acc_id in bmap:
                    bmap[acc_id] += amt  # amt is already negative for expenses or we use -abs
                    bmap[acc_id] -= abs(amt) + amt  # neutralise then subtract abs
            elif direction == FinanceTransactionDirection.TRANSFER:
                a = abs(amt)
                if acc_id in bmap:
                    bmap[acc_id] -= a
                if to_acc_id in bmap:
                    bmap[to_acc_id] += a
        return bmap

    # Simpler, correct balance computation
    def _account_balances_correct(up_to: datetime) -> Dict[int, float]:
        bmap: Dict[int, float] = {a.id: 0.0 for a in accounts}
        txns = (
            db.query(
                FinanceTransaction.account_id,
                FinanceTransaction.to_account_id,
                FinanceTransaction.direction,
                FinanceTransaction.amount,
            )
            .filter(FinanceTransaction.occurred_at <= up_to)
            .all()
        )
        for acc_id, to_acc_id, direction, amount in txns:
            amt = float(amount or 0)
            if direction == FinanceTransactionDirection.INCOME:
                if acc_id in bmap:
                    bmap[acc_id] += amt
            elif direction == FinanceTransactionDirection.EXPENSE:
                if acc_id in bmap:
                    bmap[acc_id] -= abs(amt)
            elif direction == FinanceTransactionDirection.TRANSFER:
                a = abs(amt)
                if acc_id in bmap:
                    bmap[acc_id] -= a
                if to_acc_id in bmap:
                    bmap[to_acc_id] += a
        return bmap

    # account income/expense this month only
    def _account_month_flow(start: datetime, end: datetime) -> Dict[int, Dict[str, float]]:
        fmap: Dict[int, Dict[str, float]] = {a.id: {"income": 0.0, "expense": 0.0} for a in accounts}
        txns = (
            db.query(FinanceTransaction.account_id, FinanceTransaction.direction, FinanceTransaction.amount)
            .filter(
                FinanceTransaction.occurred_at >= start,
                FinanceTransaction.occurred_at <= end,
                FinanceTransaction.direction.in_([
                    FinanceTransactionDirection.INCOME,
                    FinanceTransactionDirection.EXPENSE,
                ]),
            )
            .all()
        )
        for acc_id, direction, amount in txns:
            if acc_id in fmap:
                if direction == FinanceTransactionDirection.INCOME:
                    fmap[acc_id]["income"] += float(amount or 0)
                else:
                    fmap[acc_id]["expense"] += abs(float(amount or 0))
        return fmap

    bal_now = _account_balances_correct(period_end)
    bal_prev = _account_balances_correct(prev_end)
    month_flow = _account_month_flow(period_start, period_end)

    total_balance = round(sum(bal_now.values()), 2)
    total_balance_prev = round(sum(bal_prev.values()), 2)

    cc_accounts: List[CommandCenterAccount] = []
    for acc in accounts:
        bal = bal_now.get(acc.id, 0.0)
        prev_bal = bal_prev.get(acc.id, 0.0)
        flow = month_flow.get(acc.id, {"income": 0.0, "expense": 0.0})
        cc_accounts.append(CommandCenterAccount(
            id=acc.id,
            code=acc.code,
            name=acc.name,
            balance=round(bal, 2),
            balance_delta=round(bal - prev_bal, 2),
            income_month=round(flow["income"], 2),
            expense_month=round(flow["expense"], 2),
        ))

    # ── Project breakdown ─────────────────────────────────────────────────────
    # Get all active targets
    targets_list = db.query(FinanceTarget).filter(FinanceTarget.is_active.is_(True)).all()
    targets_map = {t.id: t for t in targets_list}

    # Get articles for cost_kind lookup
    articles_map: Dict[int, FinanceArticle] = {
        a.id: a for a in db.query(FinanceArticle).filter(FinanceArticle.is_active.isnot(False)).all()
    }

    # Aggregate transactions for the month by target + article direction + cost_kind
    proj_txns = (
        db.query(
            FinanceTransaction.target_id,
            FinanceTransaction.article_id,
            FinanceTransaction.direction,
            sqlfunc.sum(FinanceTransaction.amount).label("total"),
        )
        .filter(
            FinanceTransaction.occurred_at >= period_start,
            FinanceTransaction.occurred_at <= period_end,
            FinanceTransaction.direction.in_([
                FinanceTransactionDirection.INCOME,
                FinanceTransactionDirection.EXPENSE,
            ]),
        )
        .group_by(
            FinanceTransaction.target_id,
            FinanceTransaction.article_id,
            FinanceTransaction.direction,
        )
        .all()
    )

    proj_agg: Dict[Optional[int], Dict[str, float]] = {}
    for target_id, article_id, direction, total in proj_txns:
        key = target_id
        if key not in proj_agg:
            proj_agg[key] = {"income": 0.0, "variable": 0.0, "fixed": 0.0, "other": 0.0}
        amt = float(total or 0)
        if direction == FinanceTransactionDirection.INCOME:
            proj_agg[key]["income"] += amt
        else:
            article = articles_map.get(article_id) if article_id else None
            ck = str(article.cost_kind.value if hasattr(article.cost_kind, "value") else (article.cost_kind or "none")) if article else "none"
            if ck == "variable":
                proj_agg[key]["variable"] += abs(amt)
            elif ck == "fixed":
                proj_agg[key]["fixed"] += abs(amt)
            else:
                proj_agg[key]["other"] += abs(amt)

    cc_projects: List[CommandCenterProjectColumn] = []
    for target_id_key, agg in sorted(proj_agg.items(), key=lambda x: -(x[1]["income"])):
        target = targets_map.get(target_id_key) if target_id_key else None
        income = agg["income"]
        var_exp = agg["variable"]
        fixed_exp = agg["fixed"]
        other_exp = agg["other"]
        gross = income - var_exp
        net = income - var_exp - fixed_exp - other_exp
        margin = round(net / income * 100, 1) if income else 0.0
        cc_projects.append(CommandCenterProjectColumn(
            target_id=target_id_key,
            target_code=target.code if target else "unknown",
            target_name=target.name if target else "Без проекта",
            income=round(income, 2),
            variable_expense=round(var_exp, 2),
            fixed_expense=round(fixed_exp, 2),
            other_expense=round(other_exp, 2),
            gross_profit=round(gross, 2),
            net_profit=round(net, 2),
            net_margin=margin,
        ))

    # ── Cash flow last 6 months ───────────────────────────────────────────────
    cashflow: List[CommandCenterCashFlowPoint] = []
    for i in range(5, -1, -1):
        cm = m - i
        cy = y
        while cm <= 0:
            cm += 12
            cy -= 1
        cf_start = datetime(cy, cm, 1, tzinfo=timezone.utc)
        cf_last = _cal.monthrange(cy, cm)[1]
        cf_end = datetime(cy, cm, cf_last, 23, 59, 59, tzinfo=timezone.utc)
        cf_inc, cf_exp = _txn_agg(cf_start, cf_end)
        cashflow.append(CommandCenterCashFlowPoint(
            period=f"{cy:04d}-{cm:02d}",
            income=round(cf_inc, 2),
            expense=round(cf_exp, 2),
            net=round(cf_inc - cf_exp, 2),
        ))

    # ── Recent transactions ───────────────────────────────────────────────────
    recent_raw = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.account),
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .filter(
            FinanceTransaction.direction.in_([
                FinanceTransactionDirection.INCOME,
                FinanceTransactionDirection.EXPENSE,
                FinanceTransactionDirection.TRANSFER,
            ])
        )
        .order_by(FinanceTransaction.occurred_at.desc(), FinanceTransaction.id.desc())
        .limit(15)
        .all()
    )

    recent_transactions: List[CommandCenterTransaction] = []
    for tx in recent_raw:
        recent_transactions.append(CommandCenterTransaction(
            id=tx.id,
            occurred_at=tx.occurred_at,
            amount=float(tx.amount or 0),
            direction=str(tx.direction.value if hasattr(tx.direction, "value") else tx.direction),
            account_name=tx.account.name if tx.account else None,
            target_id=tx.target_id,
            target_name=tx.target.name if tx.target else None,
            article_name=tx.article.name if tx.article else None,
            description=tx.description_raw,
        ))

    # ── Alerts ────────────────────────────────────────────────────────────────
    unclassified_count = (
        db.query(sqlfunc.count(FinanceTransaction.id))
        .filter(
            FinanceTransaction.occurred_at >= period_start,
            FinanceTransaction.occurred_at <= period_end,
            FinanceTransaction.direction.in_([
                FinanceTransactionDirection.INCOME,
                FinanceTransactionDirection.EXPENSE,
            ]),
            (FinanceTransaction.article_id.is_(None)) | (FinanceTransaction.target_id.is_(None)),
        )
        .scalar() or 0
    )

    alerts: List[CommandCenterAlert] = []
    if unclassified_count > 0:
        alerts.append(CommandCenterAlert(
            type="unclassified",
            count=unclassified_count,
            message=f"{unclassified_count} операций без статьи или проекта",
        ))

    return CommandCenterResponse(
        period=period_str,
        kpi=CommandCenterKpi(
            total_balance=total_balance,
            income_month=round(cur_inc, 2),
            expense_month=round(cur_exp, 2),
            net_profit_month=round(cur_profit, 2),
            net_margin_month=cur_margin,
            income_delta_pct=_pct_delta(cur_inc, prv_inc),
            expense_delta_pct=_pct_delta(cur_exp, prv_exp),
            profit_delta=round(cur_profit - prv_profit, 2),
            balance_delta=round(total_balance - total_balance_prev, 2),
        ),
        accounts=cc_accounts,
        projects=cc_projects,
        cashflow=cashflow,
        recent_transactions=recent_transactions,
        alerts=alerts,
    )


# ─── P&L Multi-Period Report ─────────────────────────────────────────────────

def _generate_pl_periods(period_from: str, period_to: str) -> List[str]:
    y0, m0 = int(period_from[:4]), int(period_from[5:7])
    y1, m1 = int(period_to[:4]), int(period_to[5:7])
    periods: List[str] = []
    y, m = y0, m0
    while (y, m) <= (y1, m1):
        periods.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return periods


def _next_period_str(period: str) -> str:
    y, m = int(period[:4]), int(period[5:7])
    m += 1
    if m > 12:
        m = 1
        y += 1
    return f"{y:04d}-{m:02d}"


def _get_article_cells(
    article_id: int,
    fact_map: Dict[int, Dict[str, float]],
    plan_map: Dict[int, Dict[str, float]],
    periods: List[str],
    children_map: Dict,
) -> tuple:
    own_facts = fact_map.get(article_id, {})
    own_plans = plan_map.get(article_id, {})
    facts = {p: own_facts.get(p, 0.0) for p in periods}
    plans = {p: own_plans.get(p, 0.0) for p in periods}
    for child in children_map.get(article_id, []):
        cf, cp = _get_article_cells(child.id, fact_map, plan_map, periods, children_map)
        for p in periods:
            facts[p] += cf[p]
            plans[p] += cp[p]
    return facts, plans


def _add_article_rows(
    articles_list: List,
    result_rows: List[PLReportRow],
    fact_map: Dict,
    plan_map: Dict,
    periods: List[str],
    children_map: Dict,
    level: int = 0,
) -> tuple:
    section_facts = {p: 0.0 for p in periods}
    section_plans = {p: 0.0 for p in periods}
    for article in articles_list:
        cells, plan_cells = _get_article_cells(article.id, fact_map, plan_map, periods, children_map)
        result_rows.append(PLReportRow(
            row_type="article",
            id=article.id,
            name=article.name,
            code=article.code,
            level=level,
            direction=str(article.direction.value) if hasattr(article.direction, "value") else str(article.direction),
            cost_kind=str(article.cost_kind.value) if hasattr(article.cost_kind, "value") else str(article.cost_kind or "none"),
            color=article.color,
            cells={p: cells[p] for p in periods},
            plan_cells={p: plan_cells[p] for p in periods},
            total_fact=sum(cells.values()),
            total_plan=sum(plan_cells.values()),
        ))
        children = children_map.get(article.id, [])
        if children:
            _add_article_rows(children, result_rows, fact_map, plan_map, periods, children_map, level + 1)
        for p in periods:
            section_facts[p] += cells[p]
            section_plans[p] += plan_cells[p]
    return section_facts, section_plans


@router.get("/models/{model_id}/pl-report", response_model=PLReportResponse)
async def get_pl_report(
    model_id: int,
    period_from: str = Query(..., description="YYYY-MM"),
    period_to: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> PLReportResponse:
    _require_finance_access(current_user)
    model = db.query(FinanceModel).filter(FinanceModel.id == model_id).first()
    if not model:
        raise HTTPException(status_code=404, detail="Finance model not found")

    periods = _generate_pl_periods(period_from, period_to)

    if not model.target_id:
        return PLReportResponse(model_id=model_id, target_id=None, periods=periods, rows=[])

    target_id = model.target_id

    articles = (
        db.query(FinanceArticle)
        .filter(FinanceArticle.target_id == target_id, FinanceArticle.is_active.isnot(False))
        .order_by(FinanceArticle.sort_order, FinanceArticle.id)
        .all()
    )

    from sqlalchemy import func as sqlfunc
    p_start = period_from + "-01"
    p_end = _next_period_str(period_to) + "-01"

    fact_rows = (
        db.query(
            FinanceTransaction.article_id,
            sqlfunc.to_char(FinanceTransaction.occurred_at, "YYYY-MM").label("period"),
            sqlfunc.sum(FinanceTransaction.amount).label("total"),
        )
        .filter(
            FinanceTransaction.target_id == target_id,
            FinanceTransaction.occurred_at >= p_start,
            FinanceTransaction.occurred_at < p_end,
            FinanceTransaction.article_id.isnot(None),
            FinanceTransaction.direction != FinanceTransactionDirection.TRANSFER,
        )
        .group_by(
            FinanceTransaction.article_id,
            sqlfunc.to_char(FinanceTransaction.occurred_at, "YYYY-MM"),
        )
        .all()
    )

    fact_map: Dict[int, Dict[str, float]] = {}
    for r in fact_rows:
        if r.article_id not in fact_map:
            fact_map[r.article_id] = {}
        fact_map[r.article_id][r.period] = float(r.total or 0)

    plan_rows_q = (
        db.query(BudgetEntry.article_id, BudgetEntry.period, BudgetEntry.amount_plan)
        .filter(
            BudgetEntry.target_id == target_id,
            BudgetEntry.period >= period_from,
            BudgetEntry.period <= period_to,
        )
        .all()
    )

    plan_map: Dict[int, Dict[str, float]] = {}
    for r in plan_rows_q:
        if r.article_id not in plan_map:
            plan_map[r.article_id] = {}
        plan_map[r.article_id][r.period] = float(r.amount_plan or 0)

    children_map: Dict[int, List] = {}
    for a in articles:
        children_map.setdefault(a.parent_id, []).append(a)

    spacer = PLReportRow(row_type="spacer", name="", cells={p: 0.0 for p in periods}, plan_cells={p: 0.0 for p in periods})
    rows: List[PLReportRow] = []

    # ── Income section ────────────────────────────────────────────────────────
    root_income = [a for a in articles if str(getattr(a.direction, "value", a.direction)) == "income" and not a.parent_id]
    income_facts: Dict[str, float] = {p: 0.0 for p in periods}
    income_plans: Dict[str, float] = {p: 0.0 for p in periods}

    if root_income:
        rows.append(PLReportRow(row_type="header", name="Доходы", bold=True, direction="income", cells={p: 0.0 for p in periods}, plan_cells={p: 0.0 for p in periods}))
        sf, sp = _add_article_rows(root_income, rows, fact_map, plan_map, periods, children_map)
        income_facts, income_plans = sf, sp
        rows.append(PLReportRow(
            row_type="total", name="Итого доходы", bold=True, direction="income",
            cells=dict(income_facts), plan_cells=dict(income_plans),
            total_fact=sum(income_facts.values()), total_plan=sum(income_plans.values()),
        ))
        rows.append(spacer)

    # ── Expense sections ──────────────────────────────────────────────────────
    all_exp_facts: Dict[str, float] = {p: 0.0 for p in periods}
    all_exp_plans: Dict[str, float] = {p: 0.0 for p in periods}
    for cost_kind_key, section_name in [("variable", "Переменные расходы"), ("fixed", "Постоянные расходы"), ("none", "Прочие расходы")]:
        def _ck(a) -> str:  # noqa: E306
            return str(a.cost_kind.value) if hasattr(a.cost_kind, "value") else str(a.cost_kind or "none")

        def _dir(a) -> str:  # noqa: E306
            return str(a.direction.value) if hasattr(a.direction, "value") else str(a.direction)

        if cost_kind_key == "none":
            group_arts = [a for a in articles if _dir(a) == "expense" and not a.parent_id and _ck(a) not in ("variable", "fixed")]
        else:
            group_arts = [a for a in articles if _dir(a) == "expense" and not a.parent_id and _ck(a) == cost_kind_key]

        if not group_arts:
            continue

        rows.append(PLReportRow(row_type="header", name=section_name, bold=True, direction="expense", cells={p: 0.0 for p in periods}, plan_cells={p: 0.0 for p in periods}))
        gf, gp = _add_article_rows(group_arts, rows, fact_map, plan_map, periods, children_map)
        rows.append(PLReportRow(
            row_type="total", name=f"Итого {section_name.lower()}", bold=True, direction="expense",
            cells=dict(gf), plan_cells=dict(gp),
            total_fact=sum(gf.values()), total_plan=sum(gp.values()),
        ))
        for p in periods:
            all_exp_facts[p] += gf[p]
            all_exp_plans[p] += gp[p]

        if cost_kind_key == "variable":
            var_facts, var_plans = gf, gp
            gross_f = {p: income_facts[p] - var_facts[p] for p in periods}
            gross_p = {p: income_plans[p] - var_plans[p] for p in periods}
            rows.append(spacer)
            rows.append(PLReportRow(
                row_type="calculated", name="Валовая прибыль", bold=True, key="gross_profit",
                cells=dict(gross_f), plan_cells=dict(gross_p),
                total_fact=sum(gross_f.values()), total_plan=sum(gross_p.values()),
            ))
            gm = {p: (gross_f[p] / income_facts[p] * 100) if income_facts.get(p, 0) != 0 else 0.0 for p in periods}
            rows.append(PLReportRow(
                row_type="calculated", name="Рентабельность по валовой прибыли", is_percent=True, key="gross_margin",
                cells=dict(gm), plan_cells={p: 0.0 for p in periods},
            ))

        rows.append(spacer)

    # ── Net profit ────────────────────────────────────────────────────────────
    net_f = {p: income_facts[p] - all_exp_facts[p] for p in periods}
    net_p = {p: income_plans[p] - all_exp_plans[p] for p in periods}
    rows.append(PLReportRow(
        row_type="calculated", name="Чистая прибыль", bold=True, key="net_profit",
        cells=dict(net_f), plan_cells=dict(net_p),
        total_fact=sum(net_f.values()), total_plan=sum(net_p.values()),
    ))
    nm = {p: (net_f[p] / income_facts[p] * 100) if income_facts.get(p, 0) != 0 else 0.0 for p in periods}
    rows.append(PLReportRow(
        row_type="calculated", name="Рентабельность по чистой прибыли", is_percent=True, key="net_margin",
        cells=dict(nm), plan_cells={p: 0.0 for p in periods},
    ))

    return PLReportResponse(model_id=model_id, target_id=target_id, periods=periods, rows=rows)


@router.get("/analytics/summary", response_model=FinanceAnalyticsSummaryResponse)
async def get_finance_analytics_summary(
    date_from: Optional[date] = Query(None, description="Начало периода (включительно)"),
    date_to: Optional[date] = Query(None, description="Конец периода (включительно)"),
    group_by: str = Query("month", description="Группировка P&L: month | date"),
    db: Session = Depends(get_db),
    current_user: User = Depends(dep_require_finance_access),
) -> FinanceAnalyticsSummaryResponse:
    import traceback as _tb
    try:
        return _finance_analytics_impl(date_from, date_to, group_by, db)
    except HTTPException:
        raise
    except Exception as exc:
        import logging as _log
        _log.getLogger(__name__).error("finance analytics error: %s\n%s", exc, _tb.format_exc())
        raise HTTPException(status_code=500, detail=f"Ошибка аналитики: {type(exc).__name__}: {exc}")


def _finance_analytics_impl(
    date_from: Optional[date],
    date_to: Optional[date],
    group_by: str,
    db: Session,
) -> FinanceAnalyticsSummaryResponse:
    if date_to is None:
        date_to = date.today()
    if date_from is None:
        date_from = date(date_to.year, 1, 1)
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="date_from must be <= date_to")

    period_start = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)
    period_end = datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc)
    days_span = max(1, (date_to - date_from).days + 1)
    prev_date_to = date_from - timedelta(days=1)
    prev_date_from = prev_date_to - timedelta(days=days_span - 1)
    prev_start = datetime.combine(prev_date_from, datetime.min.time(), tzinfo=timezone.utc)
    prev_end = datetime.combine(prev_date_to, datetime.max.time(), tzinfo=timezone.utc)

    def _sum_period(start_dt: datetime, end_dt: datetime) -> Dict[str, float]:
        rows = (
            db.query(FinanceTransaction.amount, FinanceTransaction.direction)
            .filter(
                FinanceTransaction.occurred_at >= start_dt,
                FinanceTransaction.occurred_at <= end_dt,
                FinanceTransaction.direction.in_(
                    [FinanceTransactionDirection.INCOME, FinanceTransactionDirection.EXPENSE]
                ),
            )
            .all()
        )
        income_total = 0.0
        expense_total = 0.0
        for amount, direction in rows:
            amt = float(amount or 0.0)
            if direction == FinanceTransactionDirection.INCOME:
                income_total += amt
            elif direction == FinanceTransactionDirection.EXPENSE:
                expense_total += abs(amt)
        return {
            "income": round(income_total, 2),
            "expense": round(expense_total, 2),
            "profit": round(income_total - expense_total, 2),
        }

    current_totals = _sum_period(period_start, period_end)
    previous_totals = _sum_period(prev_start, prev_end)

    payment_summary = get_payment_status_summary(db, today=date_to)

    unclassified_rows = (
        db.query(FinanceTransaction.amount, FinanceTransaction.direction)
        .filter(
            FinanceTransaction.occurred_at >= period_start,
            FinanceTransaction.occurred_at <= period_end,
            (FinanceTransaction.target_id.is_(None)) | (FinanceTransaction.article_id.is_(None)),
        )
        .all()
    )
    unclassified_amount = 0.0
    for amount, direction in unclassified_rows:
        amt = float(amount or 0.0)
        if direction == FinanceTransactionDirection.EXPENSE:
            unclassified_amount += abs(amt)
        else:
            unclassified_amount += amt

    pnl_query = (
        db.query(FinanceTransaction)
        .filter(
            FinanceTransaction.occurred_at >= period_start,
            FinanceTransaction.occurred_at <= period_end,
            FinanceTransaction.direction.in_(
                [FinanceTransactionDirection.INCOME, FinanceTransactionDirection.EXPENSE]
            ),
        )
        .all()
    )
    pnl_buckets: Dict[str, Dict[str, float]] = {}
    for tx in pnl_query:
        if not tx.occurred_at:
            continue
        key = tx.occurred_at.date().isoformat() if group_by == "date" else f"{tx.occurred_at.year:04d}-{tx.occurred_at.month:02d}"
        if key not in pnl_buckets:
            pnl_buckets[key] = {"income": 0.0, "expense": 0.0}
        amt = float(tx.amount or 0.0)
        if tx.direction == FinanceTransactionDirection.INCOME:
            pnl_buckets[key]["income"] += amt
        elif tx.direction == FinanceTransactionDirection.EXPENSE:
            pnl_buckets[key]["expense"] += abs(amt)
    pnl_rows = [
        FinancePnlRow(
            period=key,
            income=round(values["income"], 2),
            expense=round(values["expense"], 2),
            profit=round(values["income"] - values["expense"], 2),
        )
        for key, values in sorted(pnl_buckets.items())
    ]

    target_rows = (
        db.query(
            FinanceTarget.id,
            FinanceTarget.code,
            FinanceTarget.name,
            FinanceTransaction.amount,
            FinanceTransaction.direction,
        )
        .join(FinanceTarget, FinanceTarget.id == FinanceTransaction.target_id, isouter=True)
        .filter(
            FinanceTransaction.occurred_at >= period_start,
            FinanceTransaction.occurred_at <= period_end,
            FinanceTransaction.direction.in_(
                [FinanceTransactionDirection.INCOME, FinanceTransactionDirection.EXPENSE]
            ),
        )
        .all()
    )
    target_buckets: Dict[str, Dict[str, object]] = {}
    for target_id, code, name, amount, direction in target_rows:
        bucket_key = str(code or "__unassigned__")
        if bucket_key not in target_buckets:
            target_buckets[bucket_key] = {
                "target_id": target_id,
                "target_code": code or "unassigned",
                "target_name": name or "Не распределено",
                "income": 0.0,
                "expense": 0.0,
            }
        amt = float(amount or 0.0)
        if direction == FinanceTransactionDirection.INCOME:
            target_buckets[bucket_key]["income"] = float(target_buckets[bucket_key]["income"]) + amt
        elif direction == FinanceTransactionDirection.EXPENSE:
            target_buckets[bucket_key]["expense"] = float(target_buckets[bucket_key]["expense"]) + abs(amt)
    target_breakdown = sorted(
        [
            FinanceAnalyticsTargetBreakdownRow(
                target_id=data["target_id"],
                target_code=str(data["target_code"]),
                target_name=str(data["target_name"]),
                income=round(float(data["income"]), 2),
                expense=round(float(data["expense"]), 2),
                profit=round(float(data["income"]) - float(data["expense"]), 2),
            )
            for data in target_buckets.values()
        ],
        key=lambda row: abs(row.profit),
        reverse=True,
    )

    expense_rows = (
        db.query(
            FinanceArticle.id,
            FinanceArticle.name,
            FinanceArticle.cost_kind,
            FinanceTransaction.amount,
        )
        .join(FinanceArticle, FinanceArticle.id == FinanceTransaction.article_id, isouter=True)
        .filter(
            FinanceTransaction.occurred_at >= period_start,
            FinanceTransaction.occurred_at <= period_end,
            FinanceTransaction.direction == FinanceTransactionDirection.EXPENSE,
        )
        .all()
    )
    expense_buckets: Dict[str, Dict[str, object]] = {}
    for article_id, name, cost_kind, amount in expense_rows:
        bucket_key = str(article_id or "__unassigned__")
        if bucket_key not in expense_buckets:
            expense_buckets[bucket_key] = {
                "article_id": article_id,
                "article_name": name or "Без статьи",
                "cost_kind": str(getattr(cost_kind, "value", cost_kind)) if cost_kind is not None else None,
                "amount": 0.0,
            }
        expense_buckets[bucket_key]["amount"] = float(expense_buckets[bucket_key]["amount"]) + abs(float(amount or 0.0))
    expense_breakdown = sorted(
        [
            FinanceAnalyticsExpenseBreakdownRow(
                article_id=data["article_id"],
                article_name=str(data["article_name"]),
                cost_kind=data["cost_kind"],
                amount=round(float(data["amount"]), 2),
            )
            for data in expense_buckets.values()
        ],
        key=lambda row: row.amount,
        reverse=True,
    )[:12]

    accounts = db.query(FinanceAccount).filter(FinanceAccount.is_active.is_(True)).all()
    balance_map: Dict[int, Dict[str, float]] = {
        acc.id: {"income": 0.0, "expense": 0.0, "balance": 0.0} for acc in accounts
    }
    balance_rows = db.query(FinanceTransaction).filter(FinanceTransaction.occurred_at <= period_end).all()
    for tx in balance_rows:
        if tx.direction == FinanceTransactionDirection.INCOME and tx.account_id in balance_map:
            amt = float(tx.amount or 0.0)
            balance_map[tx.account_id]["income"] += amt
            balance_map[tx.account_id]["balance"] += amt
        elif tx.direction == FinanceTransactionDirection.EXPENSE and tx.account_id in balance_map:
            amt = float(tx.amount or 0.0)
            balance_map[tx.account_id]["expense"] += abs(amt)
            balance_map[tx.account_id]["balance"] += amt
        elif tx.direction == FinanceTransactionDirection.TRANSFER:
            amt = abs(float(tx.amount or 0.0))
            if tx.account_id in balance_map:
                balance_map[tx.account_id]["balance"] -= amt
            if tx.to_account_id in balance_map:
                balance_map[tx.to_account_id]["balance"] += amt
    account_balances = [
        FinanceAccountBalance(
            account_id=acc.id,
            account_code=acc.code,
            account_name=acc.name,
            income_total=round(balance_map[acc.id]["income"], 2),
            expense_total=round(balance_map[acc.id]["expense"], 2),
            balance=round(balance_map[acc.id]["balance"], 2),
        )
        for acc in sorted(accounts, key=lambda item: item.code)
    ]

    return FinanceAnalyticsSummaryResponse(
        date_from=date_from,
        date_to=date_to,
        kpi=FinanceAnalyticsKpiBlock(
            income_total=current_totals["income"],
            expense_total=current_totals["expense"],
            profit_total=current_totals["profit"],
            prev_income_total=previous_totals["income"],
            prev_expense_total=previous_totals["expense"],
            prev_profit_total=previous_totals["profit"],
            income_delta=round(current_totals["income"] - previous_totals["income"], 2),
            expense_delta=round(current_totals["expense"] - previous_totals["expense"], 2),
            profit_delta=round(current_totals["profit"] - previous_totals["profit"], 2),
            overdue_payments_3_count=int(payment_summary["overdue_3_count"]),
            overdue_payments_10_count=int(payment_summary["overdue_10_count"]),
            unclassified_transactions_count=len(unclassified_rows),
            unclassified_transactions_amount=round(unclassified_amount, 2),
        ),
        pnl=pnl_rows,
        target_breakdown=target_breakdown,
        expense_breakdown=expense_breakdown,
        account_balances=account_balances,
    )


@router.post("/import")
async def import_finance_transactions(
    account_code: str = Form(..., description="Код счёта (finance_accounts.code)"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> Dict[str, int]:
    """
    Импорт операций из CSV/XLSX напрямую в единый финансовый журнал.

    Минимальный CSV-формат: колонки
      date,amount,counterparty,description,bank_operation_id(optional)
    """
    _require_finance_access(current_user)

    account = db.query(FinanceAccount).filter(FinanceAccount.code == account_code).first()
    if not account:
        raise HTTPException(status_code=400, detail="account_code не найден в finance_accounts")

    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")

    created = 0
    skipped = 0

    def _upsert_row(
        date_str: str,
        amount_val: float,
        counterparty: str,
        description: str,
        bank_operation_id: Optional[str],
        bank_source: str,
    ) -> None:
        nonlocal created, skipped
        if not date_str or not amount_val:
            skipped += 1
            return
        # Нормализуем строку даты: поддерживаем "YYYY-MM-DD", "YYYY-MM-DD HH:MM:SS" и т.п.
        try:
            ds = date_str.strip()
            # Если дата пришла как "2026-03-09 00:00:00" (из Excel), берём только часть до пробела
            if " " in ds and "T" not in ds:
                ds = ds.split(" ")[0]
            occurred_at = datetime.fromisoformat(ds + "T12:00:00")
        except Exception:
            skipped += 1
            return

        direction = "income" if amount_val > 0 else "expense"
        # v2: версионированный dedup_seed, чтобы избежать конфликтов со старыми импортами
        dedup_seed = f"v2|{bank_source}|{date_str}|{amount_val}|{counterparty.strip()}|{description.strip()}"
        dedup_hash = hashlib.sha1(dedup_seed.encode("utf-8")).hexdigest()

        # Проверка на дубликат.
        # Если есть bank_operation_id — считаем дубликатом по нему.
        # Если operation_id отсутствует — используем dedup_hash.
        if bank_operation_id:
            existing = (
                db.query(FinanceTransaction.id)
                .filter(
                    FinanceTransaction.bank_source == bank_source,
                    FinanceTransaction.bank_operation_id == bank_operation_id,
                )
                .first()
            )
        else:
            existing = (
                db.query(FinanceTransaction.id)
                .filter(
                    FinanceTransaction.bank_source == bank_source,
                    FinanceTransaction.dedup_hash == dedup_hash,
                )
                .first()
            )
        if existing:
            skipped += 1
            return

        tx = FinanceTransaction(
            occurred_at=occurred_at,
            amount=amount_val,
            direction=direction,
            account_id=account.id,
            to_account_id=None,
            transfer_group_id=None,
            counterparty_name=counterparty.strip() or None,
            counterparty_phone=None,
            description_raw=description.strip() or None,
            bank_source=bank_source,
            bank_operation_id=bank_operation_id,
            dedup_hash=dedup_hash,
            target_id=None,
            article_id=None,
            student_id=None,
            group_id=None,
            teacher_id=None,
            status=FinanceTransactionStatus.NEW,
        )
        try:
            with db_transaction(db):
                db.add(tx)
                # авто-классификация по правилам для новых импортированных операций
                apply_recognition_rules(db, tx)
            created += 1
        except IntegrityError:
            # Дубликат по уникальному индексу или другой конфликт — пропускаем строку
            skipped += 1

    if filename.endswith(".csv"):
        bank_source = "import_csv"
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        for row in reader:
            date_str = (row.get("date") or "").strip()
            amount_raw = (row.get("amount") or "").replace(" ", "").replace(",", ".")
            try:
                amount_val = float(amount_raw)
            except ValueError:
                skipped += 1
                continue
            counterparty = (row.get("counterparty") or "").strip()
            description = (row.get("description") or "").strip()
            bank_operation_id = (row.get("bank_operation_id") or "").strip() or None
            _upsert_row(date_str, amount_val, counterparty, description, bank_operation_id, bank_source)
    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl не установлен в окружении сервера")

        # Для нового импорта Excel используем отдельный bank_source,
        # чтобы не конфликтовать со старыми записями ("import_xlsx").
        bank_source = "import_xlsx_v2"
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Файл .xlsx без строк")

        # Нормализуем заголовки: поддерживаем как минимальный формат
        # (date, amount, counterparty, description, bank_operation_id),
        # так и русские выгрузки из банка (Дата, Сумма, Контрагент, Описание / источник, Тип).
        raw_header = [str(v).strip().lower() if v is not None else "" for v in rows[0]]
        rus_to_std = {
            "дата": "date",
            "сумма": "amount",
            "контрагент": "counterparty",
            "описание / источник": "description",
            "описание/источник": "description",
            "описание": "description",
            "тип": "type",
        }
        header = [rus_to_std.get(name, name) for name in raw_header]
        col_index = {name: idx for idx, name in enumerate(header)}

        for r in rows[1:]:
            if not any(r):
                continue

            def _get(col: str) -> str:
                idx = col_index.get(col)
                if idx is None or idx >= len(r):
                    return ""
                val = r[idx]
                return "" if val is None else str(val).strip()

            date_str = _get("date")
            # Поддержка формата дат dd.mm.yyyy (банковская выгрузка) и ISO yyyy-mm-dd
            if date_str and "." in date_str and "-" not in date_str:
                from datetime import datetime as _dt

                try:
                    parsed = _dt.strptime(date_str, "%d.%m.%Y")
                    date_str = parsed.date().isoformat()
                except Exception:
                    # Некорректные даты отфильтрует _upsert_row
                    pass

            amount_raw = _get("amount").replace(" ", "").replace(",", ".")
            try:
                amount_val = float(amount_raw)
            except ValueError:
                skipped += 1
                continue

            # Для банковских выгрузок, где сумма всегда положительная,
            # используем колонку "Тип" (Доход / Расход), если она есть.
            tx_type = _get("type").lower()
            if tx_type.startswith("расход"):
                amount_val = -abs(amount_val)
            elif tx_type.startswith("доход"):
                amount_val = abs(amount_val)

            counterparty = _get("counterparty")
            description = _get("description")
            bank_operation_id = _get("bank_operation_id") or None
            _upsert_row(date_str, amount_val, counterparty, description, bank_operation_id, bank_source)
    else:
        raise HTTPException(status_code=400, detail="Поддерживаются только форматы .csv и .xlsx")

    # Итоговые числа по импортированным и пропущенным строкам
    return {"imported": created, "skipped": skipped}


@router.post("/manual-transaction", response_model=FinanceLedgerBankRow)
async def create_manual_transaction(
    payload: FinanceManualTransactionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceLedgerBankRow:
    """
    Ручное добавление операции в единый журнал (например, наличные — счёт «Наличка»).
    Операция создаётся сразу с статусом classified.
    """
    _require_finance_access(current_user)

    amount = float(payload.amount or 0.0)
    if amount == 0:
        raise HTTPException(status_code=400, detail="Сумма не может быть нулевой")

    if payload.direction not in {d.value for d in FinanceTransactionDirection}:  # type: ignore[attr-defined]
        raise HTTPException(status_code=400, detail="Некорректное направление: income, expense или transfer")

    account = db.query(FinanceAccount).filter(FinanceAccount.id == payload.account_id, FinanceAccount.is_active.is_(True)).first()
    if not account:
        raise HTTPException(status_code=400, detail="Счёт не найден или неактивен")

    if payload.direction == "transfer":
        if not payload.to_account_id:
            raise HTTPException(status_code=400, detail="Для перевода необходимо указать счёт-назначения (to_account_id)")
        if payload.to_account_id == payload.account_id:
            raise HTTPException(status_code=400, detail="Счёт-источник и счёт-назначения не могут совпадать")
        to_account_obj = db.query(FinanceAccount).filter(FinanceAccount.id == payload.to_account_id, FinanceAccount.is_active.is_(True)).first()
        if not to_account_obj:
            raise HTTPException(status_code=400, detail="Счёт-назначения не найден или неактивен")
    else:
        to_account_obj = None

    if payload.article_id and not db.query(FinanceArticle.id).filter(FinanceArticle.id == payload.article_id).first():
        raise HTTPException(status_code=400, detail="Статья не найдена")
    if payload.target_id and not db.query(FinanceTarget.id).filter(FinanceTarget.id == payload.target_id).first():
        raise HTTPException(status_code=400, detail="Цель/проект не найден")

    occurred_at = datetime.combine(payload.occurred_at, datetime.min.time(), tzinfo=timezone.utc)
    description = (payload.description or "").strip() or None

    if payload.direction == "transfer":
        import uuid as _uuid
        group_id = str(_uuid.uuid4())
        # Расход со счёта-источника
        tx_out = FinanceTransaction(
            occurred_at=occurred_at,
            amount=abs(amount),
            direction=FinanceTransactionDirection.EXPENSE,
            account_id=payload.account_id,
            to_account_id=payload.to_account_id,
            transfer_group_id=group_id,
            counterparty_name=description,
            counterparty_phone=None,
            description_raw=description,
            bank_source="manual",
            bank_operation_id=None,
            dedup_hash=None,
            target_id=payload.target_id,
            article_id=None,
            student_id=None,
            group_id=None,
            teacher_id=None,
            status=FinanceTransactionStatus.CLASSIFIED,
        )
        # Приход на счёт-назначения
        tx_in = FinanceTransaction(
            occurred_at=occurred_at,
            amount=abs(amount),
            direction=FinanceTransactionDirection.INCOME,
            account_id=payload.to_account_id,
            to_account_id=payload.account_id,
            transfer_group_id=group_id,
            counterparty_name=description,
            counterparty_phone=None,
            description_raw=description,
            bank_source="manual",
            bank_operation_id=None,
            dedup_hash=None,
            target_id=payload.target_id,
            article_id=None,
            student_id=None,
            group_id=None,
            teacher_id=None,
            status=FinanceTransactionStatus.CLASSIFIED,
        )
        with db_transaction(db):
            db.add(tx_out)
            db.add(tx_in)
        db.refresh(tx_out)
        tx = tx_out  # Возвращаем расходную часть (исходящий счёт)
    else:
        tx = FinanceTransaction(
            occurred_at=occurred_at,
            amount=abs(amount),
            direction=FinanceTransactionDirection(payload.direction),  # type: ignore[call-arg]
            account_id=payload.account_id,
            to_account_id=None,
            transfer_group_id=None,
            counterparty_name=description,
            counterparty_phone=None,
            description_raw=description,
            bank_source="manual",
            bank_operation_id=None,
            dedup_hash=None,
            target_id=payload.target_id,
            article_id=payload.article_id,
            student_id=None,
            group_id=None,
            teacher_id=None,
            status=FinanceTransactionStatus.CLASSIFIED,
        )
        with db_transaction(db):
            db.add(tx)
        db.refresh(tx)

    tx = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.account),
            joinedload(FinanceTransaction.to_account),
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .filter(FinanceTransaction.id == tx.id)
        .first()
    )
    account = getattr(tx, "account", None)
    to_account = getattr(tx, "to_account", None)
    target = getattr(tx, "target", None)
    article = getattr(tx, "article", None)

    return FinanceLedgerBankRow(
        id=tx.id,
        occurred_at=tx.occurred_at,
        amount=float(tx.amount or 0.0),
        direction="transfer" if tx.transfer_group_id else str(getattr(tx.direction, "value", tx.direction)),
        status=str(getattr(tx.status, "value", tx.status)),
        account_id=tx.account_id,
        account_code=getattr(account, "code", None),
        account_name=getattr(account, "name", None),
        to_account_id=tx.to_account_id,
        to_account_code=getattr(to_account, "code", None),
        to_account_name=getattr(to_account, "name", None),
        transfer_group_id=tx.transfer_group_id,
        counterparty_name=tx.counterparty_name,
        counterparty_phone=tx.counterparty_phone,
        description=tx.description_raw,
        bank_source=tx.bank_source,
        bank_operation_id=tx.bank_operation_id,
        target_id=tx.target_id,
        target_code=getattr(target, "code", None),
        target_name=getattr(target, "name", None),
        article_id=tx.article_id,
        article_name=getattr(article, "name", None),
        student_id=tx.student_id,
    )


@router.get("/ledger/bank", response_model=List[FinanceLedgerBankRow])
async def list_finance_ledger_bank(
    status_filter: Optional[List[str]] = Query(
        None,
        description="Фильтр по статусу finance_transactions: new, classified, applied",
    ),
    unclassified_only: bool = Query(
        False,
        description="Только неразобранные операции (target_id IS NULL OR article_id IS NULL)",
    ),
    limit: int = Query(500, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceLedgerBankRow]:
    """
    Операции из единого финансового журнала, которые пришли из банков (bank_source не NULL),
    для использования на вкладке «Операции банка».
    """
    _require_finance_access(current_user)

    q = (
        db.query(FinanceTransaction)
        .outerjoin(FinanceAccount, FinanceAccount.id == FinanceTransaction.account_id)
        .outerjoin(FinanceAccount, FinanceAccount.id == FinanceTransaction.to_account_id)
    )
    # Чтобы не путать алиасы, сделаем отдельные выборки ниже через joinedload
    q = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.account),
            joinedload(FinanceTransaction.to_account),
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .filter(FinanceTransaction.bank_source.isnot(None))
        .order_by(FinanceTransaction.created_at.desc())
    )

    if status_filter:
        allowed = {s.value for s in FinanceTransactionStatus}  # type: ignore[attr-defined]
        invalid = [s for s in status_filter if s not in allowed]
        if invalid:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Некорректный status: {', '.join(invalid)}",
            )
        q = q.filter(FinanceTransaction.status.in_(status_filter))

    if unclassified_only:
        q = q.filter(
            (FinanceTransaction.target_id.is_(None))
            | (FinanceTransaction.article_id.is_(None))
        )

    items: List[FinanceTransaction] = q.limit(limit).all()

    rows: List[FinanceLedgerBankRow] = []
    for tx in items:
        account: Optional[FinanceAccount] = getattr(tx, "account", None)
        to_account: Optional[FinanceAccount] = getattr(tx, "to_account", None)
        target: Optional[FinanceTarget] = getattr(tx, "target", None)
        article: Optional[FinanceArticle] = getattr(tx, "article", None)

        rows.append(
            FinanceLedgerBankRow(
                id=tx.id,
                occurred_at=tx.occurred_at,
                amount=tx.amount,
                direction=str(getattr(tx.direction, "value", tx.direction)),
                status=str(getattr(tx.status, "value", tx.status)),
                account_id=tx.account_id,
                account_code=getattr(account, "code", None),
                account_name=getattr(account, "name", None),
                to_account_id=tx.to_account_id,
                to_account_code=getattr(to_account, "code", None),
                to_account_name=getattr(to_account, "name", None),
                transfer_group_id=tx.transfer_group_id,
                counterparty_name=tx.counterparty_name,
                counterparty_phone=tx.counterparty_phone,
                description=tx.description_raw,
                bank_source=tx.bank_source,
                bank_operation_id=tx.bank_operation_id,
                target_id=tx.target_id,
                target_code=getattr(target, "code", None),
                target_name=getattr(target, "name", None),
                article_id=tx.article_id,
                article_name=getattr(article, "name", None),
                student_id=tx.student_id,
            )
        )
    return rows


@router.get("/transactions", response_model=List[FinanceLedgerBankRow])
async def list_finance_transactions(
    account_ids: Optional[List[int]] = Query(
        None,
        description="Фильтр по счетам (finance_accounts.id). Пусто = все активные счета.",
    ),
    target_ids: Optional[List[int]] = Query(
        None,
        description="Фильтр по проектам (finance_targets.id). Пусто = все.",
    ),
    include_unassigned_targets: bool = Query(
        False,
        description="Вместе с target_ids показывать операции без проекта (target_id IS NULL).",
    ),
    article_ids: Optional[List[int]] = Query(
        None,
        description="Фильтр по статьям (finance_articles.id). Пусто = все.",
    ),
    direction: Optional[str] = Query(
        None,
        description="Фильтр по типу операции: income | expense | transfer",
    ),
    status: Optional[List[str]] = Query(
        None,
        description="Фильтр по статусу finance_transactions: new, classified, applied (если используется). Пусто = все.",
    ),
    unclassified_only: bool = Query(
        False,
        description="Только неразобранные операции (target_id IS NULL OR article_id IS NULL)",
    ),
    date_from: Optional[date] = Query(None, description="Начало периода (включительно)"),
    date_to: Optional[date] = Query(None, description="Конец периода (включительно)"),
    limit: int = Query(5000, ge=1, le=50000),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> List[FinanceLedgerBankRow]:
    """
    Каноничный список операций единого финансового журнала.

    Используется вкладкой «Все операции» в разделе Финансы → Журнал.
    Возвращает те же поля, что и /finance/ledger/bank, но без фильтра по bank_source
    (включает как банковские операции, так и ручные / импортированные записи).
    """
    _require_finance_access(current_user)

    q = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.account),
            joinedload(FinanceTransaction.to_account),
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .order_by(FinanceTransaction.occurred_at.desc(), FinanceTransaction.id.desc())
    )

    if account_ids:
        q = q.filter(FinanceTransaction.account_id.in_(account_ids))
    if target_ids:
        target_filter = FinanceTransaction.target_id.in_(target_ids)
        if include_unassigned_targets:
            target_filter = or_(target_filter, FinanceTransaction.target_id.is_(None))
        q = q.filter(target_filter)
    if article_ids:
        q = q.filter(FinanceTransaction.article_id.in_(article_ids))
    if direction:
        if direction not in {d.value for d in FinanceTransactionDirection}:  # type: ignore[attr-defined]
            raise HTTPException(status_code=400, detail="Некорректное значение direction")
        q = q.filter(FinanceTransaction.direction == FinanceTransactionDirection(direction))  # type: ignore[call-arg]
    if status:
        allowed = {s.value for s in FinanceTransactionStatus}  # type: ignore[attr-defined]
        invalid = [s for s in status if s not in allowed]
        if invalid:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Некорректный status: {', '.join(invalid)}",
            )
        q = q.filter(FinanceTransaction.status.in_(status))
    if unclassified_only:
        q = q.filter(
            (FinanceTransaction.target_id.is_(None))
            | (FinanceTransaction.article_id.is_(None))
        )
    if date_from is not None:
        q = q.filter(FinanceTransaction.occurred_at >= datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc))
    if date_to is not None:
        q = q.filter(FinanceTransaction.occurred_at <= datetime.combine(date_to, datetime.max.time(), tzinfo=timezone.utc))

    items: List[FinanceTransaction] = q.limit(limit).all()

    rows: List[FinanceLedgerBankRow] = []
    for tx in items:
        account: Optional[FinanceAccount] = getattr(tx, "account", None)
        to_account: Optional[FinanceAccount] = getattr(tx, "to_account", None)
        target: Optional[FinanceTarget] = getattr(tx, "target", None)
        article: Optional[FinanceArticle] = getattr(tx, "article", None)

        rows.append(
            FinanceLedgerBankRow(
                id=tx.id,
                occurred_at=tx.occurred_at,
                amount=float(tx.amount or 0.0),
                direction=str(getattr(tx.direction, "value", tx.direction)),
                status=str(getattr(tx.status, "value", tx.status)),
                account_id=tx.account_id,
                account_code=getattr(account, "code", None),
                account_name=getattr(account, "name", None),
                to_account_id=tx.to_account_id,
                to_account_code=getattr(to_account, "code", None),
                to_account_name=getattr(to_account, "name", None),
                transfer_group_id=tx.transfer_group_id,
                counterparty_name=tx.counterparty_name,
                counterparty_phone=tx.counterparty_phone,
                description=tx.description_raw,
                bank_source=tx.bank_source,
                bank_operation_id=tx.bank_operation_id,
                target_id=tx.target_id,
                target_code=getattr(target, "code", None),
                target_name=getattr(target, "name", None),
                article_id=tx.article_id,
                article_name=getattr(article, "name", None),
                student_id=tx.student_id,
            )
        )

    return rows


@router.patch("/transactions/{transaction_id}", response_model=FinanceLedgerBankRow)
async def update_finance_transaction(
    transaction_id: int,
    payload: FinanceTransactionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceLedgerBankRow:
    """Partial update for a finance journal transaction."""
    _require_finance_access(current_user)

    tx = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.account),
            joinedload(FinanceTransaction.to_account),
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .filter(FinanceTransaction.id == transaction_id)
        .first()
    )
    if not tx:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Transaction not found")

    fields_set = payload.model_fields_set

    if "occurred_at" in fields_set:
        tx.occurred_at = payload.occurred_at

    if "amount" in fields_set:
        amount = float(payload.amount or 0.0)
        if amount <= 0:
            raise HTTPException(status_code=400, detail="amount must be greater than 0")
        tx.amount = abs(amount)

    if payload.direction is not None:
        if payload.direction not in {d.value for d in FinanceTransactionDirection}:  # type: ignore[attr-defined]
            raise HTTPException(status_code=400, detail="Invalid direction")
        tx.direction = FinanceTransactionDirection(payload.direction)  # type: ignore[call-arg]
        if payload.direction != FinanceTransactionDirection.TRANSFER.value:
            tx.to_account_id = None
            tx.transfer_group_id = None

    if payload.status is not None:
        if payload.status not in {s.value for s in FinanceTransactionStatus}:  # type: ignore[attr-defined]
            raise HTTPException(status_code=400, detail="Invalid status")
        tx.status = FinanceTransactionStatus(payload.status)  # type: ignore[call-arg]

    if "account_id" in fields_set:
        if payload.account_id and not db.query(FinanceAccount.id).filter(FinanceAccount.id == payload.account_id).first():
            raise HTTPException(status_code=400, detail="account_id not found")
        tx.account_id = payload.account_id

    if "to_account_id" in fields_set:
        if payload.to_account_id and not db.query(FinanceAccount.id).filter(FinanceAccount.id == payload.to_account_id).first():
            raise HTTPException(status_code=400, detail="to_account_id not found")
        tx.to_account_id = payload.to_account_id

    if "target_id" in fields_set:
        if payload.target_id and not db.query(FinanceTarget.id).filter(FinanceTarget.id == payload.target_id).first():
            raise HTTPException(status_code=400, detail="target_id not found")
        tx.target_id = payload.target_id

    if "article_id" in fields_set:
        if payload.article_id and not db.query(FinanceArticle.id).filter(FinanceArticle.id == payload.article_id).first():
            raise HTTPException(status_code=400, detail="article_id not found")
        tx.article_id = payload.article_id

    if "transfer_group_id" in fields_set:
        tx.transfer_group_id = payload.transfer_group_id or None

    if "counterparty_name" in fields_set:
        tx.counterparty_name = payload.counterparty_name.strip() if payload.counterparty_name else None

    if "description" in fields_set:
        tx.description_raw = payload.description.strip() if payload.description else None

    with db_transaction(db):
        pass
    tx = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.account),
            joinedload(FinanceTransaction.to_account),
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .filter(FinanceTransaction.id == transaction_id)
        .first()
    )
    if not tx:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Transaction not found")

    account: Optional[FinanceAccount] = getattr(tx, "account", None)
    to_account: Optional[FinanceAccount] = getattr(tx, "to_account", None)
    target: Optional[FinanceTarget] = getattr(tx, "target", None)
    article: Optional[FinanceArticle] = getattr(tx, "article", None)

    return FinanceLedgerBankRow(
        id=tx.id,
        occurred_at=tx.occurred_at,
        amount=float(tx.amount or 0.0),
        direction="transfer" if tx.transfer_group_id else str(getattr(tx.direction, "value", tx.direction)),
        status=str(getattr(tx.status, "value", tx.status)),
        account_id=tx.account_id,
        account_code=getattr(account, "code", None),
        account_name=getattr(account, "name", None),
        to_account_id=tx.to_account_id,
        to_account_code=getattr(to_account, "code", None),
        to_account_name=getattr(to_account, "name", None),
        transfer_group_id=tx.transfer_group_id,
        counterparty_name=tx.counterparty_name,
        counterparty_phone=tx.counterparty_phone,
        description=tx.description_raw,
        bank_source=tx.bank_source,
        bank_operation_id=tx.bank_operation_id,
        target_id=tx.target_id,
        target_code=getattr(target, "code", None),
        target_name=getattr(target, "name", None),
        article_id=tx.article_id,
        article_name=getattr(article, "name", None),
        student_id=tx.student_id,
    )

@router.delete("/transactions/{transaction_id}")
async def delete_finance_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> Dict[str, bool]:
    """Удаление транзакции журнала (используется только в личных финансах)."""
    _require_finance_access(current_user)

    tx = db.query(FinanceTransaction).filter(FinanceTransaction.id == transaction_id).first()
    if not tx:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Транзакция не найдена")

    with db_transaction(db):
        if tx.bank_source and tx.bank_operation_id:
            bank_transaction = (
                db.query(BankTransaction)
                .filter(BankTransaction.operation_id == tx.bank_operation_id)
                .first()
            )
            if bank_transaction is not None:
                if bank_transaction.status == BankTransactionStatus.APPLIED.value:
                    raise HTTPException(
                        status_code=400,
                        detail="Зачисленную банковскую операцию нельзя скрыть без отмены зачисления",
                    )
                bank_transaction.status = BankTransactionStatus.IGNORED.value
                bank_transaction.student_id = None
                bank_transaction.student_account_id = None
        db.delete(tx)
    return {"ok": True}


@router.post("/transactions/{transaction_id}/apply-student", response_model=FinanceLedgerBankRow)
async def apply_finance_transaction_to_student(
    transaction_id: int,
    payload: FinanceTransactionApplyStudentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
) -> FinanceLedgerBankRow:
    """
    Зачислить операцию журнала на счёт ученика.

    Используется журналом (Финансы → Журнал), чтобы не идти в таб «Операции банка».
    Создаёт StudentAccountTransaction, обновляет баланс счёта и помечает транзакцию как applied.
    """
    _require_finance_access(current_user)

    tx = (
        db.query(FinanceTransaction)
        .options(
            joinedload(FinanceTransaction.account),
            joinedload(FinanceTransaction.to_account),
            joinedload(FinanceTransaction.target),
            joinedload(FinanceTransaction.article),
        )
        .filter(FinanceTransaction.id == transaction_id)
        .first()
    )
    if not tx:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Транзакция не найдена")

    if tx.direction != FinanceTransactionDirection.INCOME:
        raise HTTPException(status_code=400, detail="Зачислить ученику можно только доходную операцию")

    amount = float(tx.amount or 0.0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Сумма операции должна быть больше 0")

    try:
        pay_date = tx.occurred_at.date() if tx.occurred_at else date.today()
    except Exception:
        pay_date = date.today()
    note = tx.counterparty_name or tx.description_raw or "Платёж из журнала"
    try:
        from app.services.student_account_payment import add_payment_to_student_account
        add_payment_to_student_account(db, payload.student_id, amount, note, pay_date)
    except ValueError as e:
        if "не найден" in str(e).lower():
            raise HTTPException(status_code=404, detail=str(e))
        raise HTTPException(status_code=400, detail=str(e))

    with db_transaction(db):
        tx.status = FinanceTransactionStatus.APPLIED
        tx.student_id = payload.student_id
    db.refresh(tx)

    account_obj: Optional[FinanceAccount] = getattr(tx, "account", None)
    to_account_obj: Optional[FinanceAccount] = getattr(tx, "to_account", None)
    target_obj: Optional[FinanceTarget] = getattr(tx, "target", None)
    article_obj: Optional[FinanceArticle] = getattr(tx, "article", None)

    return FinanceLedgerBankRow(
        id=tx.id,
        occurred_at=tx.occurred_at,
        amount=float(tx.amount or 0.0),
        direction=str(getattr(tx.direction, "value", tx.direction)),
        status=str(getattr(tx.status, "value", tx.status)),
        account_id=tx.account_id,
        account_code=getattr(account_obj, "code", None),
        account_name=getattr(account_obj, "name", None),
        to_account_id=tx.to_account_id,
        to_account_code=getattr(to_account_obj, "code", None),
        to_account_name=getattr(to_account_obj, "name", None),
        transfer_group_id=tx.transfer_group_id,
        counterparty_name=tx.counterparty_name,
        counterparty_phone=tx.counterparty_phone,
        bank_source=tx.bank_source,
        bank_operation_id=tx.bank_operation_id,
        target_id=tx.target_id,
        target_code=getattr(target_obj, "code", None),
        target_name=getattr(target_obj, "name", None),
        article_id=tx.article_id,
        article_name=getattr(article_obj, "name", None),
        student_id=tx.student_id,
    )


# --- Счета учеников (StudentAccount): канонический API Finance по ТЗ ---


@router.post("/student-accounts", response_model=StudentAccountResponse, status_code=status.HTTP_201_CREATED)
async def create_student_account_finance(
    payload: FinanceStudentAccountCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(dep_require_finance_access),
) -> StudentAccountResponse:
    """Создать счёт ученику. Канонический API Finance (ТЗ этап 4)."""
    try:
        account = finance_create_student_account(db, payload.student_id, payload.name or "")
    except ValueError as e:
        msg = str(e)
        if "not found" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)
    with db_transaction(db):
        pass
    db.refresh(account)
    return StudentAccountResponse.model_validate(account)


# --- Банковские операции (BankTransaction): канонический API Finance по ТЗ ---


@router.post("/bank-transactions/{transaction_id}/apply", response_model=BankTransactionResponse)
async def apply_bank_transaction_to_student(
    transaction_id: int,
    payload: BankTransactionApplyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(dep_require_finance_access),
) -> BankTransactionResponse:
    """
    Зачислить банковскую операцию (BankTransaction) на выбранного ученика.
    Канонический API Finance; дублирует функциональность POST /api/sales/bank-transactions/{id}/apply (compatibility layer).
    """
    try:
        with db_transaction(db):
            result = bank_operation_apply(db, transaction_id, payload.student_id)
    except ValueError as e:
        msg = str(e)
        if "не найден" in msg.lower():
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=400, detail=msg)
    db.refresh(result.transaction)
    return BankTransactionResponse.model_validate(result.transaction)


@router.post("/bank-transactions/backfill-ledger")
async def backfill_bank_transactions_to_ledger(
    db: Session = Depends(get_db),
    current_user: User = Depends(dep_require_finance_access),
) -> dict:
    """
    Однократный бэкфилл: создаёт/обновляет записи в finance_transactions
    для всех существующих bank_transactions. Безопасно запускать повторно —
    ensure_finance_transaction_for_bank_transaction идемпотентна.
    """
    from app.services.finance_ledger import ensure_finance_transaction_for_bank_transaction

    all_bank_txs: list[BankTransaction] = (
        db.query(BankTransaction)
        .filter(BankTransaction.status != BankTransactionStatus.IGNORED.value)
        .all()
    )
    created = 0
    updated = 0

    for bt in all_bank_txs:
        bank_source = "tochka" if bt.tochka_account_id else "import_xlsx"
        existing_ft = (
            db.query(FinanceTransaction)
            .filter(
                FinanceTransaction.bank_source == bank_source,
                FinanceTransaction.bank_operation_id == bt.operation_id,
            )
            .first()
        )
        ensure_finance_transaction_for_bank_transaction(db, bt, bank_source=bank_source)
        if existing_ft:
            updated += 1
        else:
            created += 1

    db.commit()
    return {"ok": True, "created": created, "updated": updated, "total": len(all_bank_txs)}
