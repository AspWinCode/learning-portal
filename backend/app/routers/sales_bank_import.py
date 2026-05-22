import hashlib
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import auth
from app.database import get_db
from app.models import BankTransaction, BankTransactionStatus, User
from app.services.finance_ledger import ensure_finance_transaction_for_bank_transaction
from app.utils.phone import normalize_phone

router = APIRouter()


def _require_sales_admin_owner(user: User) -> None:
    auth.ensure_permission(user, "sales.access")


def _parse_vertical_date(text: str, today: date) -> Optional[str]:
    if not text or not str(text).strip():
        return None
    text = str(text).strip()
    year = today.year
    months_ru = {
        "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5, "июня": 6,
        "июля": 7, "августа": 8, "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
    }
    if "сегодня" in text.lower():
        return today.isoformat()
    if "вчера" in text.lower():
        return (today - timedelta(days=1)).isoformat()
    for month_name, month_num in months_ru.items():
        if month_name in text.lower():
            parts = re.findall(r"\d+", text)
            if parts:
                day = int(parts[0])
                try:
                    return date(year, month_num, day).isoformat()
                except ValueError:
                    return None
            return None
    return None


def _import_bank_transactions_vertical(rows: list, db: Session) -> dict:
    lines = []
    for row in rows:
        value = row[0] if row and len(row) > 0 else None
        lines.append(str(value).strip() if value is not None else "")
    today = date.today()
    amount_re = re.compile(r"^[+\-–]\s*([\d\s,]+)\s*[₽р]", re.IGNORECASE)
    imported = 0
    skipped = 0
    last_date_str = None
    for index, line in enumerate(lines):
        if not line:
            continue
        line_norm = line.replace("\xa0", " ")
        if "₽" not in line_norm and " р" not in line_norm.lower():
            maybe_date = _parse_vertical_date(line, today)
            if maybe_date:
                last_date_str = maybe_date
            continue
        match = amount_re.match(line_norm)
        if not match:
            continue
        amount_str = match.group(1).replace(" ", "").replace(",", ".")
        try:
            amount_val = float(amount_str)
        except ValueError:
            skipped += 1
            continue
        if amount_val <= 0:
            amount_val = -abs(amount_val)
        else:
            amount_val = abs(amount_val)
        if line.strip().startswith(("-", "–")):
            amount_val = -abs(amount_val)
        date_str = last_date_str
        counterparty = ""
        if index + 1 < len(lines):
            counterparty = (lines[index + 2] or "").strip() if index + 2 < len(lines) else ""
            time_line = (lines[index + 4] or "").strip() if index + 4 < len(lines) else ""
            if time_line:
                parsed = _parse_vertical_date(time_line, today)
                if parsed:
                    date_str = parsed
        if not date_str:
            date_str = today.isoformat()
        payer_name = counterparty or "Списание" if amount_val < 0 else "Из выписки (без ФИО)"
        payer_phone = None
        if amount_val > 0 and counterparty:
            phone_match = re.search(r"\+7\s*\(?\d{3}\)?\s*\d{3}[- ]?\d{2}[- ]?\d{2}", counterparty)
            if phone_match:
                payer_phone = normalize_phone(re.sub(r"\D", "", phone_match.group(0)))
            if "," in counterparty:
                rest = counterparty.split(",", 1)[1].strip()
                if rest and len(rest) >= 2:
                    payer_name = rest[:512]
        if amount_val > 0 and not payer_name:
            payer_name = counterparty or "Из выписки (без ФИО)"
        operation_id_source = f"vertical_xlsx|{date_str}|{amount_val}|{payer_name}|{payer_phone or ''}|{index}"
        operation_id = hashlib.sha256(operation_id_source.encode("utf-8")).hexdigest()
        if db.query(BankTransaction.id).filter(BankTransaction.operation_id == operation_id).first():
            skipped += 1
            continue
        status = BankTransactionStatus.EXPENSE.value if amount_val < 0 else BankTransactionStatus.NEW.value
        bank_transaction = BankTransaction(
            operation_id=operation_id,
            tochka_account_id=None,
            amount=amount_val,
            payer_phone=payer_phone,
            payer_name=(payer_name or "")[:512] or None,
            payment_date=date_str,
            status=status,
            expense_category=None,
        )
        db.add(bank_transaction)
        ensure_finance_transaction_for_bank_transaction(db, bank_transaction, bank_source="import_xlsx")
        imported += 1
    db.commit()
    errors = []
    if imported == 0 and skipped == 0 and len(lines) > 1:
        errors.append("В файле не найдено строк с суммой вида «+ 3 200 ₽» или «– 102 ₽».")
    return {"imported": imported, "skipped": skipped, "errors": errors}


def _parse_date_any(raw_value) -> Optional[str]:
    if raw_value is None:
        return None
    if isinstance(raw_value, datetime):
        try:
            value = raw_value.date() if hasattr(raw_value, "date") else raw_value
            return value.isoformat()
        except Exception:
            pass
    if isinstance(raw_value, (int, float)):
        try:
            base = date(1899, 12, 30)
            value = base + timedelta(days=int(float(raw_value)))
            return value.isoformat()
        except (ValueError, TypeError, OverflowError):
            pass
    text = str(raw_value).strip()
    try:
        serial = float(text)
        if 1000 < serial < 100000:
            base = date(1899, 12, 30)
            value = base + timedelta(days=int(serial))
            return value.isoformat()
    except (ValueError, TypeError):
        pass
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text[:10]).date().isoformat()
    except (ValueError, TypeError):
        return text


def _parse_amount(raw_value) -> Optional[float]:
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    text = str(raw_value).strip().replace(" ", "").replace("\u00a0", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_payment_purpose(purpose: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not purpose or not str(purpose).strip():
        return None, None
    text = str(purpose).strip()
    phone_raw = None
    phone_end = 0
    for pattern in (
        r"\+7\s*\(?\d{3}\)?\s*\d{3}[- ]?\d{2}[- ]?\d{2}",
        r"8\s*\(?\d{3}\)?\s*\d{3}[- ]?\d{2}[- ]?\d{2}",
        r"\+7\d{10}",
        r"8\d{10}",
    ):
        match = re.search(pattern, text)
        if match:
            phone_raw = re.sub(r"\D", "", match.group(0))
            if phone_raw.startswith("8") and len(phone_raw) == 11:
                phone_raw = "7" + phone_raw[1:]
            elif phone_raw.startswith("7") and len(phone_raw) == 11:
                pass
            elif len(phone_raw) == 10:
                phone_raw = "7" + phone_raw
            else:
                phone_raw = None
            if phone_raw:
                phone_end = match.end()
                break
    name = None
    for prefix in ("Получатель ", "Плательщик "):
        if prefix in text:
            start = text.find(prefix) + len(prefix)
            end = text.find(" через", start)
            if end == -1:
                end = text.find(".", start)
            if end == -1:
                end = len(text)
            name = text[start:end].strip()
            if name:
                break
    if not name and phone_end > 0:
        rest = text[phone_end:].strip()
        for stop in ("Заказ ", "Order ", " заказ ", " №", " N "):
            if stop in rest:
                idx = rest.find(stop)
                candidate = rest[:idx].strip()
                if candidate and len(candidate) >= 3 and re.search(r"[\u0400-\u04FF]", candidate):
                    name = candidate
                    break
        if not name and rest and re.search(r"[\u0400-\u04FF]", rest):
            name = rest[:80].strip()
    return name or None, phone_raw


@router.post("/bank-transactions/import-xlsx")
async def import_bank_transactions_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    _require_sales_admin_owner(current_user)
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Поддерживается только формат .xlsx")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")
    workbook = load_workbook(filename=BytesIO(data), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": ["Пустой лист"]}
    first_row = rows[0]
    is_vertical = (
        len(first_row) == 1
        and first_row[0] is not None
        and not any(
            key in str(first_row[0]).strip().lower()
            for key in ("дата", "сумма", "зачисление", "списание", "кредит", "дебет", "назначение")
        )
    )
    if is_vertical:
        return _import_bank_transactions_vertical(rows, db)

    headers = [str(header).strip().lower() if header is not None else "" for header in rows[0]]
    header_map = {name: index for index, name in enumerate(headers)}

    def col(row, keys: List[str], allow_number: bool = True):
        for key in keys:
            for name, index in header_map.items():
                if key in name and index < len(row):
                    raw = row[index] if index < len(row) else None
                    if raw is None:
                        continue
                    if allow_number and isinstance(raw, (int, float)):
                        return str(raw).strip() if raw != 0 else None
                    text = str(raw).strip()
                    if text:
                        return text
        return None

    imported = 0
    skipped = 0
    for row_index, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        date_str = _parse_date_any(col(row, ["дата", "date"]))
        if not date_str:
            for name, col_index in header_map.items():
                if "дата" in name and col_index < len(row) and row[col_index] is not None:
                    date_str = _parse_date_any(row[col_index])
                    if date_str:
                        break
        amount_raw = col(row, ["сумма", "amount", "зачисление", "кредит"])
        amount = _parse_amount(amount_raw)
        if amount is None or amount <= 0:
            amount = _parse_amount(col(row, ["списание", "дебет"]))
            if amount is not None and amount > 0:
                amount = -amount
        payer_name = col(row, ["фио", "плательщик", "payer"])
        payer_phone_raw = col(row, ["телефон", "phone"])
        purpose = col(row, ["назначение", "payment purpose", "назначение платежа"])
        if (not payer_name or not payer_phone_raw) and purpose:
            name_from_purpose, phone_from_purpose = _parse_payment_purpose(purpose)
            if not payer_name and name_from_purpose:
                payer_name = name_from_purpose
            if not payer_phone_raw and phone_from_purpose:
                payer_phone_raw = phone_from_purpose
        if not payer_name and amount and amount > 0 and date_str:
            payer_name = "Из выписки (без ФИО)"
        if amount is None or not date_str:
            skipped += 1
            continue
        is_expense = amount < 0
        if is_expense:
            payer_name = payer_name or col(row, ["контрагент", "counterparty"]) or (purpose[:512] if purpose else "Списание")
        else:
            if not payer_name:
                skipped += 1
                continue

        payer_phone = (normalize_phone(payer_phone_raw or "") or None) if not is_expense else None
        operation_id_source = f"manual_xlsx|{date_str}|{amount}|{payer_name}|{payer_phone}|{row_index}"
        operation_id = hashlib.sha256(operation_id_source.encode("utf-8")).hexdigest()

        exists = db.query(BankTransaction.id).filter(BankTransaction.operation_id == operation_id).first()
        if exists:
            skipped += 1
            continue

        bank_transaction = BankTransaction(
            operation_id=operation_id,
            tochka_account_id=None,
            amount=amount,
            payer_phone=(payer_phone or None) if not is_expense else None,
            payer_name=(payer_name or "")[:512] or None,
            payment_date=date_str,
            status=BankTransactionStatus.EXPENSE.value if is_expense else BankTransactionStatus.NEW.value,
            expense_category=None,
        )
        db.add(bank_transaction)
        imported += 1

    db.commit()
    errors: List[str] = []
    if imported == 0 and skipped > 0:
        errors.append(
            "Ни одна строка не подошла. Проверьте: в первой строке — заголовки (Дата, Зачисление/Списание или Кредит/Дебет, Назначение/Контрагент); даты в формате ДД.ММ.ГГГГ или число Excel; суммы — числа с запятой или точкой."
        )
    return {"imported": imported, "skipped": skipped, "errors": errors}
