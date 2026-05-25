from datetime import date, datetime, time as dt_time
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import Text, cast, or_, update as sa_update
from sqlalchemy.orm import Session, joinedload, selectinload

from app import auth
from app.database import get_db
from app.models import (
    Abonement,
    DiscountType,
    EventRegistration,
    EventRegistrationStatus,
    Invoice,
    InvoiceStatus,
    Lead,
    LeadActivity,
    LeadCommunication,
    LeadInfoTemplate,
    LeadSource,
    LeadStatus,
    LeadStatusOption,
    LeadTask,
    LeadTaskStatus,
    LeadTaskTemplate,
    StudentCard,
    Task,
    TaskStatus,
    User,
    UserRole,
)
from app.routers.action_log import log_action
from app.schemas.sales import (
    LeadCreate,
    LeadConvertToStudentResponse,
    LeadImportResponse,
    LeadPushStatsResponse,
    LeadResponse,
    LeadUpdate,
    SpecialistQuestionnaireRequest,
    SpecialistQuestionnaireResponse,
    TildaLeadRequest,
    TildaLeadResponse,
)
from app.services.ai_insights import build_lead_ai_insight
from app.services.lead_conversion import convert_lead_to_student as lead_conversion_convert
from app.services.person_sync import sync_lead_person, sync_student_card_person
from app.utils.phone import normalize_phone, validate_phone_for_lead

router = APIRouter()

TILDA_SOURCE_START = "Тильда_Первый Шаг"
TILDA_SOURCE_BASE = "Тильда_Специалист"
TILDA_SOURCE_PRO = "Тильда_Эксперт"
_SEND_INFO_TASK_MARKER = "отправить информацию"


def _lead_eager_options():
    return (
        joinedload(Lead.owner),
        joinedload(Lead.abonement),
        joinedload(Lead.status_option),
        joinedload(Lead.source_ref),
        selectinload(Lead.tasks).joinedload(LeadTask.owner),
        selectinload(Lead.tasks).joinedload(LeadTask.template),
        selectinload(Lead.communications).joinedload(LeadCommunication.sender),
        selectinload(Lead.communications).joinedload(LeadCommunication.template),
        selectinload(Lead.invoices),
    )


def _fix_mojibake(value: Optional[str]) -> Optional[str]:
    if not value or not isinstance(value, str):
        return value
    for encoding in ("latin1", "cp1252"):
        try:
            fixed = value.encode(encoding).decode("utf-8")
            if fixed != value:
                return fixed
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
    return value


def _fix_lead_strings(lead: Lead) -> Lead:
    text_fields = [
        "contact_name",
        "phone",
        "parent_full_name",
        "child_full_name",
        "parent_phone",
        "child_phone",
        "email",
        "city",
        "school_name",
        "school_class",
        "source",
        "referral_name",
        "comment",
        "pause_reason",
        "communication_channel",
        "lost_reason",
    ]
    for field_name in text_fields:
        if hasattr(lead, field_name):
            setattr(lead, field_name, _fix_mojibake(getattr(lead, field_name)))
    return lead


def _filter_query_by_role(query, user: User):
    if auth.resolve_effective_role(user) in (UserRole.ADMIN, UserRole.OWNER, UserRole.SALES):
        return query
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")


def _normalize_source_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    return " ".join(name.strip().split())


def _is_referral_source(name: Optional[str]) -> bool:
    return (name or "").strip().lower() == "рекомендация"


def _resolve_source(
    db: Session,
    source_id: Optional[int] = None,
    source_name: Optional[str] = None,
) -> Tuple[Optional[int], Optional[str]]:
    normalized = _normalize_source_name(source_name)
    if source_id:
        source_obj = db.query(LeadSource).filter(LeadSource.id == source_id).first()
        if not source_obj:
            raise HTTPException(status_code=404, detail="Lead source not found")
        return source_obj.id, source_obj.name
    if normalized:
        source_obj = db.query(LeadSource).filter(cast(LeadSource.name, Text).ilike(normalized)).first()
        if source_obj:
            return source_obj.id, source_obj.name
        return None, normalized
    return None, None


def _require_owner_or_admin(lead: Lead, user: User) -> None:
    if auth.resolve_effective_role(user) in (UserRole.ADMIN, UserRole.OWNER, UserRole.SALES):
        return
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")


def _get_default_lead_status_option_id(db: Session, base_status: LeadStatus) -> Optional[int]:
    status_str = base_status.value if hasattr(base_status, "value") else str(base_status)
    option = (
        db.query(LeadStatusOption)
        .filter(
            LeadStatusOption.base_status == status_str,
            LeadStatusOption.is_active.is_(True),
        )
        .order_by(LeadStatusOption.id.asc())
        .first()
    )
    return option.id if option else None


def _add_activity(
    db: Session,
    lead_id: int,
    actor_id: int,
    type: str,
    title: str,
    description: Optional[str] = None,
    channel: Optional[str] = None,
    status_effect_from: Optional[str] = None,
    status_effect_to: Optional[str] = None,
    related_task_id: Optional[int] = None,
    related_invoice_id: Optional[int] = None,
    payload_json: Optional[dict] = None,
) -> LeadActivity:
    activity = LeadActivity(
        lead_id=lead_id,
        type=type,
        title=title,
        description=description,
        channel=channel,
        created_by=actor_id,
        status_effect_from=status_effect_from,
        status_effect_to=status_effect_to,
        related_task_id=related_task_id,
        related_invoice_id=related_invoice_id,
        payload_json=payload_json,
    )
    db.add(activity)
    return activity


@router.get("/leads", response_model=List[LeadResponse])
async def list_leads(
    status_filter: Optional[LeadStatus] = None,
    questionnaire_filled: Optional[bool] = None,
    q: Optional[str] = None,
    source: Optional[str] = None,
    tag: Optional[str] = None,
    overdue_only: bool = False,
    created_from: Optional[datetime] = Query(default=None),
    created_to: Optional[datetime] = Query(default=None),
    next_contact_from: Optional[datetime] = Query(default=None),
    next_contact_to: Optional[datetime] = Query(default=None),
    limit: int = Query(default=500, ge=1, le=2000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    auth.ensure_permission(current_user, "sales.access")
    query = _filter_query_by_role(
        db.query(Lead).options(*_lead_eager_options()).order_by(Lead.created_at.desc()),
        current_user,
    )
    if status_filter:
        query = query.filter(Lead.status == status_filter)
    if questionnaire_filled is not None:
        query = query.filter(Lead.questionnaire_filled == questionnaire_filled)
    if q:
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                cast(Lead.contact_name, Text).ilike(search),
                cast(Lead.phone, Text).ilike(search),
                cast(Lead.parent_full_name, Text).ilike(search),
                cast(Lead.child_full_name, Text).ilike(search),
                cast(Lead.parent_phone, Text).ilike(search),
                cast(Lead.child_phone, Text).ilike(search),
                cast(Lead.school_name, Text).ilike(search),
                cast(Lead.school_class, Text).ilike(search),
            )
        )
    if source:
        query = query.filter(Lead.source.ilike(f"%{source.strip()}%"))
    if tag:
        query = query.filter(Lead.tags.isnot(None), cast(Lead.tags, Text).ilike(f"%{tag.strip()}%"))
    if overdue_only:
        now = datetime.utcnow()
        query = query.filter(Lead.next_contact_at.isnot(None), Lead.next_contact_at < now)
    if created_from:
        query = query.filter(Lead.created_at >= created_from)
    if created_to:
        query = query.filter(Lead.created_at <= created_to)
    if next_contact_from:
        query = query.filter(Lead.next_contact_at >= next_contact_from)
    if next_contact_to:
        query = query.filter(Lead.next_contact_at <= next_contact_to)
    leads = query.offset(offset).limit(limit).all()

    if status_filter == LeadStatus.DEMO:
        candidate_ids = [lead.id for lead in leads if not getattr(lead, "post_visit_stage", None)]
        if candidate_ids:
            came_lead_ids = {
                row[0]
                for row in db.query(EventRegistration.lead_id)
                .filter(
                    EventRegistration.lead_id.in_(candidate_ids),
                    cast(EventRegistration.note, Text).ilike("%[came]%"),
                )
                .distinct()
                .all()
            }
            if came_lead_ids:
                db.execute(
                    sa_update(Lead)
                    .where(Lead.id.in_(came_lead_ids))
                    .values(post_visit_stage="new")
                    .execution_options(synchronize_session=False)
                )
                db.commit()
                refreshed = {
                    refreshed_lead.id: refreshed_lead
                    for refreshed_lead in db.query(Lead).options(*_lead_eager_options()).filter(Lead.id.in_(came_lead_ids)).all()
                }
                leads = [refreshed.get(lead.id, lead) for lead in leads]

    result: List[Lead] = []
    for lead in leads:
        fixed = _fix_lead_strings(lead)
        fixed.ai_insight = build_lead_ai_insight(fixed)
        result.append(fixed)
    return result


@router.post("/leads/import-xlsx", response_model=LeadImportResponse)
async def import_leads_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Поддерживается только формат .xlsx")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")

    workbook = load_workbook(filename=BytesIO(data), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return LeadImportResponse(created=0, skipped=0, errors=["Пустой лист"])

    headers = [str(header).strip().lower() if header is not None else "" for header in rows[0]]
    header_map = {name: index for index, name in enumerate(headers)}

    def val(row, key_variants: List[str]) -> Optional[str]:
        for key in key_variants:
            index = header_map.get(key)
            if index is None or index >= len(row):
                continue
            raw = row[index]
            if raw is None:
                continue
            text = str(raw).strip()
            if text:
                return text
        return None

    def parse_row_datetime(raw_value) -> Optional[datetime]:
        if raw_value is None:
            return None
        if isinstance(raw_value, datetime):
            return raw_value
        text = str(raw_value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None

    def parse_row_minutes(raw_value) -> Optional[int]:
        if raw_value is None:
            return None
        text = str(raw_value).strip()
        if not text:
            return None
        try:
            minutes = int(float(text.replace(",", ".")))
            return minutes if minutes >= 0 else None
        except ValueError:
            return None

    created = 0
    skipped = 0
    errors: List[str] = []
    for row_index, row in enumerate(rows[1:], start=2):
        parent_name = val(row, ["фио родителя", "родитель", "parent_full_name"])
        child_name = val(row, ["фио ребенка", "фио ребёнка", "ребенок", "ребёнок", "child_full_name"])
        parent_phone = val(row, ["телефон родителя", "parent_phone"])
        child_phone = val(row, ["телефон школьника", "телефон ребенка", "телефон ребёнка", "child_phone"])
        source_name_raw = val(row, ["источник", "source"])
        referral_name = val(row, ["кто пригласил", "рекомендовал", "referral_name"])
        comment = val(row, ["комментарий", "comment"])
        school_name = val(row, ["школа", "school", "school_name"])
        school_class = val(row, ["класс", "class", "school_class"])
        outreach_date_raw = val(row, ["дата обхода", "outreach_date", "outreach_at"])
        outreach_minutes_raw = val(row, ["время обхода (мин)", "время обхода", "outreach_minutes"])
        outreach_at = parse_row_datetime(outreach_date_raw)
        outreach_minutes = parse_row_minutes(outreach_minutes_raw)

        if not any([parent_name, child_name, parent_phone, child_phone, source_name_raw, comment]):
            skipped += 1
            continue

        source_id, source_name = _resolve_source(db, None, source_name_raw)
        if _is_referral_source(source_name) and not (referral_name or "").strip():
            errors.append(f"Строка {row_index}: для источника 'рекомендация' не указан пригласивший")
            skipped += 1
            continue

        contact_name = parent_name or child_name or "Без имени"
        phone = parent_phone or child_phone or "не указан"
        lead = Lead(
            owner_id=current_user.id,
            contact_name=contact_name,
            phone=phone,
            phone_normalized=normalize_phone(parent_phone or child_phone or phone) or None,
            parent_full_name=parent_name,
            child_full_name=child_name,
            parent_phone=parent_phone,
            child_phone=child_phone,
            source=source_name,
            source_id=source_id,
            referral_name=referral_name,
            comment=comment,
            school_name=school_name,
            school_class=school_class,
            outreach_at=outreach_at,
            outreach_minutes=outreach_minutes,
            status=LeadStatus.NEW,
        )
        db.add(lead)
        db.flush()
        sync_lead_person(db, lead)
        created += 1

    db.commit()
    log_action(db, current_user.id, "import", "lead", None, {"created": created, "skipped": skipped})
    return LeadImportResponse(created=created, skipped=skipped, errors=errors)


@router.get("/leads/import-template")
async def download_leads_import_template(
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "LeadsImport"
    headers = [
        "ФИО родителя",
        "ФИО ребенка",
        "Телефон родителя",
        "Телефон школьника",
        "Школа",
        "Класс",
        "Дата обхода",
        "Время обхода (мин)",
        "Источник",
        "Кто пригласил",
        "Комментарий",
    ]
    worksheet.append(headers)
    worksheet.append(
        [
            "Иванова Анна Петровна",
            "Иванов Петр",
            "+7 999 111-22-33",
            "+7 900 111-22-44",
            "Школа №12",
            "7А",
            "2026-02-01",
            "35",
            "рекомендация",
            "Мария Сидорова",
            "Интерес к занятиям после пробного урока",
        ]
    )
    worksheet.freeze_panes = "A2"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = "leads_import_template.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/public/leads/specialist-questionnaire", response_model=SpecialistQuestionnaireResponse, status_code=status.HTTP_201_CREATED)
async def submit_specialist_questionnaire(
    payload: SpecialistQuestionnaireRequest,
    db: Session = Depends(get_db),
):
    owner = (
        db.query(User)
        .filter(User.role.in_([UserRole.SALES, UserRole.OWNER, UserRole.ADMIN]))
        .order_by(User.id)
        .first()
    )
    if not owner:
        raise HTTPException(status_code=500, detail="No sales/owner/admin user configured")

    card = StudentCard(
        student_full_name=payload.child_full_name,
        birth_date=payload.birth_date,
        student_phone=payload.child_phone,
        phone_normalized=normalize_phone(payload.parent_phone or payload.child_phone or "") or None,
        telegram=payload.child_telegram,
        gender=payload.gender,
        on_grant=False,
        format_type=None,
        city=payload.city,
        school=payload.school_name,
        grade=payload.school_class,
        parent_full_name=payload.parent_full_name,
        parent_phone=payload.parent_phone,
        parent_phone_2=payload.parent_phone_2,
        parent_telegram=payload.parent_telegram,
        parent_email=payload.parent_email,
        student_email=payload.student_email,
        preferred_messenger=payload.preferred_messenger,
        comment=payload.comment,
        source=payload.source or "Анкета Специалист",
        discount_type=DiscountType.NONE,
        discount_value=0.0,
        anketa_status="filled",
    )

    extra_parts: List[str] = []
    if payload.birth_date:
        extra_parts.append(f"Дата рождения: {payload.birth_date.isoformat()}")
    if payload.child_phone:
        extra_parts.append(f"Телефон ученика: {payload.child_phone}")
    if payload.child_telegram:
        extra_parts.append(f"Телеграм ученика: {payload.child_telegram}")
    if payload.gender:
        extra_parts.append(f"Пол: {payload.gender}")
    if payload.parent_phone_2:
        extra_parts.append(f"Второй телефон родителя: {payload.parent_phone_2}")
    if payload.parent_telegram:
        extra_parts.append(f"Телеграм родителя: {payload.parent_telegram}")
    if payload.student_email:
        extra_parts.append(f"Email ученика: {payload.student_email}")
    if payload.preferred_messenger:
        extra_parts.append(f"Мессенджер: {payload.preferred_messenger}")

    base_comment = payload.comment or ""
    extras_str = "\n".join(extra_parts) if extra_parts else ""
    full_comment = base_comment
    if extras_str:
        full_comment = (base_comment + "\n\n" if base_comment else "") + extras_str

    questionnaire_data = payload.model_dump(mode="json")
    lead = Lead(
        owner_id=owner.id,
        contact_name=payload.parent_full_name,
        phone=payload.parent_phone,
        phone_normalized=normalize_phone(payload.parent_phone or payload.child_phone or "") or None,
        parent_full_name=payload.parent_full_name,
        child_full_name=payload.child_full_name,
        parent_phone=payload.parent_phone,
        child_phone=payload.child_phone,
        email=payload.parent_email or payload.student_email,
        city=payload.city,
        school_name=payload.school_name,
        school_class=payload.school_class,
        comment=full_comment or None,
        source=payload.source or "Анкета Специалист",
        tags=["direction:specialist"],
        status=LeadStatus.NEW,
        questionnaire_filled=True,
        questionnaire_data=questionnaire_data,
    )
    db.add(card)
    db.add(lead)
    db.flush()
    sync_student_card_person(db, card)
    lead.student_card_id = card.id
    sync_lead_person(db, lead)
    db.commit()
    db.refresh(lead)
    return SpecialistQuestionnaireResponse(lead_id=lead.id)


@router.post("/public/leads/tilda-lead", response_model=TildaLeadResponse, status_code=status.HTTP_201_CREATED)
async def submit_tilda_lead(
    payload: TildaLeadRequest,
    db: Session = Depends(get_db),
):
    parent_name = (payload.parent_full_name or "").strip()
    child_name = (payload.child_full_name or "").strip()
    if not parent_name:
        raise HTTPException(status_code=400, detail="Укажите ФИО родителя")
    if not child_name:
        raise HTTPException(status_code=400, detail="Укажите ФИО ученика")

    normalized_phone, phone_error = validate_phone_for_lead(payload.parent_phone)
    if phone_error:
        raise HTTPException(status_code=400, detail=phone_error)

    owner = (
        db.query(User)
        .filter(User.role.in_([UserRole.SALES, UserRole.OWNER, UserRole.ADMIN]))
        .order_by(User.id)
        .first()
    )
    if not owner:
        raise HTTPException(status_code=500, detail="В системе не настроен пользователь для приёма заявок")

    kind = (payload.kind or "start").strip()
    if kind == "base":
        source_label = TILDA_SOURCE_BASE
        tag = "tilda_base_lead"
    elif kind == "pro":
        source_label = TILDA_SOURCE_PRO
        tag = "tilda_pro_lead"
    else:
        source_label = TILDA_SOURCE_START
        tag = "tilda_start_lead"

    source_id, source_name = _resolve_source(db, None, source_label)
    status_option_id = _get_default_lead_status_option_id(db, LeadStatus.NEW)

    lead = Lead(
        owner_id=owner.id,
        contact_name=parent_name,
        phone=normalized_phone,
        phone_normalized=normalized_phone,
        parent_full_name=parent_name,
        parent_phone=normalized_phone,
        child_full_name=child_name,
        source=source_name or source_label,
        source_id=source_id,
        status=LeadStatus.NEW,
        status_option_id=status_option_id,
        tags=[tag],
    )
    db.add(lead)
    db.flush()
    sync_lead_person(db, lead)
    db.commit()
    db.refresh(lead)
    return TildaLeadResponse(lead_id=lead.id)


@router.post("/leads", response_model=LeadResponse, status_code=status.HTTP_201_CREATED)
async def create_lead(
    payload: LeadCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    effective_role = auth.resolve_effective_role(current_user)
    owner_id = payload.owner_id if (effective_role in (UserRole.ADMIN, UserRole.OWNER) and payload.owner_id) else current_user.id
    source_id, source_name = _resolve_source(db, payload.source_id, payload.source)
    if _is_referral_source(source_name) and not (payload.referral_name or "").strip():
        raise HTTPException(status_code=400, detail="Для источника 'рекомендация' укажите, кто пригласил")

    if payload.abonement_id:
        abonement = db.query(Abonement).filter(Abonement.id == payload.abonement_id).first()
        if not abonement:
            raise HTTPException(status_code=404, detail="Abonement not found")

    lead = Lead(
        owner_id=owner_id,
        contact_name=payload.contact_name,
        phone=payload.phone,
        phone_normalized=normalize_phone(payload.parent_phone or payload.phone or payload.child_phone or "") or None,
        parent_full_name=payload.parent_full_name,
        child_full_name=payload.child_full_name,
        parent_phone=payload.parent_phone,
        child_phone=payload.child_phone,
        email=payload.email,
        city=payload.city,
        school_name=payload.school_name,
        school_class=payload.school_class,
        outreach_at=payload.outreach_at,
        outreach_minutes=payload.outreach_minutes,
        source=source_name,
        source_id=source_id,
        referral_name=payload.referral_name,
        tags=payload.tags,
        abonement_id=payload.abonement_id,
        desired_slot=payload.desired_slot,
        comment=payload.comment,
        next_contact_at=payload.next_contact_at,
        status=LeadStatus.NEW,
    )
    db.add(lead)
    db.flush()
    sync_lead_person(db, lead)
    _add_activity(
        db,
        lead.id,
        current_user.id,
        type="lead_created",
        title="Лид создан",
        description=f"Источник: {source_name or '—'}",
    )
    db.commit()
    db.refresh(lead)

    log_action(db, current_user.id, "create", "lead", lead.id, {"owner_id": owner_id})
    return lead


@router.get("/leads/send-info-status")
async def get_leads_send_info_status(
    lead_ids: str = Query(..., description="Comma-separated lead IDs"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    if not lead_ids.strip():
        return {}
    ids = [int(item.strip()) for item in lead_ids.split(",") if item.strip()]
    if not ids:
        return {}
    base = _filter_query_by_role(db.query(Lead), current_user)
    allowed_ids = {lead.id for lead in base.filter(Lead.id.in_(ids)).all()}
    tasks = (
        db.query(LeadTask)
        .options(joinedload(LeadTask.template))
        .outerjoin(LeadTaskTemplate, LeadTask.template_id == LeadTaskTemplate.id)
        .filter(
            LeadTask.lead_id.in_(allowed_ids),
            or_(
                cast(LeadTask.note, Text).ilike(f"%{_SEND_INFO_TASK_MARKER}%"),
                cast(LeadTaskTemplate.name, Text).ilike(f"%{_SEND_INFO_TASK_MARKER}%"),
            ),
        )
        .all()
    )
    result: Dict[str, str] = {str(item): "none" for item in ids if item in allowed_ids}
    for task in tasks:
        if task.lead_id not in allowed_ids:
            continue
        key = str(task.lead_id)
        if task.status == LeadTaskStatus.OPEN:
            result[key] = "open"
        elif result.get(key) != "open":
            result[key] = "done"

    common_tasks = db.query(Task).filter(Task.category == "leads", Task.tags.isnot(None)).all()
    per_lead_flags: Dict[int, Dict[str, bool]] = {}
    for common_task in common_tasks:
        tags = common_task.tags or []
        lead_tag = next((tag for tag in tags if isinstance(tag, str) and tag.startswith("lead:")), None)
        if not lead_tag:
            continue
        try:
            lead_id = int(lead_tag.split(":", 1)[1])
        except (ValueError, IndexError):
            continue
        if lead_id not in allowed_ids:
            continue
        flags = per_lead_flags.setdefault(lead_id, {"has_active": False, "has_any": False})
        flags["has_any"] = True
        if common_task.status == TaskStatus.ACTIVE.value:
            flags["has_active"] = True

    for lead_id, flags in per_lead_flags.items():
        key = str(lead_id)
        result[key] = "open" if flags["has_active"] else "done"
    return result


@router.get("/leads/badges")
async def get_leads_badges(
    lead_ids: str = Query(..., description="Comma-separated lead IDs"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    if not lead_ids.strip():
        return {}
    ids = [int(item.strip()) for item in lead_ids.split(",") if item.strip()]
    if not ids:
        return {}
    base = _filter_query_by_role(db.query(Lead), current_user)
    allowed_ids = {lead.id for lead in base.filter(Lead.id.in_(ids)).all()}

    today_start = datetime.combine(date.today(), dt_time.min)
    today_end = datetime.combine(date.today(), dt_time.max)
    invoice_lead_ids = {
        row[0]
        for row in db.query(Invoice.lead_id)
        .filter(Invoice.lead_id.in_(allowed_ids), Invoice.status != InvoiceStatus.PAID)
        .distinct()
        .all()
    }
    task_today_lead_ids = {
        row[0]
        for row in db.query(LeadTask.lead_id)
        .filter(
            LeadTask.lead_id.in_(allowed_ids),
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at >= today_start,
            LeadTask.due_at <= today_end,
        )
        .distinct()
        .all()
    }
    overdue_lead_ids = {
        row[0]
        for row in db.query(LeadTask.lead_id)
        .filter(
            LeadTask.lead_id.in_(allowed_ids),
            LeadTask.status == LeadTaskStatus.OPEN,
            LeadTask.due_at < today_start,
            LeadTask.due_at.isnot(None),
        )
        .distinct()
        .all()
    }

    result = {}
    for lead_id in allowed_ids:
        result[str(lead_id)] = {
            "has_invoice": lead_id in invoice_lead_ids,
            "has_task_today": lead_id in task_today_lead_ids,
            "is_overdue": lead_id in overdue_lead_ids,
        }
    return result


@router.get("/leads/no-show-ids", response_model=List[int])
async def list_no_show_lead_ids(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    rows = (
        db.query(EventRegistration.lead_id)
        .filter(
            EventRegistration.status == EventRegistrationStatus.REGISTERED,
            or_(
                cast(EventRegistration.note, Text).ilike("%[no-show]%"),
                cast(EventRegistration.note, Text).ilike("%no-show%"),
            ),
        )
        .distinct()
        .all()
    )
    return [row[0] for row in rows]


@router.get("/leads/{lead_id}", response_model=LeadResponse)
async def get_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    lead = _fix_lead_strings(lead)
    lead.ai_insight = build_lead_ai_insight(lead)
    return lead


@router.put("/leads/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: int,
    payload: LeadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)

    update_data = payload.dict(exclude_unset=True)
    old_status = lead.status.value
    if "status" in update_data and update_data["status"] is not None:
        lead.status = update_data["status"]
    if "lost_reason" in update_data:
        lead.lost_reason = update_data["lost_reason"]

    if "source_id" in update_data or "source" in update_data:
        source_id, source_name = _resolve_source(db, update_data.get("source_id"), update_data.get("source"))
        if _is_referral_source(source_name) and not (update_data.get("referral_name") or lead.referral_name or "").strip():
            raise HTTPException(status_code=400, detail="Для источника 'рекомендация' укажите, кто пригласил")
        lead.source_id = source_id
        lead.source = source_name

    for field in [
        "contact_name",
        "phone",
        "parent_full_name",
        "child_full_name",
        "parent_phone",
        "child_phone",
        "email",
        "city",
        "school_name",
        "school_class",
        "outreach_at",
        "outreach_minutes",
        "referral_name",
        "tags",
        "abonement_id",
        "desired_slot",
        "comment",
        "next_contact_at",
        "pause_reason",
        "questionnaire_filled",
        "no_answer_attempt",
        "max_user_id",
    ]:
        if field in update_data:
            setattr(lead, field, update_data[field])
    if any(field in update_data for field in ("phone", "parent_phone", "child_phone")):
        lead.phone_normalized = normalize_phone(
            update_data.get("parent_phone", lead.parent_phone)
            or update_data.get("phone", lead.phone)
            or update_data.get("child_phone", lead.child_phone)
            or ""
        ) or None
    sync_lead_person(db, lead)

    new_status = lead.status.value
    if "status" in update_data and old_status != new_status:
        _add_activity(
            db,
            lead_id,
            current_user.id,
            type="status_changed",
            title="Статус изменён",
            status_effect_from=old_status,
            status_effect_to=new_status,
        )

    db.commit()
    db.refresh(lead)
    log_action(db, current_user.id, "update", "lead", lead.id, update_data)
    return lead


@router.delete("/leads/{lead_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_permission("sales.access")),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _require_owner_or_admin(lead, current_user)
    log_action(db, current_user.id, "delete", "lead", lead.id, None)
    db.delete(lead)
    db.commit()
    return None


@router.post("/leads/{lead_id}/convert-to-student", response_model=LeadConvertToStudentResponse)
async def convert_lead_to_student(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.get_current_active_user),
):
    auth.ensure_permission(current_user, "sales.access")
    try:
        result = lead_conversion_convert(db, lead_id, actor_user_id=current_user.id)
    except ValueError as exc:
        message = str(exc)
        if "не найден" in message.lower():
            raise HTTPException(status_code=404, detail=message)
        raise HTTPException(status_code=400, detail=message)
    return LeadConvertToStudentResponse(
        student_id=result.student_id,
        lead=_fix_lead_strings(result.lead),
    )
