import csv
from datetime import date, datetime, timedelta, time as dt_time
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from openpyxl import load_workbook
from sqlalchemy import distinct, func as sa_func, nulls_last
from sqlalchemy.orm import Session, joinedload

from app import auth
from app.database import get_db
from app.models import (
    B2BSchool,
    B2BSchoolContact,
    B2BSchoolEvent,
    B2BSchoolInteraction,
    B2BProject,
    B2BSchoolPipelineStage,
    Lead,
    LeadStatus,
    Task,
    TaskStatus,
    User,
    UserRole,
)
from app.schemas import (
    B2BSchoolContactCreate,
    B2BSchoolContactResponse,
    B2BSchoolContactUpdate,
    B2BSchoolCreate,
    B2BSchoolEventCreate,
    B2BSchoolEventResponse,
    B2BSchoolImportResponse,
    B2BSchoolInteractionCreate,
    B2BSchoolInteractionResponse,
    B2BSchoolResponse,
    B2BSchoolUpdate,
    B2BLeadListItem,
    B2BProjectCreate,
    B2BProjectResponse,
    B2BProjectUpdate,
)

router = APIRouter()


def _leads_count_and_conversion(db: Session, school_id: int) -> Tuple[int, float]:
    total = db.query(Lead).filter(Lead.b2b_school_id == school_id).count()
    if total == 0:
        return 0, 0.0
    won = db.query(Lead).filter(
        Lead.b2b_school_id == school_id,
        Lead.status == LeadStatus.WON,
    ).count()
    return total, round(100.0 * won / total, 1)


def _school_to_response(db: Session, school: B2BSchool) -> B2BSchoolResponse:
    leads_count, conversion_percent = _leads_count_and_conversion(db, school.id)
    contacts = [
        B2BSchoolContactResponse(
            id=c.id,
            b2b_school_id=c.b2b_school_id,
            full_name=c.full_name,
            position=c.position,
            phone=c.phone,
            phone_extra=c.phone_extra,
            created_at=c.created_at,
        )
        for c in school.school_contacts
    ]
    manager_name = school.manager.full_name if school.manager else None
    return B2BSchoolResponse(
        id=school.id,
        name=school.name,
        director=school.director,
        city=school.city,
        address=school.address,
        student_count=school.student_count,
        friendship_degree=school.friendship_degree,
        pipeline_stage=school.pipeline_stage,
        next_step=school.next_step,
        next_step_date=school.next_step_date,
        manager_id=school.manager_id,
        manager_full_name=manager_name,
        source=getattr(school, "source", None),
        priority=getattr(school, "priority", None),
        preference=getattr(school, "preference", None),
        support_letter_status=getattr(school, "support_letter_status", None),
        partnership=getattr(school, "partnership", None),
        event_dates=school.event_dates,
        meeting_scheduled_at=school.meeting_scheduled_at,
        meeting_outcomes=school.meeting_outcomes,
        walkthrough_scheduled_at=school.walkthrough_scheduled_at,
        created_at=school.created_at,
        updated_at=school.updated_at,
        leads_count=leads_count,
        conversion_percent=conversion_percent,
        contacts=contacts,
    )


@router.get("/b2b-schools/cities", response_model=List[str])
async def list_b2b_school_cities(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """╨б╨┐╨╕╤Б╨╛╨║ ╨│╨╛╤А╨╛╨┤╨╛╨▓, ╨▓ ╨║╨╛╤В╨╛╤А╤Л╤Е ╨╡╤Б╤В╤М B2B ╤И╨║╨╛╨╗╤Л (╨┤╨╗╤П ╨▓╤Л╨▒╨╛╤А╨░ ╨┐╤А╨╕ ╨┤╨╛╨▒╨░╨▓╨╗╨╡╨╜╨╕╨╕ ╤И╨║╨╛╨╗ ╨▓ ╨▓╨╛╤А╨╛╨╜╨║╤Г)."""
    rows = db.query(distinct(B2BSchool.city)).filter(B2BSchool.city.isnot(None)).filter(B2BSchool.city != "").order_by(B2BSchool.city).all()
    return [r[0] for r in rows if r[0] and r[0].strip()]


@router.get("/b2b-schools/managers", response_model=List[Dict[str, Any]])
async def list_b2b_managers(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """List sales users for assigning as school manager."""
    users = db.query(User).filter(User.role == UserRole.SALES, User.is_active == True).order_by(User.full_name).all()
    return [{"id": u.id, "full_name": u.full_name} for u in users]


@router.get("/b2b-schools/plan-for-today", response_model=Dict[str, List[B2BSchoolResponse]])
async def plan_for_today(
    city: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Schools for owner's daily plan: overdue, no next step, find_contacts > 3 days, today, tomorrow, week, risks."""
    today = date.today()
    tomorrow = today + timedelta(days=1)
    week_end = today + timedelta(days=7)
    cutoff = today - timedelta(days=3)
    cutoff_7 = today - timedelta(days=7)
    cutoff_14 = today - timedelta(days=14)
    cutoff_48h = datetime.utcnow() - timedelta(hours=48)
    date_1d_ago = today - timedelta(days=1)
    date_2d_ago = today - timedelta(days=2)
    base = db.query(B2BSchool).options(
        joinedload(B2BSchool.school_contacts),
        joinedload(B2BSchool.manager),
    )
    if city and city.strip():
        base = base.filter(B2BSchool.city == city.strip())

    def run(q):
        return [_school_to_response(db, s) for s in q.order_by(nulls_last(B2BSchool.next_step_date.asc()), B2BSchool.name).all()]

    from sqlalchemy import or_, exists

    overdue = base.filter(B2BSchool.next_step_date.isnot(None), B2BSchool.next_step_date < today)
    no_next = base.filter((B2BSchool.next_step.is_(None)) | (B2BSchool.next_step_date.is_(None)))
    cutoff_ts = datetime.combine(cutoff, dt_time.min)
    cutoff_7_ts = datetime.combine(cutoff_7, dt_time.min)
    cutoff_14_ts = datetime.combine(cutoff_14, dt_time.min)
    find_stale = base.filter(
        B2BSchool.pipeline_stage == "find_contacts",
        or_(
            B2BSchool.updated_at < cutoff_ts,
            (B2BSchool.updated_at.is_(None)) & (B2BSchool.created_at < cutoff_ts),
        ),
    )
    # Риск: в find_contacts, без контактов, обновление/создание более 48ч назад
    no_contacts_exists = ~exists().where(B2BSchoolContact.b2b_school_id == B2BSchool.id)
    find_contacts_no_contacts_48h = base.filter(
        B2BSchool.pipeline_stage == "find_contacts",
        no_contacts_exists,
        or_(
            B2BSchool.updated_at < cutoff_48h,
            (B2BSchool.updated_at.is_(None)) & (B2BSchool.created_at < cutoff_48h),
        ),
    )
    today_q = base.filter(B2BSchool.next_step_date == today)
    tomorrow_q = base.filter(B2BSchool.next_step_date == tomorrow)
    week_q = base.filter(
        B2BSchool.next_step_date.isnot(None),
        B2BSchool.next_step_date > tomorrow,
        B2BSchool.next_step_date <= week_end,
    )
    no_contacts = base.filter(no_contacts_exists)
    no_touches_7d = base.filter(
        B2BSchool.pipeline_stage.in_(["agreement", "permission_received", "event_scheduled"]),
        or_(
            B2BSchool.updated_at < cutoff_7_ts,
            (B2BSchool.updated_at.is_(None)) & (B2BSchool.created_at < cutoff_7_ts),
        ),
    )
    event_done_no_leads = base.filter(
        B2BSchool.pipeline_stage.in_(["event_done", "walkthrough_done"]),
        ~exists().where(Lead.b2b_school_id == B2BSchool.id),
    )
    negotiations_14d = base.filter(
        B2BSchool.pipeline_stage == "agreement",
        or_(
            B2BSchool.updated_at < cutoff_14_ts,
            (B2BSchool.updated_at.is_(None)) & (B2BSchool.created_at < cutoff_14_ts),
        ),
    )

    # Риск: проведено без лидов 24–48ч — дата проведения 1–2 дня назад
    def _event_held_date(school: B2BSchool) -> Optional[date]:
        if school.pipeline_stage == "walkthrough_done" and getattr(school, "walkthrough_scheduled_at", None):
            d = school.walkthrough_scheduled_at
            return d.date() if hasattr(d, "date") else (d if isinstance(d, date) else None)
        if school.pipeline_stage == "event_done" and getattr(school, "event_dates", None) and isinstance(school.event_dates, list):
            dates_parsed = []
            for x in school.event_dates:
                if isinstance(x, str):
                    try:
                        dates_parsed.append(datetime.strptime(x[:10], "%Y-%m-%d").date())
                    except Exception:
                        pass
                elif isinstance(x, date):
                    dates_parsed.append(x)
            return max(dates_parsed) if dates_parsed else None
        return None

    event_done_no_leads_schools = event_done_no_leads.order_by(nulls_last(B2BSchool.next_step_date.asc()), B2BSchool.name).all()
    event_done_no_leads_list = [_school_to_response(db, s) for s in event_done_no_leads_schools]
    event_done_no_leads_24_48h = []
    for s in event_done_no_leads_schools:
        held = _event_held_date(s)
        if held and date_2d_ago <= held <= date_1d_ago:
            event_done_no_leads_24_48h.append(_school_to_response(db, s))

    return {
        "overdue": run(overdue),
        "no_next_step": run(no_next),
        "find_contacts_stale": run(find_stale),
        "find_contacts_no_contacts_48h": run(find_contacts_no_contacts_48h),
        "today": run(today_q),
        "tomorrow": run(tomorrow_q),
        "week": run(week_q),
        "no_contacts": run(no_contacts),
        "no_touches_7d": run(no_touches_7d),
        "event_done_no_leads": event_done_no_leads_list,
        "event_done_no_leads_24_48h": event_done_no_leads_24_48h,
        "negotiations_14d": run(negotiations_14d),
    }


class CitySummaryItem(BaseModel):
    """Сводка по одному городу для плана на сегодня."""
    city: str
    schools_in_work: int
    overdue: int
    events_this_week: int
    leads_7d: int
    partners: int

    class Config:
        frozen = True


@router.get("/b2b-schools/plan-city-summary", response_model=List[CitySummaryItem])
async def plan_city_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Сводка по городам для плана: в работе, просрочки, мероприятия на неделе, лиды за 7 дн, партнёры."""
    today = date.today()
    week_end = today + timedelta(days=7)
    cutoff_7 = today - timedelta(days=7)
    cutoff_7_ts = datetime.combine(cutoff_7, dt_time.min)

    cities_rows = db.query(distinct(B2BSchool.city)).filter(
        B2BSchool.city.isnot(None),
        B2BSchool.city != "",
    ).all()
    cities = [r[0] for r in cities_rows if r[0]]

    schools = db.query(B2BSchool).filter(
        B2BSchool.city.isnot(None),
        B2BSchool.city != "",
    ).all()

    def _event_in_week(edates: Any) -> bool:
        if not edates or not isinstance(edates, list):
            return False
        for d in edates:
            if isinstance(d, str):
                try:
                    d_parsed = datetime.strptime(d[:10], "%Y-%m-%d").date()
                    if today <= d_parsed <= week_end:
                        return True
                except ValueError:
                    pass
        return False

    def _is_partner(s: B2BSchool) -> bool:
        if s.pipeline_stage == B2BSchoolPipelineStage.PARTNERS.value:
            return True
        if s.partnership and isinstance(s.partnership, dict) and s.partnership.get("active_partner"):
            return True
        return False

    work_stages = {
        B2BSchoolPipelineStage.NEW.value,
        B2BSchoolPipelineStage.REJECTED.value,
    }

    lead_counts_by_city: Dict[str, int] = {c: 0 for c in cities}
    leads_q = db.query(B2BSchool.city, sa_func.count(Lead.id)).join(
        B2BSchool, Lead.b2b_school_id == B2BSchool.id
    ).filter(
        Lead.created_at >= cutoff_7_ts,
        B2BSchool.city.isnot(None),
        B2BSchool.city != "",
    ).group_by(B2BSchool.city).all()
    for city, cnt in leads_q:
        if city:
            lead_counts_by_city[city] = cnt

    result: List[CitySummaryItem] = []
    for city in sorted(cities):
        city_schools = [s for s in schools if s.city == city]
        schools_in_work = sum(1 for s in city_schools if s.pipeline_stage not in work_stages)
        overdue = sum(1 for s in city_schools if s.next_step_date and s.next_step_date < today)
        events_this_week = sum(1 for s in city_schools if _event_in_week(s.event_dates))
        leads_7d = lead_counts_by_city.get(city, 0)
        partners = sum(1 for s in city_schools if _is_partner(s))
        result.append(CitySummaryItem(
            city=city,
            schools_in_work=schools_in_work,
            overdue=overdue,
            events_this_week=events_this_week,
            leads_7d=leads_7d,
            partners=partners,
        ))
    return result


@router.get("/b2b-schools", response_model=List[B2BSchoolResponse])
async def list_b2b_schools(
    pipeline_stage: Optional[str] = Query(default=None),
    project_id: Optional[int] = Query(default=None),
    city: Optional[str] = Query(default=None, description="Фильтр по городу"),
    manager_id: Optional[int] = Query(default=None, description="Фильтр по ответственному"),
    overdue: Optional[bool] = Query(default=None, description="Только с просроченным next_step_date"),
    search: Optional[str] = Query(default=None, description="Поиск по названию/городу"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    from sqlalchemy import or_
    query = db.query(B2BSchool).options(
        joinedload(B2BSchool.school_contacts),
        joinedload(B2BSchool.manager),
    ).order_by(B2BSchool.created_at.desc())
    if project_id is not None:
        project = db.query(B2BProject).filter(B2BProject.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        cities: List[str] = []
        if project.main_city:
            cities.append(project.main_city)
        if isinstance(project.cities, list):
            cities.extend([c for c in project.cities if isinstance(c, str)])
        cities = [c.strip() for c in cities if c and isinstance(c, str)]
        if cities:
            query = query.filter(B2BSchool.city.in_(cities))
    if city is not None and city.strip():
        query = query.filter(B2BSchool.city == city.strip())
    if pipeline_stage is not None:
        query = query.filter(B2BSchool.pipeline_stage == pipeline_stage)
    if manager_id is not None:
        query = query.filter(B2BSchool.manager_id == manager_id)
    if overdue is True:
        today = date.today()
        query = query.filter(B2BSchool.next_step_date.isnot(None), B2BSchool.next_step_date < today)
    if search is not None and search.strip():
        term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                B2BSchool.name.ilike(term),
                (B2BSchool.city.isnot(None) & B2BSchool.city.ilike(term)),
            )
        )
    schools = query.all()
    return [_school_to_response(db, s) for s in schools]


def _parse_import_rows_from_xlsx(data: bytes) -> List[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[int]]]:
    """Returns list of (name, director, city, address, student_count). First row = header."""
    wb = load_workbook(filename=BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    name_cols = ["название", "name", "школа", "school"]
    dir_cols = ["директор", "director"]
    city_cols = ["город", "city"]
    addr_cols = ["адрес", "address"]
    count_cols = ["учеников", "student_count", "количество", "кол-во"]

    def col(key_list: List[str]) -> Optional[int]:
        for k in key_list:
            if k in headers:
                return headers.index(k)
        return None

    name_idx = col(name_cols) if any(h in headers for h in name_cols) else (0 if headers else None)
    dir_idx = col(dir_cols)
    city_idx = col(city_cols)
    addr_idx = col(addr_cols)
    count_idx = col(count_cols)

    result: List[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[int]]] = []
    for row in rows[1:]:
        row_list = list(row) if row else []
        while len(row_list) < len(headers):
            row_list.append(None)
        def v(i: Optional[int]) -> Optional[str]:
            if i is None or i >= len(row_list):
                return None
            x = row_list[i]
            if x is None:
                return None
            s = str(x).strip()
            return s if s else None
        def n(i: Optional[int]) -> Optional[int]:
            if i is None or i >= len(row_list):
                return None
            x = row_list[i]
            if x is None:
                return None
            try:
                return int(float(str(x).replace(",", ".")))
            except (ValueError, TypeError):
                return None
        name = v(name_idx) if name_idx is not None else v(0)
        result.append((name, v(dir_idx), v(city_idx), v(addr_idx), n(count_idx)))
    return result


def _parse_import_rows_from_csv(data: bytes) -> List[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[int]]]:
    """Returns list of (name, director, city, address, student_count). First row = header."""
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = data.decode("cp1251")
        except Exception:
            text = data.decode("utf-8", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [str(h).strip().lower() for h in rows[0]]
    name_cols = ["название", "name", "школа", "school"]
    dir_cols = ["директор", "director"]
    city_cols = ["город", "city"]
    addr_cols = ["адрес", "address"]
    count_cols = ["учеников", "student_count", "количество", "кол-во"]

    def col(key_list: List[str]) -> Optional[int]:
        for k in key_list:
            if k in headers:
                return headers.index(k)
        return None

    name_idx = col(name_cols) if any(h in headers for h in name_cols) else 0
    dir_idx = col(dir_cols)
    city_idx = col(city_cols)
    addr_idx = col(addr_cols)
    count_idx = col(count_cols)

    result: List[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[int]]] = []
    for row in rows[1:]:
        while len(row) < len(headers):
            row.append("")
        def v(i: Optional[int]) -> Optional[str]:
            if i is None or i >= len(row):
                return None
            s = str(row[i]).strip()
            return s if s else None
        def n(i: Optional[int]) -> Optional[int]:
            if i is None or i >= len(row):
                return None
            try:
                return int(float(str(row[i]).replace(",", ".")))
            except (ValueError, TypeError):
                return None
        name = v(name_idx) if name_idx is not None else v(0)
        result.append((name, v(dir_idx), v(city_idx), v(addr_idx), n(count_idx)))
    return result


@router.post("/b2b-schools/import", response_model=B2BSchoolImportResponse)
async def import_b2b_schools(
    file: UploadFile = File(...),
    city: Optional[str] = Query(default=None, description="Assign city to all imported schools if not in file"),
    manager_id: Optional[int] = Query(default=None, description="Assign manager to all imported schools"),
    launch_in_work: bool = Query(default=False, description="Set pipeline find_contacts and next step for created schools"),
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Import B2B schools from Excel (.xlsx) or CSV. Columns: name (required), director, city, address, student_count."""
    filename = (file.filename or "").lower().strip()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Поддерживаются только форматы .xlsx и .csv")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")

    if filename.endswith(".xlsx"):
        rows = _parse_import_rows_from_xlsx(data)
    else:
        rows = _parse_import_rows_from_csv(data)

    if launch_in_work and manager_id is None:
        return B2BSchoolImportResponse(
            created=0, skipped=0,
            errors=["Для опции «запустить в работу» укажите ответственного (менеджер) в параметрах импорта."],
        )

    tomorrow = date.today() + timedelta(days=1)
    next_step_launch = "Найти контакты (директор/завуч/информатика)"
    default_city = city.strip() if city and city.strip() else None
    created = 0
    skipped = 0
    errors: List[str] = []

    for i, (name, director, city_val, address, student_count) in enumerate(rows, start=2):
        school_name = (name or "").strip()
        if not school_name:
            skipped += 1
            continue
        school_city = (city_val or "").strip() or default_city
        school_manager_id = manager_id
        pipeline_stage = B2BSchoolPipelineStage.NEW.value
        next_step = None
        next_step_date = None
        if launch_in_work:
            pipeline_stage = B2BSchoolPipelineStage.FIND_CONTACTS.value
            next_step = next_step_launch
            next_step_date = tomorrow
        school = B2BSchool(
            name=school_name,
            director=(director or "").strip() or None,
            city=school_city,
            address=(address or "").strip() or None,
            student_count=student_count,
            pipeline_stage=pipeline_stage,
            next_step=next_step,
            next_step_date=next_step_date,
            manager_id=school_manager_id,
        )
        db.add(school)
        db.flush()
        if launch_in_work and school_manager_id:
            _on_b2b_school_created(db, school, current_user.id)
        created += 1

    db.commit()
    return B2BSchoolImportResponse(created=created, skipped=skipped, errors=errors)


def _create_b2b_task(
    db: Session,
    created_by_id: int,
    title: str,
    assigned_to_id: Optional[int] = None,
) -> Optional[Task]:
    """Create a Task for B2B automation. Title truncated to 512 chars."""
    if not title or not title.strip():
        return None
    task = Task(
        title=(title.strip()[:512]),
        created_by_id=created_by_id,
        assigned_to_id=assigned_to_id,
        status=TaskStatus.ACTIVE.value,
    )
    db.add(task)
    return task


def _on_b2b_school_created(
    db: Session,
    school: B2BSchool,
    created_by_id: int,
) -> None:
    """Trigger: school created. If find_contacts + manager → task 'Найти контакты'."""
    if school.pipeline_stage != B2BSchoolPipelineStage.FIND_CONTACTS.value:
        return
    if not school.manager_id:
        return
    _create_b2b_task(
        db,
        created_by_id=created_by_id,
        title=f"Найти контакты (директор/завуч/информатика): {school.name}",
        assigned_to_id=school.manager_id,
    )


def _on_b2b_school_stage_changed(
    db: Session,
    school: B2BSchool,
    new_stage: str,
    created_by_id: int,
) -> None:
    """Trigger: pipeline_stage changed. Create task and optionally set next_step/next_step_date."""
    day2 = date.today() + timedelta(days=2)
    manager_id = school.manager_id
    name = school.name or "Школа"

    if new_stage in (B2BSchoolPipelineStage.FIRST_CONTACT.value, B2BSchoolPipelineStage.CONTACT_FOUND.value):
        if not school.next_step or not school.next_step_date:
            school.next_step = "Дожим через 2 дня"
            school.next_step_date = day2
        _create_b2b_task(
            db, created_by_id,
            title=f"Дожим через 2 дня: {name}",
            assigned_to_id=manager_id,
        )
    elif new_stage == B2BSchoolPipelineStage.MEETING_SCHEDULED.value:
        _create_b2b_task(
            db, created_by_id,
            title=f"Подтвердить встречу за 1 день / Провести: {name}",
            assigned_to_id=manager_id,
        )
    elif new_stage == B2BSchoolPipelineStage.PERMISSION_RECEIVED.value:
        _create_b2b_task(
            db, created_by_id,
            title=f"Подготовка мероприятия: {name}",
            assigned_to_id=manager_id,
        )
    elif new_stage in (B2BSchoolPipelineStage.EVENT_DONE.value, B2BSchoolPipelineStage.WALKTHROUGH_DONE.value):
        _create_b2b_task(
            db, created_by_id,
            title=f"Занести лиды: {name}",
            assigned_to_id=manager_id,
        )
    elif new_stage == B2BSchoolPipelineStage.LETTER_SENT.value:
        if not school.next_step or not school.next_step_date:
            school.next_step = "Дожим через 2 дня"
            school.next_step_date = day2
        _create_b2b_task(
            db, created_by_id,
            title=f"Дожим через 2 дня: {name}",
            assigned_to_id=manager_id,
        )
        if school.pipeline_stage != B2BSchoolPipelineStage.AGREEMENT.value:
            school.pipeline_stage = B2BSchoolPipelineStage.AGREEMENT.value
    elif new_stage == B2BSchoolPipelineStage.EVENT_SCHEDULED.value:
        _create_b2b_task(
            db, created_by_id,
            title=f"Напоминание за 3 дня до мероприятия: {name}",
            assigned_to_id=manager_id,
        )
        _create_b2b_task(
            db, created_by_id,
            title=f"Напоминание за 1 день до мероприятия: {name}",
            assigned_to_id=manager_id,
        )
        _create_b2b_task(
            db, created_by_id,
            title=f"Подготовить материалы к мероприятию: {name}",
            assigned_to_id=manager_id,
        )
    elif new_stage == B2BSchoolPipelineStage.LEADS_RECEIVED.value:
        _create_b2b_task(
            db, created_by_id,
            title=f"Передать лиды в обработку: {name}",
            assigned_to_id=manager_id,
        )
        _create_b2b_task(
            db, created_by_id,
            title=f"Контроль конверсии через 7 дней: {name}",
            assigned_to_id=manager_id or created_by_id,
        )


# Порядок чек-листа партнёрства: следующий шаг → заголовок задачи
PARTNERSHIP_CHAIN = [
    ("invited", "Отправить соглашение школе"),
    ("agreement_sent", "Получить подпись со стороны школы"),
    ("signed_school", "Получить подпись с нашей стороны"),
    ("signed_both", "Получить оригиналы договора"),
    ("originals_received", "Разместить иконку на сайте школы"),
    ("icon_on_site", "Ввести в реестр активных партнёров"),
    ("active_partner", None),
]


def _on_b2b_partnership_updated(
    db: Session,
    school: B2BSchool,
    previous_partnership: Optional[Dict[str, Any]],
    current_user_id: int,
) -> None:
    """При изменении чек-листа партнёрства: если какой-то пункт только что отмечен — создать одну задачу на следующий шаг."""
    new_partnership = getattr(school, "partnership", None) or {}
    if not isinstance(new_partnership, dict):
        return
    prev = previous_partnership if isinstance(previous_partnership, dict) else {}
    manager_id = school.manager_id
    if not manager_id:
        return
    name = school.name or "Школа"
    # Найти самый «продвинутый» пункт, который только что стал True
    last_newly_checked: Optional[str] = None
    for i, (key, _) in enumerate(PARTNERSHIP_CHAIN):
        if key == "active_partner":
            break
        old_val = prev.get(key) if prev else False
        new_val = new_partnership.get(key)
        if new_val and not old_val:
            last_newly_checked = key
    if last_newly_checked is None:
        return
    # Следующий шаг в цепочке
    for i, (key, task_title) in enumerate(PARTNERSHIP_CHAIN):
        if key == last_newly_checked and i + 1 < len(PARTNERSHIP_CHAIN):
            next_title = PARTNERSHIP_CHAIN[i + 1][1]
            if next_title:
                _create_b2b_task(
                    db,
                    created_by_id=current_user_id,
                    title=f"{next_title}: {name}",
                    assigned_to_id=manager_id,
                )
            break


def _load_school_with_contacts(db: Session, school_id: int) -> B2BSchool:
    school = db.query(B2BSchool).options(
        joinedload(B2BSchool.school_contacts),
        joinedload(B2BSchool.manager),
    ).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    _ = school.school_contacts
    return school


@router.get("/b2b-schools/{school_id}", response_model=B2BSchoolResponse)
async def get_b2b_school(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    school = _load_school_with_contacts(db, school_id)
    return _school_to_response(db, school)


def _event_label(event: Optional[Any]) -> Optional[str]:
    """Строка-источник по мероприятию для лида."""
    if not event:
        return None
    parts = []
    if getattr(event, "format", None) == "online" and getattr(event, "online_type", None):
        label = {"webinar": "Вебинар", "olympiad": "Олимпиада", "open_doors": "День открытых дверей"}.get(
            event.online_type, event.online_type
        )
        parts.append(label)
    else:
        fmt = getattr(event, "format", None) or ""
        parts.append("Онлайн" if fmt == "online" else "Гибрид" if fmt == "hybrid" else "Офлайн")
    dates = getattr(event, "event_dates", None)
    if dates and isinstance(dates, list):
        parts.append(", ".join(dates[:3]))
    return " ".join(parts) if parts else "Мероприятие"


@router.get("/b2b-schools/{school_id}/leads", response_model=List[B2BLeadListItem])
async def list_school_leads(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Список лидов по B2B-школе для карточки."""
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    leads = (
        db.query(Lead)
        .options(joinedload(Lead.b2b_event))
        .filter(Lead.b2b_school_id == school_id)
        .order_by(Lead.created_at.desc())
        .all()
    )
    return [
        B2BLeadListItem(
            id=l.id,
            contact_name=l.contact_name,
            phone=l.phone or "",
            status=l.status.value if hasattr(l.status, "value") else str(l.status),
            source=l.source,
            source_event=_event_label(l.b2b_event) if getattr(l, "b2b_event", None) else None,
            created_at=l.created_at,
        )
        for l in leads
    ]


@router.post("/b2b-schools/{school_id}/leads/transfer")
async def transfer_school_leads(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Передать лиды школы в обработку: статус NEW → CONTACTED."""
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    updated = (
        db.query(Lead)
        .filter(Lead.b2b_school_id == school_id, Lead.status == LeadStatus.NEW)
        .update({Lead.status: LeadStatus.CONTACTED}, synchronize_session="fetch")
    )
    db.commit()
    return {"updated": updated}


@router.get("/b2b-schools/{school_id}/interactions", response_model=List[B2BSchoolInteractionResponse])
async def list_school_interactions(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Журнал взаимодействий по B2B-школе (по убыванию даты)."""
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    items = (
        db.query(B2BSchoolInteraction)
        .filter(B2BSchoolInteraction.b2b_school_id == school_id)
        .order_by(B2BSchoolInteraction.happened_at.desc())
        .all()
    )
    return [
        B2BSchoolInteractionResponse(
            id=i.id,
            b2b_school_id=i.b2b_school_id,
            type=i.type,
            happened_at=i.happened_at,
            summary=i.summary,
            created_by_id=i.created_by_id,
            created_by_name=i.created_by.full_name if i.created_by else None,
            created_at=i.created_at,
        )
        for i in items
    ]


@router.post("/b2b-schools/{school_id}/interactions", response_model=B2BSchoolInteractionResponse, status_code=201)
async def create_school_interaction(
    school_id: int,
    payload: B2BSchoolInteractionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Добавить взаимодействие; опционально обновить next_step/next_step_date школы."""
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    if payload.type not in ("call", "letter", "meeting"):
        raise HTTPException(status_code=400, detail="type must be call, letter or meeting")
    interaction = B2BSchoolInteraction(
        b2b_school_id=school_id,
        type=payload.type,
        happened_at=payload.happened_at,
        summary=payload.summary,
        created_by_id=current_user.id,
    )
    db.add(interaction)
    if payload.next_step is not None or payload.next_step_date is not None:
        if payload.next_step is not None:
            school.next_step = payload.next_step
        if payload.next_step_date is not None:
            school.next_step_date = payload.next_step_date
    db.commit()
    db.refresh(interaction)
    return B2BSchoolInteractionResponse(
        id=interaction.id,
        b2b_school_id=interaction.b2b_school_id,
        type=interaction.type,
        happened_at=interaction.happened_at,
        summary=interaction.summary,
        created_by_id=interaction.created_by_id,
        created_by_name=current_user.full_name,
        created_at=interaction.created_at,
    )


def _on_b2b_online_event_created(
    db: Session,
    school: B2BSchool,
    event: B2BSchoolEvent,
    current_user_id: int,
) -> None:
    """При создании онлайн-мероприятия — создать задачи менеджеру."""
    manager_id = school.manager_id
    if not manager_id:
        return
    name = school.name or "Школа"
    online_label = {"webinar": "Вебинар", "olympiad": "Олимпиада", "open_doors": "День открытых дверей"}.get(
        event.online_type or "", event.online_type or "онлайн"
    )
    suffix = f" ({online_label}): {name}"
    _create_b2b_task(db, current_user_id, f"Отправить ссылку{suffix}", manager_id)
    _create_b2b_task(db, current_user_id, f"Рассылка по чатам{suffix}", manager_id)
    _create_b2b_task(db, current_user_id, f"Напоминание за 2 дня до мероприятия{suffix}", manager_id)
    _create_b2b_task(db, current_user_id, f"Напоминание за 1 день до мероприятия{suffix}", manager_id)
    _create_b2b_task(db, current_user_id, f"Занести лиды{suffix}", manager_id)


@router.get("/b2b-schools/{school_id}/events", response_model=List[B2BSchoolEventResponse])
async def list_school_events(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Список мероприятий B2B-школы."""
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    events = (
        db.query(B2BSchoolEvent)
        .filter(B2BSchoolEvent.b2b_school_id == school_id)
        .order_by(B2BSchoolEvent.created_at.desc())
        .all()
    )
    return [
        B2BSchoolEventResponse(
            id=e.id,
            b2b_school_id=e.b2b_school_id,
            format=e.format,
            online_type=e.online_type,
            event_dates=e.event_dates,
            created_at=e.created_at,
        )
        for e in events
    ]


@router.post("/b2b-schools/{school_id}/events", response_model=B2BSchoolEventResponse, status_code=201)
async def create_school_event(
    school_id: int,
    payload: B2BSchoolEventCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    """Добавить мероприятие; при format=online создать авто-задачи."""
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    if payload.format not in ("offline", "online", "hybrid"):
        raise HTTPException(status_code=400, detail="format must be offline, online or hybrid")
    if payload.format == "online" and payload.online_type and payload.online_type not in ("webinar", "olympiad", "open_doors"):
        raise HTTPException(status_code=400, detail="online_type must be webinar, olympiad or open_doors")
    event = B2BSchoolEvent(
        b2b_school_id=school_id,
        format=payload.format,
        online_type=payload.online_type,
        event_dates=payload.dates,
    )
    db.add(event)
    db.flush()
    if payload.format == "online":
        _on_b2b_online_event_created(db, school, event, current_user.id)
    db.commit()
    db.refresh(event)
    return B2BSchoolEventResponse(
        id=event.id,
        b2b_school_id=event.b2b_school_id,
        format=event.format,
        online_type=event.online_type,
        event_dates=event.event_dates,
        created_at=event.created_at,
    )


@router.post("/b2b-schools", response_model=B2BSchoolResponse, status_code=201)
async def create_b2b_school(
    payload: B2BSchoolCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    school = B2BSchool(
        name=payload.name,
        director=payload.director,
        address=payload.address,
        city=payload.city,
        student_count=payload.student_count,
        friendship_degree=payload.friendship_degree.value if payload.friendship_degree else None,
        pipeline_stage=payload.pipeline_stage.value,
        next_step=payload.next_step,
        next_step_date=payload.next_step_date,
        manager_id=payload.manager_id,
        source=payload.source,
        priority=payload.priority,
        preference=payload.preference,
        event_dates=payload.event_dates,
        meeting_scheduled_at=payload.meeting_scheduled_at,
        meeting_outcomes=payload.meeting_outcomes,
        walkthrough_scheduled_at=payload.walkthrough_scheduled_at,
    )
    db.add(school)
    db.flush()
    _on_b2b_school_created(db, school, current_user.id)
    db.commit()
    db.refresh(school)
    _ = school.school_contacts
    return _school_to_response(db, school)


@router.put("/b2b-schools/{school_id}", response_model=B2BSchoolResponse)
async def update_b2b_school(
    school_id: int,
    payload: B2BSchoolUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    school = _load_school_with_contacts(db, school_id)
    old_stage = school.pipeline_stage
    old_support_letter = getattr(school, "support_letter_status", None)
    old_partnership = None
    if getattr(school, "partnership", None) is not None:
        p = getattr(school, "partnership", None)
        old_partnership = dict(p) if isinstance(p, dict) else None
    data = payload.dict(exclude_unset=True)
    if "pipeline_stage" in data and data["pipeline_stage"] is not None:
        data["pipeline_stage"] = data["pipeline_stage"].value
    if "friendship_degree" in data and data["friendship_degree"] is not None:
        data["friendship_degree"] = data["friendship_degree"].value
    for key, value in data.items():
        setattr(school, key, value)
    new_stage = school.pipeline_stage
    if new_stage != old_stage:
        _on_b2b_school_stage_changed(db, school, new_stage, current_user.id)
    if "partnership" in data and school.manager_id:
        _on_b2b_partnership_updated(db, school, old_partnership, current_user.id)
    new_support_letter = getattr(school, "support_letter_status", None)
    if new_support_letter == "requested" and old_support_letter != "requested" and school.manager_id:
        _create_b2b_task(
            db,
            created_by_id=current_user.id,
            title=f"Запросить письмо поддержки: {school.name or 'Школа'}",
            assigned_to_id=school.manager_id,
        )
    db.commit()
    db.refresh(school)
    _ = school.school_contacts
    return _school_to_response(db, school)


@router.delete("/b2b-schools/{school_id}", status_code=204)
async def delete_b2b_school(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    db.delete(school)
    db.commit()
    return None


@router.get("/b2b-schools/{school_id}/contacts", response_model=List[B2BSchoolContactResponse])
async def list_school_contacts(
    school_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    school = _load_school_with_contacts(db, school_id)
    return [
        B2BSchoolContactResponse(
            id=c.id,
            b2b_school_id=c.b2b_school_id,
            full_name=c.full_name,
            position=c.position,
            phone=c.phone,
            phone_extra=c.phone_extra,
            created_at=c.created_at,
        )
        for c in school.school_contacts
    ]


@router.post("/b2b-schools/{school_id}/contacts", response_model=B2BSchoolContactResponse, status_code=201)
async def create_school_contact(
    school_id: int,
    payload: B2BSchoolContactCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    school = db.query(B2BSchool).filter(B2BSchool.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")
    contact = B2BSchoolContact(
        b2b_school_id=school_id,
        full_name=payload.full_name,
        position=payload.position,
        phone=payload.phone,
        phone_extra=payload.phone_extra,
    )
    db.add(contact)
    db.commit()
    db.refresh(contact)
    return B2BSchoolContactResponse(
        id=contact.id,
        b2b_school_id=contact.b2b_school_id,
        full_name=contact.full_name,
        position=contact.position,
        phone=contact.phone,
        phone_extra=contact.phone_extra,
        created_at=contact.created_at,
    )


@router.put("/b2b-schools/{school_id}/contacts/{contact_id}", response_model=B2BSchoolContactResponse)
async def update_school_contact(
    school_id: int,
    contact_id: int,
    payload: B2BSchoolContactUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    contact = db.query(B2BSchoolContact).filter(
        B2BSchoolContact.id == contact_id,
        B2BSchoolContact.b2b_school_id == school_id,
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    data = payload.dict(exclude_unset=True)
    for key, value in data.items():
        setattr(contact, key, value)
    db.commit()
    db.refresh(contact)
    return B2BSchoolContactResponse(
        id=contact.id,
        b2b_school_id=contact.b2b_school_id,
        full_name=contact.full_name,
        position=contact.position,
        phone=contact.phone,
        phone_extra=contact.phone_extra,
        created_at=contact.created_at,
    )


@router.delete("/b2b-schools/{school_id}/contacts/{contact_id}", status_code=204)
async def delete_school_contact(
    school_id: int,
    contact_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    contact = db.query(B2BSchoolContact).filter(
        B2BSchoolContact.id == contact_id,
        B2BSchoolContact.b2b_school_id == school_id,
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    db.delete(contact)
    db.commit()
    return None


@router.get("/b2b-projects", response_model=List[B2BProjectResponse])
async def list_b2b_projects(
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    projects = db.query(B2BProject).order_by(B2BProject.created_at.desc()).all()
    return [
        B2BProjectResponse(
            id=p.id,
            name=p.name,
            location=p.location,
            main_city=p.main_city,
            cities=p.cities or [],
            created_at=p.created_at,
            updated_at=p.updated_at,
        )
        for p in projects
    ]


@router.post("/b2b-projects", response_model=B2BProjectResponse, status_code=201)
async def create_b2b_project(
    payload: B2BProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    project = B2BProject(
        name=payload.name,
        location=payload.location,
        main_city=payload.main_city,
        cities=payload.cities,
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    return B2BProjectResponse(
        id=project.id,
        name=project.name,
        location=project.location,
        main_city=project.main_city,
        cities=project.cities or [],
        created_at=project.created_at,
        updated_at=project.updated_at,
    )


@router.get("/b2b-projects/{project_id}", response_model=B2BProjectResponse)
async def get_b2b_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    project = db.query(B2BProject).filter(B2BProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return B2BProjectResponse(
        id=project.id,
        name=project.name,
        location=project.location,
        main_city=project.main_city,
        cities=project.cities or [],
        created_at=project.created_at,
        updated_at=project.updated_at,
    )


@router.put("/b2b-projects/{project_id}", response_model=B2BProjectResponse)
async def update_b2b_project(
    project_id: int,
    payload: B2BProjectUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    project = db.query(B2BProject).filter(B2BProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    data = payload.dict(exclude_unset=True)
    for key, value in data.items():
        setattr(project, key, value)
    db.commit()
    db.refresh(project)
    return B2BProjectResponse(
        id=project.id,
        name=project.name,
        location=project.location,
        main_city=project.main_city,
        cities=project.cities or [],
        created_at=project.created_at,
        updated_at=project.updated_at,
    )


@router.delete("/b2b-projects/{project_id}", status_code=204)
async def delete_b2b_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(auth.require_role(["owner"])),
):
    project = db.query(B2BProject).filter(B2BProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(project)
    db.commit()
    return None
